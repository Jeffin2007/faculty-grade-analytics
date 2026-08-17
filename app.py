"""
================================================================================
 FACULTY GRADE ANALYTICS PORTAL -- app.py (Production Ready)
================================================================================
 Single-file application: FastHTML + pandas + openpyxl + Plotly + xhtml2pdf
 + Ollama Cloud / Local API / Deterministic Insights Engine. Analytics and LLM
 advisory are strictly layered: the UI renders structured analytics and never
 computes academic statistics itself.

 Semester final result treated as the single source of truth. No internal /
 assessment / attendance / end-semester calculation is performed.
================================================================================
"""

from __future__ import annotations

import asyncio
import base64
import csv
import difflib
import gc
import hashlib
import html
import io
import json
import math
import os
import re
import sys
import threading
import time
import urllib.error
import urllib.request
import uuid
import warnings
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, Union
from urllib.parse import quote, unquote

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

warnings.filterwarnings("ignore")

from fasthtml.common import (
    A,
    Aside,
    Body,
    Br,
    Button,
    Div,
    FastHTML,
    Footer,
    Form,
    H1,
    H2,
    H3,
    H4,
    H5,
    Header,
    Hr,
    Img,
    Input,
    Label,
    Link,
    Main,
    Meta,
    Nav,
    NotStr,
    Ol,
    Option,
    P,
    Script,
    Section,
    Select,
    Span,
    Strong,
    Style,
    Table,
    Tbody,
    Td,
    Th,
    Thead,
    Title,
    Tr,
    Ul,
    Response,
    RedirectResponse,
    HTMLResponse,
)

try:
    from xhtml2pdf import pisa

    _PISA_OK = True
except Exception:  # pragma: no cover
    pisa = None
    _PISA_OK = False

try:
    from langchain_community.llms import Ollama

    _LANGCHAIN_OK = True
except Exception:  # pragma: no cover
    Ollama = None
    _LANGCHAIN_OK = False

try:
    import kaleido  # noqa: F401

    _KALEIDO_OK = True
except Exception:
    _KALEIDO_OK = False

# =============================================================================
# 2) CONFIGURATION & AI PROVIDER DESIGN
# =============================================================================

# AI Provider Configuration
# Supported values for AI_PROVIDER:
#   1. ollama_cloud  - (Default) Uses Ollama Cloud API with OLLAMA_API_KEY
#   2. ollama_local  - Uses local Ollama server at OLLAMA_HOST / OLLAMA_BASE_URL
#   3. deterministic - Bypasses LLM API calls; uses deterministic Python insights

def _load_env_file() -> None:
    """Load environment variables from .env file if present in workspace."""
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.isfile(env_path):
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, val = line.partition("=")
                    key = key.strip()
                    val = val.strip().strip("'\"")
                    if key and key not in os.environ:
                        os.environ[key] = val
        except Exception:
            pass

_load_env_file()

AI_PROVIDER = os.environ.get("AI_PROVIDER", "ollama_cloud").strip().lower()

CFG = {
    "app_title": "Faculty Grade Analytics Portal",
    "app_version": "Version 1.0",
    "app_subtitle": "Saranathan College of Engineering · Regulations 2024 Compatible",
    "host": os.environ.get("HOST", "0.0.0.0"),
    "port": int(os.environ.get("PORT", 8000)),
    "max_upload_mb": int(os.environ.get("MAX_UPLOAD_MB", "20")),
    "ai_provider": AI_PROVIDER,
    "ollama_api_key": os.environ.get("OLLAMA_API_KEY", "").strip(),
    "ollama_model": os.environ.get("OLLAMA_MODEL", "llama3").strip(),
    "ollama_host": os.environ.get("OLLAMA_HOST", os.environ.get("OLLAMA_BASE_URL", "")).strip(),
    "ai_timeout_seconds": float(os.environ.get("AI_TIMEOUT_SECONDS", "90")),
    "ai_enabled": os.environ.get("AI_ENABLED", "1") not in ("0", "false", "no") and AI_PROVIDER != "deterministic",
}

CLASS_AI_INSTRUCTION = (
    "You are an expert AI Academic Advisor for teachers. Review the compiled class metrics. "
    "Write a practical analysis summarizing overall strengths, problem subjects, and "
    "targeted recommendations. Format with bold keywords and crisp bullet points."
)

# =============================================================================
# 3) CONSTANTS & RESULT STATUS CLASSIFICATION
# =============================================================================

GRADE_POINTS = {
    "O": 10, "S": 10, "A+": 9, "A": 8, "B+": 7, "B": 6, "C+": 5.5, "C": 5,
    "U": 0, "RA": 0, "UA": 0, "SA": 0, "WD": 0, "MM": 0, "WH2": 0,
}
PASSING_GRADES = {"O", "S", "A+", "A", "B+", "B", "C+", "C"}
FAILING_GRADES = {"U", "RA", "UA", "SA", "WD", "MM", "WH2"}
ARREAR_GRADES = {"U", "RA", "UA"}
ATTENDANCE_GRADES = {"SA"}
WITHDRAWAL_GRADES = {"WD"}
MALPRACTICE_GRADES = {"MM", "WH2"}

GRADE_ORDER = ["O", "S", "A+", "A", "B+", "B", "C+", "C", "U", "RA", "UA", "SA", "WD", "MM", "WH2"]
PASS_GRADE_ORDER = ["O", "S", "A+", "A", "B+", "B", "C+", "C"]
FAIL_GRADE_ORDER = ["U", "RA", "UA", "SA", "WD", "MM", "WH2"]

RESULT_STATUS = {
    "U": {"type": "arrear", "description": "Reappearance required"},
    "RA": {"type": "arrear", "description": "Reappearance / Arrear"},
    "UA": {"type": "arrear", "description": "Absent / Reappearance required"},
    "SA": {"type": "attendance_issue", "description": "Shortage of Attendance"},
    "WD": {"type": "withdrawal", "description": "Withdrawal"},
    "MM": {"type": "malpractice", "description": "Malpractice"},
    "WH2": {"type": "malpractice", "description": "Malpractice (Withheld)"},
}

GRADE_COLORS = {
    "O": "#16a34a",
    "S": "#15803d",
    "A+": "#22c55e",
    "A": "#4ade80",
    "B+": "#86efac",
    "B": "#64748b",
    "C+": "#94a3b8",
    "C": "#cbd5e1",
    "U": "#dc2626",
    "RA": "#ef4444",
    "UA": "#f87171",
    "SA": "#d97706",
    "WD": "#9333ea",
    "MM": "#7c3aed",
    "WH2": "#6b21a8",
}

C_NAVY = "#0f1b33"
C_NAVY_LIGHT = "#1e3a5f"
C_GREEN = "#16a34a"
C_AMBER = "#d97706"
C_RED = "#dc2626"
C_BLUE = "#2563eb"
C_SLATE = "#64748b"
C_PURPLE = "#9333ea"
C_INDIGO = "#7c3aed"
C_BG = "#f1f5f9"

COLORWAY = [C_NAVY, C_BLUE, C_GREEN, C_AMBER, C_RED, C_SLATE, "#0ea5e9", "#a855f7"]

GRADE_ALIAS = {
    "A+": ("A+", "APLUS", "A PLUS", "A_+", "A.", "A++"),
    "B+": ("B+", "BPLUS", "B PLUS", "B_+"),
    "C+": ("C+", "CPLUS", "C PLUS", "C_+"),
}

GRADE_TOKEN_CLEAN = {
    "A+": "A+", "A PLUS": "A+", "AP": "A+", "A PLUS PLUS": "A+",
    "B+": "B+", "B PLUS": "B+", "BP": "B+",
    "C+": "C+", "C PLUS": "C+", "CP": "C+",
    "A0": "A+", "A1": "O", "O": "O", "S": "S", "A": "A", "B": "B", "C": "C",
    "U": "U", "FAIL": "U", "F": "U",
    "RA": "RA", "R.A": "RA", "R/A": "RA", "R A": "RA",
    "AB": "UA", "ABSENT": "UA", "UA": "UA", "ABS": "UA", "A B": "UA", "A.B": "UA", "A/B": "UA",
    "SA": "SA", "S A": "SA", "S.A": "SA", "NE": "SA",
    "WD": "WD", "W": "WD", "DW": "WD", "W.D": "WD", "WITHDRAWN": "WD",
    "MM": "MM", "M.M": "MM", "MALPRACTICE": "MM",
    "WH": "WH2", "WH1": "WH2", "WH2": "WH2", "WH 2": "WH2", "WH-2": "WH2", "WH.2": "WH2", "WH02": "WH2", "WITHHELD": "WH2",
}


HEADER_ALIASES: Dict[str, List[str]] = {
    "regno": [
        "register number", "student register number", "register no", "reg number",
        "reg no", "regno", "registration number", "registration no", "student number",
        "student no", "student id", "roll number", "roll no", "admission number",
        "student reg no", "register num", "regnum", "reg no.", "student. no",
    ],
    "name": [
        "student name", "name of the student", "candidate name", "student", "name",
        "full name", "student full name", "candidate", "student's name",
    ],
    "subject": [
        "subject name", "course name", "subject", "course", "paper", "course title",
        "subject title", "name of the subject", "course name subject",
    ],
    "course_code": [
        "course code", "subject code", "code", "course id", "subject no",
        "course number", "course id",
    ],
    "credits": [
        "credit", "credits", "credit point", "credit points", "credits earned", "credits sp",
    ],
    "grade": [
        "final grade", "letter grade", "final letter grade", "grade", "result grade",
        "result", "grade achieved", "final result", "grade secured", "overall grade",
    ],
    "dept": ["department", "department name", "branch", "discipline", "dept", "branch name"],
    "programme": ["programme", "program", "degree", "programme of study", "program name"],
    "batch": ["batch", "batch year", "batch no"],
    "year": ["year", "year of study", "study year"],
    "semester": ["semester", "sem", "current semester", "semester number"],
    "section": ["section", "sec", "class section"],
    "academic_year": ["academic year", "academic session", "ay", "acadyear", "aca year"],
    "instructor": ["instructor", "faculty", "faculty name", "staff", "staff name", "teacher"],
}

REQUIRED_FIELDS = ["regno", "name", "subject", "credits", "grade"]
OPTIONAL_FIELDS = ["course_code", "dept", "programme", "batch", "year", "semester", "section", "academic_year", "instructor"]

STATUS_CLEARED = "cleared"
STATUS_U = "u"
STATUS_MULTI_U = "multi-u"
STATUS_BACKLOG = "backlog"
STATUS_SA = "sa"
STATUS_WD = "wd"
STATUS_MALPRACTICE = "malpractice"

# =============================================================================
# 3.1) SYLLABUS CATALOG & DYNAMIC ALIAS RESOLVER
# =============================================================================

from syllabus.loader import (
    get_catalog,
    get_registered_departments,
    resolve_course,
    save_and_register_syllabus,
    invalidate_catalog_cache,
    normalize_course_code,
    normalize_text,
    normalize_alphanumeric,
)
from syllabus_tools.extractor import extract_syllabus_pdf
from syllabus_tools.publisher import publish_syllabus_draft

SYLLABUS_CATALOG_R2024: List[Dict[str, Any]] = []

# Pre-computed O(1) indexes for SYLLABUS_CATALOG_R2024
COURSE_CODE_INDEX: Dict[str, Dict[str, Any]] = {}
SUBJECT_NAME_INDEX: Dict[str, Dict[str, Any]] = {}
ALIAS_INDEX: Dict[str, Dict[str, Any]] = {}
SEMESTER_INDEX: Dict[int, List[Dict[str, Any]]] = {}
CATALOG_VERSION = "r2024-full-catalog-2026-08-16"

def _build_syllabus_indexes():
    global SYLLABUS_CATALOG_R2024, CATALOG_VERSION
    seen_codes = set()
    all_courses: List[Dict[str, Any]] = []

    COURSE_CODE_INDEX.clear()
    SUBJECT_NAME_INDEX.clear()
    ALIAS_INDEX.clear()
    SEMESTER_INDEX.clear()

    # 1. Primary AI_DS department catalog first
    ai_ds_cat = get_catalog("AI_DS", "R2024")
    if ai_ds_cat:
        for item in ai_ds_cat.courses:
            code = str(item.get("code", "")).strip().upper()
            if code and code not in seen_codes:
                seen_codes.add(code)
                all_courses.append(item)

    # 2. All other registered department catalogs
    for dept in get_registered_departments():
        dept_code = dept.get("code", "")
        if dept_code.upper() == "AI_DS":
            continue
        cat = get_catalog(dept_code, "R2024")
        if cat:
            for item in cat.courses:
                code = str(item.get("code", "")).strip().upper()
                if code and code not in seen_codes:
                    seen_codes.add(code)
                    all_courses.append(item)

    SYLLABUS_CATALOG_R2024 = all_courses

    for item in SYLLABUS_CATALOG_R2024:
        code_raw = str(item.get("code", "")).strip().upper()
        code_clean = re.sub(r"[^A-Z0-9]", "", code_raw)
        if code_raw:
            COURSE_CODE_INDEX[code_raw] = item
        if code_clean:
            COURSE_CODE_INDEX[code_clean] = item
            # Variant indexing for trailing 'A' equivalence (e.g. 24EN201 <-> 24EN201A)
            if not code_clean.endswith("A"):
                COURSE_CODE_INDEX[f"{code_clean}A"] = item
            elif code_clean.endswith("A") and len(code_clean) > 4 and code_clean[-2].isdigit():
                COURSE_CODE_INDEX[code_clean[:-1]] = item

        name_raw = str(item.get("name", "")).strip()
        name_upper = name_raw.upper()
        name_clean = re.sub(r"[^A-Z0-9]", "", name_upper)
        if name_upper:
            SUBJECT_NAME_INDEX[name_upper] = item
        if name_clean:
            SUBJECT_NAME_INDEX[name_clean] = item

        for alias in item.get("aliases", []):
            alias_upper = str(alias).strip().upper()
            alias_clean = re.sub(r"[^A-Z0-9]", "", alias_upper)
            if alias_upper:
                ALIAS_INDEX[alias_upper] = item
            if alias_clean:
                ALIAS_INDEX[alias_clean] = item

        sem = int(item.get("semester", 1))
        SEMESTER_INDEX.setdefault(sem, []).append(item)

    CATALOG_VERSION = f"catalog-ver-{time.time()}"
    try:
        _memoized_resolve_subject_info.cache_clear()
    except Exception:
        pass

_build_syllabus_indexes()

import functools

@functools.lru_cache(maxsize=4096)
def _memoized_resolve_subject_info(norm: str, clean_code: str, _catalog_version: str = CATALOG_VERSION) -> Tuple[str, str, float, int, str, float, bool]:
    # Stage 1: Exact / Normalized Course Code (O(1))
    if norm in COURSE_CODE_INDEX:
        item = COURSE_CODE_INDEX[norm]
        return (item["name"], item["code"], float(item.get("credits", 3.0)), int(item.get("semester", 1)), item.get("category", "Sem 1-4 Foundation"), 1.0, False)
    if clean_code in COURSE_CODE_INDEX:
        item = COURSE_CODE_INDEX[clean_code]
        return (item["name"], item["code"], float(item.get("credits", 3.0)), int(item.get("semester", 1)), item.get("category", "Sem 1-4 Foundation"), 1.0, False)
    stripped_code = re.sub(r"[A-Z]$", "", clean_code)
    if stripped_code and stripped_code in COURSE_CODE_INDEX:
        item = COURSE_CODE_INDEX[stripped_code]
        return (item["name"], item["code"], float(item.get("credits", 3.0)), int(item.get("semester", 1)), item.get("category", "Sem 1-4 Foundation"), 1.0, False)

    # Stage 2: Exact / Normalized Subject Name (O(1))
    if norm in SUBJECT_NAME_INDEX:
        item = SUBJECT_NAME_INDEX[norm]
        return (item["name"], item["code"], float(item.get("credits", 3.0)), int(item.get("semester", 1)), item.get("category", "Sem 1-4 Foundation"), 1.0, False)
    if clean_code in SUBJECT_NAME_INDEX:
        item = SUBJECT_NAME_INDEX[clean_code]
        return (item["name"], item["code"], float(item.get("credits", 3.0)), int(item.get("semester", 1)), item.get("category", "Sem 1-4 Foundation"), 1.0, False)

    # Stage 3: Exact / Normalized Alias (O(1))
    if norm in ALIAS_INDEX:
        item = ALIAS_INDEX[norm]
        return (item["name"], item["code"], float(item.get("credits", 3.0)), int(item.get("semester", 1)), item.get("category", "Sem 1-4 Foundation"), 1.0, False)

    # Stage 4: Dynamic Syllabus Resolver Check
    try:
        res_course = resolve_course(norm, context={"department": "AI_DS", "regulation": "R2024"})
        if res_course and not res_course[6]:
            return (res_course[0], res_course[1], float(res_course[2]), int(res_course[3]), res_course[4], float(res_course[5]), False)
    except Exception:
        pass

    # If the input is a course code pattern (e.g. 24EN201A, 24TA201, 24CS211, etc.), preserve authoritative code
    code_pattern = re.match(r"^(?:24[-_]?)?([A-Z]{2,4})[-_]?(\d{3,4})[A-Z]?$", clean_code)
    if code_pattern:
        dept_pfx = code_pattern.group(1)
        num_part = code_pattern.group(2)
        sem_digit = int(num_part[0]) if num_part and num_part[0].isdigit() else 1
        cred_val = 1.0 if (len(num_part) >= 3 and num_part[1] == '1') else 3.0
        return (norm, stripped_code or clean_code, cred_val, sem_digit, "Sem 1-4 Foundation", 1.0, False)

    # Stage 5: Fuzzy Match Fallback (only executed for full descriptive subject titles)
    best_item = None
    best_score = 0.0

    for item in SYLLABUS_CATALOG_R2024:
        s_name = difflib.SequenceMatcher(None, norm, str(item.get("name", "")).upper()).ratio()
        if s_name > best_score:
            best_score = s_name
            best_item = item
        for alias in item.get("aliases", []):
            s_alias = difflib.SequenceMatcher(None, norm, str(alias).upper()).ratio()
            if s_alias > best_score:
                best_score = s_alias
                best_item = item

    if best_item and best_score >= 0.80:
        return (best_item["name"], best_item["code"], float(best_item.get("credits", 3.0)), int(best_item.get("semester", 1)), best_item.get("category", "Sem 1-4 Foundation"), round(best_score, 2), False)
    elif best_item and best_score >= 0.60:
        return (best_item["name"], best_item["code"], float(best_item.get("credits", 3.0)), int(best_item.get("semester", 1)), best_item.get("category", "Sem 1-4 Foundation"), round(best_score, 2), True)

    return (norm, clean_code or norm, 3.0, 1, "Sem 1-4 Foundation", round(best_score, 2), True)


def resolve_subject_info(raw_name: Any, custom_overrides: Optional[Dict[str, str]] = None) -> Tuple[str, str, float, int, str, float, bool]:
    """
    Staged resolver for subject strings, course codes, or abbreviations against R2024 catalog.
    Uses pre-built O(1) hash indexes for course code, name, and alias lookups.
    """
    if not raw_name:
        return ("Unknown Subject", "", 0.0, 0, "Uncategorized", 0.0, True)

    clean_raw = str(raw_name).strip()
    norm = re.sub(r"\s+", " ", clean_raw).upper()
    clean_code = re.sub(r"[^A-Z0-9]", "", norm)

    # 0. Check custom manual alias overrides first
    if custom_overrides and norm in custom_overrides:
        target_name = custom_overrides[norm]
        if target_name.upper() in SUBJECT_NAME_INDEX:
            item = SUBJECT_NAME_INDEX[target_name.upper()]
            return (item["name"], item["code"], float(item.get("credits", 3.0)), int(item.get("semester", 1)), item.get("category", "Custom Subject"), 1.0, False)
        return (target_name, "", 3.0, 1, "Custom Subject", 1.0, False)

    res = _memoized_resolve_subject_info(norm, clean_code, CATALOG_VERSION)
    if res[0] == norm:
        return (clean_raw, clean_code or clean_raw, res[2], res[3], res[4], res[5], res[6])
    return res


# =============================================================================
# 3.2) DIRECT COE PDF EXTRACTION, METADATA & RECONCILIATION ENGINE
# =============================================================================

@dataclass
class DocumentMetadata:
    institution: str = "Saranathan College of Engineering"
    programme: str = "B.Tech AI & DS"
    department: str = "Department of AI & DS"
    regulation: str = "R2024"
    semester: str = "Semester III"
    academic_year: str = "2025 - 2026"
    exam_session: str = "Nov / Dec 2025"
    publication_date: str = "Unknown / Needs Review"
    page_count: int = 0
    document_type: str = "DIGITAL_TEXT_PDF"  # "DIGITAL_TEXT_PDF" | "SCANNED_PDF" | "EMPTY_OR_CORRUPT_PDF"


@dataclass
class StudentResultRecord:
    register_number: str
    student_name: str
    batch_number: str = ""
    subject_code: str = ""
    subject_name: str = ""
    original_subject_text: str = ""
    credits: float = 3.0
    result_status: str = "U"
    raw_result_status: str = ""
    result_semester: int = 3
    subject_semester: int = 3
    source_type: str = "PDF"  # "PDF" | "EXCEL" | "XLS" | "XLSX"
    source_page: int = 1
    extraction_confidence: float = 1.0
    # Legacy-XLS / multi-source provenance (additive; unused by the PDF pipeline)
    source_sheet: str = ""
    source_row: int = 0
    grade_column: str = ""
    gp_column: str = ""
    source_gp: Optional[float] = None
    data_quality_issue: str = ""  # e.g. "GRADE_POINT_MISMATCH", "" if none


@dataclass
class PDFExtractionReport:
    ok: bool = True
    doc_metadata: DocumentMetadata = field(default_factory=DocumentMetadata)
    records: List[StudentResultRecord] = field(default_factory=list)
    prev_records: List[StudentResultRecord] = field(default_factory=list)
    all_records: List[StudentResultRecord] = field(default_factory=list)
    detected_semesters: List[Dict[str, Any]] = field(default_factory=list)
    primary_semester: int = 1
    raw_inspector_items: List[Dict[str, Any]] = field(default_factory=list)
    quarantined_tokens: List[Dict[str, Any]] = field(default_factory=list)
    student_count: int = 0
    subject_count: int = 0
    result_cell_count: int = 0
    unknown_token_count: int = 0
    unresolved_subject_count: int = 0
    overall_confidence: float = 1.0
    warnings: List[str] = field(default_factory=list)
    fatal_error: str = ""


@dataclass
class ReconciliationReport:
    matched_count: int = 0
    mismatched_count: int = 0
    mismatched_records: List[Dict[str, Any]] = field(default_factory=list)
    missing_in_excel: List[Dict[str, Any]] = field(default_factory=list)
    missing_in_pdf: List[Dict[str, Any]] = field(default_factory=list)


def _resolve_id_columns(headers: List[str]) -> Tuple[int, int]:
    """
    Resolve (regno_col_idx, name_col_idx) from a table header row by label, so the
    parser isn't hostage to a fixed column position. Falls back to the known COE
    layout (S.No, Register No, Name) only when header labels can't be matched.
    """
    regno_idx = None
    name_idx = None
    for idx, h in enumerate(headers):
        h_norm = re.sub(r"\s+", " ", str(h or "")).strip().upper()
        if regno_idx is None and any(k in h_norm for k in ("REGISTER NO", "REG.NO", "REG NO", "REGISTER NUMBER", "REGNO")):
            regno_idx = idx
        if name_idx is None and ("NAME" in h_norm and "STAFF" not in h_norm and "MENTOR" not in h_norm):
            name_idx = idx
    if regno_idx is None:
        regno_idx = 1  # known COE layout: S.No, Register No, Name, ...
    if name_idx is None:
        name_idx = regno_idx + 1
    return regno_idx, name_idx


ROMAN_TO_INT_MAP = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8,
    "1": 1, "2": 2, "3": 3, "4": 4, "5": 5, "6": 6, "7": 7, "8": 8,
    "01": 1, "02": 2, "03": 3, "04": 4, "05": 5, "06": 6, "07": 7, "08": 8,
    "1ST": 1, "2ND": 2, "3RD": 3, "4TH": 4, "5TH": 5, "6TH": 6, "7TH": 7, "8TH": 8,
    "FIRST": 1, "SECOND": 2, "THIRD": 3, "FOURTH": 4, "FIFTH": 5, "SIXTH": 6, "SEVENTH": 7, "EIGHTH": 8,
}
INT_TO_ROMAN_MAP = {
    1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII"
}

def _parse_page_semester(text: str) -> Optional[int]:
    """
    Parse semester number from page text, header, or table caption across any page in the document.
    Works for any number of pages (whether arrear results span 1, 5, or 20 pages across any department).
    """
    if not text:
        return None
    upper_txt = text.upper()
    m = re.search(r"SEMESTER\s*[:\-\s]\s*([IVX0-9]+|FIRST|SECOND|THIRD|FOURTH|FIFTH|SIXTH|SEVENTH|EIGHTH)\b", upper_txt)
    if m:
        raw = m.group(1).strip()
        if raw in ROMAN_TO_INT_MAP:
            return ROMAN_TO_INT_MAP[raw]
    m_sem = re.search(r"\bSEM\s*[:\-\s]\s*([IVX0-9]+|FIRST|SECOND|THIRD|FOURTH|FIFTH|SIXTH|SEVENTH|EIGHTH)\b", upper_txt)
    if m_sem:
        raw = m_sem.group(1).strip()
        if raw in ROMAN_TO_INT_MAP:
            return ROMAN_TO_INT_MAP[raw]
    m_ord = re.search(r"\b([1-8](?:ST|ND|RD|TH)?|[IVX]{1,4}|FIRST|SECOND|THIRD|FOURTH|FIFTH|SIXTH|SEVENTH|EIGHTH)\s+SEM(?:ESTER)?\b", upper_txt)
    if m_ord:
        raw = m_ord.group(1).strip()
        if raw in ROMAN_TO_INT_MAP:
            return ROMAN_TO_INT_MAP[raw]
    return None


def extract_coe_pdf(pdf_bytes: bytes, filename: str, analysis_context: Optional[Dict[str, Any]] = None) -> PDFExtractionReport:
    """
    Universal multi-page COE PDF extraction engine using PyMuPDF with pdfplumber fallback.
    Extracts document metadata, multi-page tables, detects per-page semesters across arbitrary
    numbers of pages (e.g. 1 to N pages of arrear exams for earlier semesters, cleanly segregated
    from the target regular cohort), isolates headers per page, and tracks student provenance.
    """
    report = PDFExtractionReport()
    if not pdf_bytes:
        report.ok = False
        report.fatal_error = "Uploaded PDF file is empty (0 bytes)."
        return report

    try:
        try:
            import pymupdf as fitz
        except ImportError:
            import fitz
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        report.ok = False
        report.doc_metadata.document_type = "EMPTY_OR_CORRUPT_PDF"
        report.fatal_error = f"Could not open or parse PDF document: {e}"
        return report

    report.doc_metadata.page_count = len(doc)
    if len(doc) == 0:
        report.ok = False
        report.doc_metadata.document_type = "EMPTY_OR_CORRUPT_PDF"
        report.fatal_error = "PDF document contains 0 pages."
        return report

    # 1. Page-by-page text sampling for PDF Classification
    total_chars = 0
    page_texts = []
    for page_num in range(len(doc)):
        text = doc[page_num].get_text("text") or ""
        page_texts.append(text)
        total_chars += len(text.strip())

    avg_chars_per_page = total_chars / len(doc)
    if avg_chars_per_page < 80:
        report.doc_metadata.document_type = "SCANNED_PDF"
        report.warnings.append("PDF appears to be scanned or image-based with minimal selectable text layer.")
    else:
        report.doc_metadata.document_type = "DIGITAL_TEXT_PDF"

    # 2. Extract Document Metadata dynamically across pages
    for p_text in page_texts[:4]:
        p_upper = p_text.upper()
        if "SARANATHAN" in p_upper:
            report.doc_metadata.institution = "Saranathan College of Engineering"
        elif "COLLEGE OF ENGINEERING" in p_upper and not report.doc_metadata.institution:
            report.doc_metadata.institution = "College of Engineering"

        if not report.doc_metadata.department:
            for d in get_registered_departments():
                d_name = d["name"].upper()
                d_code = d["code"].upper()
                if d_name in p_upper or f"({d_code})" in p_upper or f"B.TECH.- {d_name}" in p_upper or f"B.E.- {d_name}" in p_upper or f"- {d_name}" in p_upper:
                    report.doc_metadata.department = d["name"]
                    prefix = "B.Tech" if ("TECH" in d_name or d_code in ("AI_DS", "IT")) else "B.E."
                    report.doc_metadata.programme = f"{prefix} {d['name']}"
                    break

        report.doc_metadata.regulation = "R2024"

        ay_m = re.search(r"(202\d\s*-\s*202\d|NOV\s*/\s*DEC\s*202\d|APR\s*/\s*MAY\s*202\d)", p_upper)
        if ay_m and not report.doc_metadata.exam_session:
            report.doc_metadata.exam_session = ay_m.group(1)

        dt_m = re.search(r"DATE\s*:\s*(\d{2}[-/\.]\d{2}[-/\.]\d{4})", p_upper)
        if dt_m and not report.doc_metadata.publication_date:
            report.doc_metadata.publication_date = dt_m.group(1)

    if analysis_context:
        if analysis_context.get("department"):
            report.doc_metadata.department = str(analysis_context["department"])
        if analysis_context.get("regulation"):
            report.doc_metadata.regulation = str(analysis_context["regulation"])

    extracted_records: List[StudentResultRecord] = []
    inspector_items: List[Dict[str, Any]] = []
    student_reg_pattern = re.compile(r"\b(8138\d{8}|\d{10,12})\b")
    last_detected_sem = 1

    # Extract all records page-by-page (handling arbitrary number of arrear & regular pages)
    for page_idx in range(len(doc)):
        page = doc[page_idx]
        src_page = page_idx + 1
        page_text = page_texts[page_idx]
        page_lines = page_text.split("\n")
        page_records: List[StudentResultRecord] = []

        p_sem = _parse_page_semester(page_text)
        if p_sem is not None:
            last_detected_sem = p_sem
        current_page_sem = p_sem if p_sem is not None else last_detected_sem

        # Scan for course codes in this page's header lines (scoped per page)
        page_course_headers = []
        for line in page_lines[:25]:
            codes = re.findall(r"\b((?:24[-_]?)?[A-Z]{2,4}[-_]?\d{3,5}[A-Z]?)\b", line.upper())
            for ccode in codes:
                can_name, code, cred, sem, cat, conf, amb = resolve_subject_info(ccode)
                if (code or ccode) not in [ch["code"] for ch in page_course_headers]:
                    page_course_headers.append({
                        "code": code or ccode,
                        "canonical_name": can_name,
                        "credits": cred,
                        "semester": sem or current_page_sem,
                        "category": cat,
                        "confidence": conf
                    })

        # 1. Extract course code headers with dynamic bounding box & spatial awareness
        words = page.get_text("words") or []
        page_course_headers = []

        if words:
            # Locate all student register numbers on the page to determine header boundary
            reg_words = [w for w in words if student_reg_pattern.search(w[4].strip())]
            min_student_y = min(w[1] for w in reg_words) if reg_words else (page.rect.height * 0.45)

            # Filter candidate words above student rows matching course code patterns
            candidate_h_words = [
                w for w in words
                if w[3] <= min_student_y and re.match(r"^(?:24[-_]?)?[A-Z]{2,6}[-_]?\d{3,5}[A-Z]?$", w[4].strip().upper())
            ]

            if candidate_h_words:
                # Group candidate course words by vertical center into bands
                y_bands: Dict[int, List[Any]] = defaultdict(list)
                for w in candidate_h_words:
                    band_key = int(round((w[1] + w[3]) / 2.0 / 6.0) * 6.0)
                    y_bands[band_key].append(w)

                # Pick the band containing the largest number of course code headers
                best_band_key = max(y_bands.keys(), key=lambda k: len(y_bands[k]))
                header_words = sorted(y_bands[best_band_key], key=lambda w: w[0])

                for cw in header_words:
                    ccode = cw[4].strip().upper()
                    if ccode not in [ch["code"] for ch in page_course_headers]:
                        can_name, code, cred, sem, cat, conf, amb = resolve_subject_info(ccode)
                        page_course_headers.append({
                            "code": code or ccode,
                            "canonical_name": can_name,
                            "credits": cred,
                            "semester": sem or current_page_sem,
                            "category": cat,
                            "confidence": conf,
                            "x0": cw[0],
                            "x1": cw[2],
                            "center": (cw[0] + cw[2]) / 2.0,
                        })

        # Fallback: scan top lines of text if word coordinates did not identify headers
        if not page_course_headers:
            for line in page_lines[:25]:
                codes = re.findall(r"\b((?:24[-_]?)?[A-Z]{2,6}[-_]?\d{3,5}[A-Z]?)\b", line.upper())
                for ccode in codes:
                    can_name, code, cred, sem, cat, conf, amb = resolve_subject_info(ccode)
                    if (code or ccode) not in [ch["code"] for ch in page_course_headers]:
                        page_course_headers.append({
                            "code": code or ccode,
                            "canonical_name": can_name,
                            "credits": cred,
                            "semester": sem or current_page_sem,
                            "category": cat,
                            "confidence": conf,
                        })

        cols = []
        for i, h in enumerate(page_course_headers):
            if "x0" in h:
                left = (page_course_headers[i-1]["x1"] + h["x0"]) / 2.0 if i > 0 else h["x0"] - 25.0
                right = (h["x1"] + page_course_headers[i+1]["x0"]) / 2.0 if i < len(page_course_headers) - 1 else h["x1"] + 25.0
                cols.append({"header": h, "x0": left, "x1": right, "center": h["center"]})

        # 2. Precision line-word extraction & multi-pass alignment
        if words:
            # Sort all words by vertical position for adaptive moving-average line clustering
            words_sorted = sorted(words, key=lambda w: (w[1] + w[3]) / 2.0)
            clustered_lines: List[List[Any]] = []
            curr_line: List[Any] = []
            curr_y_center = None

            for w in words_sorted:
                wy = (w[1] + w[3]) / 2.0
                if curr_y_center is None or abs(wy - curr_y_center) <= 4.0:
                    curr_line.append(w)
                    curr_y_center = sum((x[1] + x[3]) / 2.0 for x in curr_line) / len(curr_line)
                else:
                    clustered_lines.append(curr_line)
                    curr_line = [w]
                    curr_y_center = wy
            if curr_line:
                clustered_lines.append(curr_line)

            for line_words in clustered_lines:
                lw = sorted(line_words, key=lambda x: x[0])
                line_str = " ".join(w[4] for w in lw if w[4])
                m = student_reg_pattern.search(line_str)
                if m:
                    regno = m.group(1)
                    tokens = [w[4] for w in lw if w[4]]
                    reg_idx = next((i for i, t in enumerate(tokens) if regno in t), -1)
                    after_tokens = tokens[reg_idx + 1:] if reg_idx >= 0 else tokens

                    raw_name = ""
                    grades_found: List[Tuple[Dict[str, Any], str, str]] = []
                    n_subjs = len(page_course_headers)

                    # Pass 1: Right-anchored tail match for regular cohort rows (handles multi-part names & initials)
                    if n_subjs > 0 and len(after_tokens) >= n_subjs:
                        tail = after_tokens[-n_subjs:]
                        tail_norm = [_grade_normalize(t) for t in tail]
                        if all(tail_norm):
                            grades_found = [(page_course_headers[i], tail[i], tail_norm[i]) for i in range(n_subjs)]
                            raw_name = " ".join(after_tokens[:-n_subjs])

                    # Pass 2: Spatial Voronoi column match (handles arrear/sparse rows with blank columns)
                    if not grades_found and cols:
                        name_words = []
                        min_col_x = cols[0]["x0"] if cols else 0.0
                        for w in lw[reg_idx + 1:] if reg_idx >= 0 else lw:
                            w_center = (w[0] + w[2]) / 2.0
                            matched_col = None
                            for c in cols:
                                if c["x0"] <= w_center <= c["x1"]:
                                    matched_col = c
                                    break
                            if matched_col:
                                g_norm = _grade_normalize(w[4])
                                if g_norm:
                                    grades_found.append((matched_col["header"], w[4], g_norm))
                                else:
                                    name_words.append(w[4])
                            elif not grades_found or w_center < min_col_x:
                                name_words.append(w[4])
                        raw_name = " ".join(name_words)

                    # Pass 3: Fallback sequential matching
                    if not grades_found:
                        name_parts = []
                        for tok in after_tokens:
                            g_norm = _grade_normalize(tok)
                            if g_norm:
                                idx = len(grades_found)
                                ch = page_course_headers[idx] if idx < len(page_course_headers) else {
                                    "code": f"SUBJ_{idx+1}",
                                    "canonical_name": f"Subject {idx+1}",
                                    "credits": 3.0,
                                    "semester": current_page_sem,
                                }
                                grades_found.append((ch, tok, g_norm))
                            elif len(grades_found) == 0:
                                name_parts.append(tok)
                        raw_name = " ".join(name_parts)

                    # Post-process & sanitize raw_name
                    raw_name = re.sub(r"\s+", " ", raw_name).strip(" -_.,")

                    for ch, raw_g, norm_g in grades_found:
                        rec = StudentResultRecord(
                            register_number=regno,
                            student_name=raw_name,
                            subject_code=ch["code"],
                            subject_name=ch["canonical_name"],
                            original_subject_text=ch["code"],
                            credits=ch.get("credits", 3.0),
                            result_status=norm_g,
                            raw_result_status=raw_g,
                            result_semester=current_page_sem,
                            subject_semester=ch.get("semester", current_page_sem),
                            source_type="PDF",
                            source_page=src_page,
                            extraction_confidence=0.98,
                        )
                        page_records.append(rec)
                        inspector_items.append({
                            "source_page": src_page,
                            "raw_text": line_str,
                            "parsed_regno": regno,
                            "parsed_name": raw_name or "—",
                            "parsed_subject": ch["canonical_name"],
                            "parsed_grade": norm_g,
                            "confidence": "HIGH"
                        })

        # 3. Native PyMuPDF fast table extraction fallback
        if len(page_records) == 0:
            try:
                tabs = page.find_tables()
                if not tabs or not tabs.tables:
                    tabs = page.find_tables(strategy="text")
                if tabs and tabs.tables:
                    for tab in tabs.tables:
                        raw_matrix = tab.extract()
                        if not raw_matrix or len(raw_matrix) < 2:
                            continue
                        headers = [str(c or "").strip() for c in raw_matrix[0]]
                        regno_col, name_col = _resolve_id_columns(headers)
                        for row in raw_matrix[1:]:
                            if not row:
                                continue
                            row_str = " ".join(str(c or "") for c in row)
                            m = student_reg_pattern.search(row_str)
                            if m:
                                regno = m.group(1)
                                name_val = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ""
                                if name_val.strip() == regno.strip() and len(row) > regno_col and row[regno_col]:
                                    fallback_idx = name_col + 1
                                    if len(row) > fallback_idx and row[fallback_idx]:
                                        name_val = str(row[fallback_idx]).strip()
                                for c_idx in range(len(row)):
                                    if c_idx in (regno_col, name_col):
                                        continue
                                    cell_v = str(row[c_idx] or "").strip()
                                    norm_g = _grade_normalize(cell_v)
                                    # Handle cell spillover from name
                                    if not norm_g and cell_v and c_idx == name_col + 1:
                                        parts = cell_v.split()
                                        if parts:
                                            cand_g = _grade_normalize(parts[-1])
                                            if cand_g:
                                                norm_g = cand_g
                                                cell_v = parts[-1]
                                                name_val = f"{name_val} {' '.join(parts[:-1])}".strip()
                                    if norm_g and cell_v.upper() != regno:
                                        subj_hdr = headers[c_idx] if c_idx < len(headers) else f"Column_{c_idx}"
                                        can_name, code, cred, sem, cat, conf, amb = resolve_subject_info(subj_hdr)
                                        rec = StudentResultRecord(
                                            register_number=regno,
                                            student_name=name_val,
                                            subject_code=code or subj_hdr,
                                            subject_name=can_name,
                                            original_subject_text=subj_hdr,
                                            credits=cred if cred > 0 else 3.0,
                                            result_status=norm_g,
                                            raw_result_status=cell_v,
                                            result_semester=current_page_sem,
                                            subject_semester=sem or current_page_sem,
                                            source_type="PDF",
                                            source_page=src_page,
                                            extraction_confidence=0.95
                                        )
                                        page_records.append(rec)
                                        inspector_items.append({
                                            "source_page": src_page,
                                            "raw_text": row_str,
                                            "parsed_regno": regno,
                                            "parsed_name": name_val or "—",
                                            "parsed_subject": can_name,
                                            "parsed_grade": norm_g,
                                            "confidence": "HIGH"
                                        })
            except Exception:
                pass

        # 4. Secondary pdfplumber table fallback
        if len(page_records) == 0:
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(pdf_bytes)) as plumber_pdf:
                    if src_page <= len(plumber_pdf.pages):
                        plumber_page = plumber_pdf.pages[src_page - 1]
                        plumber_tables = plumber_page.extract_tables()
                        for ptab in plumber_tables:
                            if not ptab or len(ptab) < 2:
                                continue
                            headers = [str(c or "").strip() for c in ptab[0]]
                            regno_col, name_col = _resolve_id_columns(headers)
                            for row in ptab[1:]:
                                if not row:
                                    continue
                                row_str = " ".join(str(c or "") for c in row)
                                m = student_reg_pattern.search(row_str)
                                if m:
                                    regno = m.group(1)
                                    name_val = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ""
                                    for c_idx in range(len(row)):
                                        if c_idx in (regno_col, name_col):
                                            continue
                                        cell_v = str(row[c_idx] or "").strip()
                                        norm_g = _grade_normalize(cell_v)
                                        if norm_g and cell_v.upper() != regno:
                                            subj_hdr = headers[c_idx] if c_idx < len(headers) else f"Column_{c_idx}"
                                            can_name, code, cred, sem, cat, conf, amb = resolve_subject_info(subj_hdr)
                                            rec = StudentResultRecord(
                                                register_number=regno,
                                                student_name=name_val,
                                                subject_code=code or subj_hdr,
                                                subject_name=can_name,
                                                original_subject_text=subj_hdr,
                                                credits=cred if cred > 0 else 3.0,
                                                result_status=norm_g,
                                                raw_result_status=cell_v,
                                                result_semester=current_page_sem,
                                                subject_semester=sem or current_page_sem,
                                                source_type="PDF",
                                                source_page=src_page,
                                                extraction_confidence=0.90
                                            )
                                            page_records.append(rec)
                                            inspector_items.append({
                                                "source_page": src_page,
                                                "raw_text": row_str,
                                                "parsed_regno": regno,
                                                "parsed_name": name_val or "—",
                                                "parsed_subject": can_name,
                                                "parsed_grade": norm_g,
                                                "confidence": "HIGH"
                                            })
            except Exception as pe:
                report.warnings.append(f"pdfplumber table extraction notice for page {src_page}: {pe}")

        extracted_records.extend(page_records)

    # 4. Multi-Semester Identification and Dynamic Multi-Page Segregation
    target_sem_requested = None
    if analysis_context and analysis_context.get("semester"):
        try:
            target_sem_requested = int(analysis_context["semester"])
        except (ValueError, TypeError):
            target_sem_requested = None

    sem_counts = Counter(r.result_semester for r in extracted_records if r.result_semester)
    
    if target_sem_requested and target_sem_requested in sem_counts:
        target_semester = target_sem_requested
    elif sem_counts:
        # Dominant cohort semester (largest student/record count across all pages)
        target_semester = sem_counts.most_common(1)[0][0]
    else:
        target_semester = 1

    primary_records = [r for r in extracted_records if r.result_semester == target_semester]
    prev_records = [r for r in extracted_records if r.result_semester and r.result_semester < target_semester]

    if not primary_records and extracted_records:
        primary_records = extracted_records

    detected_semesters = []
    for s_num in sorted(set(r.result_semester for r in extracted_records if r.result_semester)):
        s_recs = [r for r in extracted_records if r.result_semester == s_num]
        s_students = len(set(r.register_number for r in s_recs))
        s_subjs = sorted(set(r.subject_code for r in s_recs))
        detected_semesters.append({
            "semester": s_num,
            "semester_label": f"Semester {INT_TO_ROMAN_MAP.get(s_num, str(s_num))}",
            "student_count": s_students,
            "record_count": len(s_recs),
            "subject_count": len(s_subjs),
            "subject_codes": s_subjs,
            "is_primary": (s_num == target_semester),
            "is_arrear": (s_num < target_semester),
        })

    report.records = primary_records
    report.prev_records = prev_records
    report.all_records = extracted_records
    report.detected_semesters = detected_semesters
    report.primary_semester = target_semester
    report.doc_metadata.semester = f"Semester {INT_TO_ROMAN_MAP.get(target_semester, str(target_semester))}"

    # Calculate overall confidence & stats for primary records
    unique_regnos = set(r.register_number for r in primary_records)
    unique_subjects = set(r.subject_code for r in primary_records)
    report.student_count = len(unique_regnos)
    report.subject_count = len(unique_subjects)
    report.result_cell_count = len(primary_records)
    report.raw_inspector_items = inspector_items

    if len(primary_records) > 0:
        report.overall_confidence = round(
            sum(r.extraction_confidence for r in primary_records) / len(primary_records), 2
        )
    else:
        report.overall_confidence = 0.0

    return report

def clean_staff_name(staff_raw: str) -> str:
    """Cleans and standardizes staff names extracted from HTML mark sheets."""
    if not staff_raw:
        return ""
    # Strip trailing digits (e.g. employee IDs like 7130, 7126, 7164)
    s = re.sub(r"\d+$", "", staff_raw).strip()
    # Strip department suffix tags (e.g. -eee, -mat, -phy, -eng, -tamil, -cse, -aids)
    s = re.sub(r"[-_](?:eee|mat|math|maths|phy|physics|tamil|eng|english|cse|it|aids|ai_ds|mech|ece|civil|chem|chemistry)$", "", s, flags=re.IGNORECASE).strip()
    # Replace dots, underscores, or hyphens with spaces
    s = re.sub(r"[._\-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return " ".join(w.capitalize() for w in s.split())


def resolve_staff_for_code(code: str, staff_map: Dict[str, str]) -> str:
    """
    Resolves staff name for a given course code from staff_map with robust fuzzy/prefix/suffix matching.
    Handles matching '24CS201' <-> '24CS201A', 'CS201' <-> '24CS201', and multiple section teachers.
    """
    if not code or not staff_map:
        return ""
    code_raw = str(code).strip().upper()
    if code_raw in staff_map and staff_map[code_raw]:
        return str(staff_map[code_raw]).strip()

    code_clean = re.sub(r"[^A-Z0-9]", "", code_raw)
    for k, v in staff_map.items():
        if not v:
            continue
        k_clean = re.sub(r"[^A-Z0-9]", "", k.upper())
        if k_clean == code_clean:
            return str(v).strip()

    # Try matching base code (stripping trailing section letter e.g. 24CS201A -> 24CS201 or 24CS201 -> 24CS201A / 24CS201B)
    base_clean = re.sub(r"([A-Z])$", "", code_clean) if re.search(r"^(?:24)?[A-Z]{2,4}\d{3}[A-Z]$", code_clean) else code_clean
    matched_names = []
    for k, v in staff_map.items():
        if not v:
            continue
        k_clean = re.sub(r"[^A-Z0-9]", "", k.upper())
        k_base = re.sub(r"([A-Z])$", "", k_clean) if re.search(r"^(?:24)?[A-Z]{2,4}\d{3}[A-Z]$", k_clean) else k_clean
        if k_base == base_clean or k_clean == base_clean:
            for name_part in [x.strip() for x in str(v).split("&")]:
                if name_part and name_part not in matched_names:
                    matched_names.append(name_part)
    if matched_names:
        return " & ".join(matched_names)

    # Try regulation prefix stripping (e.g. 24CS201 vs CS201)
    no_reg_code = re.sub(r"^(?:24|21|22|23)", "", base_clean)
    for k, v in staff_map.items():
        if not v:
            continue
        k_clean = re.sub(r"[^A-Z0-9]", "", k.upper())
        k_no_reg = re.sub(r"^(?:24|21|22|23)", "", re.sub(r"([A-Z])$", "", k_clean))
        if k_no_reg == no_reg_code:
            for name_part in [x.strip() for x in str(v).split("&")]:
                if name_part and name_part not in matched_names:
                    matched_names.append(name_part)
    if matched_names:
        return " & ".join(matched_names)

    return ""


def parse_combined_ia_marks_content(
    raw_bytes: bytes, filename: str
) -> Tuple[Dict[str, Dict[str, Dict[str, Any]]], Dict[str, str], Dict[str, str], Dict[str, Dict[str, str]]]:
    """
    Parses multi-test / section-wise Cycle Test mark summary reports (e.g. II A IA.html, II B.html).
    Returns (multi_test_marks, course_titles, course_staff, student_metadata).
    multi_test_marks schema:
      {
        "ia1": { "813825243001": { "name": ..., "batch_no": ..., "quota": ..., "marks": { "24CS201A": "96" } } },
        "mut1": { ... },
        "ia2": { ... },
        "mut2": { ... },
        "ia3": { ... }
      }
    """
    import quopri
    from bs4 import BeautifulSoup

    test_marks: Dict[str, Dict[str, Dict[str, Any]]] = {
        "ia1": {}, "ia2": {}, "ia3": {}
    }
    course_titles: Dict[str, str] = {}
    course_staff: Dict[str, str] = {}
    student_meta: Dict[str, Dict[str, str]] = {}

    if not raw_bytes:
        return test_marks, course_titles, course_staff, student_meta

    if filename.lower().endswith(('.mhtml', '.mht')):
        try:
            decoded = quopri.decodestring(raw_bytes).decode('utf-8', errors='ignore')
        except Exception:
            decoded = raw_bytes.decode('utf-8', errors='ignore')
    else:
        decoded = raw_bytes.decode('utf-8', errors='ignore')

    if "<table" not in decoded.lower():
        return test_marks, course_titles, course_staff, student_meta

    try:
        soup = BeautifulSoup(decoded, 'html.parser')
        tables = soup.find_all('table')
        regno_pattern = re.compile(r"\b(\d{10,12})\b")

        for table in tables:
            rows = table.find_all('tr')
            if len(rows) < 3:
                continue

            r0 = rows[0]
            r1 = rows[1]

            r0_cells = r0.find_all(['th', 'td'])
            r1_cells = r1.find_all(['th', 'td'])
            r1_texts = [c.get_text().strip().upper() for c in r1_cells]

            # Check if this table has multi-test subheaders (A1, A2, A3, etc.)
            has_multi_tests = any(t in ("A1", "A2", "A3", "A4", "A5", "IAT1", "IAT2", "IAT3", "IA1", "IA2", "IA3", "CT1", "CT2", "CT3") for t in r1_texts)
            if not has_multi_tests:
                continue

            col_mappings = []
            fixed_col_count = 0
            cur_data_col = 0
            r1_idx = 0

            for c in r0_cells:
                txt = c.get_text().strip()
                rowspan = int(c.get('rowspan', 1) or 1)
                colspan = int(c.get('colspan', 1) or 1)

                if rowspan >= 2 or (colspan == 1 and not re.search(r"[A-Z]{2,4}\d{3}", txt.upper())):
                    fixed_col_count += 1
                    cur_data_col += 1
                else:
                    clean_txt = re.sub(r"\s+", " ", txt)
                    parts = clean_txt.split("-", 1)
                    code_raw = parts[0].strip().upper()
                    staff_raw = parts[1].strip() if len(parts) > 1 else ""

                    staff_clean = clean_staff_name(staff_raw)
                    m_code = re.search(r"((?:\d{2})?[A-Z]{2,4}[-_]?\d{3,4}[A-Z]?)", code_raw)
                    ccode = m_code.group(1).replace("-", "").replace("_", "") if m_code else code_raw
                    base_code = re.sub(r"([A-Z])$", "", ccode) if re.search(r"^(?:\d{2})?[A-Z]{2,4}\d{3,4}[A-Z]$", ccode) else ccode

                    if ccode and staff_clean:
                        course_staff[ccode] = staff_clean
                        if base_code != ccode:
                            if base_code not in course_staff:
                                course_staff[base_code] = staff_clean
                            else:
                                existing = [x.strip() for x in course_staff[base_code].split("&")]
                                if staff_clean not in existing:
                                    course_staff[base_code] = f"{course_staff[base_code]} & {staff_clean}"

                    for _ in range(colspan):
                        sub_test = r1_texts[r1_idx] if r1_idx < len(r1_texts) else ""
                        r1_idx += 1
                        col_mappings.append((cur_data_col, ccode, sub_test, staff_clean))
                        cur_data_col += 1

            # Determine fixed metadata column positions dynamically if present
            quota_col_idx = 4
            batch_col_idx = 2
            name_col_idx = 3
            cgpa_col_idx = 6

            for c_idx, c in enumerate(r0_cells):
                txt_lower = c.get_text().strip().lower()
                if "quota" in txt_lower:
                    quota_col_idx = c_idx
                elif "batch" in txt_lower:
                    batch_col_idx = c_idx
                elif "name" in txt_lower and "staff" not in txt_lower:
                    name_col_idx = c_idx
                elif "cgpa" in txt_lower or "gpa" in txt_lower:
                    cgpa_col_idx = c_idx

            for r in rows[2:]:
                cells = [c.get_text().strip() for c in r.find_all(['th', 'td'])]
                if len(cells) < fixed_col_count:
                    continue
                row_str = " ".join(cells[:max(7, fixed_col_count)])
                m_reg = regno_pattern.search(row_str)
                if not m_reg:
                    continue

                regno = m_reg.group(1)
                batchno = cells[batch_col_idx].strip() if batch_col_idx < len(cells) else (cells[2] if len(cells) > 2 else "")
                name = cells[name_col_idx].strip() if name_col_idx < len(cells) else (cells[3] if len(cells) > 3 else "")
                raw_q = cells[quota_col_idx].strip().upper() if quota_col_idx < len(cells) else (cells[4].strip().upper() if len(cells) > 4 else "")
                quota = raw_q if raw_q and raw_q not in ("&NBSP;", "-", "—", "NA", "N/A", "NONE") else "GQ"
                cgpa = cells[cgpa_col_idx].strip() if cgpa_col_idx < len(cells) else (cells[6].strip() if len(cells) > 6 else "")

                student_meta[regno] = {
                    "quota": quota,
                    "cgpa": cgpa,
                    "batch_no": batchno,
                    "name": name,
                }

                for (col_idx, ccode, sub_test, staff_name) in col_mappings:
                    if col_idx < len(cells):
                        mark_val = cells[col_idx].strip()
                        if mark_val and mark_val not in ("&nbsp;", "-", "—"):
                            test_key = None
                            sub_clean = re.sub(r"[\s\-_]+", "", sub_test).upper()
                            if sub_clean in ("A1", "IAT1", "IA1", "CT1", "TEST1", "1"):
                                test_key = "ia1"
                            elif sub_clean in ("A2", "IAT2", "IA2", "CT2", "TEST2", "2"):
                                test_key = "ia2"
                            elif sub_clean in ("A3", "IAT3", "IA3", "CT3", "TEST3", "3"):
                                test_key = "ia3"
                            elif sub_clean in ("A4", "A5", "MODEL", "MUT", "RETEST", "TEST4", "TEST5", "4", "5"):
                                test_key = "ia4"

                            if test_key:
                                test_marks.setdefault(test_key, {})
                                st_rec = test_marks[test_key].setdefault(regno, {
                                    "name": name,
                                    "batch_no": batchno,
                                    "quota": quota,
                                    "marks": {}
                                })
                                st_rec["marks"][ccode] = mark_val
                                base_c = re.sub(r"([A-Z])$", "", ccode) if re.search(r"^(?:\d{2})?[A-Z]{2,4}\d{3,4}[A-Z]$", ccode) else ccode
                                if base_c != ccode:
                                    st_rec["marks"][base_c] = mark_val

    except Exception:
        pass

    return test_marks, course_titles, course_staff, student_meta


def parse_ia_marks_content(
    raw_bytes: bytes, filename: str, target_test_key: str = "ia1"
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, str], Dict[str, str]]:
    """
    Parses Internal Assessment / Cycle Test mark sheets uploaded as .mhtml, .html, .xlsx, or .csv.
    Handles both single-assessment mark sheets and combined all-cycle-test reports.
    Returns (student_marks, course_titles, course_staff).
    """
    # 1. Try multi-test combined parsing first
    multi_marks, c_titles, c_staff, _ = parse_combined_ia_marks_content(raw_bytes, filename)
    if any(multi_marks.values()):
        # If target test exists in multi_marks, return that; otherwise return first non-empty
        if multi_marks.get(target_test_key):
            return multi_marks[target_test_key], c_titles, c_staff
        for t_k in ("ia1", "ia2", "ia3"):
            if multi_marks.get(t_k):
                return multi_marks[t_k], c_titles, c_staff

    # 2. Standard single-table parser for single-test sheets
    import quopri
    from bs4 import BeautifulSoup

    student_marks: Dict[str, Dict[str, Any]] = {}
    course_titles: Dict[str, str] = {}
    course_staff: Dict[str, str] = {}

    if not raw_bytes:
        return student_marks, course_titles, course_staff

    if filename.lower().endswith(('.mhtml', '.mht')):
        try:
            decoded = quopri.decodestring(raw_bytes).decode('utf-8', errors='ignore')
        except Exception:
            decoded = raw_bytes.decode('utf-8', errors='ignore')
    else:
        decoded = raw_bytes.decode('utf-8', errors='ignore')

    if "<table" in decoded.lower():
        try:
            soup = BeautifulSoup(decoded, 'html.parser')
            tables = soup.find_all('table')
            regno_pattern = re.compile(r"\b(\d{10,12})\b")

            for table in tables:
                rows = table.find_all('tr')
                if not rows:
                    continue

                header_cells = [re.sub(r"\s+", " ", c.get_text()).strip() for c in rows[0].find_all(['th', 'td'])]
                header_norm = [h.upper() for h in header_cells]

                # Check if course catalog table (Subcode | Title | Faculty Name)
                if any("SUBCODE" in h or "COURSE CODE" in h for h in header_norm) and any("TITLE" in h or "SUBJECT" in h for h in header_norm):
                    subcode_col = next((i for i, h in enumerate(header_norm) if "SUBCODE" in h or "CODE" in h), 0)
                    title_col = next((i for i, h in enumerate(header_norm) if "TITLE" in h or "SUBJECT" in h), 1)
                    faculty_col = next((i for i, h in enumerate(header_norm) if "FACULTY" in h or "STAFF" in h or "TEACHER" in h or "INSTRUCTOR" in h), -1)
                    for r in rows[1:]:
                        cells = [re.sub(r"\s+", " ", c.get_text()).strip() for c in r.find_all(['th', 'td'])]
                        if len(cells) > max(subcode_col, title_col):
                            code_tok = re.sub(r"[^A-Z0-9]", "", cells[subcode_col].upper())
                            if code_tok:
                                course_titles[code_tok] = cells[title_col]
                                if faculty_col >= 0 and len(cells) > faculty_col:
                                    st_clean = clean_staff_name(cells[faculty_col])
                                    course_staff[code_tok] = st_clean
                                    base_tok = re.sub(r"([A-Z])$", "", code_tok) if re.search(r"^(?:\d{2})?[A-Z]{2,4}\d{3,4}[A-Z]$", code_tok) else code_tok
                                    if base_tok != code_tok:
                                        course_staff[base_tok] = st_clean
                    continue

                # Main Marks Table
                if any("REGNO" in h or "REGISTER" in h for h in header_norm):
                    regno_col = next((i for i, h in enumerate(header_norm) if "REGNO" in h or "REGISTER" in h), 1)
                    batch_col = next((i for i, h in enumerate(header_norm) if "BATCH" in h), -1)
                    name_col = next((i for i, h in enumerate(header_norm) if "NAME" in h and "STAFF" not in h), -1)

                    code_cols: Dict[int, str] = {}
                    for col_idx, h_text in enumerate(header_cells):
                        clean_h = re.sub(r"\s+", "", h_text.upper())
                        m_code = re.search(r"((?:\d{2}[-_]?)?[A-Z]{2,4}[-_]?\d{3,4}[A-Z]?)", clean_h)
                        if m_code and clean_h not in ("TOTAL", "AVG", "ATT%", "S.NO", "REGNO", "BATCHNO", "NAME"):
                            code_cols[col_idx] = m_code.group(1).replace("-", "").replace("_", "")

                    for r in rows[1:]:
                        cells = [re.sub(r"\s+", " ", c.get_text()).strip() for c in r.find_all(['th', 'td'])]
                        if not cells:
                            continue
                        row_str = " ".join(cells)
                        m_reg = regno_pattern.search(row_str)
                        if not m_reg:
                            continue

                        regno = m_reg.group(1)
                        batch_no = cells[batch_col] if batch_col >= 0 and batch_col < len(cells) else ""
                        student_name = cells[name_col] if name_col >= 0 and name_col < len(cells) else ""

                        marks_map = {}
                        for col_idx, ccode in code_cols.items():
                            if col_idx < len(cells):
                                marks_map[ccode] = cells[col_idx]

                        student_marks[regno] = {
                            "batch_no": batch_no,
                            "name": student_name,
                            "marks": marks_map
                        }
        except Exception:
            pass

    return student_marks, course_titles, course_staff


def build_subject_mapping_log(records: List[StudentResultRecord]) -> List[Dict[str, Any]]:
    """
    Collapse per-cell PDF records into one row per unique subject_code, in first-seen
    order, with resolution provenance. This is what repeated page headers must NOT
    fan out into duplicate subjects -- dict keying on subject_code enforces that.
    """
    seen: Dict[str, Dict[str, Any]] = {}
    for r in records:
        code = r.subject_code or r.original_subject_text or r.subject_name
        if not code or code in seen:
            continue
        can_name, canon_code, cred, sem, cat, conf, ambiguous = resolve_subject_info(
            r.subject_code or r.original_subject_text
        )
        resolved_in_catalog = bool(canon_code) and canon_code.upper() in COURSE_CODE_INDEX
        if resolved_in_catalog and conf >= 0.999:
            method, resolution_confidence = "EXACT_COURSE_CODE", 1.0
        elif resolved_in_catalog:
            method, resolution_confidence = "ALIAS_OR_NAME_MATCH", conf
        elif ambiguous:
            method, resolution_confidence = "UNRESOLVED", 0.0
        else:
            method, resolution_confidence = "FUZZY_MATCH", conf
        seen[code] = {
            "course_code": r.subject_code or code,
            "official_subject_name": can_name if resolved_in_catalog else (r.subject_name or can_name),
            "semester": sem,
            "credits": cred,
            "course_type": (COURSE_CODE_INDEX.get(canon_code.upper(), {}) or {}).get("type", ""),
            "resolution_method": method,
            "resolution_confidence": resolution_confidence,
            "source_page": r.source_page,
            "unresolved": not resolved_in_catalog,
        }
    return list(seen.values())


def validate_export_dataset(ca: "ClassAnalysis") -> Tuple[bool, List[str]]:
    """
    Pre-download validation gate. Returns (ok, issues). When issues is non-empty the
    caller must refuse to generate the Excel file rather than ship something misleading.
    """
    issues: List[str] = []

    if not SESSION.get("staff_verified", False):
        issues.append("Extracted/entered staff names require manual verification. Please review the 'Subject Staff Mapping' table below and click 'Save Staff Names' to confirm.")

    if ca.student_count <= 0 or not ca.students:
        issues.append("Student count is zero — no students found to export.")
    if ca.subject_count <= 0 or not ca.subjects:
        issues.append("Subject count is zero — no subjects found to export.")

    codes_seen = [m["course_code"] for m in (ca.subject_mappings or [])]
    dupe_codes = {c for c in codes_seen if codes_seen.count(c) > 1}
    if dupe_codes:
        issues.append(f"Duplicate course codes detected in Subject Mapping: {sorted(dupe_codes)}")

    regnos = [s.regno for s in ca.students]
    dupe_regnos = {r for r in regnos if regnos.count(r) > 1}
    if dupe_regnos:
        issues.append(f"Duplicate register numbers detected: {sorted(dupe_regnos)}")

    unresolved = [m for m in (ca.subject_mappings or []) if m.get("unresolved")]
    # Unresolved subjects are not a hard failure -- they route to the Unresolved sheet --
    # but a subject with a blank name IS a hard failure, since that would render a blank
    # column header on the export.
    for m in (ca.subject_mappings or []):
        if not m.get("official_subject_name"):
            issues.append(f"Course code {m.get('course_code')} has no resolvable subject name.")

    # Index quarantined result cells by register number so a short-count student can be
    # explained precisely (source data malformed) rather than blamed on the parser.
    quarantine_by_regno: Dict[str, List[Dict[str, Any]]] = {}
    for q in (ca.quarantined_tokens or []):
        reg_key = str(q.get("register_number") or q.get("regno") or "").strip().upper()
        if reg_key:
            quarantine_by_regno.setdefault(reg_key, []).append(q)

    for s in ca.students:
        if s.name and s.name.strip() == s.regno.strip():
            issues.append(f"Student {s.regno} has name equal to register number — parser failure, not a real name.")
        if not s.name or not s.name.strip():
            issues.append(f"Student {s.regno} has no resolved name.")
        if ca.subject_count and s.total_courses and s.total_courses < ca.subject_count:
            reg_matches = quarantine_by_regno.get(s.regno.strip().upper(), [])
            if reg_matches:
                # The shortfall is explained by a quarantined (malformed/blank) source cell --
                # this is a source-data issue, not a parser/extraction defect.
                for q in reg_matches:
                    issues.append(
                        "EXPORT BLOCKED — malformed result cell requires correction:\n"
                        f"  Register No: {q.get('register_number', s.regno)}\n"
                        f"  Student: {q.get('student_name') or s.name}\n"
                        f"  Course: {q.get('course_code', 'Unknown')}\n"
                        f"  Raw Result: {q.get('raw_value', '')}\n"
                        f"  Reason: {q.get('classification', 'SOURCE_MALFORMED')}\n"
                        "  Required Action: Verify original COE/grade register before export."
                    )
            else:
                issues.append(
                    f"Student {s.regno} has only {s.total_courses} results but {ca.subject_count} subjects are "
                    "expected, and no matching quarantined cell explains the gap — requires manual review "
                    "(source data completeness issue, not a confirmed parser defect)."
                )

    expected_cells = ca.student_count * ca.subject_count if ca.student_count and ca.subject_count else 0
    for s in ca.students:
        if s.total_courses and ca.subject_count and s.total_courses > ca.subject_count:
            issues.append(f"Student {s.regno} has {s.total_courses} results but only {ca.subject_count} subjects exist.")
            break

    return (len(issues) == 0, issues)


def build_class_analysis_excel(ca: "ClassAnalysis") -> bytes:
    """
    Build the faculty-facing 'Download Class Analysis' workbook. Course code and
    resolved official subject name are always shown side by side; nothing here
    ever falls back to using the raw course code as the subject name.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="0F1B33", end_color="0F1B33", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, headers: List[str], rows: List[List[Any]]):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
        for r_idx, row in enumerate(rows, start=2):
            ws.append(row)
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = cell_border
                h_name = headers[c_idx - 1] if c_idx - 1 < len(headers) else ""
                
                # Alignments and number formats
                if isinstance(val, float):
                    if "%" in h_name or "pct" in h_name.lower() or "rate" in h_name.lower():
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(val, int) and "Register" not in h_name and "Reg" not in h_name and "S.No" not in h_name:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif "Register" in h_name or "S.No" in h_name or "Code" in h_name:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-fit column widths based on max content length
        for c_idx, h in enumerate(headers, start=1):
            max_len = len(str(h or ""))
            for r_idx in range(2, len(rows) + 2):
                cell_v = ws.cell(row=r_idx, column=c_idx).value
                if cell_v is not None:
                    max_len = max(max_len, len(str(cell_v)))
            width = max(12, min(55, max_len + 4))
            ws.column_dimensions[get_column_letter(c_idx)].width = width


    mappings = sorted(ca.subject_mappings or [], key=lambda m: m["course_code"])
    code_to_name = {m["course_code"]: m["official_subject_name"] for m in mappings}

    def col_label(subj: "SubjectAnalysis") -> str:
        name = code_to_name.get(subj.course_code, subj.subject)
        return f"{subj.course_code} - {name}" if subj.course_code else name

    # Sheet 1: Class Summary
    ws1 = wb.active
    ws1.title = "Class Summary"
    write_sheet(ws1, ["Metric", "Value"], [
        ["File", ca.file_name],
        ["Generated At", ca.generated_at],
        ["Students", ca.student_count],
        ["Subjects", ca.subject_count],
        ["Result Cells", ca.record_count],
        ["Class GPA", ca.class_gpa],
        ["Pass Rate %", ca.pass_rate],
        ["Arrear Students", ca.arrear_student_count],
        ["Multiple-U Students", ca.multiple_u_count],
        ["Malpractice Students", ca.malpractice_student_count],
    ])

    # Sheet 2: Student Results (Course Code - Subject Name columns, one per subject)
    ws2 = wb.create_sheet("Student Results")
    subj_cols = sorted(ca.subjects, key=lambda s: s.course_code or s.subject)
    headers2 = ["S.No", "Register Number", "Student Name"] + [col_label(s) for s in subj_cols] + \
               ["GPA", "Academic Arrears", "SA", "WD", "Malpractice", "Risk Level"]
    rows2 = []
    for i, s in enumerate(ca.students, start=1):
        grade_by_code = {c.course_code: c.grade for c in s.courses}
        row = [i, s.regno, s.name] + [grade_by_code.get(sc.course_code, "—") for sc in subj_cols] + \
              [s.gpa, s.arrear_count, s.sa_count, s.wd_count, s.malpractice_count, s.risk_level]
        rows2.append(row)
    write_sheet(ws2, headers2, rows2)

    # Sheet 3: Subject Analysis
    ws3 = wb.create_sheet("Subject Analysis")
    write_sheet(ws3, ["Course Code", "Subject Name", "Students", "Avg GP", "Pass %", "U", "RA", "Arrears"],
                [[s.course_code, code_to_name.get(s.course_code, s.subject), s.student_count,
                  s.avg_gp, s.pass_pct, s.u_count, s.ra_count, s.arrear_count] for s in subj_cols])

    # Sheet 4: Subject Mapping (faculty verification sheet)
    ws4 = wb.create_sheet("Subject Mapping")
    write_sheet(ws4, ["Course Code", "Official Subject Name", "Semester", "Credits",
                       "Course Type", "Resolution Method", "Confidence", "Source Page"],
                [[m["course_code"], m["official_subject_name"], m["semester"], m["credits"],
                  m["course_type"], m["resolution_method"], m["resolution_confidence"], m["source_page"]]
                 for m in mappings if not m.get("unresolved")])

    # Sheet 5: Unresolved / Quarantined Items
    ws5 = wb.create_sheet("Unresolved Subjects")
    unresolved_rows = [[m["course_code"], "Unknown", "Requires Faculty Review", m["source_page"]]
                        for m in mappings if m.get("unresolved")]
    write_sheet(ws5, ["Course Code", "Subject Name", "Resolution", "Source Page"], unresolved_rows)
    ws5b = wb.create_sheet("Quarantined Tokens")
    write_sheet(ws5b, ["Register No", "Student Name", "Course Code", "Raw Value", "Grade Point",
                        "Classification", "Source", "Source Row", "Reason"],
                [[q.get("register_number", q.get("regno", "")), q.get("student_name", ""),
                  q.get("course_code", q.get("column", "")), q.get("raw_value", ""),
                  q.get("grade_point", ""), q.get("classification", "UNKNOWN"),
                  q.get("source", ""), q.get("source_row", q.get("row", "")), q.get("reason", "")]
                 for q in (ca.quarantined_tokens or [])])

    # Sheet 4: Student Risk (reads StudentAnalysis.attention / risk_level / is_high_performer
    # exactly as already computed in compute_subject_analytics — no re-derivation here)
    ws_risk = wb.create_sheet("Student Risk")
    write_sheet(ws_risk, ["Register Number", "Student Name", "GPA", "Risk Level", "Attention Status",
                           "High Performer", "Attention Subjects"],
                [[s.regno, s.name, s.gpa, s.risk_level, s.attention, "YES" if s.is_high_performer else "NO",
                  ", ".join(s.attention_subjects)] for s in ca.students])

    # Sheet 5: Arrear & Backlog Tracker (reads StudentAnalysis fields directly)
    ws_arrear = wb.create_sheet("Arrear & Backlog Tracker")
    write_sheet(ws_arrear, ["Register Number", "Student Name", "Current Arrear Count", "Current Arrear Subjects",
                             "Carried Previous Arrear Count", "Carried Previous Arrears",
                             "Cleared in Reappearance", "Total Active Arrear Load"],
                [[s.regno, s.name, s.arrear_count, ", ".join(s.attention_subjects) if s.arrear_count else "None",
                  s.carried_previous_arrear_count if s.arrear_count > 0 else 0,
                  ", ".join(s.carried_previous_arrears) if (s.arrear_count > 0 and s.carried_previous_arrears) else "None",
                  ", ".join(s.cleared_previous_subjects) if s.cleared_previous_subjects else "None",
                  s.total_active_arrear_count if s.arrear_count > 0 else (s.carried_previous_arrear_count if s.has_backlog_arrears else 0)]
                 for s in ca.students if s.arrear_count > 0 or s.has_backlog_arrears])

    # Sheet 9: Source Provenance (reads StudentResultRecord.source_page /
    # extraction_confidence via subject_mappings — no re-derivation here)
    ws_prov = wb.create_sheet("Source Provenance")
    write_sheet(ws_prov, ["Course Code", "Official Subject Name", "Source Page", "Resolution Method", "Confidence"],
                [[m["course_code"], m["official_subject_name"], m["source_page"], m["resolution_method"],
                  m["resolution_confidence"]] for m in mappings])
    ws_prov.append([])
    ws_prov.append(["Document metadata"])
    for k, v in (ca.metadata or {}).items():
        ws_prov.append([k, v])

    # Sheet 6: Data Quality
    ok, issues = validate_export_dataset(ca)
    ws6 = wb.create_sheet("Data Quality")
    write_sheet(ws6, ["Check", "Status"], [
        ["Students verified", ca.student_count],
        ["Subjects verified", ca.subject_count],
        ["Result cells verified", ca.record_count],
        ["Unresolved subjects", sum(1 for m in mappings if m.get("unresolved"))],
        ["Validation passed", "YES" if ok else "NO"],
    ] + [["Issue", i] for i in issues])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =============================================================================
# 3.3) PDF -> DEPARTMENT OFFICIAL EXCEL CONVERTER (Analyses 1 / 2 / 3 / 5 / 6 / 7)
# =============================================================================

_DEPT_EXPANSION = {
    "DEPARTMENT OF AI & DS": "DEPARTMENT OF ARTIFICIAL INTELLIGENCE AND DATA SCIENCE",
}


def _is_theory_course(course_code: str, subject_name: str = "", credits: float = 3.0) -> bool:
    """
    Returns True if the course is a Theory subject, False if it is a Practical / Lab / Workshop course.
    """
    code_upper = str(course_code or "").upper().strip()
    name_upper = str(subject_name or "").upper().strip()

    # Keywords in name or code indicating practical / laboratory courses
    lab_keywords = [
        "LAB", "LABORATORY", "PRACTICAL", "PRACTICE", "WORKSHOP",
        "PROJECT", "SEMINAR", "INTERNSHIP", "STUDIO", "VIVA"
    ]
    if any(re.search(rf"\b{kw}\b", name_upper) for kw in lab_keywords):
        return False
    if "LAB" in code_upper or "PRAC" in code_upper or "WORKSHOP" in code_upper:
        return False

    # Check course code digit patterns (e.g., 24CS211, 24EM211, 24PH111, CS8381, GE8261)
    # Saranathan College Of Engineering / Autonomous regulations designate 1, 8, or 9 in tens place for practical/lab courses
    m = re.search(r"(\d{3,4})[A-Z]?$", code_upper)
    if m:
        num_str = m.group(1)
        if len(num_str) >= 3 and num_str[-2] in ('1', '8', '9'):
            return False

    # Check syllabus catalog category if available
    clean_code = re.sub(r"[^A-Z0-9]", "", code_upper)
    stripped = re.sub(r"[A-Z]$", "", clean_code)
    for cat_item in (COURSE_CODE_INDEX.get(clean_code), COURSE_CODE_INDEX.get(stripped)):
        if cat_item:
            cat_str = str(cat_item.get("category", "")).upper()
            if "LAB" in cat_str or "PRACTICAL" in cat_str or "PROJECT" in cat_str:
                return False
            break

    return True

def _dept_hod_label(department: str, programme: str) -> str:
    """
    Generate the department HOD signature label, e.g.:
    'HOD / AIDS', 'HOD / CSE', 'HOD / EEE', 'HOD / ECE', 'HOD / AIML', 'HOD / MECH', 'HOD / CIVIL', 'HOD / IT', 'HOD / CSBS', 'HOD / ICE'
    """
    text = f"{department} {programme}".upper()
    if "AIML" in text or "ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING" in text:
        return "HOD / AIML"
    if "ARTIFICIAL INTELLIGENCE" in text or "AI & DS" in text or "AI AND DS" in text or "AIDS" in text:
        return "HOD / AIDS"
    if "COMPUTER SCIENCE AND BUSINESS" in text or "CSBS" in text:
        return "HOD / CSBS"
    if "COMPUTER SCIENCE" in text or "CSE" in text:
        return "HOD / CSE"
    if "ELECTRONICS AND COMMUNICATION" in text or "ECE" in text:
        return "HOD / ECE"
    if "ELECTRICAL AND ELECTRONICS" in text or "EEE" in text:
        return "HOD / EEE"
    if "INSTRUMENTATION AND CONTROL" in text or "ICE" in text:
        return "HOD / ICE"
    if "MECHANICAL" in text or "MECH" in text:
        return "HOD / MECH"
    if "CIVIL" in text:
        return "HOD / CIVIL"
    if re.search(r"\bIT\b", text) or "INFORMATION TECHNOLOGY" in text:
        return "HOD / IT"
    words = re.findall(r"[A-Z]+", department.upper())
    initials = "".join(w[0] for w in words if w not in ("OF", "AND", "THE", "DEPARTMENT"))
    return f"HOD / {initials[:6] or 'DEPT'}"

def _dept_short_code(department: str, programme: str) -> str:
    text = f"{department} {programme}".upper()
    if "ARTIFICIAL INTELLIGENCE" in text or "AI & DS" in text or "AI AND DS" in text or "AIDS" in text:
        return "AI_DS"
    if "COMPUTER SCIENCE" in text or "CSE" in text:
        return "CSE"
    if "ELECTRONICS AND COMMUNICATION" in text or "ECE" in text:
        return "ECE"
    if "ELECTRICAL AND ELECTRONICS" in text or "EEE" in text:
        return "EEE"
    if "MECHANICAL" in text:
        return "MECH"
    if "CIVIL" in text:
        return "CIVIL"
    if re.search(r"\bIT\b", text) or "INFORMATION TECHNOLOGY" in text:
        return "IT"
    words = re.findall(r"[A-Z]+", department.upper())
    initials = "".join(w[0] for w in words if w not in ("OF", "AND", "THE", "DEPARTMENT"))
    return initials[:6] or "DEPT"


def _session_short_code(exam_session: str) -> str:
    s = (exam_session or "").upper()
    months = "JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC"
    m = re.search(rf"({months})\s*/?\s*({months})?", s)
    year_m = re.search(r"(20\d{2})", s)
    year = year_m.group(1) if year_m else datetime.now().strftime("%Y")
    if m and m.group(1):
        parts = [p for p in [m.group(1), m.group(2)] if p]
        return "_".join(parts) + f"_{year}"
    return f"RESULT_{year}"


def _semester_short_code(semester_text: Any) -> str:
    """
    Extracts a clean, normalized semester code (e.g. SEM_I, SEM_II, SEM_III, SEM_IV, etc.)
    derived from roman numerals, digit numbers, or textual representation.
    """
    s_str = str(semester_text or "").strip().upper()
    m = re.search(r"(?:SEMESTER\s*|SEM\s*)?([IVX]+|\d+)", s_str)
    if m:
        tok = m.group(1)
        int_to_roman = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII"}
        roman = int_to_roman.get(int(tok)) if tok.isdigit() and int(tok) in int_to_roman else tok
        return f"SEM_{roman}"
    return "SEM"


def department_excel_filename(pdf_report: "PDFExtractionReport") -> str:
    """
    Filename format: <dept>_<exam_duration>_<semester>_Result_Analysis.xlsx
    e.g. AI_DS_NOV_DEC_2025_SEM_II_Result_Analysis.xlsx
    """
    meta = pdf_report.doc_metadata
    dept_code = _dept_short_code(meta.department, meta.programme)
    session_code = _session_short_code(meta.exam_session)
    sem_code = _semester_short_code(meta.semester)
    fname = f"{dept_code}_{session_code}_{sem_code}_Result_Analysis.xlsx"
    return re.sub(r"[^\w\.\-]", "_", fname)


def _semester_parity(semester_text: Any) -> str:
    s_str = str(semester_text or "").strip().upper()
    m = re.search(r"([IVX]+|\d+)", s_str)
    if not m:
        return ""
    tok = m.group(1)
    roman = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8}
    n = roman.get(tok) or (int(tok) if tok.isdigit() else None)
    if n is None:
        return ""
    return "EVEN" if n % 2 == 0 else "ODD"


def _grade_gp_lookup(student: "StudentAnalysis") -> Dict[str, Tuple[str, float]]:
    return {c.course_code: (c.grade, c.points) for c in student.courses}


def build_department_excel(
    ca: "ClassAnalysis",
    pdf_report: "PDFExtractionReport",
    staff_map: Dict[str, str],
    source_filename: str = "",
    ia_marks_dir: Optional[Dict[str, Dict[str, Dict[str, Any]]]] = None,
    prev_ca: Optional["ClassAnalysis"] = None,
    prev_pdf_report: Optional["PDFExtractionReport"] = None,
    mentor_map: Optional[Dict[str, str]] = None,
    student_meta_map: Optional[Dict[str, Dict[str, str]]] = None,
) -> bytes:
    """
    Build the departmental official result-analysis workbook: Analysis 1_New, 2_New,
    3_New, 4_New, 5_New, 6_New, Analysis 7 -- in that order, styled to resemble the reference
    departmental workbook.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    meta = pdf_report.doc_metadata
    institution = ("SARANATHAN COLLEGE OF ENGINEERING, TIRUCHIRAPPALLI - 620 012"
                   if "SARANATHAN" in meta.institution.upper() else meta.institution.upper())
    dept_text = meta.department.upper()
    dept_line = _DEPT_EXPANSION.get(dept_text, dept_text if dept_text.startswith("DEPARTMENT") else f"DEPARTMENT OF {dept_text}")
    programme_line = f"COURSE: {meta.programme}"
    hod_label = _dept_hod_label(meta.department, meta.programme)
    parity = _semester_parity(meta.semester)
    ay_line = f"ACADEMIC YEAR: {meta.academic_year}" + (f" ({parity})" if parity else "")
    session_line = meta.exam_session

    mappings = list(ca.subject_mappings or [])  # first-seen order == PDF course order
    code_to_staff = {m["course_code"]: resolve_staff_for_code(m["course_code"], staff_map) for m in mappings}
    code_to_subj: Dict[str, "SubjectAnalysis"] = {s.course_code: s for s in ca.subjects}

    ia_data = ia_marks_dir or {}
    ia1_dict = ia_data.get("ia1", {})
    ia2_dict = ia_data.get("ia2", {})
    ia3_dict = ia_data.get("ia3", {})
    ia4_dict = ia_data.get("ia4", {}) or ia_data.get("model", {}) or ia_data.get("retest", {})

    THIN = Side(style="thin", color="9AA5B1")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TITLE_FONT = Font(bold=True, size=12, color="0F1B33")
    SUB_FONT = Font(bold=True, size=10, color="0F1B33")
    HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
    NOTE_FONT = Font(italic=True, size=8, color="64748B")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def title_block(ws, ncols: int, extra_lines: List[str], analysis_title: str) -> int:
        r = 1
        for line in [institution, dept_line, analysis_title] + extra_lines:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(ncols, 1))
            cell = ws.cell(row=r, column=1, value=line)
            cell.font = TITLE_FONT if r <= 2 else SUB_FONT
            cell.alignment = CENTER
            r += 1
        return r + 1  # blank spacer row before the header row

    def style_header_row(ws, row_idx: int, ncols: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_body_row(ws, row_idx: int, ncols: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    
    def add_hod_signature(ws, ncols: int, cur_row: int) -> int:
        sig_r = cur_row + 3
        cell = ws.cell(row=sig_r, column=ncols, value=hod_label)
        cell.font = Font(bold=True, size=10, color="0F1B33")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return sig_r
    def page_setup(ws, ncols: int, last_row: int, landscape: bool = True, freeze: str = "A1"):
        ws.page_setup.orientation = "landscape" if landscape else "portrait"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.oddHeader.center.text = f"&B{ws.title}"
        ws.oddFooter.center.text = "Page &P of &N"
        last_col = get_column_letter(max(ncols, 1))
        ws.print_area = f"A1:{last_col}{max(last_row, 1)}"
        ws.freeze_panes = freeze

    # ---------------------------------------------------------------
    # Analysis 1 - University Examination Result Analysis
    # ---------------------------------------------------------------
    ws1 = wb.create_sheet("Analysis 1_New")
    headers1 = ["S.NO", "SUBJECT CODE", "SUBJECT TITLE / STAFF NAME", "NO.OF STUDENTS REGISTERED",
                "NO. OF STUDENTS ABSENT", "NO.OF STUDENTS FAILED", "NO. OF STUDENTS WH",
                "NO. OF STUDENTS PASSED", "PASS % (Including ABSENT)"]
    ncols1 = len(headers1)
    start_row = title_block(ws1, ncols1, [programme_line, ay_line, session_line],
                             "ANALYSIS 1 - UNIVERSITY EXAMINATION RESULT ANALYSIS")
    for c, h in enumerate(headers1, start=1):
        ws1.cell(row=start_row, column=c, value=h)
    style_header_row(ws1, start_row, ncols1)
    r = start_row + 1
    for i, m in enumerate(mappings, start=1):
        code = m["course_code"]
        subj = code_to_subj.get(code)
        staff = code_to_staff.get(code, "")
        staff_line = f"Staff: {staff}" if staff else "(Staff name not entered)"
        ws1.cell(row=r, column=1, value=i)
        ws1.cell(row=r, column=2, value=code)
        ws1.cell(row=r, column=3, value=f"{m['official_subject_name']}\n{staff_line}")
        ws1.cell(row=r, column=4, value=subj.student_count if subj else 0)
        ws1.cell(row=r, column=5, value=subj.sa_count if subj else 0)
        ws1.cell(row=r, column=6, value=subj.arrear_count if subj else 0)
        ws1.cell(row=r, column=7, value=(subj.wh2_count + subj.mm_count) if subj else 0)
        ws1.cell(row=r, column=8, value=subj.pass_count if subj else 0)
        p_cell = ws1.cell(row=r, column=9, value=round(subj.pass_pct, 2) if subj and subj.pass_pct is not None else 0.0)
        p_cell.number_format = '0.00'
        style_body_row(ws1, r, ncols1)

        ws1.cell(row=r, column=3).alignment = LEFT_WRAP
        r += 1
    # ── Note row ────────────────────────────────────────────────────────────
    ws1.cell(row=r, column=1,
             value="Note: Absent-count is not separately tracked by COE PDF extraction; shown as 0 rather than fabricated.")
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols1)
    ws1.cell(row=r, column=1).font = NOTE_FONT

    # ── Overall Pass Percentage row (matches departmental format) ────────────
    r += 1
    total_registered = sum(
        (code_to_subj[m["course_code"]].student_count if code_to_subj.get(m["course_code"]) else 0)
        for m in mappings
    )
    total_passed = sum(
        (code_to_subj[m["course_code"]].pass_count if code_to_subj.get(m["course_code"]) else 0)
        for m in mappings
    )
    overall_pct = round((total_passed / total_registered * 100), 2) if total_registered > 0 else 0.0

    # Apply full border to all columns in the row first
    for col_idx in range(1, ncols1 + 1):
        cell = ws1.cell(row=r, column=col_idx)
        cell.border = BORDER
        cell.alignment = CENTER

    # Col 3: "Overall Pass Percentage" label — bold, left-aligned
    label_cell = ws1.cell(row=r, column=3, value="Overall Pass Percentage")
    label_cell.font = Font(bold=True, size=10, color="0F1B33")
    label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    label_cell.border = BORDER

    # Col 4: numeric value — bold, centered
    val_cell = ws1.cell(row=r, column=4, value=overall_pct)
    val_cell.number_format = "0.00"
    val_cell.font = Font(bold=True, size=10, color="0F1B33")
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.border = BORDER



    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 34
    for c in range(4, ncols1 + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 14
    sig_r1 = add_hod_signature(ws1, ncols1, r)
    page_setup(ws1, ncols1, sig_r1, landscape=True, freeze=f"A{start_row + 1}")

    # ---------------------------------------------------------------
    # Analysis 2 - Comparison of Previous Batch Results
    # ---------------------------------------------------------------
    ws2 = wb.create_sheet("Analysis 2_New")
    ncols2 = 6
    start_row = title_block(ws2, ncols2, [programme_line, ay_line, session_line],
                             "ANALYSIS 2 - COMPARISON OF PREVIOUS BATCH RESULTS")
    
    def calc_batch_stats(c_analysis: Optional["ClassAnalysis"]) -> Dict[str, Any]:
        if not c_analysis:
            return {
                "total": "N/A", "appeared": "N/A", "passed": "N/A", "failed": "N/A",
                "pass_pct": "N/A", "u1": "N/A", "u2": "N/A", "u3": "N/A", "u4": "N/A",
                "u5_plus": "N/A", "wh": "N/A"
            }
        st_count = c_analysis.student_count
        app_count = max(0, st_count - getattr(c_analysis, "sa_student_count", 0))
        pass_c = getattr(c_analysis, "cleared_count", 0)
        fail_c = getattr(c_analysis, "arrear_student_count", 0)
        p_pct = round(c_analysis.pass_rate, 2) if c_analysis.pass_rate is not None else 0.0
        
        u1 = sum(1 for s in c_analysis.students if (s.u_count + s.ra_count + s.ua_count) == 1)
        u2 = sum(1 for s in c_analysis.students if (s.u_count + s.ra_count + s.ua_count) == 2)
        u3 = sum(1 for s in c_analysis.students if (s.u_count + s.ra_count + s.ua_count) == 3)
        u4 = sum(1 for s in c_analysis.students if (s.u_count + s.ra_count + s.ua_count) == 4)
        u5_plus = sum(1 for s in c_analysis.students if (s.u_count + s.ra_count + s.ua_count) >= 5)
        wh = sum(1 for s in c_analysis.students if getattr(s, "wh2_count", 0) > 0 or getattr(s, "mm_count", 0) > 0)
        
        return {
            "total": st_count, "appeared": app_count, "passed": pass_c, "failed": fail_c,
            "pass_pct": p_pct, "u1": u1, "u2": u2, "u3": u3, "u4": u4,
            "u5_plus": u5_plus, "wh": wh
        }

    curr_stats = calc_batch_stats(ca)
    prev_stats = calc_batch_stats(prev_ca)

    r = start_row
    # Summary Header
    ws2.cell(row=r, column=2, value="BATCH: previous")
    ws2.cell(row=r, column=3, value="OVERALL METRICS")
    ws2.cell(row=r, column=4, value="BATCH: current")
    style_header_row(ws2, r, ncols2)
    r += 1

    metrics_rows = [
        ("Total Strength", "total"),
        ("No. of Students appeared", "appeared"),
        ("No. of Students passed", "passed"),
        ("No. of Students failed", "failed"),
        ("Percentage of Pass", "pass_pct"),
        ("One Arrear", "u1"),
        ("Two Arrear", "u2"),
        ("Three Arrear", "u3"),
        ("Four Arrear", "u4"),
        ("Five Arrear & More", "u5_plus"),
        ("Withheld", "wh"),
    ]

    for label, key in metrics_rows:
        pv = prev_stats[key]
        cv = curr_stats[key]
        ws2.cell(row=r, column=2, value=pv)
        ws2.cell(row=r, column=3, value=label)
        ws2.cell(row=r, column=4, value=cv)
        style_body_row(ws2, r, ncols2)
        
        if key == "pass_pct":
            if isinstance(pv, float): ws2.cell(row=r, column=2).number_format = '0.00'
            if isinstance(cv, float): ws2.cell(row=r, column=4).number_format = '0.00'
            
        ws2.cell(row=r, column=3).alignment = LEFT_WRAP
        ws2.cell(row=r, column=3).font = Font(bold=True, size=9)
        r += 1

    r += 1 # Spacer row

    # 2. Subject-by-Subject Comparison Block
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws2.cell(row=r, column=1, value="BATCH: previous")
    ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws2.cell(row=r, column=3, value="Batch current")
    style_header_row(ws2, r, ncols2)
    r += 1

    sub_headers = ["Pass %", "Subject Name*", "Sub. code", "Subject Name", "Staff Name", "Pass%"]
    for c_idx, sh in enumerate(sub_headers, start=1):
        ws2.cell(row=r, column=c_idx, value=sh)
    style_header_row(ws2, r, ncols2)
    r += 1

    prev_subj_map: Dict[str, Any] = {}
    if prev_ca and prev_ca.subjects:
        for ps in prev_ca.subjects:
            code_c = re.sub(r"[^A-Z0-9]", "", ps.course_code.upper())
            name_c = re.sub(r"[^A-Z0-9]", "", ps.subject.upper())
            prev_subj_map[ps.course_code.upper()] = ps
            prev_subj_map[code_c] = ps
            prev_subj_map[name_c] = ps

    for i, m in enumerate(mappings, start=1):
        c_code = m["course_code"]
        curr_s = code_to_subj.get(c_code)
        c_name = m["official_subject_name"]
        c_staff = code_to_staff.get(c_code, "") or "(Staff name not entered)"
        c_pass = round(curr_s.pass_pct, 2) if curr_s and curr_s.pass_pct is not None else 0.0

        p_pass = "N/A"
        p_name = "N/A"
        if prev_subj_map:
            c_code_clean = re.sub(r"[^A-Z0-9]", "", c_code.upper())
            c_name_clean = re.sub(r"[^A-Z0-9]", "", c_name.upper())
            matched_ps = prev_subj_map.get(c_code.upper()) or prev_subj_map.get(c_code_clean) or prev_subj_map.get(c_name_clean)
            if matched_ps:
                p_name = f"{matched_ps.course_code} - {matched_ps.subject}"
                p_pass = round(matched_ps.pass_pct, 2) if matched_ps.pass_pct is not None else 0.0

        p_pass_cell = ws2.cell(row=r, column=1, value=p_pass)
        if isinstance(p_pass, float): p_pass_cell.number_format = '0.00'
        ws2.cell(row=r, column=2, value=p_name)
        ws2.cell(row=r, column=3, value=c_code)
        ws2.cell(row=r, column=4, value=c_name)
        ws2.cell(row=r, column=5, value=c_staff)
        c_pass_cell = ws2.cell(row=r, column=6, value=c_pass)
        if isinstance(c_pass, float): c_pass_cell.number_format = '0.00'

        style_body_row(ws2, r, ncols2)
        ws2.cell(row=r, column=2).alignment = LEFT_WRAP
        ws2.cell(row=r, column=4).alignment = LEFT_WRAP
        ws2.cell(row=r, column=5).alignment = LEFT_WRAP
        r += 1

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 32
    ws2.column_dimensions["E"].width = 22
    ws2.column_dimensions["F"].width = 12
    sig_r2 = add_hod_signature(ws2, ncols2, r)
    page_setup(ws2, ncols2, sig_r2, landscape=True, freeze=f"C{start_row + 15}")

    # ---------------------------------------------------------------
    # Analysis 3 - Result Analysis of Failed Students
    # ---------------------------------------------------------------
    ws3 = wb.create_sheet("Analysis 3_New")
    # In official format, Analysis 3 columns display Theory subjects
    analysis3_mappings = [m for m in mappings if _is_theory_course(m["course_code"], m["official_subject_name"], m.get("credits", 3.0))]
    if not analysis3_mappings:
        analysis3_mappings = mappings

    k_subjs = len(analysis3_mappings)
    ncols3 = 4 + k_subjs + 1 + 2 + 1  # 4 fixed + k subjects + 1 (No. of Arrears) + 2 (Old Arrears) + 1 (Mentor)

    # Title Block (Matching Departmental Format)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols3)
    c1 = ws3.cell(row=1, column=1, value=institution)
    c1.font = TITLE_FONT
    c1.alignment = CENTER

    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols3)
    c2 = ws3.cell(row=2, column=1, value=dept_line)
    c2.font = TITLE_FONT
    c2.alignment = CENTER

    ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols3)
    c3 = ws3.cell(row=3, column=1, value="ANALYSIS 3 - RESULT ANALYSIS OF FAILED STUDENTS")
    c3.font = SUB_FONT
    c3.alignment = CENTER

    sem_num = 1
    sem_m = re.search(r"([IVX]+|\d+)", str(meta.semester or ""))
    if sem_m:
        tok = sem_m.group(1).upper()
        roman = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8}
        sem_num = roman.get(tok) or (int(tok) if tok.isdigit() else 1)
    year_num = (sem_num + 1) // 2
    roman_year = {1: "I", 2: "II", 3: "III", 4: "IV"}.get(year_num, "I")
    roman_sem = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII"}.get(sem_num, "I")

    right_col = max(ncols3 - 2, 5)
    ws3.cell(row=4, column=1, value=programme_line).font = Font(bold=True, size=9)
    batch_str = f"BATCH: {meta.academic_year[:4]} - {int(meta.academic_year[:4])+4}" if (meta.academic_year and len(meta.academic_year)>=4 and meta.academic_year[:4].isdigit()) else "BATCH: 2024 - 2028"
    ws3.cell(row=4, column=right_col, value=batch_str).font = Font(bold=True, size=9)

    ws3.cell(row=5, column=1, value=ay_line).font = Font(bold=True, size=9)
    ws3.cell(row=5, column=right_col, value=f"YEAR / SEM: {roman_year} / {roman_sem}").font = Font(bold=True, size=9)

    ws3.cell(row=6, column=1, value=session_line).font = Font(bold=True, size=9)
    ws3.cell(row=6, column=right_col, value=f"TOTAL STRENGTH: {ca.student_count}").font = Font(bold=True, size=9)

    # 2-Row Header Layout (Row 8 and Row 9)
    h1 = 8
    h2 = 9

    # S.No. (Col 1)
    ws3.merge_cells(start_row=h1, start_column=1, end_row=h2, end_column=1)
    ws3.cell(row=h1, column=1, value="S.No.")

    # Reg.No (Col 2)
    ws3.merge_cells(start_row=h1, start_column=2, end_row=h2, end_column=2)
    ws3.cell(row=h1, column=2, value="Reg.No")

    # Student Name (Col 3)
    ws3.merge_cells(start_row=h1, start_column=3, end_row=h2, end_column=3)
    ws3.cell(row=h1, column=3, value="Student Name")

    # Quota (Col 4)
    ws3.merge_cells(start_row=h1, start_column=4, end_row=h2, end_column=4)
    ws3.cell(row=h1, column=4, value="Quota")

    # Subjects (Cols 5 to 4+k_subjs)
    if k_subjs > 0:
        ws3.merge_cells(start_row=h1, start_column=5, end_row=h1, end_column=4 + k_subjs)
        ws3.cell(row=h1, column=5, value="Subjects")
        for j, m in enumerate(analysis3_mappings, start=5):
            ws3.cell(row=h2, column=j, value=m["course_code"])

    # No. of Arrears (Col 5+k_subjs)
    ws3.merge_cells(start_row=h1, start_column=5 + k_subjs, end_row=h2, end_column=5 + k_subjs)
    ws3.cell(row=h1, column=5 + k_subjs, value="No. of\nArrears")

    # OLD ARREARS (Cols 6+k_subjs to 7+k_subjs)
    ws3.merge_cells(start_row=h1, start_column=6 + k_subjs, end_row=h1, end_column=7 + k_subjs)
    ws3.cell(row=h1, column=6 + k_subjs, value="OLD ARREARS")
    ws3.cell(row=h2, column=6 + k_subjs, value="LIST OF SUBJECT")
    ws3.cell(row=h2, column=7 + k_subjs, value="NO.OF\nARREARS")

    # NAME OF THE MENTOR (Col 8+k_subjs)
    ws3.merge_cells(start_row=h1, start_column=8 + k_subjs, end_row=h2, end_column=8 + k_subjs)
    ws3.cell(row=h1, column=8 + k_subjs, value="NAME OF THE\nMENTOR")

    style_header_row(ws3, h1, ncols3)
    style_header_row(ws3, h2, ncols3)

    # Data Rows (For current semester failed students only)
    r = h2 + 1
    n_failed = 0
    mentor_lookup = mentor_map or {}

    for s in ca.students:
        # Include ONLY students who hold active arrears in the current semester
        if s.arrear_count <= 0:
            continue
        n_failed += 1
        gp_lookup = _grade_gp_lookup(s)

        raw_q3 = (
            (student_meta_map.get(s.regno, {}).get("quota") if student_meta_map else "") or
            s.meta.get("quota") or
            ia1_dict.get(s.regno, {}).get("quota") or
            ia2_dict.get(s.regno, {}).get("quota") or
            ia3_dict.get(s.regno, {}).get("quota") or
            ""
        )
        quota3 = raw_q3.strip().upper() if raw_q3 and raw_q3.strip().upper() not in ("NA", "N/A", "NONE", "") else "GQ"

        ws3.cell(row=r, column=1, value=n_failed)
        ws3.cell(row=r, column=2, value=s.regno)
        ws3.cell(row=r, column=3, value=s.name)
        ws3.cell(row=r, column=4, value=quota3)

        for j, m in enumerate(analysis3_mappings, start=5):
            grade, _pts = gp_lookup.get(m["course_code"], ("", 0))
            ws3.cell(row=r, column=j, value=grade if grade in ARREAR_GRADES else "")

        ws3.cell(row=r, column=5 + k_subjs, value=s.arrear_count)

        # OLD ARREARS: Populate carried previous semester failed subjects for this current arrear student
        prev_arrears = getattr(s, "carried_previous_arrears", [])
        if prev_arrears:
            subj_str = ",".join(prev_arrears)
            count_str = len(prev_arrears)
        else:
            subj_str = "-"
            count_str = "-"

        ws3.cell(row=r, column=6 + k_subjs, value=subj_str)
        ws3.cell(row=r, column=7 + k_subjs, value=count_str)

        # NAME OF THE MENTOR: Prefilled or blank for manual typing
        m_name = mentor_lookup.get(s.regno, "") or s.meta.get("mentor", "")
        ws3.cell(row=r, column=8 + k_subjs, value=m_name)

        style_body_row(ws3, r, ncols3)
        ws3.cell(row=r, column=3).alignment = LEFT_WRAP
        ws3.cell(row=r, column=6 + k_subjs).alignment = LEFT_WRAP
        ws3.cell(row=r, column=8 + k_subjs).alignment = LEFT_WRAP
        r += 1

    if n_failed == 0:
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols3)
        ws3.cell(row=r, column=1, value="No students with academic arrears (U / RA / UA) were found.")
        r += 1

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 28
    ws3.column_dimensions["D"].width = 8
    for c in range(5, 5 + k_subjs):
        ws3.column_dimensions[get_column_letter(c)].width = 11
    ws3.column_dimensions[get_column_letter(5 + k_subjs)].width = 11
    ws3.column_dimensions[get_column_letter(6 + k_subjs)].width = 30
    ws3.column_dimensions[get_column_letter(7 + k_subjs)].width = 11
    ws3.column_dimensions[get_column_letter(8 + k_subjs)].width = 24

    sig_r3 = add_hod_signature(ws3, ncols3, r)
    page_setup(ws3, ncols3, sig_r3, landscape=True, freeze=f"E{h2 + 1}")

    # ---------------------------------------------------------------
    # Analysis 4 - Internal Assessment Marks of Failed Students
    # ---------------------------------------------------------------
    ws4 = wb.create_sheet("Analysis 4_New")
    headers4 = [
        "S. No", "STAFF NAME", "SUB. CODE", "SUBJECT NAME", "NO.OF FAILURES",
        "REGISTER NUMBER", "NAME OF THE STUDENT", "QUOTA",
        "IAT 1", "MUT 1", "IAT 2", "MUT 2", "IAT 3"
    ]
    ncols4 = len(headers4)
    start_row = title_block(ws4, ncols4, [programme_line, ay_line, session_line, "MARK SCALE: OUT OF 100"],
                             "ANALYSIS 4 - INTERNAL ASSESSMENT MARKS OF FAILED STUDENTS")
    for c, h in enumerate(headers4, start=1):
        ws4.cell(row=start_row, column=c, value=h)
    style_header_row(ws4, start_row, ncols4)
    r = start_row + 1

    ia_data = ia_marks_dir or {}
    ia1_dict = ia_data.get("ia1", {})
    ia2_dict = ia_data.get("ia2", {})
    ia3_dict = ia_data.get("ia3", {})

    s_no = 0
    for s in ca.students:
        arrear_courses = [c for c in s.courses if c.grade in ARREAR_GRADES]
        if not arrear_courses:
            continue

        for c in arrear_courses:
            s_no += 1
            code_key = c.course_code.upper()
            code_clean = re.sub(r"[^A-Z0-9]", "", code_key)

            def is_lab_course(ccode: str, sub_name: str) -> bool:
                s_up = (sub_name or "").upper()
                c_up = (ccode or "").upper()
                if any(w in s_up for w in ("LAB", "LABORATORY", "PRACTICAL", "SEMINAR", "PROJECT", "STUDIO", "INTERNSHIP")):
                    return True
                if re.search(r"^[A-Z0-9]{2,4}\d1\d[A-Z]?$", c_up) or re.search(r"^\d{2}[A-Z]{2,4}\d1\d[A-Z]?$", c_up):
                    return True
                return False

            def get_mark(ia_map: Dict[str, Dict[str, Any]], reg: str, ccode: str, is_lab: bool = False) -> str:
                if not ia_map or reg not in ia_map:
                    return ""
                m_dict = ia_map[reg].get("marks", {})
                if not m_dict:
                    return ""
                if ccode in m_dict:
                    val = str(m_dict[ccode]).strip()
                    return "AB" if val.upper() in ("A", "AB", "ABS") else val
                c_clean = re.sub(r"[^A-Z0-9]", "", ccode.upper())
                c_base = re.sub(r"([A-Z])$", "", c_clean) if re.search(r"^(?:\d{2})?[A-Z]{2,4}\d{3,4}[A-Z]$", c_clean) else c_clean
                for k, v in m_dict.items():
                    k_clean = re.sub(r"[^A-Z0-9]", "", k.upper())
                    k_base = re.sub(r"([A-Z])$", "", k_clean) if re.search(r"^(?:\d{2})?[A-Z]{2,4}\d{3,4}[A-Z]$", k_clean) else k_clean
                    if k_clean == c_clean or k_base == c_base or k_clean == c_base or k_base == c_clean:
                        val = str(v).strip()
                        return "AB" if val.upper() in ("A", "AB", "ABS") else val
                return ""

            is_lab = is_lab_course(c.course_code, c.subject)
            mark_iat1 = get_mark(ia1_dict, s.regno, code_key, is_lab)
            mark_iat2 = get_mark(ia2_dict, s.regno, code_key, is_lab)
            mark_iat3 = get_mark(ia3_dict, s.regno, code_key, is_lab)
            # If IAT3 was not found, check if a model / retest / A5 mark was recorded
            if not mark_iat3 and ia4_dict:
                mark_iat3 = get_mark(ia4_dict, s.regno, code_key, is_lab)
            mark_mut = get_mark(ia4_dict, s.regno, code_key, is_lab) if ia4_dict else ""

            staff_name = code_to_staff.get(c.course_code, "") or resolve_staff_for_code(c.course_code, staff_map) or "(Staff name not entered)"
            subj_obj = code_to_subj.get(c.course_code)
            num_failures = subj_obj.arrear_count if subj_obj else 0
            raw_quota = (
                (student_meta_map.get(s.regno, {}).get("quota") if student_meta_map else "") or
                s.meta.get("quota") or
                ia1_dict.get(s.regno, {}).get("quota") or
                ia2_dict.get(s.regno, {}).get("quota") or
                ia3_dict.get(s.regno, {}).get("quota") or
                (ia4_dict.get(s.regno, {}).get("quota") if ia4_dict else "") or
                ""
            )
            quota = raw_quota.strip().upper() if raw_quota and raw_quota.strip().upper() not in ("NA", "N/A", "NONE", "") else "GQ"

            ws4.cell(row=r, column=1, value=s_no)
            ws4.cell(row=r, column=2, value=staff_name)
            ws4.cell(row=r, column=3, value=c.course_code)
            ws4.cell(row=r, column=4, value=c.subject)
            ws4.cell(row=r, column=5, value=num_failures)
            ws4.cell(row=r, column=6, value=s.regno)
            ws4.cell(row=r, column=7, value=s.name)
            ws4.cell(row=r, column=8, value=quota)
            ws4.cell(row=r, column=9, value=mark_iat1)
            ws4.cell(row=r, column=10, value="")
            ws4.cell(row=r, column=11, value=mark_iat2)
            ws4.cell(row=r, column=12, value="")
            ws4.cell(row=r, column=13, value=mark_iat3)
            if mark_mut and mark_mut != mark_iat3 and ncols4 >= 14:
                ws4.cell(row=r, column=14, value=mark_mut)

            style_body_row(ws4, r, ncols4)
            ws4.cell(row=r, column=2).alignment = LEFT_WRAP
            ws4.cell(row=r, column=4).alignment = LEFT_WRAP
            ws4.cell(row=r, column=7).alignment = LEFT_WRAP
            r += 1

    if s_no == 0:
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols4)
        ws4.cell(row=r, column=1, value="No students with academic arrears (U / RA) were found.")
        r += 1
    elif not ia1_dict and not ia2_dict and not ia3_dict:
        ws4.cell(row=r, column=1, value="Note: Internal Assessment mark sheets can be uploaded on the PDF-to-Excel page to populate IAT 1, IAT 2, and IAT 3 marks.")
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols4)
        ws4.cell(row=r, column=1).font = NOTE_FONT
        r += 1

    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 30
    ws4.column_dimensions["E"].width = 14
    ws4.column_dimensions["F"].width = 16
    ws4.column_dimensions["G"].width = 26
    ws4.column_dimensions["H"].width = 10
    ws4.column_dimensions["I"].width = 10
    ws4.column_dimensions["J"].width = 10
    ws4.column_dimensions["K"].width = 10
    ws4.column_dimensions["L"].width = 10
    ws4.column_dimensions["M"].width = 10
    sig_r4 = add_hod_signature(ws4, ncols4, r)
    page_setup(ws4, ncols4, sig_r4, landscape=True, freeze=f"F{start_row + 1}")


    # ---------------------------------------------------------------
    # Analysis 5 - List of Subject Toppers
    # ---------------------------------------------------------------
    ws5 = wb.create_sheet("Analysis 5_New")
    headers5 = ["S.No", "STAFF NAME", "SUBJECT CODE", "SUBJECT NAME", "NO.OF TOPPERS",
                "REGISTER NUMBER", "NAME OF THE STUDENT", "GRADE"]
    ncols5 = len(headers5)
    start_row = title_block(ws5, ncols5, [programme_line, ay_line, session_line],
                             "ANALYSIS 5 - LIST OF SUBJECT TOPPERS")
    for c, h in enumerate(headers5, start=1):
        ws5.cell(row=start_row, column=c, value=h)
    style_header_row(ws5, start_row, ncols5)
    r = start_row + 1

    # Filter strictly for Theory subjects (no laboratory / practical courses)
    theory_mappings = [m for m in mappings if _is_theory_course(m["course_code"], m["official_subject_name"], m.get("credits", 3.0))]
    if not theory_mappings:
        theory_mappings = mappings

    for i, m in enumerate(theory_mappings, start=1):
        code = m["course_code"]
        staff = code_to_staff.get(code, "")
        staff_disp = staff if staff else "(Staff name not entered)"
        entries = [(s.regno, s.name, c_.grade, c_.points)
                   for s in ca.students for c_ in s.courses
                   if c_.course_code == code and c_.grade in PASSING_GRADES]
        if not entries:
            ws5.cell(row=r, column=1, value=i)
            ws5.cell(row=r, column=2, value=staff_disp)
            ws5.cell(row=r, column=3, value=code)
            ws5.cell(row=r, column=4, value=m["official_subject_name"])
            ws5.cell(row=r, column=5, value=0)
            ws5.cell(row=r, column=6, value="—")
            ws5.cell(row=r, column=7, value="—")
            ws5.cell(row=r, column=8, value="—")
            style_body_row(ws5, r, ncols5)
            r += 1
            continue
        best_points = max(e[3] for e in entries)
        toppers = sorted([e for e in entries if e[3] == best_points], key=lambda e: e[0])
        top_start_row = r
        for idx, (regno, name, grade, points) in enumerate(toppers):
            if idx == 0:
                ws5.cell(row=r, column=1, value=i)
                ws5.cell(row=r, column=2, value=staff_disp)
                ws5.cell(row=r, column=3, value=code)
                ws5.cell(row=r, column=4, value=m["official_subject_name"])
                ws5.cell(row=r, column=5, value=len(toppers))
            ws5.cell(row=r, column=6, value=regno)
            ws5.cell(row=r, column=7, value=name)
            ws5.cell(row=r, column=8, value=grade)
            style_body_row(ws5, r, ncols5)
            r += 1
        if len(toppers) > 1:
            for colm in (1, 2, 3, 4, 5):
                ws5.merge_cells(start_row=top_start_row, start_column=colm, end_row=r - 1, end_column=colm)

    ws5.column_dimensions["A"].width = 6
    ws5.column_dimensions["B"].width = 22
    ws5.column_dimensions["C"].width = 14
    ws5.column_dimensions["D"].width = 34
    ws5.column_dimensions["E"].width = 12
    ws5.column_dimensions["F"].width = 16
    ws5.column_dimensions["G"].width = 26
    ws5.column_dimensions["H"].width = 10
    sig_r5 = add_hod_signature(ws5, ncols5, r)
    page_setup(ws5, ncols5, sig_r5, landscape=True, freeze=f"A{start_row + 1}")

    # ---------------------------------------------------------------
    # Analysis 6 - Rank List Based on GPA (reuses the existing ranking engine)
    # ---------------------------------------------------------------
    ws6 = wb.create_sheet("Analysis 6_New")
    headers6 = ["S. No.", "REGISTER NUMBER", "NAME OF THE STUDENT", "GPA", "RANK"]
    ncols6 = len(headers6)
    start_row = title_block(ws6, ncols6, [programme_line, ay_line, session_line],
                             "ANALYSIS 6 - RANK LIST BASED ON GPA")
    for c, h in enumerate(headers6, start=1):
        ws6.cell(row=start_row, column=c, value=h)
    style_header_row(ws6, start_row, ncols6)
    ranked = sorted([s for s in ca.students if s.rank], key=lambda s: s.rank)
    r = start_row + 1
    for i, s in enumerate(ranked, start=1):
        ws6.cell(row=r, column=1, value=i)
        ws6.cell(row=r, column=2, value=s.regno)
        ws6.cell(row=r, column=3, value=s.name)
        gpa_c = ws6.cell(row=r, column=4, value=s.gpa)
        if s.gpa is not None:
            gpa_c.number_format = '0.00'
        ws6.cell(row=r, column=5, value=s.rank)
        style_body_row(ws6, r, ncols6)
        r += 1

    ws6.column_dimensions["A"].width = 6
    ws6.column_dimensions["B"].width = 16
    ws6.column_dimensions["C"].width = 28
    ws6.column_dimensions["D"].width = 10
    ws6.column_dimensions["E"].width = 10
    sig_r6 = add_hod_signature(ws6, ncols6, r)
    page_setup(ws6, ncols6, sig_r6, landscape=False, freeze=f"A{start_row + 1}")

    # ---------------------------------------------------------------
    # Analysis 7 - Provisional Results (raw, per-student per-subject)
    # ---------------------------------------------------------------
    ws7 = wb.create_sheet("Analysis 7")
    ncols7 = 3 + len(mappings) * 2
    start_row = title_block(ws7, ncols7, [programme_line, ay_line, session_line],
                             "ANALYSIS 7 - PROVISIONAL RESULTS")
    hdr_top, hdr_bot = start_row, start_row + 1
    ws7.merge_cells(start_row=hdr_top, start_column=1, end_row=hdr_bot, end_column=1)
    ws7.cell(row=hdr_top, column=1, value="S.No")
    ws7.merge_cells(start_row=hdr_top, start_column=2, end_row=hdr_bot, end_column=2)
    ws7.cell(row=hdr_top, column=2, value="Reg. No")
    ws7.merge_cells(start_row=hdr_top, start_column=3, end_row=hdr_bot, end_column=3)
    ws7.cell(row=hdr_top, column=3, value="Name of the Student")
    col = 4
    for m in mappings:
        ws7.merge_cells(start_row=hdr_top, start_column=col, end_row=hdr_top, end_column=col + 1)
        ws7.cell(row=hdr_top, column=col, value=f"{m['course_code']}\n{m['official_subject_name']}\n({m['credits']} Cr)")
        ws7.cell(row=hdr_bot, column=col, value="Grade")
        ws7.cell(row=hdr_bot, column=col + 1, value="GP")
        col += 2
    style_header_row(ws7, hdr_top, ncols7)
    style_header_row(ws7, hdr_bot, ncols7)
    ws7.row_dimensions[hdr_top].height = 42
    r = hdr_bot + 1
    for i, s in enumerate(ca.students, start=1):
        gp_lookup = _grade_gp_lookup(s)
        ws7.cell(row=r, column=1, value=i)
        ws7.cell(row=r, column=2, value=s.regno)
        ws7.cell(row=r, column=3, value=s.name)
        col = 4
        for m in mappings:
            grade, points = gp_lookup.get(m["course_code"], ("—", ""))
            ws7.cell(row=r, column=col, value=grade)
            ws7.cell(row=r, column=col + 1, value=points if points != "" else "")
            col += 2
        style_body_row(ws7, r, ncols7)
        r += 1

    # 1. Identify student row range
    stud_start = hdr_bot + 1
    stud_end = r - 1

    # 2. Write the 17 summary rows starting at row `r`
    # Row 1 (r): TOTAL REGISTRED
    ws7.cell(row=r, column=1, value="TOTAL REGISTRED")
    ws7.cell(row=r, column=3, value="TOTAL REGISTRED")
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        ws7.cell(row=r, column=col, value=f'=COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "<>—")')
        col += 2

    # Row 2 (r+1): TOTAL APPEARED
    ws7.cell(row=r+1, column=3, value="TOTAL APPEARED")
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        ws7.cell(row=r+1, column=col, value=f'={col_ltr}{r}-{col_ltr}{r+3}')
        col += 2

    # Row 3 (r+2): No. of Pass
    ws7.cell(row=r+2, column=3, value="No. of Pass")
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        # Sum of COUNTIFs for passing grades O, A+, A, B+, B, C
        pass_formula = " + ".join(f'COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "{g}")' for g in PASS_GRADE_ORDER)
        ws7.cell(row=r+2, column=col, value=f'={pass_formula}')
        col += 2

    # Row 4 (r+3): No. Of Absent
    ws7.cell(row=r+3, column=3, value="No. Of Absent")
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        ws7.cell(row=r+3, column=col, value=f'=COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "UA")')
        col += 2

    # Row 5 (r+4): No. Of Fail
    ws7.cell(row=r+4, column=3, value="No. Of Fail")
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        ws7.cell(row=r+4, column=col, value=f'=COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "RA") + COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "U")')
        col += 2

    # Row 6 (r+5): % of Pass based on Registered
    ws7.cell(row=r+5, column=1, value="Registred")
    ws7.cell(row=r+5, column=3, value="% of Pass")
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        ws7.cell(row=r+5, column=col, value=f'={col_ltr}{r+2}/{col_ltr}{r}*100')
        col += 2

    # Rows 7-15: Grade counts
    grade_rows = [
        ("O Grade", "O"),
        ("A+ Grade", "A+"),
        ("A Grade", "A"),
        ("B+ Grade", "B+"),
        ("B Grade", "B"),
        ("C Grade", "C"),
        ("RA", "RA"),
        ("UA", "UA"),
        ("WH", "WH2")  # WH in provisional results is mapped to WH2 or MM
    ]
    for idx, (label, g_token) in enumerate(grade_rows):
        row_offset = 6 + idx
        ws7.cell(row=r+row_offset, column=3, value=label)
        col = 4
        for m in mappings:
            col_ltr = get_column_letter(col)
            if g_token == "WH2":
                # WH counts both WH2 and MM
                ws7.cell(row=r+row_offset, column=col, value=f'=COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "WH2") + COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "MM")')
            else:
                ws7.cell(row=r+row_offset, column=col, value=f'=COUNTIF({col_ltr}{stud_start}:{col_ltr}{stud_end}, "{g_token}")')
            col += 2

    # Row 16 (r+15): SUM
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        ws7.cell(row=r+15, column=col, value=f'=SUM({col_ltr}{r+6}:{col_ltr}{r+14})')
        col += 2

    # Row 17 (r+16): % of Pass based on Appeared
    ws7.cell(row=r+16, column=1, value="Appeared")
    ws7.cell(row=r+16, column=3, value="% of Pass")
    col = 4
    for m in mappings:
        col_ltr = get_column_letter(col)
        ws7.cell(row=r+16, column=col, value=f'={col_ltr}{r+2}/{col_ltr}{r+1}*100')
        col += 2

    # Style all summary rows
    summary_font = Font(bold=True, size=9)
    for row_offset in range(17):
        row_idx = r + row_offset
        style_body_row(ws7, row_idx, ncols7)
        # Bold column A, B, C labels
        for c in (1, 2, 3):
            cell = ws7.cell(row=row_idx, column=c)
            cell.font = summary_font
            if c == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # Bold and align formulas
        for c in range(4, ncols7 + 1):
            cell = ws7.cell(row=row_idx, column=c)
            cell.font = summary_font
            # Format percentage cells
            if row_offset in (5, 16) and (c % 2 == 0):
                cell.number_format = '0.00'

    ws7.column_dimensions["A"].width = 16  # Widened to fit "TOTAL REGISTRED"
    ws7.column_dimensions["B"].width = 16
    ws7.column_dimensions["C"].width = 26
    for c in range(4, ncols7 + 1):
        ws7.column_dimensions[get_column_letter(c)].width = 9

    # ── Arrear count distribution table at bottom of Analysis 7 ─────────────
    r_val_7 = r + 18
    arr_1 = sum(1 for s in ca.students if s.arrear_count == 1)
    arr_2 = sum(1 for s in ca.students if s.arrear_count == 2)
    arr_3 = sum(1 for s in ca.students if s.arrear_count == 3)
    arr_4 = sum(1 for s in ca.students if s.arrear_count == 4)
    arr_5_more = sum(1 for s in ca.students if s.arrear_count >= 5)
    all_pass = sum(1 for s in ca.students if s.arrear_count == 0)
    total_students = len(ca.students)

    arrear_distribution = [
        ("1 Arr", arr_1),
        ("2 Arr", arr_2),
        ("3 Arr", arr_3),
        ("4 Arr", arr_4),
        ("5 and MORE", arr_5_more),
        ("All Pass", all_pass),
        ("Total", total_students),
    ]

    for label, count_val in arrear_distribution:
        style_body_row(ws7, r_val_7, ncols7)

        # Bold and left-align the label in Column 3
        cell_key = ws7.cell(row=r_val_7, column=3, value=label)
        cell_key.font = summary_font
        cell_key.alignment = Alignment(horizontal="left", vertical="center")
        cell_key.border = BORDER

        # Bold and center-align the value in Column 4
        cell_val = ws7.cell(row=r_val_7, column=4, value=count_val)
        cell_val.font = summary_font
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.border = BORDER

        # Style other cells in that row to have summary font & borders
        for c in range(1, ncols7 + 1):
            if c not in (3, 4):
                cell = ws7.cell(row=r_val_7, column=c)
                cell.font = summary_font
                cell.border = BORDER

        r_val_7 += 1

    last_row_7 = r_val_7 - 1

    sig_r7 = add_hod_signature(ws7, ncols7, last_row_7)
    page_setup(ws7, ncols7, sig_r7, landscape=True, freeze=f"D{hdr_bot + 1}")

    # ---------------------------------------------------------------
    # Hidden metadata sheet - PDF provenance (section 19)
    # ---------------------------------------------------------------
    ok, _issues = validate_export_dataset(ca)
    ws_meta = wb.create_sheet("Source Metadata")
    ws_meta.append(["Field", "Value"])
    for c in range(1, 3):
        cell = ws_meta.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in [
        ["Source File", source_filename or pdf_report.doc_metadata.institution],
        ["Source Type", "COE PDF"],
        ["Source Pages", meta.page_count],
        ["Extraction Timestamp", datetime.now().isoformat(timespec="seconds")],
        ["Students", ca.student_count],
        ["Subjects", ca.subject_count],
        ["Validation Status", "PASSED" if ok else "ISSUES FOUND"],
    ]:
        ws_meta.append(row)
    ws_meta.column_dimensions["A"].width = 22
    ws_meta.column_dimensions["B"].width = 44
    ws_meta.sheet_state = "hidden"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def reconcile_pdf_and_excel(
    pdf_records: List[StudentResultRecord],
    excel_records: Any
) -> ReconciliationReport:
    """
    Reconcile COE PDF extraction against uploaded Excel dataset.
    Matches by (register_number, subject_code/canonical_subject) and checks result_status.
    """
    report = ReconciliationReport()
    pdf_map = {(r.register_number.upper(), r.subject_name.strip().lower()): r for r in pdf_records}

    excel_map = {}
    if isinstance(excel_records, pd.DataFrame):
        for _, row in excel_records.iterrows():
            reg = str(row["regno"]).strip().upper()
            subj = str(row["subject"]).strip().lower()
            excel_map[(reg, subj)] = row
    elif isinstance(excel_records, list):
        for r in excel_records:
            if hasattr(r, "regno"):
                reg = str(r.regno).strip().upper()
                subj = str(r.subject).strip().lower()
                excel_map[(reg, subj)] = r
            elif isinstance(r, dict):
                reg = str(r.get("regno", "")).strip().upper()
                subj = str(r.get("subject", "")).strip().lower()
                excel_map[(reg, subj)] = r

    all_keys = set(pdf_map.keys()) | set(excel_map.keys())

    for key in all_keys:
        pdf_rec = pdf_map.get(key)
        excel_item = excel_map.get(key)

        if pdf_rec is not None and excel_item is not None:
            excel_grade = excel_item["grade"] if isinstance(excel_item, (pd.Series, dict)) else excel_item.grade
            excel_name = excel_item["name"] if isinstance(excel_item, (pd.Series, dict)) else excel_item.name
            if pdf_rec.result_status == excel_grade:
                report.matched_count += 1
            else:
                report.mismatched_count += 1
                report.mismatched_records.append({
                    "register_number": pdf_rec.register_number,
                    "student_name": pdf_rec.student_name or excel_name,
                    "subject": pdf_rec.subject_name,
                    "pdf_grade": pdf_rec.result_status,
                    "excel_grade": excel_grade,
                    "status": "MISMATCH"
                })
        elif pdf_rec is not None and excel_item is None:
            report.missing_in_excel.append({
                "register_number": pdf_rec.register_number,
                "student_name": pdf_rec.student_name,
                "subject": pdf_rec.subject_name,
                "pdf_grade": pdf_rec.result_status,
            })
        elif excel_item is not None and pdf_rec is None:
            excel_reg = excel_item["regno"] if isinstance(excel_item, (pd.Series, dict)) else excel_item.regno
            excel_name = excel_item["name"] if isinstance(excel_item, (pd.Series, dict)) else excel_item.name
            excel_subj = excel_item["subject"] if isinstance(excel_item, (pd.Series, dict)) else excel_item.subject
            excel_grade = excel_item["grade"] if isinstance(excel_item, (pd.Series, dict)) else excel_item.grade
            report.missing_in_pdf.append({
                "register_number": excel_reg,
                "student_name": excel_name,
                "subject": excel_subj,
                "excel_grade": excel_grade,
            })

    return report


def pdf_records_to_dataframe(pdf_records: List[StudentResultRecord]) -> pd.DataFrame:
    """Convert StudentResultRecord list into normalized pandas DataFrame for compute_class_analysis.
    Generic over source_type ("PDF" | "XLS" | "XLSX") -- this is the single conversion function
    the analytics engine consumes from, regardless of which ingestion path produced the records."""
    rows = []
    for idx, r in enumerate(pdf_records):
        rows.append({
            "regno": r.register_number,
            "name": r.student_name,
            "subject": r.subject_name,
            "course_code": r.subject_code,
            "credits": r.credits if (r.credits and r.credits > 0) else 3.0,
            "grade": r.result_status,
            "points": GRADE_POINTS.get(r.result_status, 0.0),
            "src_row": r.source_page,
            "source_type": r.source_type,
            "batch": r.batch_number,
            "source_gp": r.source_gp,
            "data_quality_issue": r.data_quality_issue,
        })
    return pd.DataFrame(rows)


# Alias: the analytics engine deliberately has ONE ingestion-agnostic conversion function.
# This name is provided so PDF/XLS/XLSX call sites can each read as intent-revealing without
# implying "PDF-only" or duplicating the conversion logic per source type.
normalized_records_to_dataframe = pdf_records_to_dataframe


# =============================================================================
# 3.3) LEGACY XLS / MULTI-SOURCE NORMALIZATION LAYER
# =============================================================================
#
#   PDF ---\
#   XLS  ---+---> extract_*(...) -> List[StudentResultRecord] (+ quarantined_tokens)
#   XLSX ---/                              |
#                                          v
#                          normalized_records_to_dataframe()
#                                          |
#                                          v
#                              compute_class_analysis()   <-- UNCHANGED analytics engine
#                                          |
#                          Dashboard / PDF report / Excel export
#
# Every ingestion path produces the SAME StudentResultRecord shape and feeds the SAME
# compute_class_analysis(). No analytics/GPA/ranking logic is duplicated per source type.

_COURSE_CODE_RE = re.compile(r"^\d{2}[A-Z]{2,4}\d{3}$")

_SUMMARY_ROW_KEYWORDS = (
    "TOTAL", "REGISTRED", "REGISTERED", "APPEARED", "NO. OF PASS", "NO OF PASS",
    "NO. OF ABSENT", "NO OF ABSENT", "% OF PASS", "PASS %", "AVERAGE", "MEDIAN",
    "SIGNATURE", "PRINCIPAL", "CONTROLLER", "HOD", "CLASS ADVISOR", "PREPARED BY",
    "VERIFIED BY", "APPROVED BY", "1 ARR", "2 ARR", "3 ARR", "4 ARR", "ALL PASS",
)


@dataclass
class LegacyXLSExtractionReport:
    ok: bool = True
    fatal_error: str = ""
    warnings: List[str] = field(default_factory=list)
    reader_method: str = ""           # "xlrd" | "libreoffice" | ""
    sheet_name: str = ""
    header_row: int = -1
    records: List[StudentResultRecord] = field(default_factory=list)
    quarantined_tokens: List[Dict[str, Any]] = field(default_factory=list)
    gp_mismatches: List[Dict[str, Any]] = field(default_factory=list)
    subject_descriptors: List[Dict[str, Any]] = field(default_factory=list)
    student_count: int = 0
    subject_count: int = 0
    expected_cell_count: int = 0
    valid_cell_count: int = 0
    malformed_cell_count: int = 0
    blank_cell_count: int = 0
    unresolved_subject_count: int = 0
    unknown_token_count: int = 0
    duplicate_key_count: int = 0
    overall_confidence: float = 1.0


def _normalize_course_code_token(token: str) -> str:
    """'24-AD-401' / '24 AD 401' -> '24AD401'. Never semantically reorders the code."""
    return re.sub(r"[\s\-\./]", "", str(token or "")).upper()


def _looks_like_course_code(token: str) -> bool:
    return bool(_COURSE_CODE_RE.match(_normalize_course_code_token(token)))


def _is_summary_row(cells: List[str]) -> bool:
    joined = " ".join(str(c).strip().upper() for c in cells if str(c).strip())
    return any(kw in joined for kw in _SUMMARY_ROW_KEYWORDS)


def _open_legacy_workbook_sheets(data: bytes) -> Tuple[Dict[str, List[List[str]]], str]:
    """
    Open a legacy .xls (BIFF/OLE) workbook and return {sheet_name: rows} with every
    cell stringified. Preferred order:
      1. xlrd (pure python) if importable -- no external process required.
      2. LibreOffice headless conversion, if the 'soffice'/'libreoffice' binary exists.
      3. Raise a clear, actionable error -- never silently degrade to guessing.
    """
    # 1. Pure-python reader, if available on this host.
    try:
        import xlrd  # type: ignore
        book = xlrd.open_workbook(file_contents=data)
        sheets: Dict[str, List[List[str]]] = {}
        for sheet in book.sheets():
            rows = []
            for r in range(sheet.nrows):
                rows.append([str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)])
            sheets[sheet.name] = rows
        return sheets, "xlrd"
    except ImportError:
        pass
    except Exception as e:
        raise RuntimeError(f"xlrd could not open the workbook: {e}")

    # 2. LibreOffice headless conversion fallback.
    import shutil
    import subprocess
    import tempfile

    soffice_bin = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice_bin:
        raise RuntimeError(
            "Cannot read legacy .xls: no Python .xls reader (xlrd) is installed and no "
            "LibreOffice ('soffice'/'libreoffice') binary was found on this host. "
            "Install xlrd, install LibreOffice, or re-save the file as .xlsx."
        )

    with tempfile.TemporaryDirectory() as tmpdir:
        src_path = os.path.join(tmpdir, "workbook.xls")
        with open(src_path, "wb") as f:
            f.write(data)
        try:
            subprocess.run(
                [
                    soffice_bin, "--headless", "--convert-to",
                    'csv:"Text - txt - csv (StarCalc)":44,34,0,1,,,,,,,,-1',
                    "--outdir", tmpdir, src_path,
                ],
                check=True, capture_output=True, timeout=90,
            )
        except subprocess.TimeoutExpired:
            raise RuntimeError("LibreOffice conversion of the .xls file timed out.")
        except subprocess.CalledProcessError as e:
            raise RuntimeError(f"LibreOffice could not convert the .xls file: {e.stderr!r}")

        sheets = {}
        for fname in sorted(os.listdir(tmpdir)):
            if not fname.lower().endswith(".csv"):
                continue
            sheet_name = fname[len("workbook-"):-4] if fname.startswith("workbook-") else fname[:-4]
            with open(os.path.join(tmpdir, fname), newline="", encoding="utf-8", errors="replace") as f:
                rows = [row for row in csv.reader(f)]
            sheets[sheet_name] = rows
        if not sheets:
            raise RuntimeError("LibreOffice produced no readable sheets from this .xls file.")
        return sheets, "libreoffice"


def _select_best_result_sheet(sheets: Dict[str, List[List[str]]]) -> Tuple[str, List[List[str]], int, int]:
    """
    Score every sheet+row for how much it looks like a per-student x per-subject grade
    matrix (a row with several course-code-shaped header cells), and return the best
    (sheet_name, rows, header_row_idx, course_code_hits).
    """
    best = ("", [], -1, 0)
    for sheet_name, rows in sheets.items():
        limit = min(20, len(rows))
        for idx in range(limit):
            row = rows[idx]
            hits = sum(1 for cell in row if _looks_like_course_code(cell))
            # Prefer catalog-known codes even more strongly, but don't require them --
            # another department's catalog may not be loaded.
            catalog_hits = sum(
                1 for cell in row
                if _normalize_course_code_token(cell) in COURSE_CODE_INDEX
            )
            score = hits + catalog_hits * 2
            if score > best[3]:
                best = (sheet_name, rows, idx, score)
    return best


def _detect_subject_descriptors(rows: List[List[str]], header_row_idx: int) -> List[Dict[str, Any]]:
    """
    From the course-code header row, build subject descriptors {course_code, grade_col,
    gp_col}. GP columns are detected structurally (the column immediately to the right
    of a course-code cell, when that column's header is blank -- the standard COE
    merged-header layout) and verified against a sample of the data rows actually
    looking GP-like (small numbers 0-10 or blank), never assumed.
    """
    header_row = rows[header_row_idx]
    sample_rows = rows[header_row_idx + 1: header_row_idx + 8]
    descriptors = []
    for c, cell in enumerate(header_row):
        if not _looks_like_course_code(cell):
            continue
        code = _normalize_course_code_token(cell)
        gp_col: Optional[int] = None
        if c + 1 < len(header_row) and not str(header_row[c + 1]).strip():
            candidate = c + 1
            gp_like = 0
            total = 0
            for r in sample_rows:
                if candidate >= len(r):
                    continue
                v = str(r[candidate]).strip()
                if not v:
                    continue
                total += 1
                try:
                    fv = float(v)
                    if 0 <= fv <= 10:
                        gp_like += 1
                except ValueError:
                    pass
            if total == 0 or gp_like / total >= 0.6:
                gp_col = candidate
        descriptors.append({"course_code": code, "raw_header": cell, "grade_col": c, "gp_col": gp_col})
    return descriptors


def _detect_id_columns_and_student_rows(
    rows: List[List[str]], header_row_idx: int, subject_cols: set
) -> Dict[str, Any]:
    """
    Find S.No / Register No / Name columns (scanning the header row and the row(s)
    immediately below it, since COE sheets commonly have a 2-row header: course codes
    on one row, {S.No, Reg. No, Name, <short subject label>...} on the next), then
    find where the actual student rows start and end (stopping before any
    totals/statistics section).
    """
    sno_col: Optional[int] = None
    regno_col: Optional[int] = None
    name_col: Optional[int] = None

    search_rows = [header_row_idx]
    if header_row_idx + 1 < len(rows):
        search_rows.append(header_row_idx + 1)

    for ridx in search_rows:
        row = rows[ridx]
        for idx, h in enumerate(row):
            if idx in subject_cols:
                continue
            h_norm = re.sub(r"\s+", " ", str(h or "")).strip().upper()
            if sno_col is None and h_norm in ("S.NO", "SL.NO", "SNO", "SLNO", "S NO", "SL NO"):
                sno_col = idx
            if regno_col is None and any(
                k in h_norm for k in ("REGISTER NO", "REG.NO", "REG NO", "REGISTER NUMBER", "REGNO", "REG. NO")
            ):
                regno_col = idx
            if name_col is None and "NAME" in h_norm and "STAFF" not in h_norm and "MENTOR" not in h_norm:
                name_col = idx
        if regno_col is not None:
            break

    if regno_col is None:
        regno_col = 1
    if name_col is None:
        name_col = regno_col + 1
    if sno_col is None:
        sno_col = max(0, regno_col - 1)

    # Data typically starts 1-2 rows after the course-code header row (past any sub-label row).
    data_start = header_row_idx + 1
    if data_start < len(rows):
        first_candidate = rows[data_start]
        looks_like_label_row = (
            regno_col < len(first_candidate)
            and not re.match(r"^\d{6,}$", str(first_candidate[regno_col]).strip())
        )
        if looks_like_label_row:
            data_start += 1

    start_idx = None
    end_idx = None
    expected_sno = None
    for idx in range(data_start, len(rows)):
        row = rows[idx]
        if len(row) <= max(regno_col, sno_col, name_col):
            if start_idx is not None:
                end_idx = idx
                break
            continue
        if _is_summary_row(row):
            if start_idx is not None:
                end_idx = idx
                break
            continue
        regno_val = str(row[regno_col]).strip()
        sno_val = str(row[sno_col]).strip()
        regno_ok = bool(re.match(r"^\d{6,}$", regno_val))
        sno_ok = sno_val.isdigit()
        if regno_ok and (not sno_val or sno_ok):
            if start_idx is None:
                start_idx = idx
                expected_sno = int(sno_val) if sno_ok else None
            elif sno_ok and expected_sno is not None:
                expected_sno += 1
                if int(sno_val) != expected_sno:
                    # sequence broke -- likely entered a different block; stop here.
                    end_idx = idx
                    break
            end_idx = idx + 1
        else:
            if start_idx is not None:
                end_idx = idx
                break
            # else: still scanning for the first valid student row
    if start_idx is None:
        start_idx, end_idx = data_start, data_start

    return {
        "sno_col": sno_col, "regno_col": regno_col, "name_col": name_col,
        "student_start": start_idx, "student_end": end_idx,
    }


def extract_legacy_xls(data: bytes, filename: str) -> LegacyXLSExtractionReport:
    """
    First-class legacy .xls ingestion. Produces the SAME StudentResultRecord shape the
    PDF pipeline produces, so it flows through the unchanged analytics engine. Never
    fabricates a grade for a malformed cell -- malformed/ambiguous cells are quarantined,
    not resolved.
    """
    report = LegacyXLSExtractionReport()
    if not data:
        report.ok = False
        report.fatal_error = "Uploaded .xls file is empty (0 bytes)."
        return report

    try:
        sheets, reader_method = _open_legacy_workbook_sheets(data)
    except Exception as e:
        report.ok = False
        report.fatal_error = str(e)
        return report
    report.reader_method = reader_method

    sheet_name, rows, header_row_idx, score = _select_best_result_sheet(sheets)
    if header_row_idx < 0 or score < 3:
        report.ok = False
        report.fatal_error = (
            "Could not locate a per-student result matrix in this workbook (no row with "
            "several course-code-shaped column headers was found in any sheet)."
        )
        return report
    report.sheet_name = sheet_name
    report.header_row = header_row_idx

    descriptors = _detect_subject_descriptors(rows, header_row_idx)
    if not descriptors:
        report.ok = False
        report.fatal_error = f"Sheet '{sheet_name}' has no detectable subject columns."
        return report
    report.subject_descriptors = descriptors
    subject_col_idxs = {d["grade_col"] for d in descriptors} | {d["gp_col"] for d in descriptors if d["gp_col"] is not None}

    ids = _detect_id_columns_and_student_rows(rows, header_row_idx, subject_col_idxs)
    regno_col, name_col = ids["regno_col"], ids["name_col"]
    student_start, student_end = ids["student_start"], ids["student_end"]

    if student_start >= student_end:
        report.ok = False
        report.fatal_error = (
            f"Sheet '{sheet_name}' — could not identify the student data region "
            "(no rows with a valid register number were found below the header)."
        )
        return report

    resolved_subjects = []
    for d in descriptors:
        can_name, code, cred, sem, cat, conf, is_amb = resolve_subject_info(d["course_code"])
        resolved_in_catalog = bool(code) and code.upper() in COURSE_CODE_INDEX
        resolved_subjects.append({
            **d, "canonical_name": can_name, "canonical_code": code or d["course_code"],
            "credits": cred if cred > 0 else 3.0, "semester": sem,
            "confidence": conf, "unresolved": (not resolved_in_catalog) and is_amb,
        })
        if (not resolved_in_catalog) and is_amb:
            report.unresolved_subject_count += 1

    seen_keys: Dict[Tuple[str, str], bool] = {}
    student_regnos = set()

    for ridx in range(student_start, student_end):
        row = rows[ridx]
        if len(row) <= max(regno_col, name_col):
            continue
        raw_reg = str(row[regno_col]).strip()
        raw_name = str(row[name_col]).strip() if name_col < len(row) else ""
        if not re.match(r"^\d{6,}$", raw_reg):
            continue
        student_regnos.add(raw_reg)

        for subj in resolved_subjects:
            gcol = subj["grade_col"]
            gpcol = subj["gp_col"]
            raw_grade = str(row[gcol]).strip() if gcol < len(row) else ""

            if not raw_grade:
                report.blank_cell_count += 1
                continue

            normalized = _grade_normalize(raw_grade)
            key = (raw_reg.upper(), subj["canonical_code"].upper())

            if not normalized or normalized not in GRADE_POINTS:
                classification = _classify_result_cell_issue(raw_grade)
                report.quarantined_tokens.append({
                    "row": str(ridx + 1),
                    "regno": raw_reg,
                    "column": subj["course_code"],
                    "raw_value": raw_grade,
                    "reason": f"Unrecognized result token '{raw_grade}' quarantined for manual department review.",
                    "register_number": raw_reg,
                    "student_name": raw_name,
                    "course_code": subj["canonical_code"],
                    "grade_point": None,
                    "classification": classification,
                    "source": "XLS",
                    "source_row": ridx + 1,
                })
                report.malformed_cell_count += 1
                continue

            if key in seen_keys:
                report.duplicate_key_count += 1
                continue
            seen_keys[key] = True

            source_gp: Optional[float] = None
            data_quality_issue = ""
            if gpcol is not None and gpcol < len(row):
                raw_gp = str(row[gpcol]).strip()
                if raw_gp:
                    try:
                        source_gp = float(raw_gp)
                        expected_gp = GRADE_POINTS[normalized]
                        if abs(source_gp - expected_gp) > 0.01:
                            data_quality_issue = "GRADE_POINT_MISMATCH"
                            report.gp_mismatches.append({
                                "register_number": raw_reg, "student_name": raw_name,
                                "course_code": subj["canonical_code"], "grade": normalized,
                                "source_gp": source_gp, "expected_gp": expected_gp,
                                "source_row": ridx + 1,
                            })
                    except ValueError:
                        pass

            report.records.append(StudentResultRecord(
                register_number=raw_reg,
                student_name=raw_name,
                subject_code=subj["canonical_code"],
                subject_name=subj["canonical_name"],
                original_subject_text=subj["course_code"],
                credits=subj["credits"],
                result_status=normalized,
                raw_result_status=raw_grade,
                subject_semester=subj["semester"],
                source_type="XLS",
                source_page=0,
                extraction_confidence=subj["confidence"],
                source_sheet=sheet_name,
                source_row=ridx + 1,
                grade_column=subj["course_code"],
                gp_column=(str(gpcol) if gpcol is not None else ""),
                source_gp=source_gp,
                data_quality_issue=data_quality_issue,
            ))
            report.valid_cell_count += 1

    report.student_count = len(student_regnos)
    report.subject_count = len(resolved_subjects)
    report.expected_cell_count = report.student_count * report.subject_count
    report.unknown_token_count = len(report.quarantined_tokens)

    if report.valid_cell_count == 0:
        report.ok = False
        report.fatal_error = f"No valid result cells could be extracted from sheet '{sheet_name}'."
        return report

    base_conf = 0.96
    if report.unknown_token_count:
        base_conf -= min(0.20, report.unknown_token_count * 0.03)
    if report.gp_mismatches:
        base_conf -= min(0.10, len(report.gp_mismatches) * 0.01)
    report.overall_confidence = round(max(0.50, base_conf), 2)
    return report


def reconcile_by_course_code(
    records_a: List[StudentResultRecord], records_b: List[StudentResultRecord],
    label_a: str = "Source A", label_b: str = "Source B"
) -> ReconciliationReport:
    """
    Generalized cross-source reconciliation keyed by (register_number, course_code) --
    used for PDF<->XLS / XLS<->XLSX comparisons. Never decides which source is correct;
    only reports MATCH / MISMATCH / SOURCE_MALFORMED-aware discrepancies for faculty review.
    """
    report = ReconciliationReport()
    map_a = {(r.register_number.strip().upper(), (r.subject_code or "").strip().upper()): r for r in records_a}
    map_b = {(r.register_number.strip().upper(), (r.subject_code or "").strip().upper()): r for r in records_b}

    for key in set(map_a) | set(map_b):
        rec_a = map_a.get(key)
        rec_b = map_b.get(key)
        if rec_a is not None and rec_b is not None:
            if rec_a.result_status == rec_b.result_status:
                report.matched_count += 1
            else:
                report.mismatched_count += 1
                report.mismatched_records.append({
                    "register_number": rec_a.register_number,
                    "student_name": rec_a.student_name or rec_b.student_name,
                    "subject": rec_a.subject_code,
                    f"{label_a.lower()}_grade": rec_a.result_status,
                    f"{label_b.lower()}_grade": rec_b.result_status,
                    "status": "MISMATCH",
                })
        elif rec_a is not None:
            report.missing_in_excel.append({
                "register_number": rec_a.register_number, "student_name": rec_a.student_name,
                "subject": rec_a.subject_code, f"{label_a.lower()}_grade": rec_a.result_status,
            })
        else:
            report.missing_in_pdf.append({
                "register_number": rec_b.register_number, "student_name": rec_b.student_name,
                "subject": rec_b.subject_code, f"{label_b.lower()}_grade": rec_b.result_status,
            })
    return report


# =============================================================================
# 4) DATA MODELS
# =============================================================================


@dataclass
class ValidationIssue:
    severity: str  # "error" | "warning"
    row: str  # spreadsheet row label
    field: str
    value: str
    reason: str

    def to_dict(self) -> Dict[str, str]:
        return {"severity": self.severity, "row": self.row, "field": self.field,
                "value": self.value, "reason": self.reason}


@dataclass
class ValidationReport:
    sheet_name: str = ""
    header_row: int = 0
    total_input_rows: int = 0
    valid_records: int = 0
    dropped_rows: int = 0
    duplicates_removed: int = 0
    issues: List[ValidationIssue] = field(default_factory=list)
    mapped_columns: Dict[str, str] = field(default_factory=dict)
    discovered_metadata: Dict[str, str] = field(default_factory=dict)
    fatal_error: str = ""
    format_detected: str = "long"  # "wide" | "long"
    subject_mappings: List[Dict[str, Any]] = field(default_factory=list)
    quarantined_tokens: List[Dict[str, Any]] = field(default_factory=list)
    copy_paste_cleaned_count: int = 0

    def add(self, severity: str, row: str, field: str, value: str, reason: str) -> None:
        self.issues.append(ValidationIssue(severity, row, field, str(value), reason))

    def error_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == "error")

    def warning_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == "warning")

    def has_fatal(self) -> bool:
        return bool(self.fatal_error)


@dataclass
class GradeRecord:
    regno: str
    name: str
    subject: str
    course_code: str
    credits: float
    grade: str
    points: float
    src_row: int
    meta: Dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        d = {"regno": self.regno, "name": self.name, "subject": self.subject,
             "course_code": self.course_code, "credits": self.credits,
             "grade": self.grade, "points": self.points, "src_row": self.src_row}
        d.update(self.meta)
        return d


@dataclass
class SubjectAnalysis:
    subject: str
    course_code: str
    credits: float
    student_count: int
    avg_gp: float
    median_gp: float
    pass_count: int
    pass_pct: float
    u_count: int
    ra_count: int
    ua_count: int
    sa_count: int
    wd_count: int
    mm_count: int
    wh2_count: int
    u_pct: float
    priority_level: str  # "High Attention" | "Moderate Attention" | "Normal"
    gp_diff_vs_class: float
    grade_counts: Dict[str, int]
    top_students: List[Dict[str, Any]] = field(default_factory=list)
    u_students: List[Dict[str, Any]] = field(default_factory=list)
    semester: int = 0
    category: str = "Sem 1-4 Foundation"
    rank: Optional[int] = None
    difficulty_score: float = 0.0
    grade_tier_counts: Dict[str, int] = field(default_factory=dict)

    @property
    def arrear_count(self) -> int:
        return self.u_count + self.ra_count + self.ua_count

    @property
    def failure_concentration(self) -> int:
        return self.arrear_count

    def to_dict(self) -> Dict[str, Any]:
        return {"subject": self.subject, "course_code": self.course_code,
                "credits": self.credits, "student_count": self.student_count,
                "avg_gp": self.avg_gp, "median_gp": self.median_gp,
                "pass_count": self.pass_count, "pass_pct": self.pass_pct,
                "u_count": self.u_count, "ra_count": self.ra_count,
                "ua_count": self.ua_count,
                "sa_count": self.sa_count, "wd_count": self.wd_count,
                "mm_count": self.mm_count, "wh2_count": self.wh2_count,
                "priority_level": self.priority_level,
                "gp_diff_vs_class": self.gp_diff_vs_class,
                "u_pct": self.u_pct, "grade_counts": self.grade_counts,
                "semester": self.semester, "category": self.category, "rank": self.rank,
                "difficulty_score": self.difficulty_score, "grade_tier_counts": self.grade_tier_counts,
                "top_students": [(s["regno"], s["name"], s["grade"], s["points"]) for s in self.top_students],
                "u_students": [(s["regno"], s["name"]) for s in self.u_students]}


@dataclass
class StudentSubjectResult:
    subject: str
    course_code: str
    credits: float
    grade: str
    points: float


@dataclass
class StudentAnalysis:
    regno: str
    name: str
    courses: List[StudentSubjectResult] = field(default_factory=list)
    total_courses: int = 0
    credits_completed: float = 0.0
    credits_attempted: float = 0.0
    quality_points: float = 0.0
    gpa: Optional[float] = None
    rank: Optional[int] = None
    percentile: Optional[float] = None
    grade_counts: Dict[str, int] = field(default_factory=dict)
    passed_courses: int = 0
    u_count: int = 0
    ra_count: int = 0
    ua_count: int = 0
    sa_count: int = 0
    wd_count: int = 0
    mm_count: int = 0
    wh2_count: int = 0
    attention: str = STATUS_CLEARED
    is_high_performer: bool = False
    strongest_subjects: List[str] = field(default_factory=list)
    attention_subjects: List[str] = field(default_factory=list)
    backlog_arrear_count: int = 0
    has_backlog_arrears: bool = False
    backlog_subjects: List[str] = field(default_factory=list)
    carried_previous_arrears: List[str] = field(default_factory=list)
    carried_previous_arrear_count: int = 0
    cleared_previous_subjects: List[str] = field(default_factory=list)
    total_active_arrear_count: int = 0
    risk_level: str = "Low Risk / Cleared"
    previous_semester_results: List[StudentSubjectResult] = field(default_factory=list)
    meta: Dict[str, str] = field(default_factory=dict)

    @property
    def arrear_count(self) -> int:
        return self.u_count + self.ra_count + self.ua_count

    @property
    def malpractice_count(self) -> int:
        return self.mm_count + self.wh2_count

    @property
    def total_active_arrears(self) -> int:
        return self.arrear_count + self.carried_previous_arrear_count

    def to_dict(self) -> Dict[str, Any]:
        d = {"regno": self.regno, "name": self.name, "total_courses": self.total_courses,
             "credits_completed": self.credits_completed,
             "credits_attempted": self.credits_attempted,
             "quality_points": self.quality_points, "gpa": self.gpa,
             "rank": self.rank, "percentile": self.percentile,
             "grade_counts": self.grade_counts, "passed_courses": self.passed_courses,
             "u_count": self.u_count, "ra_count": self.ra_count, "ua_count": self.ua_count,
             "sa_count": self.sa_count, "wd_count": self.wd_count,
             "mm_count": self.mm_count, "wh2_count": self.wh2_count,
             "arrear_count": self.arrear_count, "malpractice_count": self.malpractice_count,
             "attention": self.attention, "is_high_performer": self.is_high_performer,
             "strongest_subjects": self.strongest_subjects,
             "attention_subjects": self.attention_subjects,
             "backlog_arrear_count": self.backlog_arrear_count,
             "has_backlog_arrears": self.has_backlog_arrears,
             "backlog_subjects": self.backlog_subjects,
             "carried_previous_arrears": self.carried_previous_arrears,
             "carried_previous_arrear_count": self.carried_previous_arrear_count,
             "cleared_previous_subjects": self.cleared_previous_subjects,
             "total_active_arrear_count": self.total_active_arrear_count,
             "risk_level": self.risk_level,
             "courses": [{"subject": c.subject, "course_code": c.course_code,
                          "credits": c.credits, "grade": c.grade, "points": c.points}
                         for c in self.courses]}
        d.update(self.meta)
        return d


@dataclass
class ClassAnalysis:
    file_name: str = ""
    generated_at: str = ""
    student_count: int = 0
    subject_count: int = 0
    record_count: int = 0
    class_gpa: Optional[float] = None
    highest_gpa: Optional[float] = None
    lowest_gpa: Optional[float] = None
    median_gpa: Optional[float] = None
    cleared_count: int = 0
    u_student_count: int = 0
    ra_student_count: int = 0
    ua_student_count: int = 0
    arrear_student_count: int = 0
    single_u_count: int = 0
    multiple_u_count: int = 0
    sa_student_count: int = 0
    wd_student_count: int = 0
    malpractice_student_count: int = 0
    backlog_student_count: int = 0
    format_detected: str = "long"
    pass_rate: Optional[float] = None
    record_pass_rate: Optional[float] = None
    grade_distribution: Dict[str, int] = field(default_factory=dict)
    subject_distribution: Dict[str, int] = field(default_factory=dict)
    students: List[StudentAnalysis] = field(default_factory=list)
    subjects: List[SubjectAnalysis] = field(default_factory=list)
    subject_mappings: List[Dict[str, Any]] = field(default_factory=list)
    quarantined_tokens: List[Dict[str, Any]] = field(default_factory=list)
    copy_paste_cleaned_count: int = 0
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def attention_students(self) -> List[StudentAnalysis]:
        return [s for s in self.students if s.attention != STATUS_CLEARED]

    def to_dict(self) -> Dict[str, Any]:
        return {"file_name": self.file_name, "generated_at": self.generated_at,
                "student_count": self.student_count, "subject_count": self.subject_count,
                "record_count": self.record_count, "class_gpa": self.class_gpa,
                "highest_gpa": self.highest_gpa, "lowest_gpa": self.lowest_gpa,
                "median_gpa": self.median_gpa, "cleared_count": self.cleared_count,
                "u_student_count": self.u_student_count, "ra_student_count": self.ra_student_count,
                "arrear_student_count": self.arrear_student_count, "multiple_u_count": self.multiple_u_count,
                "sa_student_count": self.sa_student_count, "wd_student_count": self.wd_student_count,
                "malpractice_student_count": self.malpractice_student_count,
                "backlog_student_count": self.backlog_student_count,
                "format_detected": self.format_detected,
                "pass_rate": self.pass_rate, "record_pass_rate": self.record_pass_rate,
                "grade_distribution": self.grade_distribution,
                "subject_distribution": self.subject_distribution,
                "subject_mappings": self.subject_mappings,
                "students": [s.to_dict() for s in self.students],
                "subjects": [s.to_dict() for s in self.subjects],
                "metadata": self.metadata}



@dataclass
class AnalyticsResult:
    ok: bool
    error: str = ""
    records: Optional[pd.DataFrame] = None
    report: ValidationReport = field(default_factory=ValidationReport)
    class_analysis: Optional[ClassAnalysis] = None


# =============================================================================
# 5) VALIDATION / EXCEL INGESTION
# =============================================================================


def _norm_key(s: Any) -> str:
    """Normalize an arbitrary spreadsheet value for comparison."""
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    return t


def _norm_header(s: Any) -> str:
    t = _norm_key(s)
    return re.sub(r"[^a-z0-9]+", " ", t.lower()).strip()


def _classify_result_cell_issue(raw: Any) -> str:
    """
    Classify a result cell that failed _grade_normalize(). This is deliberately
    separate from subject/course-code resolution ("unresolved subject" in
    subject_mappings) -- this classifies the RESULT VALUE itself.

    - "BLANK": cell has no content at all (missing result, not a bad one).
    - "SOURCE_MALFORMED": cell has content but it isn't a single recognizable
      grade token (e.g. two grade letters concatenated/space-joined like
      "A B", a stray number, or unrecognized text). This never guesses which
      of multiple tokens was "intended" -- it only classifies, never resolves.
    """
    val = _norm_key(raw)
    if not val:
        return "BLANK"
    return "SOURCE_MALFORMED"


def _grade_normalize(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    val = _norm_key(raw).upper()
    val = re.sub(r"\s+", " ", val).strip()
    if val in GRADE_POINTS:
        return val
    key = re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9+/.]", " ", val)).strip()
    if key in GRADE_TOKEN_CLEAN:
        return GRADE_TOKEN_CLEAN[key]
    compact = val.replace(" ", "").replace(".", "").replace("/", "").replace("-", "")
    if compact in GRADE_TOKEN_CLEAN:
        return GRADE_TOKEN_CLEAN[compact]
    return None


def _columns_to_targets(columns: List[str]) -> Dict[str, str]:
    """Map each canonical field to a matching spreadsheet column (or None)."""
    mapping: Dict[str, str] = {}

    taken: set = set()
    for target, aliases in HEADER_ALIASES.items():
        for a in aliases:
            want = _norm_header(a)
            for col in columns:
                if col in taken:
                    continue
                if _norm_header(col) == want:
                    mapping[target] = col
                    taken.add(col)
                    break
            if target in mapping:
                break

    key_rules = [
        ("academic_year", ["academic year", "academic session", "aca year", "acad", "session"]),
        ("regno", ["register number", "reg number", "register", "reg no", "regnum", "student number",
                   "student id", "student no", "roll", "admission", "matriculation"]),
        ("credits", ["credit"]),
        ("course_code", ["course code", "subject code", "sub code", "course id"]),
        ("subject", ["subject", "paper", "course", "title"]),
        ("grade", ["grade", "letter", "result", "secured", "achieved"]),
        ("semester", ["sem"]),
        ("programme", ["programme", "program", "degree"]),
        ("section", ["section", "sec"]),
        ("batch", ["batch"]),
        ("year", ["year"]),
        ("dept", ["dept", "department", "branch", "discipline"]),
        ("instructor", ["instructor", "faculty", "staff", "teacher"]),
        ("name", ["name"]),
    ]
    for col in columns:
        if col in taken:
            continue
        n = _norm_header(col)
        if not n:
            continue
        best = None
        best_len = -1
        for target, kws in key_rules:
            if target in mapping:
                continue
            for kw in kws:
                if kw in n and len(kw) > best_len:
                    best = target
                    best_len = len(kw)
        if best is not None:
            mapping[best] = col
            taken.add(col)

    if "course_code" not in mapping:
        for col in columns:
            if col in taken:
                continue
            n = _norm_header(col)
            if "code" in n and ("course" in n or "subject" in n):
                mapping["course_code"] = col
                taken.add(col)
                break

    return mapping

def _clean_copy_paste_df(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    """Sanitize noisy COE PDF copy-paste dumps, split merged RegNo+Name cells, and strip headers/footers."""
    if df is None or len(df) == 0:
        return df, 0

    df = df.copy().fillna("")
    cleaned_rows_count = 0
    noise_keywords = [
        "SARANATHAN COLLEGE", "CONTROLLER OF EXAMINATIONS", "END SEMESTER",
        "RESULT SHEET", "PAGE ", "SIGNATURE OF", "TOTAL PASSED", "PASS PERCENTAGE",
        "DEPT OF AI", "CONTROLLER OF EXAM"
    ]

    valid_row_indices = []
    for idx in range(len(df)):
        row_str = " ".join(str(v).strip().upper() for v in df.iloc[idx].tolist())
        if any(kw in row_str for kw in noise_keywords):
            cleaned_rows_count += 1
            continue
        valid_row_indices.append(idx)

    if valid_row_indices and len(valid_row_indices) < len(df):
        df = df.iloc[valid_row_indices].reset_index(drop=True)

    reg_name_pattern = re.compile(r"^(8138\d{8}|\d{12})\s*([A-Za-z\.\s]{2,})$")
    split_regno = []
    split_name = []

    for cidx in range(min(3, len(df.columns))):
        col_vals = [str(v).strip() for v in df.iloc[:, cidx].tolist()]
        matches = [reg_name_pattern.match(v) for v in col_vals if v and reg_name_pattern.match(v)]
        if len(matches) >= max(1, int(len(df) * 0.2)):
            for v in col_vals:
                m = reg_name_pattern.match(v)
                if m:
                    split_regno.append(m.group(1))
                    split_name.append(m.group(2).strip())
                else:
                    split_regno.append(v)
                    split_name.append("")
            df.insert(0, "Register Number (Cleaned)", split_regno)
            df.insert(1, "Student Name (Cleaned)", split_name)
            cleaned_rows_count += len(matches)
            break

    return df, cleaned_rows_count


def _read_workbook(data: bytes) -> Tuple[Optional[pd.DataFrame], ValidationReport]:
    """Locate the most populated sheet and the best header row from an .xlsx."""
    report = ValidationReport()
    bio = io.BytesIO(data)
    try:
        import openpyxl

        wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    except Exception as e:
        report.fatal_error = f"Corrupted or unreadable XLSX workbook: {e}"
        return None, report

    sheet_names = wb.sheetnames
    if not sheet_names:
        report.fatal_error = "Workbook contains no worksheets."
        return None, report

    chosen = max(
        sheet_names,
        key=lambda sn: _sheet_row_count(wb, sn),
    )
    report.sheet_name = chosen

    try:
        scan = pd.read_excel(io.BytesIO(data), sheet_name=chosen, header=None,
                             engine="openpyxl", dtype=str)
    except Exception as e:
        report.fatal_error = f"Could not read sheet '{chosen}': {e}"
        return None, report

    scan = scan.fillna("")
    scan, cleaned_cnt = _clean_copy_paste_df(scan)
    report.copy_paste_cleaned_count = cleaned_cnt

    best_row = 0
    best_score = 0
    limit = min(16, len(scan))
    for idx in range(limit):
        row = scan.iloc[idx].tolist()
        nonempty = sum(1 for c in row if str(c).strip())
        if nonempty == 0:
            continue
        mapped = _columns_to_targets(list(map(str, row)))
        score = nonempty + len(mapped or {}) * 10
        if score > best_score:
            best_score = score
            best_row = idx
    report.header_row = best_row

    # Re-slice scan starting from best_row as header
    if best_row < len(scan):
        headers = [str(c).strip() or f"Column_{i}" for i, c in enumerate(scan.iloc[best_row].tolist())]
        df = scan.iloc[best_row + 1:].copy()
        df.columns = headers
    else:
        df = scan.copy()

    df = df.fillna("")
    return df, report


def _sheet_row_count(wb, sheet_name: str) -> int:
    try:
        ws = wb[sheet_name]
        ws.calculate_dimension()
        return ws.max_row or 0
    except Exception:
        return 0


def _clean_value(v: Any) -> str:
    s = _norm_key(v)
    return re.sub(r"\s+", " ", s).strip()


def _parse_credits(raw: Any) -> Optional[float]:
    if raw is None or str(raw).strip() == "":
        return None
    try:
        v = float(str(raw).strip())
        if 0.0 <= v <= 100.0 and math.isfinite(v):
            return v
    except Exception:
        pass
    return None


def _readable_required_columns() -> str:
    inner = []
    for f in REQUIRED_FIELDS:
        aliases = HEADER_ALIASES.get(f, [])
        inner.append(f"{f.replace('_', ' ')} (e.g. '{aliases[0] if aliases else f}')")
    return ", ".join(inner)


def validate_and_clean(data: bytes, file_name: str, custom_mapping: Optional[Dict[str, str]] = None) -> AnalyticsResult:
    """Validate + normalize the uploaded workbook into clean GradeRecords (supporting wide and long formats)."""
    report = ValidationReport()

    df, report = _read_workbook(data)
    if report.has_fatal():
        return AnalyticsResult(ok=False, report=report)

    cols = [str(c) for c in df.columns.tolist()]
    report.total_input_rows = len(df)

    # 1. Determine format type: Wide or Long
    mapping = custom_mapping or _columns_to_targets(cols) or {}
    report.mapped_columns = mapping

    has_regno = "regno" in mapping and bool(mapping["regno"])
    has_subject_col = "subject" in mapping and bool(mapping["subject"])
    has_grade_col = "grade" in mapping and bool(mapping["grade"])

    is_wide_format = False
    if has_regno and not (has_subject_col and has_grade_col):
        is_wide_format = True
    elif has_regno and has_subject_col and has_grade_col:
        is_wide_format = False

    regno_col = mapping.get("regno")
    name_col = mapping.get("name")
    
    ignored_headers = {
        "s.no", "sl.no", "sno", "slno", "serial", "serial no", "serial number",
        "regno", "register no", "register number", "reg number", "roll no", "student id",
        "name", "student name", "candidate name", "batch", "dept", "department",
        "section", "sec", "total", "total marks", "remarks", "result", "cgpa", "gpa",
        "arrears", "arrear count", "passed", "failed", "status", "attendance", "attendance %"
    }

    if not is_wide_format and not (has_subject_col and has_grade_col):
        candidate_grade_cols = []
        for col in cols:
            norm_col = _norm_header(col)
            if norm_col in ignored_headers or col == regno_col or col == name_col:
                continue
            sample_vals = [df[col].iloc[i] for i in range(min(15, len(df)))]
            valid_grade_hits = sum(1 for v in sample_vals if _grade_normalize(v) in GRADE_POINTS)
            if valid_grade_hits >= 2 or resolve_subject_info(col)[5] >= 0.60:
                candidate_grade_cols.append(col)
        if candidate_grade_cols and has_regno:
            is_wide_format = True

    records: List[GradeRecord] = []
    seen: Dict[Tuple[str, str], GradeRecord] = {}
    dropped = 0
    dup = 0
    subject_mappings_log: List[Dict[str, Any]] = []

    if is_wide_format:
        report.format_detected = "wide"
        if not regno_col:
            for col in cols:
                if _norm_header(col) in ignored_headers and ("reg" in _norm_header(col) or "roll" in _norm_header(col)):
                    regno_col = col
                    break
            if not regno_col:
                regno_col = cols[0]
        
        subj_cols = []
        for col in cols:
            norm_col = _norm_header(col)
            if col == regno_col or col == name_col:
                continue
            if norm_col in ignored_headers:
                continue
            can_name, code, cred, sem, cat, conf, is_amb = resolve_subject_info(col)
            sample_vals = [df[col].iloc[i] for i in range(min(15, len(df)))]
            grade_count = sum(1 for v in sample_vals if _grade_normalize(v) in GRADE_POINTS)
            if grade_count > 0 or conf >= 0.60:
                subj_cols.append(col)
                subject_mappings_log.append({
                    "raw_header": col,
                    "canonical_subject": can_name,
                    "course_code": code,
                    "credits": cred if cred > 0 else 3.0,
                    "semester": sem,
                    "category": cat,
                    "confidence": conf,
                    "is_ambiguous": is_amb
                })

        report.subject_mappings = subject_mappings_log

        if not subj_cols:
            report.fatal_error = (
                "Wide-format Excel detected, but could not locate valid subject columns with declared grades. "
                "Please verify column headers."
            )
            return AnalyticsResult(ok=False, report=report)

        for ridx in range(len(df)):
            row = df.iloc[ridx]
            rno = ridx + report.header_row + 2
            raw_reg = _clean_value(row[regno_col]) if regno_col in df.columns else ""
            raw_name = _clean_value(row[name_col]) if name_col and name_col in df.columns else ""

            if not raw_reg:
                dropped += 1
                report.add("error", str(rno), "regno", "", "Missing student register number.")
                continue

            for col in subj_cols:
                raw_g = _clean_value(row[col])
                if not raw_g:
                    continue
                grade = _grade_normalize(raw_g)
                if not grade or grade not in GRADE_POINTS:
                    can_name_q, code_q, _c, _s, _cat, _conf, _amb = resolve_subject_info(col, custom_mapping)
                    classification = _classify_result_cell_issue(raw_g)
                    report.quarantined_tokens.append({
                        # Legacy fields (kept for backward compatibility with existing readers)
                        "row": str(rno),
                        "regno": raw_reg,
                        "column": col,
                        "raw_value": raw_g,
                        "reason": f"Unrecognized result token '{raw_g}' quarantined for manual department review.",
                        # Structured Data Quality fields
                        "register_number": raw_reg,
                        "student_name": raw_name,
                        "course_code": code_q or col,
                        "grade_point": None,
                        "classification": classification,
                        "source": "EXCEL",
                        "source_row": rno,
                    })
                    report.add("warning", str(rno), col, raw_g, f"Ignored & quarantined unrecognized result token '{raw_g}'.")
                    continue

                can_name, code, cred, sem, cat, conf, is_amb = resolve_subject_info(col, custom_mapping)
                credit_val = cred if cred > 0 else 3.0

                key = (raw_reg.strip().upper(), can_name.strip().lower())
                if key in seen:
                    dup += 1
                    report.add("warning", str(rno), "duplicate", raw_reg, f"Duplicate grade for subject '{can_name}' removed.")
                    continue

                meta = {}
                for opt in OPTIONAL_FIELDS:
                    opt_col = mapping.get(opt)
                    if opt_col and opt_col in df.columns:
                        meta[opt] = _clean_value(row[opt_col])

                rec = GradeRecord(
                    regno=raw_reg,
                    name=raw_name,
                    subject=can_name,
                    course_code=code,
                    credits=credit_val,
                    grade=grade,
                    points=GRADE_POINTS[grade],
                    src_row=rno,
                    meta=meta,
                )
                seen[key] = rec
                records.append(rec)

    else:
        report.format_detected = "long"
        missing = [f for f in REQUIRED_FIELDS if f not in mapping or not mapping[f]]
        if missing:
            report.fatal_error = (
                "Could not locate required columns in the detected header row "
                f"(row {report.header_row + 1}). Missing: {', '.join(missing)}. "
                "Expected columns such as: " + _readable_required_columns() + "."
            )
            return AnalyticsResult(ok=False, report=report)

        col_index = {t: (cols.index(v) if v and v in cols else None) for t, v in mapping.items()}
        for ridx in range(len(df)):
            row = df.iloc[ridx]
            rno = ridx + report.header_row + 2
            vals: Dict[str, str] = {}
            for target, ci in col_index.items():
                if ci is None:
                    vals[target] = ""
                else:
                    try:
                        vals[target] = _clean_value(row.iloc[ci])
                    except Exception:
                        vals[target] = ""

            if not vals["regno"]:
                dropped += 1
                report.add("error", str(rno), "regno", "", "Missing student/register number.")
                continue
            if not vals["subject"]:
                dropped += 1
                report.add("error", str(rno), "subject", "", "Missing subject/course name.")
                continue
            grade = _grade_normalize(vals["grade"])
            if not grade or grade not in GRADE_POINTS:
                classification = _classify_result_cell_issue(vals["grade"])
                report.quarantined_tokens.append({
                    # Legacy fields (kept for backward compatibility with existing readers)
                    "row": str(rno),
                    "regno": vals["regno"],
                    "column": mapping.get("grade", "grade"),
                    "raw_value": vals["grade"],
                    "reason": f"Unrecognized grade token '{vals['grade']}' quarantined.",
                    # Structured Data Quality fields
                    "register_number": vals["regno"],
                    "student_name": vals.get("name", ""),
                    "course_code": vals.get("subject", mapping.get("grade", "grade")),
                    "grade_point": None,
                    "classification": classification,
                    "source": "EXCEL",
                    "source_row": rno,
                })
                dropped += 1
                report.add("error", str(rno), "grade", vals["grade"],
                           f"Unknown grade '{vals['grade']}' (expecting O, A+, A, B+, B, C, U, RA, SA, WD, MM, WH2).")
                continue

            can_name, code, cred, sem, cat, conf, is_amb = resolve_subject_info(vals["subject"], custom_mapping)
            course_code_val = vals["course_code"] or code
            credit_parsed = _parse_credits(vals["credits"])
            credit_val = credit_parsed if credit_parsed is not None else (cred if cred > 0 else 3.0)

            norm_regno = vals["regno"].strip().upper()
            norm_subject = can_name.strip().lower()
            key = (norm_regno, norm_subject)

            if key in seen:
                pre = seen[key]
                if pre.credits == credit_val and pre.grade == grade:
                    dup += 1
                    report.add("warning", str(rno), "duplicate", vals["regno"],
                               "Duplicate record (same student + subject exported twice) removed.")
                    continue
                dropped += 1
                report.add("warning", str(rno), "conflict", vals["regno"],
                           f"Conflicting repeat grade (same student+subject seen at row {pre.src_row}). Kept first occurrence only.")
                continue

            meta = {opt: vals[opt] for opt in OPTIONAL_FIELDS if vals.get(opt)}
            rec = GradeRecord(
                regno=vals["regno"],
                name=vals["name"] or "",
                subject=can_name,
                course_code=course_code_val,
                credits=credit_val,
                grade=grade,
                points=GRADE_POINTS[grade],
                src_row=rno,
                meta=meta,
            )
            seen[key] = rec
            records.append(rec)

    report.valid_records = len(records)
    report.dropped_rows = dropped
    report.duplicates_removed = dup

    if not records:
        report.fatal_error = (
            "No valid result records remain after validation (all rows were dropped). "
            "Please review the validation issues above."
        )
        if len(report.issues) <= 12:
            report.fatal_error += " Issues: " + "; ".join(i.reason for i in report.issues[:12])
        return AnalyticsResult(ok=False, report=report)

    out = pd.DataFrame([r.to_dict() for r in records])

    meta_disc: Dict[str, str] = {}
    for opt in OPTIONAL_FIELDS:
        series = out.get(opt)
        if series is not None:
            present = [str(v) for v in series.tolist() if str(v).strip()]
            if present:
                meta_disc[opt] = present[0]
    report.discovered_metadata = meta_disc

    return AnalyticsResult(ok=True, records=out, report=report)


def _round(x: Optional[float], ndigits: int = 2) -> Optional[float]:
    if x is None:
        return None
    return round(float(x), ndigits)


def _pct(num: Union[int, float], den: Union[int, float]) -> Optional[float]:
    if not den:
        return None
    return round(100.0 * num / den, 1)


def compute_subject_analytics(records: pd.DataFrame, class_gpa_ref: Optional[float] = None) -> List[SubjectAnalysis]:
    """Subject-level metrics from the validated records dataframe with syllabus categories & subject rankings."""
    subjects: List[SubjectAnalysis] = []
    for subject, grp in records.groupby("subject", sort=True):
        grp = grp.reset_index(drop=True)
        students = grp["regno"].astype(str).nunique()
        pts = grp["points"].astype(float)
        avg_gp = _round(float(pts.mean()), 2) if students else None
        median_gp = _round(float(pts.median()), 2) if students else None

        counts = {g: int((grp["grade"] == g).sum()) for g in GRADE_ORDER}
        pass_count = int(grp["grade"].isin(PASSING_GRADES).sum())
        u_count = int((grp["grade"] == "U").sum())
        ra_count = int((grp["grade"] == "RA").sum())
        ua_count = int((grp["grade"] == "UA").sum())
        sa_count = int((grp["grade"] == "SA").sum())
        wd_count = int((grp["grade"] == "WD").sum())
        mm_count = int((grp["grade"] == "MM").sum())
        wh2_count = int((grp["grade"] == "WH2").sum())

        arrear_total = u_count + ra_count + ua_count
        pass_pct = _pct(pass_count, len(grp))
        u_pct = _pct(arrear_total, len(grp))

        gp_diff = (avg_gp - class_gpa_ref) if (avg_gp is not None and class_gpa_ref is not None) else 0.0

        if (pass_pct is not None and pass_pct < 75.0) or arrear_total >= 5 or gp_diff <= -1.0:
            priority = "High Attention"
        elif (pass_pct is not None and pass_pct < 85.0) or arrear_total >= 2:
            priority = "Moderate Attention"
        else:
            priority = "Normal"

        # Resolve syllabus metadata
        can_name, code_syl, cred_syl, sem_syl, cat_syl, _, _ = resolve_subject_info(subject)

        course_code = ""
        cc = grp["course_code"].dropna().astype(str)
        cc = cc[cc.str.strip() != ""]
        if len(cc):
            course_code = cc.mode().iloc[0]
        elif code_syl:
            course_code = code_syl

        credits_series = grp["credits"].dropna().astype(float)
        credits = float(credits_series.mode().iloc[0]) if len(credits_series) else (cred_syl if cred_syl > 0 else 3.0)

        top = grp.sort_values(["points", "regno"], ascending=[False, True]).head(6)
        top_students = [
            {"regno": str(r["regno"]), "name": str(r["name"]),
             "grade": str(r["grade"]), "points": float(r["points"])}
            for _, r in top.iterrows()
        ]
        u_df = grp[grp["grade"].isin(ARREAR_GRADES)].sort_values("regno")
        u_students = [
            {"regno": str(r["regno"]), "name": str(r["name"]),
             "grade": str(r["grade"]), "points": float(r["points"])}
            for _, r in u_df.iterrows()
        ]

        # Compute Difficulty Score & Grade Tiers
        u_pct_val = u_pct or 0.0
        avg_gp_gap = max(0.0, 8.0 - (avg_gp or 0.0))
        diff_score = _round(u_pct_val * 0.6 + avg_gp_gap * 4.0 + credits * 1.5, 1) or 0.0

        distinction_cnt = counts.get("O", 0) + counts.get("A+", 0)
        high_pass_cnt = counts.get("A", 0) + counts.get("B+", 0)
        pass_cnt = counts.get("B", 0) + counts.get("C", 0)
        fail_cnt = u_count + ra_count + ua_count + sa_count + wd_count + mm_count + wh2_count
        grade_tiers = {
            "Distinction (O/A+)": distinction_cnt,
            "High Pass (A/B+)": high_pass_cnt,
            "Pass (B/C)": pass_cnt,
            "Fail / Arrear": fail_cnt,
        }

        subjects.append(SubjectAnalysis(
            subject=str(subject), course_code=course_code, credits=credits,
            student_count=int(students), avg_gp=avg_gp, median_gp=median_gp,
            pass_count=pass_count, pass_pct=pass_pct, u_count=u_count, ra_count=ra_count,
            ua_count=ua_count,
            sa_count=sa_count, wd_count=wd_count, mm_count=mm_count, wh2_count=wh2_count,
            u_pct=u_pct, priority_level=priority, gp_diff_vs_class=gp_diff,
            grade_counts=counts, top_students=top_students, u_students=u_students,
            semester=sem_syl, category=cat_syl,
            difficulty_score=diff_score, grade_tier_counts=grade_tiers,
        ))

    # Rank subjects by arrear rate (descending), arrear count (descending), and avg GP (ascending)
    subjects_sorted = sorted(
        subjects,
        key=lambda s: (-(s.u_pct or 0.0), -s.arrear_count, (s.avg_gp if s.avg_gp is not None else 10.0))
    )
    for idx, subj in enumerate(subjects_sorted):
        subj.rank = idx + 1

    return subjects


def compute_student_analytics(records: pd.DataFrame, prev_records: Optional[Union[pd.DataFrame, List[Any]]] = None) -> List[StudentAnalysis]:
    """Student-level metrics with credit-weighted GPA, risk scoring & Sem 1-4 backlog tracking."""
    prev_by_student: Dict[str, List[StudentSubjectResult]] = defaultdict(list)
    if prev_records is not None:
        if isinstance(prev_records, pd.DataFrame):
            for _, pr in prev_records.iterrows():
                p_reg = str(pr.get("regno", "")).strip().upper()
                if p_reg:
                    prev_by_student[p_reg].append(StudentSubjectResult(
                        subject=str(pr.get("subject", "")),
                        course_code=str(pr.get("course_code", "") or ""),
                        credits=float(pr.get("credits", 3.0) or 3.0),
                        grade=str(pr.get("grade", "")),
                        points=float(pr.get("points", 0.0) or 0.0),
                    ))
        elif isinstance(prev_records, list):
            for prec in prev_records:
                p_reg = getattr(prec, "register_number", "") or (prec.get("regno") if isinstance(prec, dict) else "")
                p_reg = str(p_reg).strip().upper()
                p_subj = getattr(prec, "subject_name", "") or (prec.get("subject") if isinstance(prec, dict) else "")
                p_code = getattr(prec, "subject_code", "") or (prec.get("course_code") if isinstance(prec, dict) else "")
                p_cred = getattr(prec, "credits", 3.0) or (prec.get("credits") if isinstance(prec, dict) else 3.0)
                p_grade = getattr(prec, "result_status", "") or (prec.get("grade") if isinstance(prec, dict) else "")
                p_pts = GRADE_POINTS.get(p_grade, 0.0)
                if p_reg and p_subj:
                    prev_by_student[p_reg].append(StudentSubjectResult(
                        subject=str(p_subj),
                        course_code=str(p_code),
                        credits=float(p_cred or 3.0),
                        grade=str(p_grade),
                        points=float(p_pts),
                    ))

    students: List[StudentAnalysis] = []
    for regno, grp in records.groupby("regno", sort=True):
        grp = grp.reset_index(drop=True)
        regno_str = str(regno).strip()
        name = ""
        nm = grp["name"].dropna().astype(str)
        nm = nm[nm.str.strip() != ""]
        if len(nm):
            name = nm.mode().iloc[0]

        courses = []
        backlog_subjects = []
        carried_previous_arrears = []
        cleared_previous_subjects = []

        for _, r in grp.iterrows():
            subj_title = str(r["subject"])
            grade_val = str(r["grade"])
            courses.append(StudentSubjectResult(
                subject=subj_title, course_code=str(r["course_code"] or ""),
                credits=float(r["credits"] or 0.0), grade=grade_val,
                points=float(r["points"] or 0.0),
            ))

        # 1. Evaluate Previous Semester Results (from multi-semester pages or uploaded previous batch PDF)
        student_prev = prev_by_student.get(regno_str.upper(), [])
        for p_sub in student_prev:
            subj_tag = p_sub.course_code or p_sub.subject
            if p_sub.grade in ARREAR_GRADES:
                if subj_tag not in carried_previous_arrears:
                    carried_previous_arrears.append(subj_tag)
                if subj_tag not in backlog_subjects:
                    backlog_subjects.append(subj_tag)
            elif p_sub.grade in PASSING_GRADES:
                if subj_tag not in cleared_previous_subjects:
                    cleared_previous_subjects.append(subj_tag)

        # 2. Evaluate Current Semester Foundation Backlogs
        for c in courses:
            if c.grade in ARREAR_GRADES:
                can_name, code, cred, sem, cat, conf, amb = resolve_subject_info(c.subject)
                subj_tag = code or can_name
                if sem in [1, 2, 3, 4] or cat == "Sem 1-4 Foundation":
                    if subj_tag not in backlog_subjects:
                        backlog_subjects.append(subj_tag)

        courses.sort(key=lambda c: (c.subject.lower(), c.course_code))

        passing = [c for c in courses if c.grade in PASSING_GRADES]
        attempted_credits = sum(c.credits for c in courses)
        completed_credits = sum(c.credits for c in passing)
        gpa_courses = [c for c in courses if c.grade not in WITHDRAWAL_GRADES]
        gpa_attempted_credits = sum(c.credits for c in gpa_courses)
        quality_points = sum(c.credits * c.points for c in courses)
        gpa = (quality_points / gpa_attempted_credits) if gpa_attempted_credits > 0 else None

        counts = {g: sum(1 for c in courses if c.grade == g) for g in GRADE_ORDER}
        u_count = counts.get("U", 0)
        ra_count = counts.get("RA", 0)
        ua_count = counts.get("UA", 0)
        sa_count = counts.get("SA", 0)
        wd_count = counts.get("WD", 0)
        mm_count = counts.get("MM", 0)
        wh2_count = counts.get("WH2", 0)

        arrear_total = u_count + ra_count + ua_count
        malpractice_total = mm_count + wh2_count
        is_current_arrear = (arrear_total > 0)
        carried_previous_arrear_count = len(carried_previous_arrears)
        backlog_arrear_count = len(backlog_subjects)
        has_backlog = (backlog_arrear_count > 0 or carried_previous_arrear_count > 0)

        # Total Active Arrear Load:
        # For current arrear students, this is the sum of current semester arrears plus carried previous arrears.
        total_active_arrear_count = arrear_total + (carried_previous_arrear_count if is_current_arrear else 0)

        if arrear_total >= 2 or (is_current_arrear and carried_previous_arrear_count > 0):
            attention = STATUS_MULTI_U
        elif has_backlog:
            attention = STATUS_BACKLOG
        elif arrear_total == 1:
            attention = STATUS_U
        elif malpractice_total > 0:
            attention = STATUS_MALPRACTICE
        elif sa_count > 0:
            attention = STATUS_SA
        elif wd_count > 0:
            attention = STATUS_WD
        else:
            attention = STATUS_CLEARED

        # Determine Student Risk Level
        if arrear_total >= 3 or total_active_arrear_count >= 3 or malpractice_total > 0:
            risk_level = "Critical Risk"
        elif arrear_total == 2 or total_active_arrear_count >= 2 or has_backlog:
            risk_level = "High Risk"
        elif arrear_total == 1 or total_active_arrear_count == 1 or sa_count > 0 or wd_count > 0:
            risk_level = "Moderate Risk"
        else:
            risk_level = "Low Risk / Cleared"

        is_high = (
            gpa is not None and gpa >= 8.5 and attention == STATUS_CLEARED
        )

        strongest = [c.subject for c in sorted(
            (c for c in passing), key=lambda c: (-c.points, c.subject.lower()))]
        attention_subjects = [c.subject for c in courses if c.grade in FAILING_GRADES]

        meta = {}
        for opt in OPTIONAL_FIELDS:
            s = grp.get(opt)
            if s is not None:
                present = [str(v) for v in s.tolist() if str(v).strip()]
                if present:
                    meta[opt] = present[0]

        students.append(StudentAnalysis(
            regno=str(regno), name=name, courses=courses,
            total_courses=len(courses),
            credits_completed=completed_credits, credits_attempted=attempted_credits,
            quality_points=quality_points, gpa=gpa,
            grade_counts=counts, passed_courses=len(passing),
            u_count=u_count, ra_count=ra_count, ua_count=ua_count, sa_count=sa_count, wd_count=wd_count,
            mm_count=mm_count, wh2_count=wh2_count,
            attention=attention, is_high_performer=is_high,
            strongest_subjects=strongest, attention_subjects=attention_subjects,
            backlog_arrear_count=backlog_arrear_count,
            has_backlog_arrears=has_backlog,
            backlog_subjects=backlog_subjects,
            carried_previous_arrears=carried_previous_arrears,
            carried_previous_arrear_count=carried_previous_arrear_count,
            cleared_previous_subjects=cleared_previous_subjects,
            total_active_arrear_count=total_active_arrear_count,
            risk_level=risk_level,
            previous_semester_results=student_prev,
            meta=meta,
        ))
    return students


def _rank_students(students: List[StudentAnalysis]) -> None:
    """Competition-style ranking: ties share the best rank; assign percentiles."""
    n = len(students)
    if n == 0:
        return
    order = sorted(
        students,
        key=lambda s: (0 if s.gpa is not None else 1, -(_round(s.gpa, 4) or 0.0), s.regno.lower()),
    )
    prev_key = None
    prev_rank = 1
    for i, s in enumerate(order):
        key = (s.gpa is not None, _round(s.gpa, 4))
        if key == prev_key:
            s.rank = prev_rank
        else:
            s.rank = i + 1
            prev_rank = s.rank
            prev_key = key
        if s.gpa is None:
            s.percentile = None
        else:
            s.percentile = _round((n - s.rank + 1) / n * 100.0, 1)


def generate_deterministic_insights(ca: ClassAnalysis) -> Dict[str, Any]:
    """Compute rich, deterministic academic insights from the ClassAnalysis object."""
    insights = []
    
    if ca.pass_rate is not None:
        if ca.pass_rate >= 85.0:
            health_status = "Excellent"
            health_color = "text-green-700"
            health_bg = "bg-green-100 border-green-300"
        elif ca.pass_rate >= 70.0:
            health_status = "Good"
            health_color = "text-blue-700"
            health_bg = "bg-blue-100 border-blue-300"
        else:
            health_status = "Needs Attention"
            health_color = "text-amber-700"
            health_bg = "bg-amber-100 border-amber-300"
    else:
        health_status = "N/A"
        health_color = "text-slate-700"
        health_bg = "bg-slate-100 border-slate-300"

    sorted_by_pass = sorted(ca.subjects, key=lambda s: (s.pass_pct or 0, s.avg_gp or 0), reverse=True)
    best_subject = sorted_by_pass[0] if sorted_by_pass else None
    
    sorted_by_u = sorted(ca.subjects, key=lambda s: (s.arrear_count, -(s.pass_pct or 100)), reverse=True)
    weakest_subject = sorted_by_u[0] if sorted_by_u and sorted_by_u[0].arrear_count > 0 else None

    if ca.arrear_student_count > 0:
        insights.append(f"{ca.arrear_student_count} student(s) hold active arrears (U / RA grades).")
    else:
        insights.append("100% of students cleared all registered subjects with zero arrears.")

    if ca.multiple_u_count > 0:
        insights.append(f"{ca.multiple_u_count} student(s) hold multiple arrears requiring immediate intervention.")

    if ca.backlog_student_count > 0:
        insights.append(f"⚠️ {ca.backlog_student_count} student(s) carry backlog arrears from Semesters I-IV foundation courses.")

    if ca.malpractice_student_count > 0:
        insights.append(f"🚨 {ca.malpractice_student_count} student(s) have recorded malpractice status (MM / WH2).")

    if best_subject and best_subject.pass_pct is not None:
        insights.append(f"'{best_subject.subject}' achieved highest pass percentage at {best_subject.pass_pct:.1f}%.")

    if weakest_subject and weakest_subject.arrear_count > 0:
        diff_str = f"{abs(weakest_subject.gp_diff_vs_class):.1f} lower average GP" if weakest_subject.gp_diff_vs_class < 0 else "comparable GP"
        insights.append(f"'{weakest_subject.subject}' has highest concentration of arrears ({weakest_subject.arrear_count} students, {diff_str} than class average).")

    top_distinction = ca.grade_distribution.get("O", 0) + ca.grade_distribution.get("A+", 0)
    if top_distinction > 0:
        insights.append(f"{top_distinction} result entries achieved top distinction (O or A+).")

    if ca.sa_student_count > 0:
        insights.append(f"{ca.sa_student_count} student(s) carry Shortage of Attendance (SA) status.")

    if ca.wd_student_count > 0:
        insights.append(f"{ca.wd_student_count} student(s) carry Withdrawal (WD) status.")

    return {
        "health_status": health_status,
        "health_color": health_color,
        "health_bg": health_bg,
        "best_subject": best_subject,
        "weakest_subject": weakest_subject,
        "bullet_insights": insights,
    }


def compute_class_analysis(records: pd.DataFrame, file_name: str, prev_records: Optional[Union[pd.DataFrame, List[Any]]] = None) -> ClassAnalysis:
    """Compute the full deterministic class analysis."""
    ca = ClassAnalysis(file_name=file_name,
                       generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ca.student_count = int(records["regno"].nunique())
    ca.subject_count = int(records["subject"].nunique())
    ca.record_count = int(len(records))
    ca.metadata["semester_records"] = ca.record_count

    ca.students = compute_student_analytics(records, prev_records=prev_records)
    _rank_students(ca.students)

    gpas = [s.gpa for s in ca.students if s.gpa is not None]
    if gpas:
        ca.class_gpa = _round(float(np.mean(gpas)), 2)
        ca.highest_gpa = _round(float(np.max(gpas)), 2)
        ca.lowest_gpa = _round(float(np.min(gpas)), 2)
        ca.median_gpa = _round(float(np.median(gpas)), 2)

    ca.subjects = compute_subject_analytics(records, ca.class_gpa)

    ca.cleared_count = sum(1 for s in ca.students if s.attention == STATUS_CLEARED)
    ca.u_student_count = sum(1 for s in ca.students if s.u_count > 0)
    ca.ra_student_count = sum(1 for s in ca.students if s.ra_count > 0)
    ca.ua_student_count = sum(1 for s in ca.students if s.ua_count > 0)
    ca.arrear_student_count = sum(1 for s in ca.students if s.arrear_count > 0)
    ca.single_u_count = sum(1 for s in ca.students if s.arrear_count == 1)
    ca.multiple_u_count = sum(1 for s in ca.students if s.arrear_count >= 2)
    ca.backlog_student_count = sum(1 for s in ca.students if s.has_backlog_arrears)
    ca.sa_student_count = sum(1 for s in ca.students if s.sa_count > 0)
    ca.wd_student_count = sum(1 for s in ca.students if s.wd_count > 0)
    ca.malpractice_student_count = sum(1 for s in ca.students if s.malpractice_count > 0)

    ca.pass_rate = _pct(ca.cleared_count, ca.student_count)
    ca.record_pass_rate = _pct(
        int((records["grade"].isin(PASSING_GRADES)).sum()), len(records))

    ca.grade_distribution = {g: int((records["grade"] == g).sum()) for g in GRADE_ORDER}
    ca.subject_distribution = {
        s.subject: s.student_count for s in ca.subjects
    }
    return ca


def get_student(ca: ClassAnalysis, regno: str) -> Optional[StudentAnalysis]:
    regno_clean = unquote(str(regno)).strip().upper()
    for s in ca.students:
        if s.regno.strip().upper() == regno_clean:
            return s
    return None


def get_subject(ca: ClassAnalysis, subject: str) -> Optional[SubjectAnalysis]:
    subj_clean = unquote(str(subject)).strip().lower()
    for s in ca.subjects:
        if s.subject.strip().lower() == subj_clean:
            return s
    return None


# =============================================================================
# 7) CHART GENERATION (Plotly Express + Heatmap + Performance Distribution)
# =============================================================================

_PLOTLY_INJECTED = {"flag": False}


def reset_plotly_flag() -> None:
    _PLOTLY_INJECTED["flag"] = False


def _default_layout(fig: go.Figure, height: Optional[int] = None) -> go.Figure:
    fig.update_layout(
        template="plotly_dark",
        font=dict(family="Inter, JetBrains Mono, system-ui, sans-serif", size=12, color="#cbd5e1"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=40, r=20, t=48, b=40),
        height=height or 360,
        title_font=dict(family="Inter", size=15, color="#f8fafc"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0, font=dict(color="#94a3b8")),
        hoverlabel=dict(bgcolor="#0f172a", font_color="#00f0ff", font_size=12, bordercolor="rgba(0, 240, 255, 0.4)"),
        xaxis=dict(gridcolor="rgba(255, 255, 255, 0.08)", zeroline=False, tickfont=dict(color="#94a3b8")),
        yaxis=dict(gridcolor="rgba(255, 255, 255, 0.08)", zeroline=False, tickfont=dict(color="#94a3b8")),
    )
    return fig


def fig_grade_distribution(ca: ClassAnalysis) -> go.Figure:
    vals = [ca.grade_distribution.get(g, 0) for g in GRADE_ORDER]
    fig = px.bar(x=GRADE_ORDER, y=vals, color=GRADE_ORDER, color_discrete_map=GRADE_COLORS)
    fig.update_traces(marker_line_color="#0f1b33", marker_line_width=0.6, showlegend=False)
    fig.update_xaxes(categoryorder="array", categoryarray=GRADE_ORDER)
    fig.update_layout(title="Grade distribution (all result records)", yaxis_title="Records")
    return _default_layout(fig, 340)


def fig_subject_avg_gp(ca: ClassAnalysis) -> go.Figure:
    rows = [s for s in ca.subjects if s.avg_gp is not None]
    labels = [s.subject for s in rows]
    vals = [s.avg_gp or 0.0 for s in rows]
    fig = px.bar(x=labels, y=vals, color=vals, color_continuous_scale=["#dc2626", "#fbbf24", "#16a34a"],
                 range_color=[0, 10])
    fig.update_layout(title="Average grade point by subject", yaxis_title="Avg grade point")
    fig.update_yaxes(range=[0, 10])
    return _default_layout(fig, 360)


def fig_subject_pass_pct(ca: ClassAnalysis) -> go.Figure:
    labels = [s.subject for s in ca.subjects]
    vals = [s.pass_pct if s.pass_pct is not None else 0.0 for s in ca.subjects]
    fig = px.bar(x=labels, y=vals, color=vals, color_continuous_scale=["#dc2626", "#f59e0b", "#16a34a"],
                 range_color=[0, 100])
    fig.update_layout(title="Pass percentage by subject", yaxis_title="Pass %")
    fig.update_yaxes(range=[0, 100])
    return _default_layout(fig, 330)


def fig_failure_concentration(ca: ClassAnalysis, top: int = 8) -> go.Figure:
    subj = sorted(ca.subjects, key=lambda s: s.arrear_count, reverse=True)[:top]
    labels = [s.subject for s in subj]
    vals = [s.arrear_count for s in subj]
    colors = [C_RED if v > 0 else C_SLATE for v in vals]
    fig = px.bar(x=labels, y=vals, color=labels, color_discrete_map={
        l: c for l, c in zip(labels, colors)})
    fig.update_layout(title="Arrear concentration (U / RA grades by subject)",
                      yaxis_title="Arrear count", showlegend=False)
    return _default_layout(fig, 320)


def fig_student_performance_distribution(ca: ClassAnalysis) -> go.Figure:
    """Requirement 7: Student Performance Distribution bar chart."""
    high = sum(1 for s in ca.students if s.gpa is not None and s.gpa >= 8.5 and s.arrear_count == 0)
    avg = sum(1 for s in ca.students if s.gpa is not None and 7.0 <= s.gpa < 8.5 and s.arrear_count == 0)
    moderate = sum(1 for s in ca.students if s.gpa is not None and 5.0 <= s.gpa < 7.0 and s.arrear_count == 0)
    attention = sum(1 for s in ca.students if s.attention != STATUS_CLEARED)
    
    categories = ["High Performers (GPA ≥8.5)", "Solid Performers (7.0-8.4)", "Average (5.0-6.9)", "Needs Attention (Arrears/SA)"]
    counts = [high, avg, moderate, attention]
    colors = [C_GREEN, C_BLUE, C_AMBER, C_RED]
    
    fig = px.bar(x=counts, y=categories, orientation="h", color=categories,
                 color_discrete_map={cat: col for cat, col in zip(categories, colors)})
    fig.update_traces(showlegend=False)
    fig.update_layout(title="Student Cohort Performance Distribution", xaxis_title="Students")
    return _default_layout(fig, 320)


def fig_grade_heatmap(ca: ClassAnalysis) -> go.Figure:
    """Requirement 7: Grade Heatmap across subjects."""
    subj_list = ca.subjects[:8]
    subjects = [s.subject for s in subj_list]
    grades = ["O", "A+", "A", "B+", "B", "C", "U", "RA", "SA", "WD"]
    
    z = []
    for g in grades:
        row = [s.grade_counts.get(g, 0) for s in subj_list]
        z.append(row)
        
    fig = go.Figure(data=go.Heatmap(
        z=z, x=subjects, y=grades,
        colorscale="Viridis",
        reversescale=True,
        hovertemplate="Subject: %{x}<br>Grade: %{y}<br>Count: %{z}<extra></extra>"
    ))
    fig.update_layout(title="Grade Matrix Heatmap (Subjects vs Grade Tiers)", height=350)
    return _default_layout(fig, 350)


def fig_student_subjects(student: StudentAnalysis) -> go.Figure:
    labels = [c.subject for c in student.courses]
    pts = [c.points for c in student.courses]
    grades = [c.grade for c in student.courses]
    colors = [GRADE_COLORS.get(g, C_SLATE) for g in grades]
    fig = px.bar(x=labels, y=pts, color=labels, color_discrete_map={l: c for l, c in zip(labels, colors)})
    fig.update_layout(title="Subject-wise grade points", yaxis_title="Grade point")
    fig.update_yaxes(range=[0, 10])
    return _default_layout(fig, 320)


def fig_student_vs_class(student: StudentAnalysis, ca: ClassAnalysis) -> go.Figure:
    labels = [c.subject for c in student.courses]
    s_pts = [c.points for c in student.courses]
    subj_by = {s.subject: s for s in ca.subjects}
    class_pts = [subj_by.get(l).avg_gp if subj_by.get(l) else 0.0 for l in labels]
    x = labels + labels
    y = s_pts + class_pts
    series = ["Student"] * len(labels) + ["Class average"] * len(labels)
    fig = px.bar(x=x, y=y, color=series, color_discrete_map={"Student": C_NAVY, "Class average": C_AMBER},
                 barmode="group")
    fig.update_layout(title="Student vs class average (by subject)", yaxis_title="Grade point")
    fig.update_yaxes(range=[0, 10])
    return _default_layout(fig, 340)


def chart_html(fig: go.Figure) -> str:
    """Render a figure as an inline HTML fragment; inline plotly.js once per page."""
    include = not _PLOTLY_INJECTED["flag"]
    _PLOTLY_INJECTED["flag"] = True
    return fig.to_html(
        full_html=False,
        include_plotlyjs=include,
        config={"displayModeBar": False, "responsive": True, "locale": "en"},
    )


def chart_card(fig: go.Figure, note: Optional[str] = None) -> Div:
    body = [Div(NotStr(chart_html(fig)), cls="w-full overflow-hidden")]
    if note:
        body.append(P(note, cls="mt-1 text-xs text-slate-500"))
    return Div(*body, cls="card p-4 sm:p-5")


def fig_to_png(fig: go.Figure, width: int = 900) -> Optional[bytes]:
    """Static PNG via Kaleido; returns None when kaleido is unavailable."""
    if not _KALEIDO_OK:
        return None
    try:
        return fig.to_image(format="png", width=width, height=420, scale=1.5)
    except Exception:
        return None


def png_data_url(png: Optional[bytes]) -> str:
    if not png:
        return ""
    return "data:image/png;base64," + base64.b64encode(png).decode("ascii")


# =============================================================================
# 8) AI FUNCTIONS & OLLAMA CLOUD / LOCAL / DETERMINISTIC PROVIDER ENGINE
# =============================================================================

_AI_CACHE: Dict[str, Dict[str, Any]] = {}

AI_GUARDRAILS = (
    "CRITICAL CONSTRAINTS: Base every statement exclusively on the metrics "
    "provided. Result status meanings: U / RA indicate academic arrear or reappearance requirement; "
    "SA indicates shortage of attendance; WD indicates withdrawal; MM / WH2 indicate malpractice-related result status. "
    "Do not speculate about reasons or circumstances. Only report the status provided. "
    "Never infer details or accuse students of cheating. "
    "Never invent: attendance figures, internal or assessment marks, medical reasons, "
    "personal problems, or faculty misconduct."
)

PROHIBITIONS = (
    "Do not invent attendance figures, internal marks, assessment marks, "
    "reasons for failure, medical information, historical performance, or "
    "teacher misconduct. "
)


def _http_post_json(url: str, payload: dict, headers: dict, timeout: float) -> Optional[dict]:
    """Helper: execute JSON HTTP POST request with timeout and error handling."""
    try:
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            if resp.status == 200:
                raw_body = resp.read().decode("utf-8")
                return json.loads(raw_body)
    except Exception:
        return None
    return None


def _ai_invoke(system: str, user: str) -> Optional[str]:
    """Execute LLM prompt against configured AI provider (ollama_cloud, ollama_local, or deterministic)."""
    provider = CFG.get("ai_provider", "ollama_cloud")
    if provider == "deterministic" or not CFG.get("ai_enabled", True):
        return None

    if not (system and user):
        return None

    timeout = CFG.get("ai_timeout_seconds", 90.0)
    model = CFG.get("ollama_model") or "llama3"

    if provider == "ollama_cloud":
        api_key = CFG.get("ollama_api_key", "")
        if not api_key:
            # Missing API key safely triggers deterministic fallback
            return None

        host = (CFG.get("ollama_host") or "https://api.ollama.com").rstrip("/")

        def _call_cloud() -> Optional[str]:
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            }
            payload = {
                "model": model,
                "messages": [
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                "temperature": 0.1,
            }
            # Primary: OpenAI-compatible /v1/chat/completions endpoint
            res = _http_post_json(f"{host}/v1/chat/completions", payload, headers, timeout)
            if res and "choices" in res and res["choices"]:
                msg = res["choices"][0].get("message", {}).get("content", "")
                if msg and msg.strip():
                    return msg.strip()

            # Secondary: Native /api/chat endpoint
            chat_payload = {
                "model": model,
                "messages": [
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                "stream": False,
            }
            res_chat = _http_post_json(f"{host}/api/chat", chat_payload, headers, timeout)
            if res_chat:
                if "message" in res_chat and isinstance(res_chat["message"], dict):
                    msg = res_chat["message"].get("content", "")
                    if msg and msg.strip():
                        return msg.strip()
                elif "response" in res_chat and str(res_chat["response"]).strip():
                    return str(res_chat["response"]).strip()

            return None

        try:
            with ThreadPoolExecutor(max_workers=1) as ex:
                fut = ex.submit(_call_cloud)
                return fut.result(timeout=timeout)
        except Exception:
            return None

    elif provider == "ollama_local":
        host = (CFG.get("ollama_host") or "http://localhost:11434").rstrip("/")

        def _call_local() -> Optional[str]:
            if _LANGCHAIN_OK and Ollama is not None:
                try:
                    llm = Ollama(model=model, base_url=host, temperature=0.1, top_p=0.9, verbose=False)
                    res = str(llm.invoke(f"{system}\n\n{user}")).strip()
                    if res:
                        return res
                except Exception:
                    pass

            headers = {"Content-Type": "application/json"}
            payload = {
                "model": model,
                "prompt": f"{system}\n\n{user}",
                "stream": False,
            }
            res = _http_post_json(f"{host}/api/generate", payload, headers, timeout)
            if res and "response" in res and str(res["response"]).strip():
                return str(res["response"]).strip()
            return None

        try:
            with ThreadPoolExecutor(max_workers=1) as ex:
                fut = ex.submit(_call_local)
                return fut.result(timeout=timeout)
        except Exception:
            return None

    return None


def _cache_get(key: str) -> Optional[str]:
    entry = _AI_CACHE.get(key)
    return entry["text"] if entry else None


def _cache_set(key: str, text: str, live: bool) -> None:
    _AI_CACHE[key] = {"text": text, "live": "live" if live else "fallback"}


def _ai_hash(*parts: Any) -> str:
    blob = "|".join(str(p) for p in parts)
    return str(abs(hash(blob)))


def _metrics_json(payload: Dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, default=str, indent=2)


def generate_class_ai_insight(ca: ClassAnalysis) -> Dict[str, str]:
    """Class-level AI insight. Returns {'text', 'live'} (live in {'live','fallback'})."""
    key = "class:" + _ai_hash(ca.file_name, ca.record_count, ca.class_gpa,
                               sorted(ca.grade_distribution.items()))
    cached = _cache_get(key)
    if cached is not None:
        return {"text": cached, "live": _AI_CACHE[key]["live"]}

    payload = {
        "class_average_gpa": ca.class_gpa,
        "pass_percentage": ca.pass_rate,
        "best_subject": ca.subjects[0].subject if ca.subjects else "None",
        "weakest_subject": ca.subjects[-1].subject if ca.subjects else "None",
        "arrear_students": ca.arrear_student_count,
        "multiple_u_students": ca.multiple_u_count,
        "malpractice_students": ca.malpractice_student_count,
        "subject_failures": {s.subject: s.arrear_count for s in ca.subjects if s.arrear_count > 0},
    }
    system = (PROHIBITIONS + CLASS_AI_INSTRUCTION
              + "\n\n" + AI_GUARDRAILS)
    user = ("Here are the structured class metrics for a single semester of declared final "
            "grades. Write the class analysis.\n\n" + _metrics_json(payload))
    text = _ai_invoke(system, user)
    if not text:
        return {"text": fallback_class_insight(ca), "live": "fallback"}
    _cache_set(key, text, True)
    return {"text": text, "live": "live"}


def clarified_top_subjects(ca: ClassAnalysis) -> List[Dict[str, Any]]:
    top = sorted(ca.subjects, key=lambda s: (s.arrear_count, -(s.avg_gp or 0)), reverse=True)[:6]
    return [{"subject": s.subject, "avg_gp": s.avg_gp, "pass_pct": s.pass_pct, "u_count": s.u_count, "ra_count": s.ra_count}
            for s in top]


def generate_subject_ai_insight(subj: SubjectAnalysis, ca: ClassAnalysis) -> Dict[str, str]:
    key = "subject:" + _ai_hash(subj.subject, subj.avg_gp, subj.arrear_count)
    cached = _cache_get(key)
    if cached is not None:
        return {"text": cached, "live": _AI_CACHE[key]["live"]}

    payload = {
        "subject": subj.subject,
        "course_code": subj.course_code,
        "credits": subj.credits,
        "students_enrolled": subj.student_count,
        "average_grade_point": subj.avg_gp,
        "median_grade_point": subj.median_gp,
        "pass_percentage": subj.pass_pct,
        "u_count": subj.u_count,
        "ra_count": subj.ra_count,
        "sa_count": subj.sa_count,
        "wd_count": subj.wd_count,
        "mm_count": subj.mm_count,
        "wh2_count": subj.wh2_count,
        "u_percentage": subj.u_pct,
        "grade_distribution": subj.grade_counts,
        "class_overall": {
            "class_gpa": ca.class_gpa,
            "class_pass_percent": ca.pass_rate,
        },
    }
    system = (
        "You are an expert AI Academic Advisor for teachers. Review the subject "
        "metrics. Write a practical analysis summarizing overall strengths, problem "
        "subjects, and targeted recommendations. Format with bold keywords and crisp bullet points.\n"
        + PROHIBITIONS + AI_GUARDRAILS
    )
    user = ("Here are single-semester subject metrics. Produce the subject analysis.\n\n"
            + _metrics_json(payload))
    text = _ai_invoke(system, user)
    if not text:
        return {"text": fallback_subject_insight(subj, ca), "live": "fallback"}
    _cache_set(key, text, True)
    return {"text": text, "live": "live"}


def generate_student_brief(student: StudentAnalysis, ca: ClassAnalysis) -> Dict[str, str]:
    """Concise per-student AI academic brief with current & carried previous arrears."""
    is_curr_arrear = (student.arrear_count > 0)
    key = "student:" + _ai_hash(
        student.regno, student.gpa, sorted(student.grade_counts.items()),
        student.carried_previous_arrear_count, is_curr_arrear
    )
    cached = _cache_get(key)
    if cached is not None:
        return {"text": cached, "live": _AI_CACHE[key]["live"]}

    current_arrears = [
        {"subject": c.subject, "code": c.course_code, "grade": c.grade, "credits": c.credits}
        for c in student.courses if c.grade in ARREAR_GRADES
    ]

    payload = {
        "regno": student.regno,
        "name": student.name,
        "gpa": student.gpa,
        "rank": student.rank,
        "percentile": student.percentile,
        "is_current_arrear_student": is_curr_arrear,
        "current_semester_arrears": {
            "count": student.arrear_count,
            "subjects": current_arrears,
        },
        "carried_previous_semester_arrears": (
            {
                "count": student.carried_previous_arrear_count,
                "subjects": student.carried_previous_arrears,
            }
            if is_curr_arrear else None
        ),
        "cleared_previous_arrears": student.cleared_previous_subjects,
        "total_active_arrear_load": student.total_active_arrear_count if is_curr_arrear else 0,
        "attendance_shortage": student.sa_count,
        "withdrawal": student.wd_count,
        "malpractice": {"mm": student.mm_count, "wh2": student.wh2_count},
        "courses": [
            {"subject": c.subject, "code": c.course_code, "grade": c.grade, "points": c.points}
            for c in student.courses
        ],
    }
    system = (
        "You are an expert AI Academic Advisor for engineering faculty. Write a concise, "
        "evidence-based student academic briefing using bold keywords and crisp bullet points.\n"
        "SPECIAL INSTRUCTION: For current arrear students (students holding active arrears in the current semester), "
        "explicitly evaluate both their current semester arrears AND any previous semester arrears still carried over. "
        "Detail their total active arrear burden and recommend a prioritized remedial plan to clear both foundation and current subjects.\n"
        "Use only the provided metrics. " + PROHIBITIONS + AI_GUARDRAILS
    )
    user = "Provide the student academic brief.\n\n" + _metrics_json(payload)
    text = _ai_invoke(system, user)
    if not text:
        return {"text": fallback_student_brief(student, ca), "live": "fallback"}
    _cache_set(key, text, True)
    return {"text": text, "live": "live"}


def generate_ptm_brief(student: StudentAnalysis, ca: ClassAnalysis) -> Dict[str, str]:
    """Parent–Teacher Meeting brief for one student with carried backlog intelligence."""
    is_curr_arrear = (student.arrear_count > 0)
    key = "ptm:" + _ai_hash(
        student.regno, student.gpa, tuple(sorted(student.grade_counts.items())),
        student.carried_previous_arrear_count, is_curr_arrear
    )
    cached = _cache_get(key)
    if cached is not None:
        return {"text": cached, "live": _AI_CACHE[key]["live"]}

    current_arrears = [
        {"subject": c.subject, "code": c.course_code, "grade": c.grade, "credits": c.credits}
        for c in student.courses if c.grade in ARREAR_GRADES
    ]

    payload = {
        "regno": student.regno,
        "name": student.name,
        "gpa": student.gpa,
        "rank": student.rank,
        "percentile": student.percentile,
        "total_courses": student.total_courses,
        "passed": student.passed_courses,
        "is_current_arrear_student": is_curr_arrear,
        "current_semester_arrears": {
            "count": student.arrear_count,
            "subjects": current_arrears,
        },
        "carried_previous_semester_arrears": (
            {
                "count": student.carried_previous_arrear_count,
                "subjects": student.carried_previous_arrears,
            }
            if is_curr_arrear else None
        ),
        "cleared_previous_arrears": student.cleared_previous_subjects,
        "total_active_arrear_load": student.total_active_arrear_count if is_curr_arrear else 0,
        "sa": student.sa_count,
        "wd": student.wd_count,
        "malpractice": {"mm": student.mm_count, "wh2": student.wh2_count},
        "courses": [
            {"subject": c.subject, "grade": c.grade, "points": c.points}
            for c in student.courses
        ],
        "class_gpa": ca.class_gpa,
    }
    system = (
        "You are an expert AI Academic Advisor helping a teacher prepare a "
        "Parent-Teacher Meeting (PTM) discussion brief. Write structured discussion points using bold keywords.\n"
        "SPECIAL INSTRUCTION: For current arrear students, discuss both their current semester arrears AND any "
        "previous semester arrears still carried over by the student. Present their total active arrear load and "
        "suggest concrete mentoring and study schedules for parents to support at home.\n"
        "Use only provided metrics. Do not invent causes or medical details. " + AI_GUARDRAILS
    )
    user = "Prepare the PTM brief for this student.\n\n" + _metrics_json(payload)
    text = _ai_invoke(system, user)
    if not text:
        return {"text": fallback_ptm_brief(student, ca), "live": "fallback"}
    _cache_set(key, text, True)
    return {"text": text, "live": "live"}


def ai_source_banner(live: str) -> str:
    if live == "live":
        provider = CFG.get("ai_provider", "ollama_cloud")
        model = CFG.get("ollama_model", "llama3")
        label = "Ollama Cloud" if provider == "ollama_cloud" else "local Ollama"
        return f"AI advisory generated by {label} ({model}) · treat as advisory, verify against official records."
    return "AI insights are currently unavailable. Analytics calculations are still available."


def fallback_class_insight(ca: ClassAnalysis) -> str:
    worst = [s for s in ca.subjects if s.arrear_count > 0]
    worst.sort(key=lambda s: (s.arrear_count, -(s.avg_gp or 0)), reverse=True)
    lines = []
    lines.append("## Overall Performance")
    if ca.pass_rate is not None and ca.pass_rate >= 75:
        lines.append(f"- Pass rate is **{ca.pass_rate:.1f}%** across {ca.student_count} students — solid cohort performance.")
    else:
        lines.append(f"- Pass rate is **{ca.pass_rate:.1f}%** — significant scope for academic recovery.")
    lines.append(f"- Class GPA average is **{fmt_gpa(ca.class_gpa)}** (median: **{fmt_gpa(ca.median_gpa)}**).")
    
    lines.append("\n## Strengths")
    if ca.cleared_count > 0:
        lines.append(f"- **{ca.cleared_count} student(s)** cleared all declared subjects without any arrears.")
    top_dist = ca.grade_distribution.get("O", 0) + ca.grade_distribution.get("A+", 0)
    if top_dist > 0:
        lines.append(f"- **{top_dist} result entries** achieved top distinction grades (O or A+).")

    lines.append("\n## Areas Requiring Attention")
    if worst:
        for s in worst[:3]:
            lines.append(f"- **{s.subject}**: {s.arrear_count} arrear(s) ({fmt_pct(s.u_pct)} of cohort).")
    else:
        lines.append("- No subjects show arrear grades in this semester evaluation.")
    if ca.multiple_u_count:
        lines.append(f"- **{ca.multiple_u_count} student(s)** hold multiple arrears requiring immediate intervention.")
    if ca.backlog_student_count:
        lines.append(f"- ⚠️ **{ca.backlog_student_count} student(s)** carry foundation backlog arrears from Semesters I-IV.")
    if ca.malpractice_student_count:
        lines.append(f"- 🚨 **{ca.malpractice_student_count} student(s)** carry malpractice status entries (MM/WH2).")

    lines.append("\n## Recommended Faculty Actions")
    lines.append(f"- Organize targeted remedial sessions for the {ca.arrear_student_count} student(s) with active arrears.")
    lines.append("- Review subject delivery for subjects with elevated failure concentrations.")

    lines.append("\n## PTM Discussion Points")
    lines.append("- Share student rank and credit completion metrics with parents.")
    lines.append("- Discuss academic support plans for attention subjects.")
    return "\n".join(lines)


def primary_grade(dist: Dict[str, int]) -> str:
    return max(dist.items(), key=lambda kv: kv[1])[0] if dist else ""


def fmt_pct(v: Optional[float]) -> str:
    return "—" if v is None else f"{v:.1f}%"


def fmt_none_pct(v: Optional[float]) -> str:
    return "—" if v is None else f"{v:.1f}%"


def summary_pct(counts: Dict[str, int]) -> Dict[str, float]:
    total = sum(counts.values())
    if not total:
        return {}
    return {g: round(100.0 * c / total, 1) for g, c in counts.items() if c and c > 0}


def grade_mix(pcts: Dict[str, float]) -> str:
    parts = [f"{g} {round(p)}%" for g, p in pcts.items() if p]
    return ", ".join(parts[:6]) if parts else "no records"


def fallback_subject_insight(subj: SubjectAnalysis, ca: ClassAnalysis) -> str:
    lines = []
    lines.append("## Overall Performance")
    lines.append(f"- **{subj.subject}** average grade point: **{fmt_gp(subj.avg_gp)}**, pass rate: **{fmt_none_pct(subj.pass_pct)}**.")
    lines.append(f"- Total enrolled students: **{subj.student_count}**.")
    
    lines.append("\n## Strengths")
    lines.append(f"- **{subj.pass_count} student(s)** successfully passed this course.")
    lines.append(f"- Grade mix distribution: {grade_mix(summary_pct(subj.grade_counts))}.")

    lines.append("\n## Areas Requiring Attention")
    if subj.arrear_count:
        lines.append(f"- **{subj.arrear_count} student(s)** hold active arrears (U: {subj.u_count}, RA: {subj.ra_count}).")
    else:
        lines.append("- Zero arrears recorded for this subject.")
    if subj.mm_count or subj.wh2_count:
        lines.append(f"- 🚨 **{subj.mm_count + subj.wh2_count} student(s)** have recorded malpractice status (MM: {subj.mm_count}, WH2: {subj.wh2_count}).")

    lines.append("\n## Recommended Faculty Actions")
    lines.append("- Plan targeted revision and tutorial sessions for students holding arrear grades.")
    lines.append("- Compare topic difficulty against class performance average.")

    lines.append("\n## PTM Discussion Points")
    lines.append(f"- Review student's grade achieved in {subj.subject} relative to class median ({fmt_gp(subj.median_gp)}).")
    return "\n".join(lines)


def fallback_student_brief(student: StudentAnalysis, ca: ClassAnalysis) -> str:
    lines = []
    lines.append("## Overall Performance")
    lines.append(f"- Student **{student.name} ({student.regno})** achieved GPA **{fmt_gpa(student.gpa)}** (Class Avg: **{fmt_gpa(ca.class_gpa)}**).")
    lines.append(f"- Rank: **{rank_text(student, ca.student_count)}** (Percentile: **{fmt_np_pct(student.percentile)}**).")

    lines.append("\n## Strengths")
    if student.strongest_subjects:
        lines.append("- Strongest performance in: **" + ", ".join(student.strongest_subjects[:3]) + "**.")
    else:
        lines.append("- Maintained passing grade standards across courses.")

    lines.append("\n## Areas Requiring Attention")
    if student.arrear_count > 0:
        lines.append(f"- **Current Semester Arrears ({student.arrear_count})**: **{', '.join(student.attention_subjects)}**.")
        if student.carried_previous_arrears:
            lines.append(f"- ⚠️ **Carried Previous Semester Arrears ({student.carried_previous_arrear_count})**: **{', '.join(student.carried_previous_arrears)}** (still carried over by this student).")
            lines.append(f"- **Total Active Arrear Load**: **{student.total_active_arrear_count} subjects** ({student.arrear_count} current + {student.carried_previous_arrear_count} carried from earlier terms).")
        if student.cleared_previous_subjects:
            lines.append(f"- ✓ **Cleared in This Exam Session**: **{', '.join(student.cleared_previous_subjects)}** (reappearance passed).")
    elif student.attention_subjects:
        lines.append("- Courses requiring immediate attention: **" + ", ".join(student.attention_subjects) + "**.")
    else:
        lines.append("- Cleared all registered semester subjects with zero active arrears.")

    lines.append("\n## Recommended Faculty Actions")
    if student.arrear_count > 0:
        lines.append(f"- Formulate structured remedial study schedule covering current ({', '.join(student.attention_subjects[:2])})" +
                     (f" and carried backlog ({', '.join(student.carried_previous_arrears[:2])})" if student.carried_previous_arrears else "") + ".")
        lines.append("- Monitor weekly cycle test benchmarks and conduct peer mentoring reviews.")
    else:
        lines.append("- Monitor academic progress and encourage advanced elective options.")

    lines.append("\n## PTM Discussion Points")
    lines.append("- Present academic summary and credit completion status to parent.")
    if student.arrear_count > 0 and student.carried_previous_arrears:
        lines.append(f"- Detail both current arrears and carried previous subjects ({', '.join(student.carried_previous_arrears)}) with recovery roadmap.")
    lines.append("- Align on targeted home study schedule.")
    return "\n".join(lines)


def fallback_ptm_brief(student: StudentAnalysis, ca: ClassAnalysis) -> str:
    lines = []
    lines.append("## Overall Performance")
    lines.append(f"- **{student.name} ({student.regno})** holds GPA **{fmt_gpa(student.gpa)}** with rank **{rank_text(student, ca.student_count)}**.")
    lines.append(f"- Credits Completed: **{student.credits_completed:.1f} / {student.credits_attempted:.1f}**.")

    lines.append("\n## Strengths")
    lines.append("- " + (", ".join(f"**{s}**" for s in student.strongest_subjects[:3]) if student.strongest_subjects
                         else "Cleared declared result papers."))

    lines.append("\n## Areas Requiring Attention")
    if student.arrear_count > 0:
        lines.append(f"- **Current Semester Arrears ({student.arrear_count})**:")
        for c in student.courses:
            if c.grade in FAILING_GRADES:
                lines.append(f"  - **{c.subject}** (Grade: {c.grade}, Credits: {c.credits})")
        if student.carried_previous_arrears:
            lines.append(f"- ⚠️ **Carried Previous Semester Arrears ({student.carried_previous_arrear_count})**: **{', '.join(student.carried_previous_arrears)}** (still uncleared).")
            lines.append(f"- **Total Active Arrear Burden**: **{student.total_active_arrear_count} subjects**.")
        if student.cleared_previous_subjects:
            lines.append(f"- ✓ **Previous Arrears Successfully Cleared**: **{', '.join(student.cleared_previous_subjects)}**.")
    elif student.attention_subjects:
        for c in student.courses:
            if c.grade in FAILING_GRADES:
                lines.append(f"- **{c.subject}** (Grade: {c.grade})")
    else:
        lines.append("- No active arrears or failure statuses.")

    lines.append("\n## Recommended Faculty Actions")
    if student.arrear_count > 0:
        lines.append("- Provide specialized remedial material for both current and carried backlog subjects.")
        lines.append("- Schedule bi-weekly mentor checks to ensure consistent preparation.")
    else:
        lines.append("- Maintain regular academic guidance and praise consistent effort.")

    lines.append("\n## PTM Discussion Points")
    lines.append("- Review student effort alignment with declared course credits.")
    if student.arrear_count > 0:
        lines.append("- Review specific arrear subjects and parental support strategies at home.")
    return "\n".join(lines)


def fmt_gpa(v: Optional[float]) -> str:
    return "—" if v is None else f"{v:.2f}"


def fmt_gp(v: Optional[float]) -> str:
    return "—" if v is None else f"{v:.2f}"


def fmt_np_pct(v: Optional[float]) -> str:
    return "—" if v is None else f"{v:.1f}%"


def rank_text(s: StudentAnalysis, total_students: int = 0) -> str:
    if s.rank is None:
        return "—"
    return f"{s.rank} of {total_students}" if total_students else f"{s.rank}"


# =============================================================================
# 9) PDF GENERATION (xhtml2pdf.pisa -- Professional University Grade Level)
# =============================================================================

PDF_CSS = """
<style>
@page { size: A4; margin: 18mm 14mm 16mm 14mm; }
body { font-family: Helvetica, Arial, sans-serif; font-size: 10px; color: #1e293b; line-height: 1.45; }
h1 { font-size: 18px; color: #0f1b33; margin: 0 0 4px 0; font-weight: bold; }
h2 { font-size: 13px; color: #0f1b33; margin: 14px 0 6px 0; border-bottom: 1.5px solid #cbd5e1; padding-bottom: 3px; font-weight: bold; }
h3 { font-size: 11px; color: #1e3a5f; margin: 10px 0 4px 0; font-weight: bold; }
.header-box { background: #0f1b33; color: #ffffff; padding: 14px; border-radius: 6px; margin-bottom: 16px; text-align: center; }
.header-box h1 { color: #ffffff; margin: 0; font-size: 18px; letter-spacing: 0.5px; text-transform: uppercase; }
.header-box p { color: #cbd5e1; margin: 3px 0 0 0; font-size: 10px; }
.meta-box { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px; padding: 8px 12px; margin-bottom: 14px; font-size: 9px; }
table { width: 100%; border-collapse: collapse; margin: 6px 0 10px 0; }
th { background: #0f1b33; color: #fff; font-size: 9px; text-align: left; padding: 5px 6px; font-weight: bold; }
td { font-size: 9px; padding: 4px 6px; border-bottom: 1px solid #e2e8f0; }
tr:nth-child(even) td { background: #f8fafc; }
.card { border: 1px solid #e2e8f0; border-radius: 6px; padding: 8px; margin: 8px 0; background: #ffffff; }
.badge-pass { color: #16a34a; font-weight: bold; }
.badge-fail { color: #dc2626; font-weight: bold; }
.grade-o { background: #16a34a; } .grade-aplus { background: #22c55e; }
.grade-a { background: #4ade80; } .grade-bplus { background: #86efac; }
.grade-b { background: #64748b; } .grade-c { background: #94a3b8; }
.grade-u { background: #dc2626; } .grade-ra { background: #ef4444; }
.grade-sa { background: #d97706; } .grade-wd { background: #9333ea; }
.grade-mm { background: #7c3aed; } .grade-wh2 { background: #6b21a8; }
.alert-red { color: #dc2626; font-weight: bold; }
.note { font-size: 8px; color: #94a3b8; margin-top: 16px; border-top: 1px solid #e2e8f0; padding-top: 6px; text-align: center; }
</style>
"""


def _css_bar(label: str, value: int, max_val: int, color_cls: str) -> str:
    if not max_val:
        return ""
    pct = min(100.0 * value / max_val, 100)
    label_esc = html.escape(str(label))
    return (f'<div class="bar-row">'
            f'<span class="bar-label">{label_esc}</span>'
            f'<span class="bar-track"><span class="bar-fill {color_cls}" '
            f'style="width:{pct:.1f}%"></span></span>'
            f'<span class="bar-val">{value}</span></div>')


def _css_bars(title: str, items: List[Tuple[str, int, str]]) -> str:
    if not items:
        return ""
    max_val = max(v for _, v, _ in items if v > 0) or 1
    rows = "".join(_css_bar(l, v, max_val, c) for l, v, c in items if v > 0)
    title_esc = html.escape(title)
    return f"<h3>{title_esc}</h3>{rows}" if rows else ""


def _grade_color_cls(grade: str) -> str:
    mapping = {"O": "grade-o", "A+": "grade-aplus", "A": "grade-a",
               "B+": "grade-bplus", "B": "grade-b", "C": "grade-c",
               "U": "grade-u", "RA": "grade-ra", "SA": "grade-sa",
               "WD": "grade-wd", "MM": "grade-mm", "WH2": "grade-wh2"}
    return mapping.get(grade, "")


def _md_to_pdf_html(text: str) -> str:
    if not text:
        return ""
    lines = text.split("\n")
    out = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            out.append("<br/>")
            continue
        if stripped.startswith("## "):
            inner = html.escape(stripped[3:])
            out.append(f"<h2>{inner}</h2>")
        elif stripped.startswith("- "):
            inner = html.escape(stripped[2:])
            inner = re.sub(r"&lt;strong&gt;(.+?)&lt;/strong&gt;", r"<b>\1</b>", inner)
            inner = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", inner)
            out.append(f"<li>{inner}</li>")
        elif stripped.startswith("**") and stripped.endswith("**"):
            inner = html.escape(stripped.strip("*"))
            out.append(f"<h3>{inner}</h3>")
        else:
            inner = html.escape(stripped)
            inner = re.sub(r"&lt;strong&gt;(.+?)&lt;/strong&gt;", r"<b>\1</b>", inner)
            inner = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", inner)
            out.append(f"<p>{inner}</p>")
    return "\n".join(out)


def _pdf_wrap(title: str, body_html: str) -> str:
    ts = datetime.now().strftime("%d %B %Y, %H:%M")
    title_esc = html.escape(title)
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"/>{PDF_CSS}
<title>{title_esc}</title></head>
<body>
<div class="header-box">
<h1>SARANATHAN COLLEGE OF ENGINEERING</h1>
<p>FACULTY GRADE ANALYTICS PORTAL · REGULATIONS 2024 · OFFICIAL ACADEMIC REPORT</p>
</div>
<div class="meta-box">
<b>Report Title:</b> {title_esc} &nbsp;|&nbsp; <b>Generated Date:</b> {ts} &nbsp;|&nbsp; <b>Status:</b> Official Declared Results
</div>
{body_html}
<p class="note">Official University Executive Grade Analytics Report generated under Regulations 2024.<br/>
AI Advisory components are derived from declared metrics; verify against official registrar records.</p>
</body></html>"""


def pdf_from_html(html_str: str, title: str) -> Optional[bytes]:
    if not _PISA_OK:
        return None
    bio = io.BytesIO()
    try:
        status = pisa.CreatePDF(
            src=html_str,
            dest=bio,
            encoding="utf-8",
        )
        if status.err:
            return None
        return bio.getvalue()
    except Exception:
        return None


def build_class_pdf(ca: ClassAnalysis, ai_text: str) -> Optional[bytes]:
    parts = []

    # Requirement 5: Cover / Executive Summary Section
    parts.append("<h2>Executive Summary & Cohort Overview</h2>")
    parts.append("<table><tr><th>Metric</th><th>Value</th><th>Metric</th><th>Value</th></tr>")
    parts.append(
        f"<tr><td>Total Students</td><td><b>{ca.student_count}</b></td>"
        f"<td>Class GPA Average</td><td><b>{fmt_gpa(ca.class_gpa)}</b></td></tr>"
        f"<tr><td>Total Evaluated Subjects</td><td><b>{ca.subject_count}</b></td>"
        f"<td>Cleared Pass Rate</td><td><b>{fmt_np_pct(ca.pass_rate)}</b></td></tr>"
        f"<tr><td>Students Cleared</td><td><b>{ca.cleared_count}</b></td>"
        f"<td>Students with Arrears (U/RA)</td><td><b class='alert-red'>{ca.arrear_student_count}</b></td></tr>"
        f"<tr><td>Multiple Arrears</td><td><b class='alert-red'>{ca.multiple_u_count}</b></td>"
        f"<td>Malpractice Records (MM/WH2)</td><td><b>{ca.malpractice_student_count}</b></td></tr>"
    )
    parts.append("</table>")

    # Static PNG chart via Kaleido if available, otherwise pure CSS bar fallback
    gd_png = fig_to_png(fig_grade_distribution(ca))
    if gd_png:
        parts.append(f"<h3>Grade Distribution Chart</h3><p><img src='{png_data_url(gd_png)}' style='width:100%; max-height:240px;' /></p>")
    else:
        gd_items = [(g, ca.grade_distribution.get(g, 0), _grade_color_cls(g))
                    for g in GRADE_ORDER]
        parts.append(_css_bars("Grade Distribution Summary", gd_items))

    # Subject Performance Ranking Section
    parts.append("<h2>Subject Performance Ranking</h2><table>"
                 "<tr><th>Rank</th><th>Subject</th><th>Code</th><th>Credits</th>"
                 "<th>Students</th><th>Avg GP</th><th>Pass %</th><th>Priority Rating</th></tr>")
    ranked_subjects = sorted(ca.subjects, key=lambda s: (s.pass_pct or 0, s.avg_gp or 0), reverse=True)
    for idx, s in enumerate(ranked_subjects, start=1):
        p_cls = "alert-red" if s.priority_level == "High Attention" else ""
        parts.append(
            f"<tr><td>{idx}</td><td>{html.escape(s.subject)}</td><td>{html.escape(s.course_code or '—')}</td>"
            f"<td>{s.credits}</td><td>{s.student_count}</td>"
            f"<td>{fmt_gp(s.avg_gp)}</td><td>{fmt_np_pct(s.pass_pct)}</td>"
            f"<td><span class='{p_cls}'>{s.priority_level}</span></td></tr>"
        )
    parts.append("</table>")

    # High Priority Student Support Section
    attn = [s for s in ca.students if s.attention != STATUS_CLEARED]
    if attn:
        parts.append("<h2>High Priority Student Support List</h2><table>"
                     "<tr><th>Reg No</th><th>Name</th><th>Academic Status</th><th>GPA</th><th>Arrears</th></tr>")
        for s in attn[:25]:
            parts.append(
                f"<tr><td>{html.escape(s.regno)}</td><td>{html.escape(s.name)}</td>"
                f"<td>{html.escape(s.attention.upper())}</td><td>{fmt_gpa(s.gpa)}</td>"
                f"<td>{s.arrear_count}</td></tr>"
            )
        parts.append("</table>")

    # Top Rankings
    ranked = [s for s in ca.students if s.gpa is not None]
    ranked.sort(key=lambda s: (s.rank or 9999, s.regno))
    if ranked:
        parts.append("<h2>Class Rankings (Top 25)</h2><table>"
                     "<tr><th>Rank</th><th>Reg No</th><th>Name</th><th>GPA</th><th>Percentile</th></tr>")
        for s in ranked[:25]:
            parts.append(
                f"<tr><td>{s.rank}</td><td>{html.escape(s.regno)}</td><td>{html.escape(s.name)}</td>"
                f"<td>{fmt_gpa(s.gpa)}</td><td>{fmt_np_pct(s.percentile)}</td></tr>"
            )
        parts.append("</table>")

    if ai_text:
        parts.append("<h2>AI Academic Advisory Insight</h2>")
        parts.append(f"<div class='card'>{_md_to_pdf_html(ai_text)}</div>")

    body = "\n".join(parts)
    return pdf_from_html(_pdf_wrap("Executive Class Analytics Report", body), "Executive Class Analytics Report")


def build_subject_pdf(subj: SubjectAnalysis, ca: ClassAnalysis, ai_text: str) -> Optional[bytes]:
    parts = []

    parts.append(f"<h2>Subject Performance Analysis: {html.escape(subj.subject)}</h2>")
    if subj.course_code:
        parts.append(f"<p><b>Course Code:</b> {html.escape(subj.course_code)}</p>")
        staff_name = resolve_staff_for_code(subj.course_code, SESSION.get("staff_directory", {}))
        if staff_name:
            parts.append(f"<p><b>Faculty In-Charge:</b> {html.escape(staff_name)}</p>")
    parts.append("<table><tr><th>Metric</th><th>Value</th></tr>")
    for label, val in [
        ("Course Credits", subj.credits),
        ("Students Enrolled", subj.student_count),
        ("Average Grade Point", fmt_gp(subj.avg_gp)),
        ("Median Grade Point", fmt_gp(subj.median_gp)),
        ("Pass Count", f"{subj.pass_count} ({fmt_np_pct(subj.pass_pct)})"),
        ("U Count (Arrear)", subj.u_count),
        ("RA Count (Arrear)", subj.ra_count),
        ("SA Count (Attendance)", subj.sa_count),
        ("WD Count (Withdrawal)", subj.wd_count),
        ("MM / WH2 (Malpractice)", subj.mm_count + subj.wh2_count),
        ("Priority Rating", subj.priority_level),
    ]:
        parts.append(f"<tr><td>{html.escape(str(label))}</td><td>{html.escape(str(val))}</td></tr>")
    parts.append("</table>")

    gd_items = [(g, subj.grade_counts.get(g, 0), _grade_color_cls(g))
                for g in GRADE_ORDER if subj.grade_counts.get(g, 0) > 0]
    parts.append(_css_bars("Subject Grade Distribution", gd_items))

    if subj.top_students:
        parts.append("<h3>Top Performers</h3><table>"
                     "<tr><th>Reg No</th><th>Name</th><th>Grade</th><th>Points</th></tr>")
        for st in subj.top_students[:10]:
            parts.append(
                f"<tr><td>{html.escape(st['regno'])}</td><td>{html.escape(st['name'])}</td>"
                f"<td>{html.escape(st['grade'])}</td><td>{st['points']}</td></tr>"
            )
        parts.append("</table>")

    if subj.u_students:
        parts.append("<h3 class='alert-red'>Students with Arrear Grade (U / RA)</h3><table>"
                     "<tr><th>Reg No</th><th>Name</th><th>Grade</th></tr>")
        for st in subj.u_students:
            parts.append(f"<tr><td>{html.escape(st['regno'])}</td><td>{html.escape(st['name'])}</td><td>{html.escape(st['grade'])}</td></tr>")
        parts.append("</table>")

    parts.append("<h3>Subject vs Class Benchmark</h3><table>"
                 "<tr><th>Metric</th><th>This Subject</th><th>Class Overall Average</th></tr>")
    parts.append(
        f"<tr><td>Average GP</td><td>{fmt_gp(subj.avg_gp)}</td>"
        f"<td>{fmt_gp(ca.class_gpa)}</td></tr>"
    )
    parts.append(
        f"<tr><td>Pass Rate</td><td>{fmt_np_pct(subj.pass_pct)}</td>"
        f"<td>{fmt_np_pct(ca.pass_rate)}</td></tr>"
    )
    parts.append("</table>")

    if ai_text:
        parts.append("<h2>AI Subject Advisory</h2>")
        parts.append(f"<div class='card'>{_md_to_pdf_html(ai_text)}</div>")

    body = "\n".join(parts)
    return pdf_from_html(_pdf_wrap(f"Subject Report – {subj.subject}", body),
                         f"Subject Report – {subj.subject}")


def build_student_pdf(student: StudentAnalysis, ca: ClassAnalysis, ai_text: str) -> Optional[bytes]:
    parts = []

    parts.append(f"<h2>Student Academic Profile: {html.escape(student.name)}</h2>")
    parts.append(f"<p><b>Register Number:</b> {html.escape(student.regno)}</p>")
    if student.meta.get("dept"):
        parts.append(f"<p><b>Department:</b> {html.escape(student.meta['dept'])}</p>")

    parts.append("<h3>Academic Snapshot</h3><table><tr><th>Metric</th><th>Value</th></tr>")
    for label, val in [
        ("Semester GPA", fmt_gpa(student.gpa)),
        ("Class GPA Average", fmt_gpa(ca.class_gpa)),
        ("Class Rank", rank_text(student, ca.student_count)),
        ("Percentile Position", fmt_np_pct(student.percentile)),
        ("Total Registered Courses", student.total_courses),
        ("Credits Completed", f"{student.credits_completed:.1f}"),
        ("Credits Attempted", f"{student.credits_attempted:.1f}"),
        ("Quality Points Earned", f"{student.quality_points:.1f}"),
        ("Courses Passed", student.passed_courses),
        ("Arrears (U / RA)", f"{student.arrear_count} (U:{student.u_count}, RA:{student.ra_count})"),
        ("SA (Attendance Shortage)", student.sa_count),
        ("WD (Withdrawn)", student.wd_count),
        ("Malpractice (MM / WH2)", f"{student.malpractice_count} (MM:{student.mm_count}, WH2:{student.wh2_count})"),
        ("Academic Status", student.attention.upper()),
    ]:
        parts.append(f"<tr><td>{html.escape(str(label))}</td><td>{html.escape(str(val))}</td></tr>")
    parts.append("</table>")

    s_png = fig_to_png(fig_student_subjects(student))
    if s_png:
        parts.append(f"<h3>Subject-wise Grade Points Chart</h3><p><img src='{png_data_url(s_png)}' style='width:100%; max-height:240px;' /></p>")
    else:
        gp_items = [(g, student.grade_counts.get(g, 0), _grade_color_cls(g))
                    for g in GRADE_ORDER if student.grade_counts.get(g, 0) > 0]
        parts.append(_css_bars("Student Grade Profile", gp_items))

    if student.courses:
        parts.append("<h3>Semester Course Results & Quality Points</h3><table>"
                     "<tr><th>Subject</th><th>Credits</th><th>Grade</th>"
                     "<th>Points</th><th>Quality Points</th></tr>")
        for c in student.courses:
            status = "✓" if c.grade in PASSING_GRADES else "✗"
            parts.append(
                f"<tr><td>{html.escape(c.subject)}</td><td>{c.credits}</td>"
                f"<td>{html.escape(c.grade)}</td><td>{c.points}</td>"
                f"<td>{c.credits * c.points:.1f} {status}</td></tr>"
            )
        if student.credits_completed > 0:
            parts.append(
                f"<tr><td><b>Total (passing)</b></td>"
                f"<td>{student.credits_completed:.1f}</td>"
                f"<td>—</td><td>—</td>"
                f"<td>{student.quality_points:.1f}</td></tr>"
            )
            parts.append(
                f"<tr><td><b>Semester GPA</b></td><td colspan='3'>({student.quality_points:.1f} / "
                f"{student.credits_completed:.1f})</td>"
                f"<td><b>{fmt_gpa(student.gpa)}</b></td></tr>"
            )
        parts.append("</table>")

    if student.strongest_subjects:
        strong_esc = html.escape(", ".join(student.strongest_subjects[:6]))
        parts.append(f"<h3>Strongest Performance Subjects</h3><p>{strong_esc}</p>")
    if student.attention_subjects:
        attn_esc = html.escape(", ".join(student.attention_subjects))
        parts.append(f"<h3 class='alert-red'>Subjects Requiring Attention</h3><p>{attn_esc}</p>")

    if ai_text:
        parts.append("<h2>AI Academic Brief & PTM Discussion Points</h2>")
        parts.append(f"<div class='card'>{_md_to_pdf_html(ai_text)}</div>")

    body = "\n".join(parts)
    return pdf_from_html(
        _pdf_wrap(f"Student Report – {student.name} ({student.regno})", body),
        f"Student Report – {student.name}"
    )


# =============================================================================
# 10) REPORT CACHE & UUID VALIDATION
# =============================================================================

REPORT_CACHE: Dict[str, Dict[str, Any]] = {}


def is_valid_uuid(val: str) -> bool:
    try:
        uuid.UUID(str(val))
        return True
    except (ValueError, TypeError, AttributeError):
        return False


def create_report_id(kind: str, payload: Dict[str, Any]) -> str:
    rid = str(uuid.uuid4())
    REPORT_CACHE[rid] = {
        "kind": kind,
        "payload": payload,
        "created": datetime.now().isoformat(),
    }
    return rid


def get_report(rid: str) -> Optional[Dict[str, Any]]:
    if not is_valid_uuid(rid):
        return None
    return REPORT_CACHE.get(rid)


def generate_report_pdf(rid: str) -> Optional[bytes]:
    entry = get_report(rid)
    if not entry:
        return None
    kind = entry["kind"]
    pl = entry["payload"]
    ca = pl.get("class_analysis")
    if ca is None:
        return None

    if kind == "class":
        ai = generate_class_ai_insight(ca)
        return build_class_pdf(ca, ai["text"])
    elif kind == "subject":
        subj = get_subject(ca, pl.get("subject", ""))
        if not subj:
            return None
        ai = generate_subject_ai_insight(subj, ca)
        return build_subject_pdf(subj, ca, ai["text"])
    elif kind == "student":
        student = get_student(ca, pl.get("regno", ""))
        if not student:
            return None
        ai = generate_student_brief(student, ca)
        return build_student_pdf(student, ca, ai["text"])
    return None


# =============================================================================
# 11) REUSABLE FASTHTML COMPONENTS + SESSION STATE
# =============================================================================

SESSION: Dict[str, Any] = {
    "records": None,
    "analytics": None,
    "file_name": "",
    "validation": None,
    "preview_raw_bytes": None,
    "preview_filename": "",
    "preview_cols": [],
    "preview_report": None,
    "ptm_briefs": {},
    "staff_directory": {},      # Course Code -> Staff Name, remembered across PDF conversions
    "pdf_to_excel_cache": None,  # (cache_key, ClassAnalysis) so /pdf-to-excel doesn't recompute every GET
    "staff_verified": False,
}

_ALERTS: List[Tuple[str, str]] = []


def push_alert(message: str, kind: str = "blue") -> None:
    _ALERTS.append((kind, message))


def pop_alerts() -> List[Tuple[str, str]]:
    out = list(_ALERTS)
    _ALERTS.clear()
    return out


def session_ready() -> bool:
    return SESSION["analytics"] is not None


COLORS = {
    "blue": "bg-blue-50 border-blue-300 text-blue-800",
    "green": "bg-green-50 border-green-300 text-green-800",
    "red": "bg-red-50 border-red-300 text-red-800",
    "amber": "bg-amber-50 border-amber-300 text-amber-800",
    "slate": "bg-slate-50 border-slate-300 text-slate-700",
}


def md_to_html(text: str) -> str:
    if not text:
        return ""
    lines = text.split("\n")
    out = []
    for line in lines:
        s = line.strip()
        if not s:
            out.append("<br/>")
            continue
        if s.startswith("## "):
            inner = html.escape(s[3:])
            out.append(f"<h3 class='text-base font-bold text-slate-800 mt-4 mb-2 border-b pb-1'>{inner}</h3>")
        elif s.startswith("- "):
            inner = html.escape(s[2:])
            inner = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", inner)
            out.append(f"<li class='my-1 ml-4 list-disc'>{inner}</li>")
        elif s.startswith("**") and s.endswith("**"):
            inner = html.escape(s.strip("*"))
            out.append(f"<h4 class='mt-3 mb-1 font-semibold text-slate-700'>{inner}</h4>")
        else:
            inner = html.escape(s)
            inner = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", inner)
            out.append(f"<p class='my-1'>{inner}</p>")
    return "".join(out)


def grade_badge_cls(grade: str) -> str:
    g = str(grade).upper().strip()
    if g in ("O", "A+", "A", "B+", "B", "C"):
        return "bg-green-100 text-green-800"
    elif g in ("U", "RA"):
        return "bg-red-100 text-red-800"
    elif g in ("SA", "WD"):
        return "bg-amber-100 text-amber-800"
    elif g in ("MM", "WH2"):
        return "bg-purple-100 text-purple-800"
    return "bg-slate-100 text-slate-800"


def explain_insight_box(title: str, reasons: List[str], id_suffix: str = "") -> Div:
    box_id = f"explain-box-{id_suffix}" if id_suffix else "explain-box"
    return Div(
        Button(
            "❓ Why is this flagged?",
            type="button",
            onclick=f"document.getElementById('{box_id}').classList.toggle('hidden')",
            cls="px-2 py-0.5 text-xs font-semibold text-blue-700 bg-blue-50 hover:bg-blue-100 rounded border border-blue-200 transition-colors inline-flex items-center gap-1 cursor-pointer"
        ),
        Div(
            Div(
                P(title, cls="text-xs font-bold text-slate-900 mb-0.5"),
                P("Calculated from result data", cls="text-[10px] font-semibold text-slate-500 uppercase tracking-wider mb-2 block"),
                *[Div(
                    Span("•", cls="text-blue-600 font-bold mr-1.5 text-xs"),
                    Span(r, cls="text-xs text-slate-700"),
                    cls="flex items-start py-0.5"
                ) for r in reasons],
                cls="p-3 bg-blue-50/90 border border-blue-200 rounded-lg mt-2 shadow-xs"
            ),
            id=box_id,
            cls="hidden mt-1"
        ),
        cls="inline-block"
    )


def grade_badge(grade: str) -> Span:
    css_class = grade.replace("+", "PLUS").replace(" ", "")
    return Span(grade, cls=f"grade-badge grade-{css_class}")


def status_badge(status: str) -> Span:
    m = {
        STATUS_CLEARED: ("Cleared", "badge-green"),
        STATUS_U: ("Arrear (1)", "badge-red"),
        STATUS_MULTI_U: ("Multi-Arrear", "badge-red"),
        STATUS_BACKLOG: ("Sem 1-4 Backlog", "badge-amber"),
        STATUS_SA: ("SA (Attendance)", "badge-amber"),
        STATUS_WD: ("WD (Withdrawn)", "badge-purple"),
        STATUS_MALPRACTICE: ("🚨 Malpractice", "badge-indigo"),
    }
    label, cls = m.get(status, ("—", "badge-slate"))
    return Span(label, cls=f"badge {cls}")


def gpa_cell(gpa: Optional[float]) -> Span:
    if gpa is None:
        return Span("—", cls="text-slate-300")
    if gpa >= 8.5:
        color = "text-green-600"
    elif gpa >= 7.0:
        color = "text-blue-600"
    elif gpa >= 5.0:
        color = "text-amber-600"
    else:
        color = "text-red-600"
    return Span(f"{gpa:.2f}", cls=f"font-semibold {color}")


def alert_bar(kind: str, message: str) -> Div:
    icon_map = {"blue": "ℹ", "green": "✓", "red": "✕", "amber": "⚠"}
    icon = icon_map.get(kind, "ℹ")
    return Div(
        Span(icon, cls="text-lg"),
        Span(message, cls="text-sm font-medium"),
        cls=f"alert alert-{kind}",
    )


def card(*children, cls: str = "") -> Div:
    return Div(*children, cls=f"card {cls}")


def stat_card(label: str, value: str, accent: str = C_BLUE, sub: str = "") -> Div:
    inner = [
        P(label, cls="text-xs font-semibold text-slate-400 uppercase tracking-wider mb-1"),
        P(value, cls="text-3xl font-bold text-slate-800 tracking-tight"),
    ]
    if sub:
        inner.append(P(sub, cls="text-xs text-slate-500 mt-1"))
    return Div(*inner, cls="stat-card", style=f"--accent-color: {accent};")


def chart_container(fig, height: str = "h-80") -> Div:
    return Div(NotStr(chart_html(fig)), cls=f"chart-container {height} overflow-hidden")


def md_block(text: str, live: str = "fallback") -> Div:
    if live == "live":
        provider_name = "Ollama Cloud" if CFG.get("ai_provider") == "ollama_cloud" else "Local Ollama"
        banner_text = f"AI advisory generated by {provider_name} ({CFG.get('ollama_model', 'llama3')})."
        banner_cls = "bg-blue-50 text-blue-700 border-blue-200"
    else:
        banner_text = "AI insights are currently unavailable. Analytics calculations are still available."
        banner_cls = "bg-amber-50 text-amber-800 border-amber-200"

    return Div(
        Div(
            Span(banner_text, cls="text-xs font-medium"),
            cls=f"px-3 py-2 border-b {banner_cls}"
        ),
        Div(NotStr(md_to_html(text)), cls="prose p-4 sm:p-5"),
        cls="card overflow-hidden"
    )


def data_table(headers: List[str], rows: List, sortable: bool = False, table_id: str = "") -> Div:
    ths = [Th(h, scope="col") for h in headers]
    tbody_rows = []
    for row in rows:
        if hasattr(row, "tag") and getattr(row, "tag", "") == "tr":
            tbody_rows.append(row)
        elif isinstance(row, (list, tuple)):
            tds = []
            for cell in row:
                if hasattr(cell, "tag") and getattr(cell, "tag", "") == "td":
                    tds.append(cell)
                else:
                    tds.append(Td(cell if cell is not None else "—"))
            tbody_rows.append(Tr(*tds))
    tbody_id = f"{table_id}-tbody" if table_id else ""
    return Div(
        Table(
            Thead(Tr(*ths)),
            Tbody(*tbody_rows, id=tbody_id),
            id=table_id,
            cls="w-full"
        ),
        cls="overflow-x-auto"
    )


def sidebar(active: str, ca: Optional[ClassAnalysis] = None) -> Div:
    ICONS = {
        "upload": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><path stroke-linecap="round" stroke-linejoin="round" d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M15 13l-3-3m0 0l-3 3m3-3v12"/></svg>',
        "dashboard": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7" rx="1" stroke-linecap="round" stroke-linejoin="round"/><rect x="14" y="3" width="7" height="7" rx="1" stroke-linecap="round" stroke-linejoin="round"/><rect x="3" y="14" width="7" height="7" rx="1" stroke-linecap="round" stroke-linejoin="round"/><rect x="14" y="14" width="7" height="7" rx="1" stroke-linecap="round" stroke-linejoin="round"/></svg>',
        "students": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><path stroke-linecap="round" stroke-linejoin="round" d="M17 20h5v-2a3 3 0 00-5.356-1.857M17 20H7m10 0v-2c0-.656-.126-1.283-.356-1.857M7 20H2v-2a3 3 0 015.356-1.857M7 20v-2c0-.656.126-1.283.356-1.857m0 0a5.002 5.002 0 019.288 0M15 7a3 3 0 11-6 0 3 3 0 016 0z"/></svg>',
        "subjects": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><path stroke-linecap="round" stroke-linejoin="round" d="M12 6.253v13m0-13C10.832 5.477 9.246 5 7.5 5S4.168 5.477 3 6.253v13C4.168 18.477 5.754 18 7.5 18s3.332.477 4.5 1.253m0-13C13.168 5.477 14.754 5 16.5 5c1.747 0 3.332.477 4.5 1.253v13C19.832 18.477 18.247 18 16.5 18c-1.746 0-3.332.477-4.5 1.253"/></svg>',
        "attention": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><path stroke-linecap="round" stroke-linejoin="round" d="M12 9v2m0 4h.01m-6.938 4h13.856c1.54 0 2.502-1.667 1.732-3L13.732 4c-.77-1.333-2.694-1.333-3.464 0L3.34 16c-.77 1.333.192 3 1.732 3z"/></svg>',
        "rankings": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><path stroke-linecap="round" stroke-linejoin="round" d="M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z"/></svg>',
        "ai-insights": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><path stroke-linecap="round" stroke-linejoin="round" d="M9.663 17h4.673M12 3v1m6.364 1.636l-.707.707M21 12h-1M4 12H3m3.343-5.657l-.707-.707m2.828 9.9a5 5 0 117.072 0l-.548.547A3.374 3.374 0 0014 18.469V19a2 2 0 11-4 0v-.531c0-.895-.356-1.754-.988-2.386l-.548-.547z"/></svg>',
        "reports": '<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2"><path stroke-linecap="round" stroke-linejoin="round" d="M9 17v-2m3 2v-4m3 4v-6m2 10H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>',
    }

    items = [
        ("/upload", "Upload", "upload"),
        ("/dashboard", "Dashboard", "dashboard"),
        ("/students", "Students", "students"),
        ("/subjects", "Subjects", "subjects"),
        ("/attention", "Attention", "attention"),
        ("/rankings", "Rankings", "rankings"),
        ("/ai-insights", "AI Insights", "ai-insights"),
        ("/reports", "Reports", "reports"),
    ]
    nav_items = []
    for href, label, key in items:
        is_active = active == key
        active_cls = "active" if is_active else ""
        nav_items.append(
            A(
                NotStr(ICONS.get(key, "")),
                Span(label, cls="sidebar-text"),
                href=href,
                cls=f"sidebar-link {active_cls}",
                aria_label=f"Navigate to {label}"
            )
        )

    return Div(
        Div(
            # College Brand Header
            Div(
                Img(src="/logo", alt="Saranathan College of Engineering", cls="sidebar-brand-logo"),
                Span("SARANATHAN", cls="sidebar-brand-name"),
                Span("COLLEGE OF ENGINEERING", cls="sidebar-brand-sub"),
                Span("TIRUCHIRAPPALLI", cls="sidebar-brand-city"),
                Div(
                    Span(CFG["app_title"], cls="portal-name"),
                    Span(f"{CFG['app_version']} · R2024 Compatible", cls="portal-ver"),
                    cls="sidebar-portal-label"
                ),
                cls="sidebar-brand"
            ),
            # Navigation
            Div(
                Span("NAVIGATION", cls="sidebar-nav-label"),
                *nav_items,
                cls="sidebar-nav-section"
            ),
            # User Profile Footer
            Div(
                Div(
                    Div("F", cls="sidebar-avatar"),
                    Div(
                        Span("Faculty", cls="sidebar-user-name"),
                        Span("AI & Data Science", cls="sidebar-user-role"),
                    ),
                    cls="sidebar-user"
                ),
                cls="sidebar-footer"
            ),
            cls="sidebar hidden lg:flex flex-col w-64 min-h-screen fixed left-0 top-0"
        ),
    )



def mobile_header(active: str) -> Div:
    items = [
        ("/upload", "Upload", "upload"),
        ("/dashboard", "Dashboard", "dashboard"),
        ("/students", "Students", "students"),
        ("/subjects", "Subjects", "subjects"),
        ("/attention", "Attention", "attention"),
        ("/rankings", "Rankings", "rankings"),
        ("/ai-insights", "AI Insights", "ai-insights"),
        ("/reports", "Reports", "reports"),
    ]
    links = []
    for href, label, key in items:
        is_active = active == key
        active_cls = "bg-blue-50 text-blue-700 font-semibold" if is_active else "text-slate-600"
        links.append(A(label, href=href, cls=f"block px-4 py-2.5 text-sm font-medium {active_cls} hover:bg-slate-50 hover:text-slate-900 transition-colors"))
    return Div(
        Div(
            Button(
                NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2" style="width:20px;height:20px"><path stroke-linecap="round" stroke-linejoin="round" d="M4 6h16M4 12h16M4 18h16"/></svg>'),
                onclick="document.getElementById('mobnav').classList.toggle('hidden')",
                cls="p-2 hover:bg-slate-100 rounded-lg text-slate-600 border-0 bg-transparent shadow-none",
                aria_label="Toggle Navigation"
            ),
            Div(
                Img(src="/logo", alt="SCE", style="width:32px;height:32px;border-radius:50%;border:1px solid #e5eaf3;object-fit:contain;"),
                Div(
                    Span(CFG["app_title"], cls="text-sm font-bold text-slate-900 block leading-tight"),
                    Span("Saranathan College of Engineering", cls="text-[10px] text-slate-500 block"),
                ),
                cls="flex items-center gap-2 flex-1"
            ),
            cls="flex items-center gap-3 px-4 py-3 mobile-header-bar mobile-nav"
        ),
        Div(*links, id="mobnav", cls="hidden bg-white border-b border-slate-200 pb-2 lg:hidden shadow-sm"),
        cls="lg:hidden"
    )



def layout(title: str, active: str, content, ca: Optional[ClassAnalysis] = None) -> Tuple:
    reset_plotly_flag()
    alerts = pop_alerts()
    alert_divs = [alert_bar(k, m) for k, m in alerts]

    app_footer = Div(
        P("Developed by Jeffin Josva S and Srimuthukrishnan S from AI&DS 2024-2028 Batch",
          cls="text-xs font-semibold text-slate-500 text-center tracking-wide"),
        cls="border-t border-slate-200/80 pt-6 mt-12 pb-4 w-full"
    )

    main_content = Div(*alert_divs, content, app_footer,
                       cls="flex-1 max-w-full p-4 sm:p-6 lg:p-8 ml-0 lg:ml-64 min-h-screen bg-slate-50 main-content")

    return (
        Title(title + " – " + CFG["app_title"]),
        mobile_header(active),
        sidebar(active, ca),
        main_content,
        Script("""
        document.addEventListener('DOMContentLoaded', function() {
            document.querySelectorAll('[data-filter]').forEach(function(btn) {
                btn.addEventListener('click', function() {
                    var target = this.getAttribute('data-filter-target') || 'student-row';
                    var val = this.getAttribute('data-filter');
                    document.querySelectorAll('tr[data-row-type="' + target + '"]').forEach(function(tr) {
                        if (val === 'all') {
                            tr.style.display = '';
                        } else if (val === 'high') {
                            tr.style.display = (tr.getAttribute('data-high') === 'true') ? '' : 'none';
                        } else {
                            tr.style.display = (tr.getAttribute('data-status') === val) ? '' : 'none';
                        }
                    });
                    this.closest('.filter-group').querySelectorAll('.filter-btn').forEach(function(b) {
                        b.classList.remove('active');
                    });
                    this.classList.add('active');
                });
            });
            var mobToggle = document.getElementById('mobnav');
            if (mobToggle) {
                document.addEventListener('click', function(e) {
                    if (!e.target.closest('.mobile-nav') && !mobToggle.contains(e.target)) {
                        mobToggle.classList.add('hidden');
                    }
                });
            }
        });
        """),
    )


# =============================================================================
# 12) PAGE FUNCTIONS
# =============================================================================

def page_upload() -> Tuple:
    return layout("Upload", "upload", Div(
        # ── TOP HEADER BAR ──────────────────────────────────────────────
        Div(
            Div(
                Span("FACULTY GRADE ANALYTICS PORTAL", cls="text-[11px] font-extrabold uppercase tracking-wider text-blue-700 bg-blue-50 px-3.5 py-1 rounded-full border border-blue-200/80 inline-block shadow-2xs"),
                cls="flex items-center"
            ),
            Div(
                # Light/Dark Theme Toggle
                Button(
                    NotStr('<svg id="theme-toggle-icon" xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2" style="width:15px;height:15px;"><path stroke-linecap="round" stroke-linejoin="round" d="M12 3v1m0 16v1m9-9h-1M4 12H3m15.364 6.364l-.707-.707M6.343 6.343l-.707-.707m12.728 0l-.707.707M6.343 17.657l-.707.707M16 12a4 4 0 11-8 0 4 4 0 018 0z"/></svg>'),
                    id="theme-toggle-btn",
                    onclick="toggleTheme()",
                    title="Toggle dark / light mode",
                    cls="w-8 h-8 rounded-full border border-slate-200 bg-white hover:bg-slate-50 flex items-center justify-center cursor-pointer transition-all"
                ),
                cls="flex items-center gap-3"
            ),
            cls="flex items-center justify-between mb-8 pb-3 border-b border-slate-200/60"
        ),

        # ── HERO SECTION ────────────────────────────────────────────────
        Div(
            # Left illustration — classical institution building + college logo
            Div(
                NotStr("""<svg viewBox="0 0 200 200" xmlns="http://www.w3.org/2000/svg" style="width:100%;height:auto;">
  <!-- Soft Cloud Background -->
  <path d="M 20,130 Q 10,90 50,70 Q 70,40 110,40 Q 150,40 170,75 Q 195,95 180,130 Z" fill="#dbeafe" opacity="0.65"/>
  <ellipse cx="100" cy="120" rx="85" ry="45" fill="#bfdbfe" opacity="0.45"/>
  <!-- College Logo Badge above Building -->
  <circle cx="100" cy="30" r="19" fill="#ffffff" stroke="#93c5fd" stroke-width="2"/>
  <image href="/logo" x="83" y="13" width="34" height="34"/>
  <!-- Institution Temple Roof / Pediment -->
  <polygon points="100,54 38,80 162,80" fill="#0052cc"/>
  <polygon points="100,58 46,78 154,78" fill="#1d4ed8"/>
  <rect x="42" y="80" width="116" height="9" fill="#ffffff" stroke="#93c5fd" stroke-width="1"/>
  <!-- Columns -->
  <rect x="50" y="89" width="13" height="53" rx="2" fill="#ffffff" stroke="#93c5fd" stroke-width="1.5"/>
  <rect x="73" y="89" width="13" height="53" rx="2" fill="#ffffff" stroke="#93c5fd" stroke-width="1.5"/>
  <rect x="94" y="89" width="13" height="53" rx="2" fill="#ffffff" stroke="#93c5fd" stroke-width="1.5"/>
  <rect x="115" y="89" width="13" height="53" rx="2" fill="#ffffff" stroke="#93c5fd" stroke-width="1.5"/>
  <rect x="137" y="89" width="13" height="53" rx="2" fill="#ffffff" stroke="#93c5fd" stroke-width="1.5"/>
  <!-- Entrance Steps / Base -->
  <rect x="34" y="142" width="132" height="14" rx="3" fill="#2563eb"/>
  <rect x="26" y="156" width="148" height="8" rx="2" fill="#1d4ed8"/>
  <!-- Doorway -->
  <path d="M 89,142 L 89,108 A 11,11 0 0,1 111,108 L 111,142 Z" fill="#1e3a8a"/>
  <!-- Windows behind columns -->
  <rect x="57" y="99" width="9" height="15" rx="1" fill="#60a5fa" opacity="0.6"/>
  <rect x="135" y="99" width="9" height="15" rx="1" fill="#60a5fa" opacity="0.6"/>
  <!-- Side Foliage / Trees -->
  <circle cx="22" cy="138" r="14" fill="#0052cc" opacity="0.85"/>
  <circle cx="14" cy="148" r="10" fill="#1d4ed8" opacity="0.85"/>
  <circle cx="178" cy="138" r="14" fill="#0052cc" opacity="0.85"/>
  <circle cx="186" cy="148" r="10" fill="#1d4ed8" opacity="0.85"/>
</svg>"""),
                cls="hidden lg:block w-40 xl:w-48 flex-shrink-0"
            ),
            # Center content
            Div(
                H1(
                    Span("Turn COE Results into ", cls="text-slate-900 block sm:inline"),
                    Span("Actionable Academic Insights", cls="text-blue-600 block sm:inline"),
                    cls="text-3xl sm:text-4xl font-extrabold tracking-tight mb-3"
                ),
                P("Upload official COE result PDFs or Excel spreadsheets. Automatically parse records, identify U/RA arrears, map subjects to the R2024 syllabus, and generate faculty-ready analytics.",
                  cls="text-slate-600 text-base max-w-xl mx-auto leading-relaxed mb-6"),
                # Stats row
                Div(
                    Div(Span("R2024", cls="text-lg font-extrabold text-blue-700 block"), Span("Regulation", cls="text-xs text-slate-500"), cls="text-center px-4"),
                    Div(cls="w-px bg-slate-200 self-stretch"),
                    Div(Span("PDF", cls="text-lg font-extrabold text-blue-700 block"), Span("Extraction", cls="text-xs text-slate-500"), cls="text-center px-4"),
                    Div(cls="w-px bg-slate-200 self-stretch"),
                    cls="flex items-center justify-center gap-1 p-3 bg-white border border-slate-200 rounded-2xl shadow-sm w-fit mx-auto"
                ),
                cls="text-center flex-1"
            ),
            # Right illustration — analytics dashboard + magnifying glass
            Div(
                NotStr("""<svg viewBox="0 0 200 200" xmlns="http://www.w3.org/2000/svg" style="width:100%;height:auto;">
  <!-- Soft Cloud Background -->
  <path d="M 20,130 Q 10,90 50,70 Q 70,40 110,40 Q 150,40 170,75 Q 195,95 180,130 Z" fill="#dbeafe" opacity="0.65"/>
  <!-- Dashboard Window -->
  <rect x="25" y="35" width="145" height="110" rx="10" fill="#ffffff" stroke="#93c5fd" stroke-width="2"/>
  <!-- Header Bar -->
  <path d="M 25,45 A 10,10 0 0,1 35,35 L 160,35 A 10,10 0 0,1 170,45 L 170,56 L 25,56 Z" fill="#edf4ff"/>
  <circle cx="40" cy="45.5" r="3.5" fill="#ef4444"/>
  <circle cx="51" cy="45.5" r="3.5" fill="#f59e0b"/>
  <circle cx="62" cy="45.5" r="3.5" fill="#10b981"/>
  <!-- Line Chart -->
  <path d="M 40,115 L 65,90 L 85,100 L 115,70 L 140,85 L 155,60" fill="none" stroke="#0052cc" stroke-width="3" stroke-linecap="round"/>
  <!-- Bar Charts -->
  <rect x="42" y="115" width="10" height="20" rx="2" fill="#93c5fd"/>
  <rect x="62" y="95" width="10" height="40" rx="2" fill="#60a5fa"/>
  <rect x="82" y="105" width="10" height="30" rx="2" fill="#3b82f6"/>
  <rect x="102" y="75" width="10" height="60" rx="2" fill="#2563eb"/>
  <rect x="122" y="90" width="10" height="45" rx="2" fill="#1d4ed8"/>
  <rect x="142" y="65" width="10" height="70" rx="2" fill="#1e3a8a"/>
  <!-- Small Floating Donut Chart Card -->
  <rect x="10" y="95" width="60" height="55" rx="8" fill="#ffffff" stroke="#bfdbfe" stroke-width="1.5"/>
  <circle cx="40" cy="122.5" r="16" fill="none" stroke="#0052cc" stroke-width="7"/>
  <circle cx="40" cy="122.5" r="16" fill="none" stroke="#60a5fa" stroke-width="7" stroke-dasharray="35 100"/>
  <!-- Large Floating Blue Magnifying Glass -->
  <g transform="translate(125, 80) rotate(-20)">
    <circle cx="20" cy="20" r="22" fill="#ffffff" stroke="#0052cc" stroke-width="5.5"/>
    <circle cx="20" cy="20" r="16" fill="#dbeafe" opacity="0.45"/>
    <rect x="16" y="42" width="8" height="26" rx="4" fill="#0052cc"/>
  </g>
</svg>"""),
                cls="hidden lg:block w-40 xl:w-48 flex-shrink-0"
            ),
            cls="flex items-center gap-6 max-w-5xl mx-auto mb-10 py-4"
        ),

        # ── UPLOAD & ACADEMIC CONTEXT CARD ───────────────────────────────
        Div(
            # Mode A: PDF Direct Upload (Primary)
            Form(
                # Step 1: Academic Department, Semester & Regulation Selection (MODERNIZED & DARK-MODE AWARE)
                Div(
                    # Step header with gradient accent bar
                    Div(
                        Div(
                            Div(
                                NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="2" stroke="currentColor" style="width:16px;height:16px"><path stroke-linecap="round" stroke-linejoin="round" d="M4.5 12a7.5 7.5 0 0 0 15 0m-15 0a7.5 7.5 0 1 1 15 0m-15 0H3m16.5 0H21m-1.5 0H12m-8.457 3.077 1.41-.513m14.095-5.13 1.41-.513M5.106 17.785l1.15-.964m11.49-9.642 1.149-.964M7.501 19.795l.75-1.3m7.5-12.99.75-1.3m-6.063 16.658.26-1.477m2.605-14.772.26-1.477m0 17.726-.26-1.477M10.698 4.614l-.26-1.477M16.5 19.794l-.75-1.299M7.5 4.205 6.75 2.906m-.63 17.668L5.7 19.3M4.5 12H3"/></svg>'),
                                cls="ctx-icon-bg w-8 h-8 rounded-xl flex items-center justify-center flex-shrink-0 text-blue-600 shadow-2xs"
                            ),
                            Div(
                                Div(
                                    Span("01", cls="text-[9px] font-black text-blue-500 tracking-widest mr-1.5 font-mono"),
                                    Span("ACADEMIC CONTEXT", cls="text-[9px] font-black uppercase tracking-widest text-blue-600 dark:text-blue-400"),
                                    cls="flex items-center mb-0.5"
                                ),
                                Span("Configure Target Department & Cohort", cls="ctx-title text-sm font-extrabold leading-tight block"),
                            ),
                            cls="flex items-center gap-3"
                        ),
                        P("Set your department, semester and regulation before uploading — enables precision syllabus mapping and cohort analytics.", cls="ctx-desc text-xs mt-3 leading-relaxed pl-11"),
                        cls="mb-5"
                    ),
                    # Fields grid — modernized
                    Div(
                        # 1. Department field
                        Div(
                            Div(
                                NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="2" stroke="currentColor" style="width:13px;height:13px"><path stroke-linecap="round" stroke-linejoin="round" d="M3.75 21h16.5M4.5 3h15M5.25 3v18m13.5-18v18M9 6.75h1.5m-1.5 3h1.5m-1.5 3h1.5m3-6H15m-1.5 3H15m-1.5 3H15M9 21v-3.375c0-.621.504-1.125 1.125-1.125h3.75c.621 0 1.125.504 1.125 1.125V21"/></svg>'),
                                Label("Department", cls="ctx-label text-[11px] font-bold uppercase tracking-wide"),
                                cls="flex items-center gap-1.5 mb-2"
                            ),
                            Div(
                                Select(
                                    Option("— Select Department —", value="", disabled=True, selected=True),
                                    *[
                                        Option(
                                            f"{d.get('name', d.get('code'))} ({d.get('code')})",
                                            value=d.get('code')
                                        )
                                        for d in get_registered_departments()
                                    ],
                                    id="select_department", name="department", required=True,
                                    cls="ctx-select w-full text-sm font-semibold rounded-xl px-3.5 py-2.5 appearance-none focus:outline-none transition-all cursor-pointer",
                                    style="padding-right:2.5rem"
                                ),
                                NotStr('<svg class="ctx-chevron" xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="2.5" stroke="#94a3b8" style="width:14px;height:14px;position:absolute;right:12px;top:50%;transform:translateY(-50%);pointer-events:none"><path stroke-linecap="round" stroke-linejoin="round" d="m19.5 8.25-7.5 7.5-7.5-7.5"/></svg>'),
                                style="position:relative"
                            ),
                            Span("Official department code", cls="ctx-hint text-[10px] mt-1.5 block"),
                            cls="flex-1"
                        ),
                        # 2. Semester field
                        Div(
                            Div(
                                NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="2" stroke="currentColor" style="width:13px;height:13px"><path stroke-linecap="round" stroke-linejoin="round" d="M6.75 3v2.25M17.25 3v2.25M3 18.75V7.5a2.25 2.25 0 0 1 2.25-2.25h13.5A2.25 2.25 0 0 1 21 7.5v11.25m-18 0A2.25 2.25 0 0 0 5.25 21h13.5A2.25 2.25 0 0 0 21 18.75m-18 0v-7.5A2.25 2.25 0 0 1 5.25 9h13.5A2.25 2.25 0 0 1 21 11.25v7.5"/></svg>'),
                                Label("Semester", cls="ctx-label text-[11px] font-bold uppercase tracking-wide"),
                                cls="flex items-center gap-1.5 mb-2"
                            ),
                            Div(
                                Select(
                                    Option("— Select Semester —", value="", disabled=True, selected=True),
                                    Option("Semester I", value="1"),
                                    Option("Semester II", value="2"),
                                    Option("Semester III", value="3"),
                                    Option("Semester IV", value="4"),
                                    Option("Semester V", value="5"),
                                    Option("Semester VI", value="6"),
                                    Option("Semester VII", value="7"),
                                    Option("Semester VIII", value="8"),
                                    id="select_semester", name="semester", required=True,
                                    cls="ctx-select w-full text-sm font-semibold rounded-xl px-3.5 py-2.5 appearance-none focus:outline-none transition-all cursor-pointer",
                                    style="padding-right:2.5rem"
                                ),
                                NotStr('<svg class="ctx-chevron" xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="2.5" stroke="#94a3b8" style="width:14px;height:14px;position:absolute;right:12px;top:50%;transform:translateY(-50%);pointer-events:none"><path stroke-linecap="round" stroke-linejoin="round" d="m19.5 8.25-7.5 7.5-7.5-7.5"/></svg>'),
                                style="position:relative"
                            ),
                            Span("Current examination cycle", cls="ctx-hint text-[10px] mt-1.5 block"),
                            cls="flex-1"
                        ),
                        # 3. Regulation field
                        Div(
                            Div(
                                NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="2" stroke="currentColor" style="width:13px;height:13px"><path stroke-linecap="round" stroke-linejoin="round" d="M9 12h3.75M9 15h3.75M9 18h3.75m3 .75H18a2.25 2.25 0 0 0 2.25-2.25V6.108c0-1.135-.845-2.098-1.976-2.192a48.424 48.424 0 0 0-1.123-.08m-5.801 0c-.065.21-.1.433-.1.664 0 .414.336.75.75.75h4.5a.75.75 0 0 0 .75-.75 2.25 2.25 0 0 0-.1-.664m-5.8 0A2.251 2.251 0 0 1 13.5 2.25H15c1.012 0 1.867.668 2.15 1.586m-5.8 0c-.376.023-.75.05-1.124.08C9.095 4.01 8.25 4.973 8.25 6.108V8.25m0 0H4.875c-.621 0-1.125.504-1.125 1.125v11.25c0 .621.504 1.125 1.125 1.125h9.75c.621 0 1.125-.504 1.125-1.125V9.375c0-.621-.504-1.125-1.125-1.125H8.25Z"/></svg>'),
                                Label("Regulation", cls="ctx-label text-[11px] font-bold uppercase tracking-wide"),
                                cls="flex items-center gap-1.5 mb-2"
                            ),
                            Div(
                                Select(
                                    Option("R2024 & R2026 (Autonomous Unified)", value="R2024", selected=True),
                                    Option("➕ Add NEW Regulation / Upload Syllabus...", value="NEW", cls="text-blue-700 font-bold bg-blue-50 dark:bg-slate-800 dark:text-blue-400"),
                                    id="select_regulation", name="regulation", required=True,
                                    cls="ctx-select w-full text-sm font-semibold rounded-xl px-3.5 py-2.5 appearance-none focus:outline-none transition-all cursor-pointer",
                                    style="padding-right:2.5rem",
                                    onchange="handleRegulationChange(this)"
                                ),
                                NotStr('<svg class="ctx-chevron" xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="2.5" stroke="#94a3b8" style="width:14px;height:14px;position:absolute;right:12px;top:50%;transform:translateY(-50%);pointer-events:none"><path stroke-linecap="round" stroke-linejoin="round" d="m19.5 8.25-7.5 7.5-7.5-7.5"/></svg>'),
                                style="position:relative"
                            ),
                            Span("R24 & R26 unified course codes", cls="ctx-hint text-[10px] mt-1.5 block"),
                            cls="flex-1"
                        ),
                        cls="grid grid-cols-1 sm:grid-cols-3 gap-5",
                        style="align-items:start"
                    ),
                    # Embedded style block with full light & dark mode tokens
                    NotStr("""<style>
                    .academic-context-card {
                        background: linear-gradient(135deg, #ffffff 0%, #f0f7ff 60%, #e8f1ff 100%);
                        border: 1.5px solid #bfdbfe;
                        border-radius: 1.25rem;
                        padding: 1.5rem;
                        box-shadow: 0 2px 12px 0 rgba(59,130,246,0.07), 0 1px 3px 0 rgba(0,0,0,0.04);
                    }
                    .academic-context-card .ctx-title { color: #0f172a; }
                    .academic-context-card .ctx-desc { color: #64748b; }
                    .academic-context-card .ctx-label { color: #475569; }
                    .academic-context-card .ctx-hint { color: #94a3b8; }
                    .academic-context-card .ctx-icon-bg {
                        background: linear-gradient(135deg, #eff6ff, #dbeafe);
                        color: #2563eb;
                    }
                    .academic-context-card .ctx-select {
                        background-color: #ffffff;
                        color: #0f172a;
                        border: 1.5px solid #cbd5e1;
                    }
                    .academic-context-card .ctx-select:focus {
                        border-color: #3b82f6 !important;
                        box-shadow: 0 0 0 3px rgba(59,130,246,0.15);
                    }

                    /* Dark Mode overrides for Academic Context */
                    body.dark .academic-context-card {
                        background: linear-gradient(135deg, #111827 0%, #172238 60%, #1b2842 100%) !important;
                        border-color: #2d456b !important;
                        box-shadow: 0 4px 20px 0 rgba(0,0,0,0.5), 0 0 15px rgba(59,130,246,0.1) !important;
                    }
                    body.dark .academic-context-card .ctx-title { color: #f8fafc !important; }
                    body.dark .academic-context-card .ctx-desc { color: #94a3b8 !important; }
                    body.dark .academic-context-card .ctx-label { color: #cbd5e1 !important; }
                    body.dark .academic-context-card .ctx-hint { color: #64748b !important; }
                    body.dark .academic-context-card .ctx-icon-bg {
                        background: linear-gradient(135deg, #1e3a8a, #1d4ed8) !important;
                        color: #93c5fd !important;
                    }
                    body.dark .academic-context-card .ctx-select {
                        background-color: #1a2333 !important;
                        color: #f1f5f9 !important;
                        border-color: #334155 !important;
                    }
                    body.dark .academic-context-card .ctx-select option {
                        background-color: #1a2333 !important;
                        color: #f1f5f9 !important;
                    }
                    body.dark .academic-context-card .ctx-select:focus {
                        border-color: #60a5fa !important;
                        box-shadow: 0 0 0 3px rgba(96,165,250,0.25) !important;
                    }
                    body.dark .academic-context-card .ctx-chevron { stroke: #64748b !important; }
                    </style>"""),
                    cls="academic-context-card mb-6"
                ),

                # Step 2: PDF Document Upload Section
                Div(
                    Div(
                        # Card header
                        Div(
                            Div(
                                Div(
                                    NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#1a56db" stroke-width="2" style="width:22px;height:22px"><path stroke-linecap="round" stroke-linejoin="round" d="M9 12h6m-6 4h6m2 5H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>'),
                                    cls="w-10 h-10 rounded-xl bg-blue-50 border border-blue-100 flex items-center justify-center flex-shrink-0"
                                ),
                                Div(
                                    Div(
                                        Span("STEP 2", cls="inline-flex items-center text-[10px] font-extrabold text-blue-700 bg-blue-100 px-2.5 py-0.5 rounded-full uppercase tracking-wider mr-2"),
                                        Span("RECOMMENDED FORMAT", cls="inline-flex items-center text-[10px] font-bold text-slate-600 bg-slate-100 px-2 py-0.5 rounded-full tracking-wider uppercase"),
                                    ),
                                    H3("Upload Official COE Result PDF", cls="text-base font-extrabold text-slate-900 leading-tight mt-1"),
                                ),
                                cls="flex items-start gap-3"
                            ),
                            cls="mb-3"
                        ),
                        P("Drop your official result PDF here or click to browse. Automatically extracts student records, segregates arrears, & maps syllabus.", cls="text-xs text-slate-500 mb-4 leading-relaxed"),

                        # PDF Input & Select
                        Div(
                            # Primary Batch PDF Input Dropzone
                            Div(
                                Div(
                                    Div(
                                        NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#1a56db" stroke-width="2" style="width:22px;height:22px;flex-shrink:0"><path stroke-linecap="round" stroke-linejoin="round" d="M9 12h6m-6 4h6m2 5H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>'),
                                        Div(
                                            Label("Current Batch COE Result PDF:", cls="block text-xs font-bold text-slate-800 leading-snug"),
                                            Span("Select official Saranathan College Of Engineering grade sheet PDF", cls="block text-[11px] text-slate-500 font-normal"),
                                        ),
                                        cls="flex items-center gap-2.5 mb-2"
                                    ),
                                    Input(type="file", name="file_pdf", accept=".pdf", required=True, id="file_pdf_input",
                                          cls="block w-full text-xs text-slate-600 bg-white border border-slate-200 rounded-lg p-1.5 cursor-pointer shadow-sm",
                                          onchange="handlePdfFileSelect(this)"),
                                ),
                                cls="fancy-file-dropzone mb-3"
                            ),
                            # Optional Previous Batch PDF Dropzone
                            Div(
                                Div(
                                    Div(
                                        NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#64748b" stroke-width="2" style="width:20px;height:20px;flex-shrink:0"><path stroke-linecap="round" stroke-linejoin="round" d="M8 7v8a2 2 0 002 2h6M8 7V5a2 2 0 012-2h4.586a1 1 0 01.707.293l4.414 4.414a1 1 0 01.293.707V15a2 2 0 01-2 2h-2M8 7H6a2 2 0 00-2 2v10a2 2 0 002 2h8a2 2 0 002-2v-2"/></svg>'),
                                        Div(
                                            Label("Previous Batch COE Result PDF (Optional):", cls="block text-xs font-bold text-slate-700 leading-snug"),
                                            Span("Optional for Sheet 2 comparison", cls="block text-[11px] text-slate-400 font-normal"),
                                        ),
                                        cls="flex items-center gap-2 mb-2"
                                    ),
                                    Input(type="file", name="file_prev_pdf", accept=".pdf", id="file_prev_pdf_input",
                                          cls="block w-full text-xs text-slate-600 bg-white border border-slate-200 rounded-lg p-1.5 cursor-pointer shadow-sm"),
                                ),
                                cls="fancy-file-dropzone mb-3"
                            ),
                            Div(
                                Span("✓ Selected PDF:", cls="text-xs font-bold text-blue-700 block mb-0.5"),
                                Span("", id="pdf_file_name", cls="text-xs font-mono font-semibold text-slate-800 break-all block"),
                                id="pdf_file_selected_info",
                                cls="hidden mt-2 p-2.5 bg-blue-50 border border-blue-200 rounded-lg text-center"
                            ),
                            cls="mb-4"
                        ),

                        Button("Analyze PDF & Preview Extraction →", type="submit", id="btn_analyze_pdf",
                               cls="w-full bg-blue-600 hover:bg-blue-700 text-white font-semibold py-3 px-4 rounded-xl text-sm transition-all shadow-md hover:shadow-lg flex items-center justify-center gap-2"),

                        # Stage Progress Container (Hidden by default)
                        Div(
                            Div(Span("⏳ Processing COE PDF...", cls="text-xs font-bold text-blue-800 block mb-2")),
                            Div(
                                Div(Span("✓", cls="text-green-600 font-bold mr-2"), Span("Reading document stream", cls="text-xs text-slate-700"), cls="flex items-center text-xs py-0.5"),
                                Div(Span("•", cls="text-blue-500 font-bold mr-2 animate-pulse"), Span("Detecting student records & subjects...", cls="text-xs text-slate-700"), cls="flex items-center text-xs py-0.5"),
                                Div(Span("•", cls="text-slate-300 font-bold mr-2"), Span("Mapping syllabus catalog...", cls="text-xs text-slate-400"), cls="flex items-center text-xs py-0.5"),
                                Div(Span("•", cls="text-slate-300 font-bold mr-2"), Span("Preparing preflight review...", cls="text-xs text-slate-400"), cls="flex items-center text-xs py-0.5"),
                                cls="bg-blue-50 p-3 rounded-lg border border-blue-100 mt-3"
                            ),
                            id="pdf_progress_card", cls="hidden mt-3"
                        ),
                        cls="card p-6 border-t-4 border-t-blue-600 hover:border-blue-600/70 transition-all shadow-sm flex flex-col justify-between h-full"
                    ),
                ),
                action="/upload-pdf", method="POST", enctype="multipart/form-data",
                onsubmit="return handleUploadFormSubmit(this, 'pdf_progress_card', 'Analyzing COE PDF...')"
            ),
            cls="max-w-2xl mx-auto mb-14"
        ),

        # ── NEW REGULATION & SYLLABUS UPLOAD MODAL ──────────────────────
        Div(
            Div(
                # Modal Header
                Div(
                    Div(
                        H3("Upload & Register New Regulation Syllabus", cls="text-base font-extrabold text-slate-900"),
                        P("Extract course codes, names, credits, and semesters into the authoritative JSON catalog.", cls="text-xs text-slate-500 mt-0.5"),
                    ),
                    Button(
                        "✕", type="button", onclick="closeNewSyllabusModal()",
                        cls="text-slate-400 hover:text-slate-700 font-bold text-lg p-1 bg-transparent border-0 cursor-pointer"
                    ),
                    cls="flex items-center justify-between pb-3 mb-4 border-b border-slate-200"
                ),

                # Modal Inputs Grid
                Div(
                    Div(
                        Label("Target Department *", cls="block text-xs font-bold text-slate-800 mb-1"),
                        Select(
                            *[
                                Option(
                                    f"{d.get('name', d.get('code'))} ({d.get('code')})",
                                    value=d.get('code')
                                )
                                for d in get_registered_departments()
                            ],
                            id="modal_dept_select",
                            cls="w-full text-xs font-medium text-slate-800 bg-white border border-slate-300 rounded-lg p-2 shadow-2xs"
                        ),
                        cls="flex-1"
                    ),
                    Div(
                        Label("New Regulation Code *", cls="block text-xs font-bold text-slate-800 mb-1"),
                        Input(
                            type="text", id="modal_reg_input", placeholder="e.g. R2028 or R2026-REV",
                            cls="w-full text-xs font-medium text-slate-800 bg-white border border-slate-300 rounded-lg p-2 shadow-2xs uppercase"
                        ),
                        cls="flex-1"
                    ),
                    cls="grid grid-cols-1 sm:grid-cols-2 gap-3 mb-3"
                ),

                Div(
                    Label("Syllabus File (.pdf or .json) *", cls="block text-xs font-bold text-slate-800 mb-1"),
                    Input(
                        type="file", id="modal_syllabus_file", accept=".pdf,.json",
                        cls="block w-full text-xs text-slate-600 bg-white border border-slate-300 rounded-lg p-2 cursor-pointer shadow-2xs"
                    ),
                    cls="mb-4"
                ),

                Button(
                    "⚡ Extract & Preview Syllabus Courses", type="button", id="btn_extract_syllabus", onclick="extractSyllabusDraft()",
                    cls="w-full bg-blue-600 hover:bg-blue-700 text-white font-semibold py-2.5 px-4 rounded-xl text-xs transition-all shadow-sm flex items-center justify-center gap-2 mb-4"
                ),

                # Preview & Confirmation Container (Hidden initially)
                Div(
                    Div(
                        Span("✓ Extracted Courses Preview & Staff Verification", cls="text-xs font-bold text-blue-800"),
                        Span("", id="modal_course_badge", cls="text-[10px] font-bold text-blue-700 bg-blue-100 px-2 py-0.5 rounded-full"),
                        cls="flex items-center justify-between mb-2"
                    ),
                    Div(
                        Table(
                            Thead(
                                Tr(
                                    Th("Code", cls="px-2 py-1.5 text-left text-[11px] font-bold text-slate-700 bg-slate-100"),
                                    Th("Subject Title", cls="px-2 py-1.5 text-left text-[11px] font-bold text-slate-700 bg-slate-100"),
                                    Th("Sem", cls="px-2 py-1.5 text-center text-[11px] font-bold text-slate-700 bg-slate-100 w-12"),
                                    Th("Credits", cls="px-2 py-1.5 text-center text-[11px] font-bold text-slate-700 bg-slate-100 w-16"),
                                    Th("Type", cls="px-2 py-1.5 text-left text-[11px] font-bold text-slate-700 bg-slate-100 w-20"),
                                )
                            ),
                            Tbody(id="syllabus_courses_tbody"),
                            cls="w-full border-collapse"
                        ),
                        cls="max-h-60 overflow-y-auto border border-slate-200 rounded-lg mb-4"
                    ),

                    # Staff Verification Section
                    Div(
                        Div(
                            Label("Confirmed By (Staff / Committee Name):", cls="block text-[11px] font-bold text-slate-700 mb-1"),
                            Input(type="text", id="modal_staff_name", value="Faculty Academic Committee",
                                  cls="w-full text-xs text-slate-800 bg-white border border-slate-300 rounded-lg p-1.5"),
                            cls="mb-2"
                        ),
                        Div(
                            Input(type="checkbox", id="modal_staff_confirm", checked=True, cls="w-4 h-4 text-blue-600 rounded mr-2"),
                            Label("I confirm that these course codes and credits match the official approved syllabus.",
                                  for_="modal_staff_confirm", cls="text-[11px] text-slate-700 font-medium cursor-pointer"),
                            cls="flex items-center mb-3"
                        ),
                        Button(
                            "✓ Confirm & Save to Syllabus Catalog (JSON)", type="button", id="btn_confirm_syllabus", onclick="saveConfirmedSyllabus()",
                            cls="w-full bg-green-600 hover:bg-green-700 text-white font-semibold py-2.5 px-4 rounded-xl text-xs transition-all shadow-sm flex items-center justify-center gap-2"
                        ),
                        cls="p-3 bg-slate-50 border border-slate-200 rounded-xl"
                    ),
                    id="syllabus_preview_container", cls="hidden"
                ),

                cls="bg-white rounded-2xl p-6 max-w-xl w-full shadow-2xl border border-slate-200 max-h-[90vh] overflow-y-auto"
            ),
            id="new_syllabus_modal",
            cls="hidden fixed inset-0 z-50 bg-slate-900/60 backdrop-blur-xs flex items-center justify-center p-4"
        ),

        # ── PLATFORM CAPABILITIES ────────────────────────────────────────
        Div(
            Div(
                Div(cls="flex-1 h-px bg-slate-200"),
                H3("Platform Capabilities & Academic Features", cls="text-xs font-bold uppercase tracking-wider text-slate-400 px-4 whitespace-nowrap"),
                Div(cls="flex-1 h-px bg-slate-200"),
                cls="flex items-center mb-8"
            ),
            Div(
                # Card 1 — Smart Extraction
                Div(
                    NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#1a56db" stroke-width="2" style="width:28px;height:28px;margin-bottom:10px"><path stroke-linecap="round" stroke-linejoin="round" d="M7 21h10a2 2 0 002-2V9.414a1 1 0 00-.293-.707l-5.414-5.414A1 1 0 0012.586 3H7a2 2 0 00-2 2v14a2 2 0 002 2z"/></svg>'),
                    H4("Smart Extraction", cls="text-sm font-bold text-slate-800 mb-1"),
                    P("Extracts student & subject results directly from official result PDFs with page provenance traceability.", cls="text-xs text-slate-500 leading-relaxed"),
                    cls="card p-5 border border-slate-100 hover:border-blue-200 hover:shadow-md transition-all flex flex-col"
                ),
                # Card 2 — R2024 / R2026 Mapping
                Div(
                    NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#1a56db" stroke-width="2" style="width:28px;height:28px;margin-bottom:10px"><path stroke-linecap="round" stroke-linejoin="round" d="M9 20l-5.447-2.724A1 1 0 013 16.382V5.618a1 1 0 011.447-.894L9 7m0 13l6-3m-6 3V7m6 10l4.553 2.276A1 1 0 0021 18.382V7.618a1 1 0 00-.553-.894L15 4m0 13V4m0 0L9 7"/></svg>'),
                    H4("Autonomous Syllabus Mapping", cls="text-sm font-bold text-slate-800 mb-1"),
                    P("Automatically identifies course codes, titles, credits, and semester categories against authoritative catalogs.", cls="text-xs text-slate-500 leading-relaxed"),
                    cls="card p-5 border border-slate-100 hover:border-blue-200 hover:shadow-md transition-all flex flex-col"
                ),
                # Card 4 — Academic Dashboards
                Div(
                    NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#1a56db" stroke-width="2" style="width:28px;height:28px;margin-bottom:10px"><path stroke-linecap="round" stroke-linejoin="round" d="M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z"/></svg>'),
                    H4("Academic Dashboards", cls="text-sm font-bold text-slate-800 mb-1"),
                    P("Generates credit-weighted GPA, cohort rankings, risk groups, and PTM advisory briefs.", cls="text-xs text-slate-500 leading-relaxed"),
                    cls="card p-5 border border-slate-100 hover:border-blue-200 hover:shadow-md transition-all flex flex-col"
                ),
                # Card 5 — Export & Reports
                Div(
                    NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#16a34a" stroke-width="2" style="width:28px;height:28px;margin-bottom:10px"><path stroke-linecap="round" stroke-linejoin="round" d="M12 10v6m0 0l-3-3m3 3l3-3m2 8H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>'),
                    H4("Export & Reports", cls="text-sm font-bold text-slate-800 mb-1"),
                    P("Faculty-ready PDF and Excel export with class summaries, subject analytics, and remedial plans.", cls="text-xs text-slate-500 leading-relaxed"),
                    cls="card p-5 border border-slate-100 hover:border-green-200 hover:shadow-md transition-all flex flex-col"
                ),
                # Card 6 — Secure & Reliable
                Div(
                    NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="#7c3aed" stroke-width="2" style="width:28px;height:28px;margin-bottom:10px"><path stroke-linecap="round" stroke-linejoin="round" d="M9 12l2 2 4-4m5.618-4.016A11.955 11.955 0 0112 2.944a11.955 11.955 0 01-8.618 3.04A12.02 12.02 0 003 9c0 5.591 3.824 10.29 9 11.622 5.176-1.332 9-6.03 9-11.622 0-1.042-.133-2.052-.382-3.016z"/></svg>'),
                    H4("Secure & Reliable", cls="text-sm font-bold text-slate-800 mb-1"),
                    P("Results are processed locally. AI receives only structured metrics. Student PII never leaves your system.", cls="text-xs text-slate-500 leading-relaxed"),
                    cls="card p-5 border border-slate-100 hover:border-purple-200 hover:shadow-md transition-all flex flex-col"
                ),
                cls="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 gap-4 mb-12 max-w-5xl mx-auto"
            ),
        ),

        # ── HOW IT WORKS ────────────────────────────────────────────────
        Div(
            H3("How It Works", cls="text-sm font-bold uppercase tracking-wider text-slate-400 text-center mb-2"),
            P("From result document to faculty action in minutes.", cls="text-xs text-slate-500 text-center mb-6"),
            Div(
                Div(
                    Span("01", cls="text-blue-600 font-extrabold text-lg block mb-1"),
                    H4("Configure", cls="text-xs font-bold text-slate-800"),
                    P("Select Dept & Sem", cls="text-[11px] text-slate-500"),
                    cls="text-center p-3 card border border-slate-100"
                ),
                Div(Span("→", cls="text-slate-300 hidden sm:block text-xl self-center")),
                Div(
                    Span("02", cls="text-blue-600 font-extrabold text-lg block mb-1"),
                    H4("Upload", cls="text-xs font-bold text-slate-800"),
                    P("Upload COE PDF", cls="text-[11px] text-slate-500"),
                    cls="text-center p-3 card border border-slate-100"
                ),
                Div(Span("→", cls="text-slate-300 hidden sm:block text-xl self-center")),
                Div(
                    Span("03", cls="text-blue-600 font-extrabold text-lg block mb-1"),
                    H4("Verify", cls="text-xs font-bold text-slate-800"),
                    P("Review extraction", cls="text-[11px] text-slate-500"),
                    cls="text-center p-3 card border border-slate-100"
                ),
                Div(Span("→", cls="text-slate-300 hidden sm:block text-xl self-center")),
                Div(
                    Span("04", cls="text-blue-600 font-extrabold text-lg block mb-1"),
                    H4("Act", cls="text-xs font-bold text-slate-800"),
                    P("Actionable metrics", cls="text-[11px] text-slate-500"),
                    cls="text-center p-3 card border border-slate-100"
                ),
                cls="grid grid-cols-2 sm:grid-cols-7 gap-2 max-w-4xl mx-auto items-center mb-12"
            ),
        ),

        # ── DATA PRIVACY ─────────────────────────────────────────────────
        Div(
            Div(
                NotStr('<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2" style="width:18px;height:18px;flex-shrink:0;color:#475569"><path stroke-linecap="round" stroke-linejoin="round" d="M12 15v2m-6 4h12a2 2 0 002-2v-6a2 2 0 00-2-2H6a2 2 0 00-2 2v6a2 2 0 002 2zm10-10V7a4 4 0 00-8 0v4h8z"/></svg>'),
                Span("Academic Data Protection", cls="font-bold text-slate-800 text-xs mr-2"),
                Span("• Results are processed locally for academic analysis. AI receives only structured metrics required for advisory generation.", cls="text-xs text-slate-500"),
                cls="flex items-center justify-center flex-wrap gap-2 p-3 bg-slate-100 border border-slate-200 rounded-xl max-w-3xl mx-auto text-center"
            ),
            cls="mb-8"
        ),

        Script("""
        var _extractedCourses = [];

        function handlePdfFileSelect(input) {
            if (input.files && input.files[0]) {
                var file = input.files[0];
                var info = document.getElementById('pdf_file_selected_info');
                if (info) {
                    info.classList.remove('hidden');
                    document.getElementById('pdf_file_name').innerText = file.name;
                }
            }
        }

        function handleRegulationChange(select) {
            if (select.value === 'NEW') {
                openNewSyllabusModal();
                select.value = 'R2024';
            }
        }

        function openNewSyllabusModal() {
            var deptSel = document.getElementById('select_department');
            if (deptSel && deptSel.value) {
                var mDept = document.getElementById('modal_dept_select');
                if (mDept) mDept.value = deptSel.value;
            }
            var modal = document.getElementById('new_syllabus_modal');
            if (modal) modal.classList.remove('hidden');
        }

        function closeNewSyllabusModal() {
            var modal = document.getElementById('new_syllabus_modal');
            if (modal) modal.classList.add('hidden');
        }

        function extractSyllabusDraft() {
            var fileInput = document.getElementById('modal_syllabus_file');
            var deptCode = document.getElementById('modal_dept_select').value;
            var regCode = document.getElementById('modal_reg_input').value.trim();

            if (!regCode) {
                alert('Please enter a Regulation Name (e.g. R2028).');
                document.getElementById('modal_reg_input').focus();
                return;
            }
            if (!fileInput.files || !fileInput.files[0]) {
                alert('Please select a syllabus PDF or JSON file.');
                return;
            }

            var btn = document.getElementById('btn_extract_syllabus');
            var origText = btn.innerHTML;
            btn.innerHTML = '<span class="animate-spin inline-block mr-1">⏳</span> Extracting syllabus courses...';
            btn.style.pointerEvents = 'none';

            var formData = new FormData();
            formData.append('file_syllabus', fileInput.files[0]);
            formData.append('department_code', deptCode);
            formData.append('regulation_code', regCode);

            fetch('/api/syllabus/extract-draft', {
                method: 'POST',
                body: formData
            })
            .then(function(res) { return res.json(); })
            .then(function(data) {
                btn.innerHTML = origText;
                btn.style.pointerEvents = 'auto';

                if (!data.ok) {
                    alert('Extraction error: ' + (data.error || 'Unknown error'));
                    return;
                }

                _extractedCourses = data.courses || [];
                renderSyllabusPreviewTable(_extractedCourses, data.department_code, data.regulation_code);
            })
            .catch(function(err) {
                btn.innerHTML = origText;
                btn.style.pointerEvents = 'auto';
                alert('Network or server error: ' + err);
            });
        }

        function renderSyllabusPreviewTable(courses, dept, reg) {
            var container = document.getElementById('syllabus_preview_container');
            var tbody = document.getElementById('syllabus_courses_tbody');
            var badge = document.getElementById('modal_course_badge');

            if (!container || !tbody) return;

            badge.innerText = courses.length + ' Courses Extracted';
            tbody.innerHTML = '';

            courses.forEach(function(c, idx) {
                var tr = document.createElement('tr');
                tr.className = 'border-b border-slate-100 text-xs';
                tr.innerHTML = `
                    <td class="p-2 font-mono font-bold text-blue-700">${c.code || ''}</td>
                    <td class="p-2 font-medium text-slate-800">${c.name || ''}</td>
                    <td class="p-2 text-center">${c.semester || 1}</td>
                    <td class="p-2 text-center font-semibold">${c.credits || 3.0}</td>
                    <td class="p-2 text-slate-500">${c.type || 'THEORY'}</td>
                `;
                tbody.appendChild(tr);
            });

            container.classList.remove('hidden');
        }

        function saveConfirmedSyllabus() {
            var chk = document.getElementById('modal_staff_confirm');
            if (chk && !chk.checked) {
                alert('Please check the confirmation box to verify the syllabus.');
                return;
            }

            var deptCode = document.getElementById('modal_dept_select').value;
            var regCode = document.getElementById('modal_reg_input').value.trim().toUpperCase();
            var actor = document.getElementById('modal_staff_name').value.trim() || 'Faculty Academic Committee';

            if (!regCode || _extractedCourses.length === 0) {
                alert('No extracted courses to save.');
                return;
            }

            var btn = document.getElementById('btn_confirm_syllabus');
            var origText = btn.innerHTML;
            btn.innerHTML = '<span class="animate-spin inline-block mr-1">⏳</span> Publishing to JSON catalog...';
            btn.style.pointerEvents = 'none';

            fetch('/api/syllabus/confirm-and-save', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    department_code: deptCode,
                    regulation_code: regCode,
                    courses: _extractedCourses,
                    actor: actor
                })
            })
            .then(function(res) { return res.json(); })
            .then(function(data) {
                btn.innerHTML = origText;
                btn.style.pointerEvents = 'auto';

                if (!data.ok) {
                    alert('Save error: ' + (data.error || 'Unknown error'));
                    return;
                }

                // Add to main regulation dropdown & select it
                var regSel = document.getElementById('select_regulation');
                if (regSel) {
                    var newOpt = document.createElement('option');
                    newOpt.value = regCode;
                    newOpt.text = regCode + ' (Staff Confirmed)';
                    newOpt.selected = true;
                    regSel.add(newOpt, 0);
                }

                // Also select the department in the main dropdown
                var deptSel = document.getElementById('select_department');
                if (deptSel) deptSel.value = deptCode;

                closeNewSyllabusModal();
                alert('✓ ' + data.message);
            })
            .catch(function(err) {
                btn.innerHTML = origText;
                btn.style.pointerEvents = 'auto';
                alert('Save error: ' + err);
            });
        }

        function handleUploadFormSubmit(form, cardId, btnText) {
            var deptSel = document.getElementById('select_department');
            var semSel = document.getElementById('select_semester');

            if (!deptSel || !deptSel.value) {
                alert('⚠️ Please select a Department before uploading.');
                if (deptSel) deptSel.focus();
                return false;
            }
            if (!semSel || !semSel.value) {
                alert('⚠️ Please select a Semester before uploading.');
                if (semSel) semSel.focus();
                return false;
            }

            if (cardId) {
                var card = document.getElementById(cardId);
                if (card) {
                    card.classList.remove('hidden');
                }
            }
            var btn = form.querySelector('button[type="submit"]');
            if (btn) {
                btn.innerHTML = '<span class="animate-spin inline-block mr-2">⏳</span> ' + btnText;
                btn.style.pointerEvents = 'none';
                btn.classList.add('opacity-75');
            }
            return true;
        }

        document.addEventListener('DOMContentLoaded', function() {
            var dropzone = document.getElementById('excel_dropzone');
            if (dropzone) {
                ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(function(eventName) {
                    dropzone.addEventListener(eventName, function(e) {
                        e.preventDefault();
                        e.stopPropagation();
                    }, false);
                });
                ['dragenter', 'dragover'].forEach(function(eventName) {
                    dropzone.addEventListener(eventName, function() {
                        dropzone.classList.add('border-emerald-600', 'bg-emerald-100/50');
                    }, false);
                });
                ['dragleave', 'drop'].forEach(function(eventName) {
                    dropzone.addEventListener(eventName, function() {
                        dropzone.classList.remove('border-emerald-600', 'bg-emerald-100/50');
                    }, false);
                });
                dropzone.addEventListener('drop', function(e) {
                    var dt = e.dataTransfer;
                    var files = dt.files;
                    if (files && files.length > 0) {
                        var fileInput = document.getElementById('file_excel_input');
                        fileInput.files = files;
                        handleExcelFileSelect(fileInput);
                    }
                }, false);
            }
        });
        """),
        cls="max-w-5xl mx-auto py-4"
    ))


def page_upload_mapping() -> Tuple:
    """Requirements 1, 2, 3: Structure Preview, Format Detection, Column Mapping & Data Quality Check."""
    cols = SESSION.get("preview_cols", [])
    report: ValidationReport = SESSION.get("preview_report") or ValidationReport()
    filename = SESSION.get("preview_filename", "")
    mapping = report.mapped_columns or {}
    format_type = getattr(report, "format_detected", "long")
    subject_maps = getattr(report, "subject_mappings", [])

    def col_select(target: str, label: str) -> Div:
        options = [Option("-- Unmapped --", value="")]
        selected = mapping.get(target, "")
        for c in cols:
            is_sel = (c == selected)
            options.append(Option(f"{c} {'✓' if is_sel else ''}", value=c, selected=is_sel))
        return Div(
            Label(f"{label}:", cls="block text-xs font-semibold text-slate-600 mb-1"),
            Select(*options, name=f"map_{target}", cls="w-full px-3 py-2 text-sm border border-slate-200 rounded-lg bg-white focus:ring-2 focus:ring-blue-500"),
            cls="mb-3"
        )

    format_badge = (
        Span("Wide Format (Student Rows × Subject Grade Columns)", cls="px-3 py-1 bg-blue-100 text-blue-800 text-xs font-semibold rounded-full")
        if format_type == "wide"
        else Span("Long Format (Row per Grade Entry)", cls="px-3 py-1 bg-green-100 text-green-800 text-xs font-semibold rounded-full")
    )

    mapping_rows = []
    for sm in subject_maps:
        conf_pct = int(sm["confidence"] * 100)
        badge = (
            Span(f"✓ Verified ({conf_pct}%)", cls="text-xs font-semibold text-green-700 bg-green-50 border border-green-200 px-2 py-0.5 rounded")
            if sm["confidence"] >= 0.80
            else Span(f"⚠ Review Required ({conf_pct}%)", cls="text-xs font-semibold text-amber-700 bg-amber-50 border border-amber-200 px-2 py-0.5 rounded")
        )
        mapping_rows.append(Tr(
            Td(sm["raw_header"], cls="px-4 py-2.5 text-xs font-mono text-slate-700 border-b"),
            Td(sm["canonical_subject"], cls="px-4 py-2.5 text-xs font-medium text-slate-900 border-b"),
            Td(sm["course_code"] or "—", cls="px-4 py-2.5 text-xs text-slate-500 border-b"),
            Td(f"Sem {sm['semester']} ({sm['category']})", cls="px-4 py-2.5 text-xs text-slate-600 border-b"),
            Td(badge, cls="px-4 py-2.5 text-xs border-b"),
        ))

    subject_mapping_card = (
        Div(
            H3("Detected Subject Columns & Abbreviation Resolution (R2024 AI & DS Catalog)", cls="text-sm font-bold text-slate-800 mb-2"),
            P("The deterministic parser mapped the following columns to official syllabus subjects:", cls="text-xs text-slate-500 mb-3"),
            Div(
                Table(
                    Thead(Tr(
                        Th("Raw Header", cls="px-4 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Resolved Canonical Subject", cls="px-4 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Code", cls="px-4 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Curriculum Group", cls="px-4 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Match Confidence", cls="px-4 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    )),
                    Tbody(*mapping_rows),
                    cls="w-full border border-slate-200 rounded-lg overflow-hidden"
                ),
                cls="overflow-x-auto mb-6"
            ),
            cls="card p-5 mb-6"
        ) if mapping_rows else None
    )

    # Quarantined Tokens Card
    quarantined = getattr(report, "quarantined_tokens", [])
    quarantine_card = (
        Div(
            Div(
                Span("⚠️", cls="text-amber-500 text-lg mr-2"),
                H3("Unrecognized Result Token Quarantine Report", cls="text-sm font-bold text-slate-800"),
                Span(f"{len(quarantined)} token(s) quarantined", cls="ml-auto text-xs badge badge-amber"),
                cls="flex items-center mb-3"
            ),
            P("The following unknown result tokens were encountered and quarantined for department review:", cls="text-xs text-slate-500 mb-3"),
            Div(
                Table(
                    Thead(Tr(
                        Th("Row", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Register No", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Column / Subject", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Raw Value", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Status / Action", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    )),
                    Tbody(*[Tr(
                        Td(q["row"], cls="px-3 py-2 text-xs font-mono border-b"),
                        Td(q["regno"] or "—", cls="px-3 py-2 text-xs font-mono border-b"),
                        Td(q["column"], cls="px-3 py-2 text-xs font-mono border-b"),
                        Td(Span(q["raw_value"], cls="px-2 py-0.5 bg-amber-100 text-amber-800 font-mono font-bold rounded text-xs"), cls="px-3 py-2 border-b"),
                        Td(q["reason"], cls="px-3 py-2 text-xs text-slate-600 border-b"),
                    ) for q in quarantined]),
                    cls="w-full border border-slate-200 rounded-lg overflow-hidden"
                ),
                cls="overflow-x-auto mb-3"
            ),
            cls="card p-5 border-l-4 border-l-amber-500 mb-6"
        ) if quarantined else None
    )

    # Department Manual Alias Override Form
    alias_override_card = Div(
        H3("Department Manual Subject Alias Override", cls="text-sm font-bold text-slate-800 mb-2"),
        P("Map custom non-standard classroom subject abbreviations directly to canonical syllabus subjects:", cls="text-xs text-slate-500 mb-3"),
        Form(
            Div(
                Input(type="text", name="raw_token", placeholder="e.g. DSEA-LAB, ML-THEORY", required=True,
                      cls="px-3 py-2 text-sm border border-slate-200 rounded-lg focus:ring-2 focus:ring-blue-500"),
                Select(
                    *[Option(item["name"], value=item["name"]) for item in SYLLABUS_CATALOG_R2024],
                    name="canonical_target", required=True,
                    cls="px-3 py-2 text-sm border border-slate-200 rounded-lg focus:ring-2 focus:ring-blue-500"
                ),
                Button("Save Alias Override", type="submit", cls="px-4 py-2 text-sm font-semibold bg-slate-800 text-white rounded-lg hover:bg-slate-900 transition-colors"),
                cls="grid grid-cols-1 sm:grid-cols-3 gap-3"
            ),
            action="/alias-override", method="POST",
        ),
        cls="card p-5 mb-6"
    )

    # Copy-Paste Sanitization Alert Banner
    cleaned_cnt = getattr(report, "copy_paste_cleaned_count", 0)
    copy_paste_alert = (
        Div(
            Span("✨", cls="text-blue-500 text-lg mr-2"),
            Span(f"PDF Copy-Paste Sanitizer active: Cleaned {cleaned_cnt} merged/shifted COE formatting anomalies.", cls="text-xs font-semibold text-blue-800"),
            cls="flex items-center px-4 py-3 bg-blue-50 border border-blue-200 rounded-xl mb-6"
        ) if cleaned_cnt > 0 else None
    )

    return layout("Excel Structure & Column Mapping", "upload", Div(
        Div(
            H1("Excel File Analysis & Preflight Review", cls="text-2xl font-bold text-slate-800 mb-1"),
            Div(
                Span(f"File: {html.escape(filename)}", cls="text-sm text-slate-500 mr-3"),
                format_badge,
                cls="flex items-center gap-2 mb-6"
            ),
        ),

        copy_paste_alert,

        # Structure Preview Cards
        Div(
            stat_card("Detected Rows", str(report.total_input_rows), "#3b82f6"),
            stat_card("Valid Records", str(report.valid_records), "#16a34a"),
            stat_card("Sheet Name", report.sheet_name or "Sheet1", "#64748b"),
            stat_card("Quality Status", "Clean" if not report.issues else f"{len(report.issues)} Issues", "#16a34a" if not report.issues else "#d97706"),
            cls="grid grid-cols-2 lg:grid-cols-4 gap-4 mb-6"
        ),

        # Data Quality Check Report
        Div(
            H3("Excel Data Quality & Preflight Check", cls="text-sm font-bold text-slate-800 mb-3"),
            Div(
                Div(Span("✓", cls="text-green-600 font-bold mr-2"), Span(f"{report.valid_records} valid result records detected.", cls="text-sm text-slate-700"), cls="py-1 flex items-center"),
                Div(Span("✓", cls="text-green-600 font-bold mr-2"), Span(f"Header row located automatically at row {report.header_row + 1}.", cls="text-sm text-slate-700"), cls="py-1 flex items-center"),
                Div(Span("ℹ", cls="text-blue-600 font-bold mr-2"), Span(f"Structure recognized as {format_type.upper()} format.", cls="text-sm text-slate-700"), cls="py-1 flex items-center"),
                Div(Span("⚠", cls="text-amber-600 font-bold mr-2"), Span(f"{report.duplicates_removed} duplicate row(s) removed during validation.", cls="text-sm text-slate-700"), cls="py-1 flex items-center") if report.duplicates_removed else None,
                Div(Span("⚠", cls="text-amber-600 font-bold mr-2"), Span(f"{report.dropped_rows} row(s) dropped due to missing mandatory fields.", cls="text-sm text-slate-700"), cls="py-1 flex items-center") if report.dropped_rows else None,
            ),
            cls="card p-5 mb-6"
        ),

        quarantine_card,
        subject_mapping_card,
        alias_override_card,

        # Column Mapping Form
        Form(
            H3("Student Identity Column Assignments", cls="text-sm font-bold text-slate-800 mb-3"),
            P("Confirm or adjust student identity spreadsheet column assignments:", cls="text-xs text-slate-500 mb-4"),
            Div(
                col_select("regno", "Register Number"),
                col_select("name", "Student Name"),
                col_select("subject", "Subject Name (Long format only)"),
                col_select("credits", "Course Credits (Long format only)"),
                col_select("grade", "Final Result Grade (Long format only)"),
                col_select("course_code", "Course Code (Optional)"),
                cls="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 gap-4 mb-6"
            ),
            Div(
                A("← Upload Different File", href="/upload", cls="px-4 py-2.5 text-sm font-medium text-slate-600 hover:text-slate-800"),
                Button("Confirm & Process Full Analytics →", type="submit",
                       cls="px-6 py-2.5 text-sm font-semibold bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors shadow-sm"),
                cls="flex items-center justify-between border-t pt-4"
            ),
            action="/upload-confirm", method="POST",
            cls="card p-6 mb-8"
        ),

        cls="max-w-4xl mx-auto"
    ))


def page_dashboard(ca: ClassAnalysis) -> Tuple:
    insights_data = generate_deterministic_insights(ca)
    
    # Non-blocking AI insight: Check cache first, otherwise load asynchronously via HTMX/fetch
    ai_key = "class:" + _ai_hash(ca.file_name, ca.record_count, ca.class_gpa, sorted(ca.grade_distribution.items()))
    cached_ai = _cache_get(ai_key)
    if cached_ai is not None:
        ai_block = md_block(cached_ai, _AI_CACHE[ai_key]["live"])
    else:
        ai_fallback = fallback_class_insight(ca)
        ai_block = Div(
            md_block(ai_fallback, "fallback"),
            hx_get="/ai-insights/class-summary",
            hx_trigger="load delay:100ms",
            hx_swap="outerHTML",
            id="class-ai-advisory"
        )

    grade_dist_fig = fig_grade_distribution(ca)
    subj_avg_fig = fig_subject_avg_gp(ca)
    fail_fig = fig_failure_concentration(ca)
    dist_fig = fig_student_performance_distribution(ca)
    heatmap_fig = fig_grade_heatmap(ca)

    attn = sorted(ca.attention_students, key=lambda s: (s.arrear_count, s.malpractice_count, s.sa_count, s.wd_count), reverse=True)[:6]

    best_subj = insights_data["best_subject"]
    weakest_subj = insights_data["weakest_subject"]
    duration = SESSION.get("analysis_duration")

    return layout("Dashboard", "dashboard", Div(
        # Page Title & Interactive Report Shortcut (Requirement 6)
        Div(
            Div(
                Div(
                    H1("Executive Dashboard", cls="text-2xl font-bold text-slate-800 inline-block"),
                    Span(f"⚡ Analysis completed in {duration}s", cls="text-xs font-semibold text-emerald-700 bg-emerald-50 px-2.5 py-1 rounded-full border border-emerald-200 ml-3 inline-block") if duration else None,
                ),
                P(f"Semester results summary · {ca.student_count} students · {ca.subject_count} subjects evaluated",
                  cls="text-slate-500 text-sm mt-1"),
                cls="flex-1"
            ),
            Div(
                A("🌐 Open Interactive Web Report", href="/reports/interactive",
                  cls="px-4 py-2 text-sm font-semibold bg-indigo-600 hover:bg-indigo-700 text-white rounded-lg shadow-sm transition-colors mr-2"),
                Form(Button("📄 Download PDF", type="submit",
                            cls="px-4 py-2 text-sm font-medium bg-slate-100 hover:bg-slate-200 text-slate-700 rounded-lg transition-colors"),
                     action="/report/class", method="POST", cls="inline-block"),
                cls="flex items-center gap-2"
            ),
            cls="flex flex-col sm:flex-row sm:items-center justify-between gap-4 mb-6"
        ),

        # 1. Executive Story & Semester Performance Overview
        Div(
            Div(
                Div(
                    H3("Semester Performance Overview", cls="text-xs font-bold uppercase tracking-wider text-slate-400 mb-1"),
                    Div(
                        Span("Class Health:", cls="text-xs font-medium text-slate-500 mr-2"),
                        Span(insights_data["health_status"],
                             cls=f"px-3 py-1 rounded-full text-xs font-bold border shadow-xs {insights_data['health_bg']} {insights_data['health_color']}"),
                        cls="flex items-center mb-4"
                    ),
                    Div(
                        Div(
                            P("Average GPA", cls="text-xs text-slate-500 font-medium"),
                            P(fmt_gpa(ca.class_gpa), cls="text-3xl font-extrabold text-slate-900"),
                            cls="bg-slate-50 p-4 rounded-xl border border-slate-100"
                        ),
                        Div(
                            P("Pass Rate", cls="text-xs text-slate-500 font-medium"),
                            P(fmt_np_pct(ca.pass_rate), cls="text-3xl font-extrabold " + ("text-green-600" if (ca.pass_rate or 0) >= 75 else "text-amber-600")),
                            cls="bg-slate-50 p-4 rounded-xl border border-slate-100"
                        ),
                        Div(
                            P("Students Needing Attention", cls="text-xs text-slate-500 font-medium"),
                            P(str(len(ca.attention_students)), cls="text-3xl font-extrabold text-red-600"),
                            cls="bg-slate-50 p-4 rounded-xl border border-slate-100"
                        ),
                        cls="grid grid-cols-1 sm:grid-cols-3 gap-4 mb-4"
                    ),
                    # Key Observations Card
                    Div(
                        H4("Key Observations", cls="text-xs font-bold text-slate-700 uppercase tracking-wider mb-2"),
                        Div(
                            Div(Span("⚠", cls="text-amber-500 font-bold mr-2 text-base"), Span(f"{ca.multiple_u_count} student(s) have multiple arrears requiring immediate faculty intervention.", cls="text-xs text-slate-700 font-medium"), cls="flex items-center py-1"),
                            Div(Span("📘", cls="text-blue-500 font-bold mr-2 text-base"), Span(f"'{weakest_subj.subject}' has highest arrear concentration ({weakest_subj.arrear_count} arrears).", cls="text-xs text-slate-700 font-medium"), cls="flex items-center py-1") if weakest_subj else None,
                            Div(Span("⭐", cls="text-yellow-500 font-bold mr-2 text-base"), Span(f"{ca.grade_distribution.get('O',0) + ca.grade_distribution.get('A+',0)} student(s) achieved distinction (O/A+ grade).", cls="text-xs text-slate-700 font-medium"), cls="flex items-center py-1"),
                        ),
                        cls="bg-blue-50/50 p-4 rounded-xl border border-blue-100"
                    ),
                    cls="card p-6"
                ),
            ),
            cls="mb-6"
        ),

        # 2. Faculty Action Center
        Div(
            H3("Faculty Action Center", cls="text-sm font-bold uppercase tracking-wider text-slate-500 mb-3 flex items-center gap-2"),
            Div(
                # Card 1: Immediate Attention
                Div(
                    Div(
                        Span("🚨", cls="text-2xl mr-2"),
                        H4("Immediate Attention", cls="text-sm font-bold text-slate-900"),
                        cls="flex items-center mb-1"
                    ),
                    P(f"{len(ca.attention_students)} Student(s) requiring academic support", cls="text-xs text-slate-500 mb-4"),
                    A("View Students →", href="/attention", cls="w-full text-center block px-3 py-2 text-xs font-bold bg-red-600 hover:bg-red-700 text-white rounded-lg transition-colors shadow-xs"),
                    cls="card p-5 border-l-4 border-l-red-600 flex flex-col justify-between"
                ),

                # Card 2: Subject Concern
                Div(
                    Div(
                        Span("📘", cls="text-2xl mr-2"),
                        H4("Subject Concern", cls="text-sm font-bold text-slate-900"),
                        cls="flex items-center mb-1"
                    ),
                    P(f"'{weakest_subj.subject}' ({weakest_subj.arrear_count} arrears)" if weakest_subj else "No subject concern", cls="text-xs text-slate-500 mb-4"),
                    A("Analyze Subject →", href=f"/subjects/{quote(weakest_subj.subject, safe='')}" if weakest_subj else "/subjects", cls="w-full text-center block px-3 py-2 text-xs font-bold bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors shadow-xs"),
                    cls="card p-5 border-l-4 border-l-blue-600 flex flex-col justify-between"
                ),

                # Card 3: PTM Required
                Div(
                    Div(
                        Span("👨‍👩‍👦", cls="text-2xl mr-2"),
                        H4("PTM Required", cls="text-sm font-bold text-slate-900"),
                        cls="flex items-center mb-1"
                    ),
                    P(f"{ca.multiple_u_count + ca.backlog_student_count} Student(s) scheduled for parent meeting", cls="text-xs text-slate-500 mb-4"),
                    A("Generate PTM Brief →", href="/attention", cls="w-full text-center block px-3 py-2 text-xs font-bold bg-indigo-600 hover:bg-indigo-700 text-white rounded-lg transition-colors shadow-xs"),
                    cls="card p-5 border-l-4 border-l-indigo-600 flex flex-col justify-between"
                ),
                cls="grid grid-cols-1 sm:grid-cols-3 gap-6 mb-6"
            )
        ),

        # 3. Academic Attention Summary Grid
        Div(
            stat_card("🔴 Arrears (U/RA)", str(ca.arrear_student_count), "#dc2626"),
            stat_card("🟠 Attendance (SA)", str(ca.sa_student_count), "#d97706"),
            stat_card("🟣 Withdrawal (WD)", str(ca.wd_student_count), "#9333ea"),
            stat_card("🔮 Malpractice (MM/WH2)", str(ca.malpractice_student_count), "#7c3aed"),
            cls="grid grid-cols-2 lg:grid-cols-4 gap-4 mb-6"
        ),

        # 3. Deterministic Important Insights Section
        Div(
            Div(
                H3("Important Calculated Insights", cls="text-sm font-bold text-slate-800 mb-3 flex items-center gap-2"),
                Div(*[
                    Div(
                        Span("•", cls="text-blue-600 font-bold mr-2 text-lg"),
                        Span(item, cls="text-sm text-slate-700"),
                        cls="flex items-start py-1"
                    ) for item in insights_data["bullet_insights"]
                ]),
                cls="card p-5"
            ),
            cls="mb-6"
        ),

        # Requirement 7: Advanced Charts (Performance Distribution & Heatmap)
        Div(
            Div(
                H3("Cohort Performance Distribution", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(dist_fig, "h-72"),
                cls="card p-5"
            ),
            Div(
                H3("Grade Heatmap Matrix", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(heatmap_fig, "h-72"),
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        # Interactive Standard Charts
        Div(
            Div(
                H3("Grade Distribution", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(grade_dist_fig, "h-72"),
                cls="card p-5"
            ),
            Div(
                H3("Average Grade Point by Subject", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(subj_avg_fig, "h-72"),
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        # 5. Attention & Failure
        Div(
            Div(
                Div(
                    H3("Students Requiring Attention", cls="text-sm font-semibold text-slate-700"),
                    A("View All →", href="/attention", cls="text-xs text-blue-600 hover:underline font-medium"),
                    cls="flex items-center justify-between mb-4"
                ),
                *[Div(
                    Div(
                        A(s.name or "—", href=f"/students/{quote(s.regno, safe='')}",
                          cls="text-sm font-medium text-slate-800 hover:text-blue-600"),
                        P(s.regno, cls="text-xs text-slate-400 font-mono"),
                        cls="flex-1 min-w-0"
                    ),
                    Div(
                        status_badge(s.attention),
                        gpa_cell(s.gpa),
                        Span(f"{s.arrear_count} Arrear", cls="text-xs text-red-600 font-semibold") if s.arrear_count else Span("", cls="text-xs text-slate-300"),
                        cls="flex items-center gap-2"
                    ),
                    cls="flex items-center justify-between py-3 border-b border-slate-100 last:border-0"
                ) for s in attn] or [
                    Div(P("All students cleared.", cls="text-sm text-green-600"), cls="py-4 text-center")
                ],
                cls="card p-5"
            ),
            Div(
                H3("Failure / Arrear Concentration", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(fail_fig, "h-64"),
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        # 6. Subject Performance Table
        Div(
            Div(
                H3("Subject Performance Overview", cls="text-sm font-semibold text-slate-700"),
                A("Subject Details →", href="/subjects", cls="text-xs text-blue-600 hover:underline font-medium"),
                cls="flex items-center justify-between mb-4"
            ),
            data_table(
                ["Subject", "Credits", "Students", "Avg GP", "Pass %", "Arrears (U/RA)", "Priority Rating", "Action"],
                [[
                    A(s.subject, href=f"/subjects/{quote(s.subject, safe='')}", cls="text-blue-600 hover:text-blue-700 font-medium"),
                    s.credits,
                    s.student_count,
                    fmt_gp(s.avg_gp),
                    fmt_np_pct(s.pass_pct),
                    Span(str(s.arrear_count), cls="text-red-600 font-semibold") if s.arrear_count else "0",
                    Span(s.priority_level, cls="text-xs font-bold text-red-600") if s.priority_level == "High Attention" else Span(s.priority_level, cls="text-xs font-medium text-slate-500"),
                    A("View →", href=f"/subjects/{quote(s.subject, safe='')}", cls="text-xs text-slate-500 hover:text-blue-600 font-medium"),
                ] for s in ca.subjects],
                table_id="subject-summary"
            ),
            cls="card p-5 mb-6"
        ),

        # 7. AI Advisory Section
        Div(
            H3("AI Academic Advisory Insight", cls="text-sm font-semibold text-slate-700 mb-3"),
            ai_block,
        ),

        cls="max-w-7xl mx-auto"
    ), ca=ca)


def page_students(ca: ClassAnalysis) -> Tuple:
    students = ca.students
    rows = []
    for s in students:
        rows.append(Tr(
            Td(str(s.rank or "—"), cls="font-medium text-slate-600"),
            Td(s.regno, cls="font-mono text-xs text-slate-500"),
            Td(A(s.name or "—", href=f"/students/{quote(s.regno, safe='')}",
                 cls="text-blue-600 hover:text-blue-700 font-medium")),
            Td(gpa_cell(s.gpa)),
            Td(str(s.arrear_count), cls="text-red-600 font-semibold") if s.arrear_count else Td("0"),
            Td(status_badge(s.attention)),
            Td(f"{s.percentile:.1f}%" if s.percentile else "—", cls="text-slate-500"),
            data_row_type="student-row",
            data_status=s.attention,
            data_high=str(s.is_high_performer).lower(),
        ))

    return layout("Students", "students", Div(
        Div(
            Div(
                H1("Students", cls="text-2xl font-bold text-slate-800"),
                P(f"{len(students)} students · sorted by class rank",
                  cls="text-slate-500 text-sm mt-1"),
                cls="flex-1"
            ),
            Div(
                Input(type="text", placeholder="Search by name or register number...",
                      hx_get="/students/search", hx_trigger="keyup changed delay:300ms",
                      hx_target="#student-table-tbody",
                      cls="w-full sm:w-72 px-4 py-2 text-sm border border-slate-200 rounded-lg "
                          "focus:ring-2 focus:ring-blue-500 focus:border-blue-500 bg-white",
                      aria_label="Search students by name or register number"),
                cls="flex-shrink-0"
            ),
            cls="flex flex-col sm:flex-row sm:items-center gap-4 mb-6"
        ),

        # Analysis Summary Before Table (Student Attention Summary)
        Div(
            H3("Student Attention Summary", cls="text-sm font-bold text-slate-800 mb-3"),
            Div(
                Div(
                    P("Total Enrolled", cls="text-xs text-slate-500 font-medium"),
                    P(str(len(students)), cls="text-2xl font-bold text-slate-800"),
                    cls="bg-slate-50 p-4 rounded-xl border border-slate-100"
                ),
                Div(
                    P("🟢 Cleared Cleanly", cls="text-xs font-semibold text-green-700"),
                    P(str(ca.cleared_count), cls="text-2xl font-bold text-green-600"),
                    cls="bg-green-50/50 p-4 rounded-xl border border-green-100"
                ),
                Div(
                    P("🟡 Needs Support (1 Arrear)", cls="text-xs font-semibold text-amber-700"),
                    P(str(ca.single_u_count), cls="text-2xl font-bold text-amber-600"),
                    cls="bg-amber-50/50 p-4 rounded-xl border border-amber-100"
                ),
                Div(
                    P("🔴 Critical (Multi-Arrear)", cls="text-xs font-semibold text-red-700"),
                    P(str(ca.multiple_u_count), cls="text-2xl font-bold text-red-600"),
                    cls="bg-red-50/50 p-4 rounded-xl border border-red-100"
                ),
                cls="grid grid-cols-2 md:grid-cols-4 gap-4"
            ),
            cls="card p-5 mb-6"
        ),

        Div(
            Button("All", data_filter="all", data_filter_target="student-row",
                   cls="filter-btn active", aria_label="Filter all students"),
            Button("Cleared", data_filter="cleared", data_filter_target="student-row",
                   cls="filter-btn", aria_label="Filter cleared students"),
            Button("Arrear (1)", data_filter="u", data_filter_target="student-row",
                   cls="filter-btn", aria_label="Filter single arrear students"),
            Button("Multi-Arrear", data_filter="multi-u", data_filter_target="student-row",
                   cls="filter-btn", aria_label="Filter multiple arrear students"),
            Button("SA", data_filter="sa", data_filter_target="student-row",
                   cls="filter-btn", aria_label="Filter SA students"),
            Button("WD", data_filter="wd", data_filter_target="student-row",
                   cls="filter-btn", aria_label="Filter WD students"),
            Button("Malpractice", data_filter="malpractice", data_filter_target="student-row",
                   cls="filter-btn", aria_label="Filter Malpractice students"),
            Button("High Performers", data_filter="high", data_filter_target="student-row",
                   cls="filter-btn", aria_label="Filter high performing students"),
            cls="flex flex-wrap gap-2 mb-6 filter-group"
        ),

        data_table(
            ["Rank", "Reg No", "Name", "GPA", "Arrear Count", "Status", "%ile"],
            rows,
            table_id="student-table"
        ),

        cls="max-w-7xl mx-auto"
    ), ca=ca)


def page_student_detail(ca: ClassAnalysis, regno: str) -> Tuple:
    s = get_student(ca, regno)
    if s is None:
        return layout("Student Not Found", "", Div(
            Div(
                Span("🔍", cls="text-4xl mb-3 block"),
                H1("Student Not Found", cls="text-xl font-bold text-slate-800 mb-2"),
                P(f"No student with register number '{html.escape(regno)}' was found.", cls="text-slate-500 mb-4"),
                A("← Back to Students", href="/students",
                  cls="inline-flex items-center gap-2 text-blue-600 hover:text-blue-700 font-medium"),
                cls="text-center py-16"
            ),
            cls="max-w-lg mx-auto"
        ))

    subj_fig = fig_student_subjects(s)
    vs_fig = fig_student_vs_class(s, ca)
    ai = generate_student_brief(s, ca)
    ptm = SESSION.get("ptm_briefs", {}).get(s.regno)

    return layout(f"Student – {s.name}", "students", Div(
        # Navigation bar & Student Header with Prominent Download Button
        Div(
            A("← Back to Students List", href="/students", cls="text-xs text-blue-600 hover:underline font-medium mb-3 inline-block"),
            Div(
                Div(
                    H1(s.name or "—", cls="text-2xl font-bold text-slate-800"),
                    Div(
                        Span(s.regno, cls="text-sm text-slate-500 font-mono"),
                        Span("·", cls="text-slate-300"),
                        status_badge(s.attention),
                        cls="flex items-center gap-2 mt-1"
                    ),
                    cls="flex-1"
                ),
                Div(
                    Form(Button("📥 Download Student PDF Report", type="submit",
                                cls="px-4 py-2.5 text-sm font-semibold bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors shadow-sm"),
                         action="/report/student/" + s.regno, method="POST"),
                    cls="flex-shrink-0"
                ),
                cls="flex flex-col sm:flex-row sm:items-center gap-4 mb-6"
            ),
        ),

        # Academic Status Classification Tree Breakdown
        Div(
            H3("Academic Status Breakdown", cls="text-sm font-bold text-slate-800 mb-3"),
            Div(
                Div(
                    Span("🔴 Arrears", cls="font-semibold text-red-700 text-xs block mb-1"),
                    Div(
                        P(f"├── U (Reappearance Required): {s.u_count}", cls="text-xs text-slate-700 font-mono"),
                        P(f"└── RA (Reappearance / Arrear): {s.ra_count}", cls="text-xs text-slate-700 font-mono"),
                        cls="pl-2 border-l-2 border-red-300"
                    ),
                    cls="bg-red-50 p-3.5 rounded-lg border border-red-200"
                ),
                Div(
                    Span("🟠 Attendance Issues", cls="font-semibold text-amber-700 text-xs block mb-1"),
                    Div(
                        P(f"└── SA (Shortage of Attendance): {s.sa_count}", cls="text-xs text-slate-700 font-mono"),
                        cls="pl-2 border-l-2 border-amber-300"
                    ),
                    cls="bg-amber-50 p-3.5 rounded-lg border border-amber-200"
                ),
                Div(
                    Span("🔵 Withdrawal", cls="font-semibold text-purple-700 text-xs block mb-1"),
                    Div(
                        P(f"└── WD (Withdrawal): {s.wd_count}", cls="text-xs text-slate-700 font-mono"),
                        cls="pl-2 border-l-2 border-purple-300"
                    ),
                    cls="bg-purple-50 p-3.5 rounded-lg border border-purple-200"
                ),
                Div(
                    Span("🟣 Malpractice Record", cls="font-semibold text-indigo-700 text-xs block mb-1"),
                    Div(
                        P(f"├── MM (Malpractice): {s.mm_count}", cls="text-xs text-slate-700 font-mono"),
                        P(f"└── WH2 (Malpractice Withheld): {s.wh2_count}", cls="text-xs text-slate-700 font-mono"),
                        cls="pl-2 border-l-2 border-indigo-300"
                    ),
                    cls="bg-indigo-50 p-3.5 rounded-lg border border-indigo-200"
                ),
                cls="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-4 gap-4"
            ),
            cls="card p-5 mb-6"
        ),

        # Recommended Faculty Action Card
        Div(
            Div(
                Div(
                    H3("Recommended Faculty Action", cls="text-sm font-bold text-slate-900 mb-1 flex items-center gap-2"),
                    P("Discuss backlog recovery plan and monitor foundation subject progress.", cls="text-xs text-slate-600 mb-3") if s.arrear_count else P("Good academic standing. Encourage advanced elective options.", cls="text-xs text-slate-600 mb-3"),
                    cls="flex-1"
                ),
                Div(
                    Form(Button("📥 Download Student Report", type="submit",
                                cls="px-4 py-2 text-xs font-semibold bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors shadow-xs mr-2 inline-block"),
                         action="/report/student/" + s.regno, method="POST", cls="inline-block"),
                    Button("👨‍👩‍👦 Generate PTM Brief", type="button",
                           hx_post=f"/student/{s.regno}/ptm",
                           hx_target="#ptm-result",
                           hx_swap="innerHTML",
                           cls="px-4 py-2 text-xs font-semibold bg-indigo-600 hover:bg-indigo-700 text-white rounded-lg transition-colors shadow-xs inline-block"),
                    cls="flex items-center gap-2"
                ),
                cls="flex flex-col sm:flex-row sm:items-center justify-between gap-4"
            ),
            cls="card p-5 border-l-4 border-l-indigo-600 mb-6 bg-indigo-50/30"
        ),

        # Dedicated Active Arrear Burden & Carried Previous Arrears Banner (for current arrear students)
        *( [Div(
            Div(
                Span("⚠️", cls="text-xl mr-3 flex-shrink-0"),
                Div(
                    H3(f"Active Arrear Burden: {s.total_active_arrear_count} Total Subjects ({s.arrear_count} Current Sem + {s.carried_previous_arrear_count} Carried Previous)", cls="text-sm font-bold text-amber-900"),
                    P("This student currently holds active arrears and continues to carry uncleared previous semester subjects. Both current and foundation backlogs require immediate faculty counseling and remedial scheduling:", cls="text-xs text-amber-800 mt-0.5 mb-2"),
                    Div(
                        Span("Carried Previous Arrears:", cls="text-xs font-bold text-amber-950 mr-2"),
                        Span(", ".join(s.carried_previous_arrears), cls="text-xs font-mono font-semibold text-amber-900 bg-amber-100/90 px-2 py-0.5 rounded border border-amber-300"),
                        cls="flex items-center flex-wrap gap-1"
                    ),
                ),
                cls="flex items-start"
            ),
            cls="card p-4 mb-6 border-l-4 border-l-amber-500 bg-amber-50/60 shadow-2xs"
        )] if (s.arrear_count > 0 and s.carried_previous_arrears) else [] ),

        # Academic Snapshot Grid
        Div(
            H3("Academic Snapshot", cls="text-sm font-bold text-slate-800 mb-3"),
            Div(
                stat_card("Semester GPA", fmt_gpa(s.gpa), "#16a34a" if (s.gpa or 0) >= 7 else "#d97706"),
                stat_card("Class Rank", rank_text(s, ca.student_count), "#3b82f6"),
                stat_card("Percentile", f"{s.percentile:.1f}%" if s.percentile else "—", "#16a34a"),
                stat_card("Cleared Courses", f"{s.passed_courses}/{s.total_courses}", "#16a34a"),
                stat_card("Current Arrears", str(s.arrear_count), "#dc2626" if s.arrear_count else "#16a34a"),
                stat_card("Carried Backlogs", str(s.carried_previous_arrear_count) if s.arrear_count > 0 else "0", "#d97706" if (s.arrear_count > 0 and s.carried_previous_arrear_count > 0) else "#64748b"),
                stat_card("Total Arrear Load", str(s.total_active_arrear_count) if s.arrear_count > 0 else "0", "#dc2626" if (s.arrear_count > 0 and s.total_active_arrear_count > 0) else "#16a34a"),
                stat_card("SA / WD / Malpractice", f"{s.sa_count} / {s.wd_count} / {s.malpractice_count}", "#7c3aed" if (s.sa_count or s.wd_count or s.malpractice_count) else "#64748b"),
                cls="grid grid-cols-2 md:grid-cols-4 gap-4"
            ),
            cls="mb-6"
        ),

        # Course Performance Table
        Div(
            H3("Semester Course Results", cls="text-sm font-semibold text-slate-700 mb-3"),
            data_table(
                ["Subject", "Code", "Credits", "Grade", "Points", "Subject Page"],
                [[
                    c.subject, c.course_code or "—", c.credits,
                    grade_badge(c.grade), f"{c.points}",
                    A("View Subject →", href=f"/subjects/{quote(c.subject, safe='')}", cls="text-xs text-blue-600 hover:underline font-medium")
                ] for c in s.courses],
                table_id="student-result"
            ),
            cls="card p-5 mb-6"
        ),

        # Optional Previous Semester / Reappearance History Section (if student has past attempts across any number of pages)
        *( [Div(
            Div(
                Span("📚", cls="text-base mr-2"),
                Div(
                    H3("Previous Semester Reappearance & Arrear Clearance History (Optional Reference)", cls="text-sm font-bold text-purple-900"),
                    P("Records from earlier semester reappearance exams written during this COE examination session (segregated from current semester GPA calculation):", cls="text-xs text-purple-700 mt-0.5"),
                ),
                cls="flex items-start mb-3"
            ),
            data_table(
                ["Subject", "Course Code", "Credits", "Reappearance Grade", "Grade Points", "Clearance Status"],
                [[
                    c.subject, c.course_code or "—", c.credits,
                    grade_badge(c.grade), f"{c.points}",
                    Span("✓ Arrear Cleared", cls="text-xs font-bold text-green-700 bg-green-50 px-2 py-0.5 rounded border border-green-200") if c.grade in PASSING_GRADES else Span("⚠ Reappearance Required", cls="text-xs font-bold text-red-700 bg-red-50 px-2 py-0.5 rounded border border-red-200")
                ] for c in s.previous_semester_results],
                table_id="student-prev-result"
            ),
            cls="card p-5 mb-6 border-l-4 border-l-purple-500 bg-purple-50/20"
        )] if getattr(s, "previous_semester_results", None) else [] ),

        # GPA Calculation Breakdown
        Div(
            H3("GPA Calculation Breakdown", cls="text-sm font-semibold text-slate-700 mb-3"),
            Div(
                Div(
                    P("Credits Completed (passed)", cls="text-xs text-slate-500 mb-1"),
                    P(f"{s.credits_completed:.1f}", cls="text-lg font-bold text-slate-800"),
                    cls="flex-1"
                ),
                Div(
                    P("Quality Points Earned", cls="text-xs text-slate-500 mb-1"),
                    P(f"{s.quality_points:.1f}", cls="text-lg font-bold text-slate-800"),
                    cls="flex-1"
                ),
                Div(
                    P("Semester GPA", cls="text-xs text-slate-500 mb-1"),
                    P(fmt_gpa(s.gpa), cls="text-2xl font-bold text-green-600"),
                    cls="flex-1"
                ),
                cls="flex gap-6 p-4 bg-slate-50 rounded-lg"
            ),
            P("Per Regulation 2024: GPA = Total Quality Points / Credits Completed (U, RA, SA, WD, MM, and WH2 excluded).", cls="text-xs text-slate-400 mt-3"),
            cls="card p-5 mb-6"
        ),

        # Student vs Class Visualizations
        Div(
            Div(
                H3("Subject-wise Grade Points", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(subj_fig, "h-64"),
                cls="card p-5"
            ),
            Div(
                H3("Student vs Class Average", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(vs_fig, "h-64"),
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        # Performance Story: Strengths & Attention Areas
        Div(
            Div(
                H3("Strengths", cls="text-sm font-bold text-slate-800 mb-3"),
                Div(*[
                    Span(g, cls="inline-flex items-center px-3 py-1.5 bg-green-50 text-green-700 text-sm font-medium rounded-lg mr-2 mb-2")
                    for g in s.strongest_subjects[:5]
                ] or [P("Maintained declared passing standards across subjects.", cls="text-slate-500 text-sm")]),
                cls="card p-5"
            ),
            Div(
                H3("Attention Areas", cls="text-sm font-bold text-slate-800 mb-3"),
                *[Div(
                    Span(c.subject, cls="text-sm font-medium text-slate-700 flex-1"),
                    grade_badge(c.grade),
                    A("View Subject →", href=f"/subjects/{quote(c.subject, safe='')}", cls="text-xs text-blue-600 hover:underline font-medium ml-3"),
                    cls="flex items-center gap-2 py-1.5 border-b border-slate-100 last:border-0"
                ) for c in s.courses if c.grade in FAILING_GRADES] or [
                    P("Cleared all subjects cleanly with no U, RA, SA, WD, or Malpractice status.", cls="text-sm text-green-600 font-medium")
                ],
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        # AI Academic Brief
        Div(
            H3("AI Academic Brief", cls="text-sm font-semibold text-slate-700 mb-3"),
            md_block(ai["text"], ai["live"]),
        ),

        # PTM Brief Action Section
        Div(
            H3("Parent-Teacher Meeting (PTM) Brief", cls="text-sm font-semibold text-slate-700 mb-3"),
            Form(
                Button("Generate / Refresh PTM Brief", type="submit",
                       hx_post=f"/student/{s.regno}/ptm",
                       hx_target="#ptm-result",
                       hx_swap="innerHTML",
                       cls="px-4 py-2 text-sm font-medium bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors mb-4"),
            ),
            Div(
                NotStr(md_to_html(ptm["text"])) if ptm else
                P("Click the button above to generate structured PTM discussion points.",
                  cls="text-sm text-slate-400"),
                id="ptm-result",
                cls="mt-2"
            ),
            cls="card p-5 mb-6"
        ),

        cls="max-w-7xl mx-auto"
    ), ca=ca)


def page_subjects(ca: ClassAnalysis) -> Tuple:
    sorted_by_pass = sorted(ca.subjects, key=lambda s: (s.pass_pct or 0, s.avg_gp or 0), reverse=True)
    sorted_by_u = sorted(ca.subjects, key=lambda s: (s.arrear_count, -(s.pass_pct or 100)), reverse=True)

    best_subj = sorted_by_pass[0] if sorted_by_pass else None
    weakest_subj = sorted_by_u[0] if sorted_by_u and sorted_by_u[0].arrear_count > 0 else None

    fail_fig = fig_failure_concentration(ca)
    avg_fig = fig_subject_avg_gp(ca)
    pass_fig = fig_subject_pass_pct(ca)

    return layout("Subject Analysis", "subjects", Div(
        Div(
            H1("Subject Analysis", cls="text-2xl font-bold text-slate-800"),
            P(f"{len(ca.subjects)} subjects evaluated · comprehensive cohort comparison",
              cls="text-slate-500 text-sm mt-1"),
            cls="mb-6"
        ),

        # Subject Ranking Cards (Dashboard Analytics)
        Div(
            Div(
                Div(
                    Span("🏆 Best Performing Subject", cls="text-sm font-bold text-green-700 mb-1 block"),
                    P(best_subj.subject if best_subj else "—", cls="text-lg font-bold text-slate-800"),
                    P(f"Pass Rate: {fmt_np_pct(best_subj.pass_pct)} · Avg GP: {fmt_gp(best_subj.avg_gp)}" if best_subj else "", cls="text-xs text-slate-500 mt-1"),
                    A("View Subject Details →", href=f"/subjects/{quote(best_subj.subject, safe='')}", cls="text-xs text-blue-600 hover:underline mt-2 inline-block font-medium") if best_subj else None,
                    cls="card p-5 border-l-4 border-l-green-500"
                ),
                Div(
                    Span("⚠️ Weakest Subject (Highest Arrear Count)", cls="text-sm font-bold text-red-700 mb-1 block"),
                    P(weakest_subj.subject if weakest_subj else "None", cls="text-lg font-bold text-slate-800"),
                    P(f"{weakest_subj.arrear_count} arrears · Pass Rate: {fmt_np_pct(weakest_subj.pass_pct)}" if weakest_subj else "All subjects cleared cleanly", cls="text-xs text-slate-500 mt-1"),
                    A("View Subject Details →", href=f"/subjects/{quote(weakest_subj.subject, safe='')}", cls="text-xs text-blue-600 hover:underline mt-2 inline-block font-medium") if weakest_subj else None,
                    cls="card p-5 border-l-4 border-l-red-500"
                ),
                cls="grid grid-cols-1 md:grid-cols-2 gap-6 mb-6"
            ),
            P("Subject rankings are generated directly from calculated cohort metrics.", cls="text-xs text-slate-400 mb-6 italic"),
        ),

        # Key Metrics Grid
        Div(
            stat_card("Total Subjects", str(len(ca.subjects)), "#3b82f6"),
            stat_card("Total Arrears (U/RA)", str(sum(s.arrear_count for s in ca.subjects)), "#dc2626"),
            stat_card("Average GP Range", f"{min(s.avg_gp or 0 for s in ca.subjects):.1f} – {max(s.avg_gp or 0 for s in ca.subjects):.1f}" if ca.subjects else "—", "#64748b"),
            stat_card("Record Pass Rate", fmt_np_pct(ca.record_pass_rate), "#16a34a"),
            cls="grid grid-cols-2 lg:grid-cols-4 gap-4 mb-8"
        ),

        # Visual Charts
        Div(
            Div(
                H3("Average Grade Point by Subject", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(avg_fig, "h-72"),
                cls="card p-5"
            ),
            Div(
                H3("Pass Percentage by Subject", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(pass_fig, "h-72"),
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        Div(
            H3("Failure / Arrear Concentration across Subjects", cls="text-sm font-semibold text-slate-700 mb-3"),
            chart_container(fail_fig, "h-64"),
            cls="card p-5 mb-6"
        ),

        # All Subjects Health Table with Smarter Priority Rating (Requirement 4)
        Div(
            H3("All Subjects Health Table", cls="text-sm font-semibold text-slate-700 mb-4"),
            data_table(
                ["Subject", "Code", "Credits", "Students", "Avg GP", "Pass %", "Arrears", "Priority Rating", "Action"],
                [[
                    A(s.subject, href=f"/subjects/{quote(s.subject, safe='')}", cls="text-blue-600 hover:text-blue-700 font-medium"),
                    s.course_code or "—",
                    s.credits,
                    s.student_count,
                    fmt_gp(s.avg_gp),
                    fmt_np_pct(s.pass_pct),
                    Span(str(s.arrear_count), cls="text-red-600 font-semibold") if s.arrear_count else "0",
                    Span(s.priority_level, cls="text-xs font-bold text-red-600") if s.priority_level == "High Attention" else Span(s.priority_level, cls="text-xs font-medium text-slate-500"),
                    A("Details →", href=f"/subjects/{quote(s.subject, safe='')}", cls="text-xs text-blue-600 hover:underline font-medium"),
                ] for s in ca.subjects],
                table_id="subject-table"
            ),
            cls="card p-5"
        ),

        cls="max-w-7xl mx-auto"
    ), ca=ca)


def page_subject_detail(ca: ClassAnalysis, subject: str) -> Tuple:
    subj = get_subject(ca, subject)
    if subj is None:
        return layout("Subject Not Found", "", Div(
            Div(
                Span("📚", cls="text-4xl mb-3 block"),
                H1("Subject Not Found", cls="text-xl font-bold text-slate-800 mb-2"),
                P(f"No subject '{html.escape(subject)}' found.", cls="text-slate-500 mb-4"),
                A("← Back to Subjects", href="/subjects",
                  cls="inline-flex items-center gap-2 text-blue-600 hover:text-blue-700 font-medium"),
                cls="text-center py-16"
            ),
            cls="max-w-lg mx-auto"
        ))

    ai = generate_subject_ai_insight(subj, ca)

    diff_str = f"{abs(subj.gp_diff_vs_class):.2f} lower than class avg" if subj.gp_diff_vs_class < 0 else f"+{subj.gp_diff_vs_class:.2f} above class avg"

    return layout(f"Subject – {subj.subject}", "subjects", Div(
        A("← Back to All Subjects", href="/subjects", cls="text-xs text-blue-600 hover:underline font-medium mb-3 inline-block"),
        Div(
            Div(
                H1(subj.subject, cls="text-2xl font-bold text-slate-800"),
                Div(
                    Span(subj.course_code or "—", cls="text-sm text-slate-500 font-mono") if subj.course_code else None,
                    Span(f"Credits: {subj.credits}", cls="text-sm text-slate-500"),
                    Span(f"Priority: {subj.priority_level}", cls="text-xs font-bold text-red-600" if subj.priority_level == "High Attention" else "text-xs font-medium text-slate-500"),
                    explain_insight_box(
                        f"Subject Flagged: {subj.subject}",
                        [
                            f"{subj.arrear_count} student(s) received U/RA grade",
                            f"Failure rate ({fmt_np_pct(subj.u_pct)}) is above class average",
                            f"Average grade point ({subj.avg_gp:.2f}) is lower than cohort average" if subj.gp_diff_vs_class < 0 else "Average grade point is at or near cohort average"
                        ],
                        id_suffix=f"subj-{quote(subj.subject, safe='')}"
                    ) if subj.priority_level == "High Attention" or subj.arrear_count > 0 else None,
                    cls="flex flex-wrap items-center gap-3 mt-1"
                ),
                cls="flex-1"
            ),
            Div(
                Form(Button("Download Report", type="submit",
                            cls="px-4 py-2 text-sm font-medium bg-slate-100 hover:bg-slate-200 text-slate-700 rounded-lg transition-colors"),
                     action="/report/subject/" + subj.subject, method="POST"),
                cls="flex-shrink-0"
            ),
            cls="flex flex-col sm:flex-row sm:items-center gap-4 mb-6"
        ),

        # Issue Detected Alert Box
        Div(
            Div(
                Span("🚨", cls="text-2xl mr-3"),
                Div(
                    H4(f"Issue Detected in {subj.subject}", cls="text-sm font-bold text-red-900"),
                    P(f"{subj.arrear_count} student(s) received U/RA grade ({fmt_np_pct(subj.u_pct)} failure rate). Remedial coaching advised.", cls="text-xs text-red-700 mt-0.5"),
                    cls="flex-1"
                ),
                cls="flex items-center"
            ),
            cls="bg-red-50 p-4 rounded-xl border border-red-200 mb-6"
        ) if subj.arrear_count > 0 else None,

        # Requirement 4: Smarter Deterministic Subject Analysis Card
        Div(
            H3(f"{subj.subject} Academic Priority & Comparison", cls="text-sm font-bold text-slate-800 mb-3"),
            Div(
                Div(
                    P("Arrear Count", cls="text-xs text-slate-500"),
                    P(f"{subj.arrear_count}", cls="text-xl font-bold text-red-600"),
                    P(f"{fmt_np_pct(subj.u_pct)} failure rate", cls="text-xs text-slate-400 mt-1"),
                    cls="bg-slate-50 p-4 rounded-lg"
                ),
                Div(
                    P("Pass Percentage", cls="text-xs text-slate-500"),
                    P(fmt_np_pct(subj.pass_pct), cls="text-xl font-bold text-slate-800"),
                    P(f"{subj.pass_count} passed", cls="text-xs text-slate-400 mt-1"),
                    cls="bg-slate-50 p-4 rounded-lg"
                ),
                Div(
                    P("Average GP Difference", cls="text-xs text-slate-500"),
                    P(f"{subj.avg_gp:.2f}" if subj.avg_gp else "—", cls="text-xl font-bold text-slate-800"),
                    P(diff_str, cls="text-xs font-medium text-amber-600" if subj.gp_diff_vs_class < 0 else "text-xs font-medium text-green-600"),
                    cls="bg-slate-50 p-4 rounded-lg"
                ),
                Div(
                    P("Attention Priority", cls="text-xs text-slate-500"),
                    P(subj.priority_level, cls="text-xl font-bold text-red-600" if subj.priority_level == "High Attention" else "text-xl font-bold text-slate-800"),
                    P("Deterministic Rating", cls="text-xs text-slate-400 mt-1"),
                    cls="bg-slate-50 p-4 rounded-lg"
                ),
                cls="grid grid-cols-2 md:grid-cols-4 gap-4"
            ),
            cls="card p-5 mb-6"
        ),

        # Subject Result Status Classification Table Breakdown
        Div(
            H3("Result Status Breakdown", cls="text-sm font-semibold text-slate-700 mb-3"),
            Div(
                data_table(
                    ["Passed", "U (Arrear)", "RA (Arrear)", "SA (Attendance)", "WD (Withdrawal)", "MM (Malpractice)", "WH2 (Withheld)"],
                    [[
                        Span(str(subj.pass_count), cls="font-bold text-green-600"),
                        Span(str(subj.u_count), cls="font-bold text-red-600") if subj.u_count else "0",
                        Span(str(subj.ra_count), cls="font-bold text-red-600") if subj.ra_count else "0",
                        Span(str(subj.sa_count), cls="font-bold text-amber-600") if subj.sa_count else "0",
                        Span(str(subj.wd_count), cls="font-bold text-purple-600") if subj.wd_count else "0",
                        Span(str(subj.mm_count), cls="font-bold text-indigo-600") if subj.mm_count else "0",
                        Span(str(subj.wh2_count), cls="font-bold text-indigo-600") if subj.wh2_count else "0",
                    ]]
                ),
                cls="card p-4 mb-6"
            ),
        ),

        # Grade Distribution Breakdown Bars
        Div(
            H3("Detailed Grade Distribution", cls="text-sm font-semibold text-slate-700 mb-4"),
            *[Div(
                Span(g, cls="w-12 text-xs font-bold text-slate-600"),
                Div(
                    Div(style=f"width:{round(100*subj.grade_counts.get(g, 0)/max(subj.student_count,1))}%; background:{GRADE_COLORS.get(g,'#999')};",
                        cls="h-6 rounded-md transition-all"),
                    cls="flex-1 bg-slate-100 rounded-md h-6 overflow-hidden"
                ),
                Span(str(subj.grade_counts.get(g, 0)), cls="w-10 text-xs text-right font-medium text-slate-600"),
                cls="flex items-center gap-3 mb-2"
            ) for g in GRADE_ORDER if subj.grade_counts.get(g, 0) > 0],
            cls="card p-5 mb-6"
        ),

        # Subject vs Class Comparison Table
        Div(
            H3("Subject vs Class Overall Benchmark", cls="text-sm font-semibold text-slate-700 mb-3"),
            data_table(
                ["Metric", "This Subject", "Class Overall Average", "Difference"],
                [
                    ["Avg GP", fmt_gp(subj.avg_gp), fmt_gp(ca.class_gpa),
                     f"+{(subj.avg_gp or 0) - (ca.class_gpa or 0):.2f}" if subj.avg_gp and ca.class_gpa else "—"],
                    ["Pass Rate", fmt_np_pct(subj.pass_pct), fmt_np_pct(ca.pass_rate),
                     f"+{(subj.pass_pct or 0) - (ca.pass_rate or 0):.1f}%" if subj.pass_pct and ca.pass_rate else "—"],
                    ["Arrears", str(subj.arrear_count), str(ca.arrear_student_count),
                     f"+{subj.arrear_count - ca.arrear_student_count}" if subj.arrear_count else "—"],
                ],
            ),
            cls="card p-5 mb-6"
        ),

        # Top Performers Section (UI Visible)
        Div(
            H3("Subject Top Performers", cls="text-sm font-semibold text-slate-700 mb-3"),
            data_table(
                ["Reg No", "Name", "Grade", "Points", "Profile Action"],
                [[st["regno"], st["name"], grade_badge(st["grade"]), st["points"],
                  A("View Student →", href=f"/students/{quote(st['regno'], safe='')}", cls="text-xs text-blue-600 hover:underline font-medium")]
                 for st in subj.top_students[:8]],
            ),
            cls="card p-5 mb-6"
        ),

        # Affected Students (Students with Arrear Grade U / RA)
        Div(
            H3("Students with Arrear Grade (U / RA)", cls="text-sm font-semibold text-red-700 mb-3"),
            *[Div(
                A(st["regno"], href=f"/students/{quote(st['regno'], safe='')}", cls="font-mono text-blue-600 hover:text-blue-700 text-sm font-medium"),
                Span(" — ", cls="text-slate-400"),
                Span(st["name"], cls="text-sm text-slate-700 flex-1"),
                grade_badge(st["grade"]),
                A("View Student Profile →", href=f"/students/{quote(st['regno'], safe='')}", cls="text-xs text-blue-600 hover:underline font-medium ml-3"),
                cls="flex items-center gap-2 py-2 border-b border-slate-100 last:border-0"
            ) for st in subj.u_students] or [
                P("No students with U / RA grade for this subject.", cls="text-sm text-green-600 py-2")
            ],
            cls="card p-5 mb-6"
        ),

        # AI Insight
        Div(
            H3("AI Subject Insight", cls="text-sm font-semibold text-slate-700 mb-3"),
            md_block(ai["text"], ai["live"]),
        ),

        cls="max-w-7xl mx-auto"
    ), ca=ca)


def page_interactive_report(ca: ClassAnalysis) -> Tuple:
    """Requirement 6: Interactive Web Report Mode."""
    ai = generate_class_ai_insight(ca)
    grade_dist_fig = fig_grade_distribution(ca)
    dist_fig = fig_student_performance_distribution(ca)
    heatmap_fig = fig_grade_heatmap(ca)

    return layout("Interactive Web Report", "reports", Div(
        # University Report Header Banner
        Div(
            Div(
                H1("SARANATHAN COLLEGE OF ENGINEERING", cls="text-xl font-extrabold text-white tracking-wide text-center"),
                P("FACULTY GRADE ANALYTICS EXECUTIVE WEB REPORT · REGULATIONS 2024", cls="text-slate-300 text-xs text-center mt-1"),
                Div(
                    Span(f"Generated: {datetime.now().strftime('%d %B %Y')}", cls="text-xs text-slate-400"),
                    Span("·", cls="text-slate-500"),
                    Span(f"Students: {ca.student_count}", cls="text-xs text-slate-400"),
                    Span("·", cls="text-slate-500"),
                    Span(f"Class GPA: {fmt_gpa(ca.class_gpa)}", cls="text-xs text-slate-400"),
                    cls="flex items-center justify-center gap-3 mt-3 border-t border-white/10 pt-2"
                ),
                cls="bg-navy-800 p-6 rounded-xl text-white shadow-lg mb-6"
            ),
        ),

        # Quick Action Buttons
        Div(
            A("📄 Download PDF Report", href="/reports",
              cls="px-4 py-2 text-sm font-semibold bg-blue-600 hover:bg-blue-700 text-white rounded-lg shadow-sm transition-colors"),
            A("← Back to Dashboard", href="/dashboard",
              cls="px-4 py-2 text-sm font-medium bg-slate-200 hover:bg-slate-300 text-slate-700 rounded-lg transition-colors"),
            cls="flex items-center gap-3 mb-6"
        ),

        # Executive Metrics Cards
        Div(
            stat_card("Total Enrolled", str(ca.student_count), "#3b82f6"),
            stat_card("Class GPA Average", fmt_gpa(ca.class_gpa), "#16a34a"),
            stat_card("Cleared Pass Rate", fmt_np_pct(ca.pass_rate), "#16a34a" if (ca.pass_rate or 0) >= 75 else "#d97706"),
            stat_card("Active Arrears (U/RA)", str(ca.arrear_student_count), "#dc2626"),
            cls="grid grid-cols-2 lg:grid-cols-4 gap-4 mb-6"
        ),

        # Interactive Visual Analytics Grid
        Div(
            Div(
                H3("Cohort Performance Distribution", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(dist_fig, "h-72"),
                cls="card p-5"
            ),
            Div(
                H3("Grade Heatmap Matrix", cls="text-sm font-semibold text-slate-700 mb-3"),
                chart_container(heatmap_fig, "h-72"),
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        # Subject Performance Ranking Table
        Div(
            H3("Subject Performance Ranking", cls="text-sm font-bold text-slate-800 mb-3"),
            data_table(
                ["Rank", "Subject", "Course Code", "Credits", "Students", "Avg GP", "Pass %", "Priority Rating"],
                [[
                    idx,
                    A(s.subject, href=f"/subjects/{quote(s.subject, safe='')}", cls="text-blue-600 hover:underline font-medium"),
                    s.course_code or "—",
                    s.credits,
                    s.student_count,
                    fmt_gp(s.avg_gp),
                    fmt_np_pct(s.pass_pct),
                    Span(s.priority_level, cls="text-xs font-bold text-red-600") if s.priority_level == "High Attention" else Span(s.priority_level, cls="text-xs font-medium text-slate-500"),
                ] for idx, s in enumerate(sorted(ca.subjects, key=lambda x: (x.pass_pct or 0, x.avg_gp or 0), reverse=True), start=1)],
            ),
            cls="card p-5 mb-6"
        ),

        # AI Advisory Summary
        Div(
            H3("AI Advisory Executive Summary", cls="text-sm font-bold text-slate-800 mb-3"),
            md_block(ai["text"], ai["live"]),
            cls="mb-6"
        ),

        cls="max-w-7xl mx-auto"
    ), ca=ca)


def page_attention(ca: ClassAnalysis) -> Tuple:
    multi_u = [s for s in ca.students if s.attention == STATUS_MULTI_U]
    backlog_u = [s for s in ca.students if s.has_backlog_arrears]
    u_only = [s for s in ca.students if s.attention == STATUS_U]
    sa = [s for s in ca.students if s.attention == STATUS_SA]
    wd = [s for s in ca.students if s.attention == STATUS_WD]
    malpractice = [s for s in ca.students if s.attention == STATUS_MALPRACTICE]

    def attn_section(title, students, badge_cls, empty_msg) -> Div:
        if not students:
            return Div(
                Div(
                    Span("✓", cls="text-2xl mb-2 block text-green-400"),
                    P(empty_msg, cls="text-sm text-slate-500"),
                    cls="text-center py-8"
                ),
                cls="card"
            )
        return Div(
            Div(
                Span(title, cls="text-sm font-semibold text-slate-700"),
                Span(f"{len(students)}", cls="ml-2 badge badge-slate"),
                cls="flex items-center mb-4"
            ),
            *[Div(
                Div(
                    A(s.name or "—", href=f"/students/{quote(s.regno, safe='')}",
                      cls="text-sm font-medium text-slate-800 hover:text-blue-600"),
                    P(
                        f"{s.regno}" +
                        (f" · Carried Prev Arrears: {', '.join(s.carried_previous_arrears)}" if (s.arrear_count > 0 and s.carried_previous_arrears) else (f" · Backlogs: {', '.join(s.backlog_subjects)}" if s.backlog_subjects else "")),
                        cls="text-xs text-slate-400 font-mono mt-0.5"
                    ),
                    cls="flex-1 min-w-0"
                ),
                Div(
                    gpa_cell(s.gpa),
                    (Span(f"{s.arrear_count} Current + {s.carried_previous_arrear_count} Carried", cls="text-xs text-red-700 font-semibold bg-red-50 px-2 py-0.5 rounded border border-red-200") if (s.arrear_count > 0 and s.carried_previous_arrears)
                     else (Span(f"{s.arrear_count} Arrear", cls="text-xs text-red-600 font-semibold") if s.arrear_count else None)),
                    A("Profile & PTM →", href=f"/students/{quote(s.regno, safe='')}", cls="text-xs text-blue-600 hover:underline font-medium ml-2"),
                    cls="flex items-center gap-3"
                ),
                cls="flex items-center justify-between py-3 border-b border-slate-100 last:border-0"
            ) for s in students],
            cls="card p-5"
        )

    return layout("Attention", "attention", Div(
        Div(
            H1("Students Requiring Academic Attention", cls="text-2xl font-bold text-slate-800"),
            P(f"{len(multi_u) + len(backlog_u) + len(u_only) + len(sa) + len(wd) + len(malpractice)} students need targeted attention across academic & status categories",
              cls="text-slate-500 text-sm mt-1"),
            cls="mb-6"
        ),

        Div(
            stat_card("Multiple Arrears", str(len(multi_u)), "#dc2626"),
            stat_card("Sem 1-4 Backlogs", str(len(backlog_u)), "#d97706"),
            stat_card("Single Arrear", str(len(u_only)), "#ef4444"),
            stat_card("SA (Attendance)", str(len(sa)), "#d97706"),
            stat_card("Malpractice (MM/WH2)", str(len(malpractice)), "#7c3aed"),
            cls="grid grid-cols-2 lg:grid-cols-5 gap-4 mb-8"
        ),

        Div(
            attn_section("Multiple Arrear Students (U/RA)", multi_u, "badge-red", "No students with multiple arrears."),
            attn_section("⚠️ Sem 1-4 Backlog Subject Arrears", backlog_u, "badge-amber", "No students carrying Sem 1-4 foundation backlogs."),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),
        Div(
            attn_section("Single Current Semester Arrear Students", u_only, "badge-red", "No students with single arrear."),
            attn_section("SA – Shortage of Attendance", sa, "badge-amber", "No students with SA status."),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),
        Div(
            attn_section("WD – Withdrawal Status", wd, "badge-purple", "No students with WD status."),
            attn_section("🚨 Malpractice Record (MM / WH2)", malpractice, "badge-indigo", "No students with malpractice record."),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        cls="max-w-7xl mx-auto"
    ), ca=ca)


def page_rankings(ca: ClassAnalysis) -> Tuple:
    ranked = [s for s in ca.students if s.gpa is not None]
    ranked.sort(key=lambda s: (s.rank or 9999, s.regno))

    rows = []
    for s in ranked:
        medal = ""
        if s.rank == 1:
            medal = "🥇"
        elif s.rank == 2:
            medal = "🥈"
        elif s.rank == 3:
            medal = "🥉"
        rows.append(Tr(
            Td(f"{medal} {s.rank}" if medal else str(s.rank), cls="font-medium text-slate-600"),
            Td(s.regno, cls="font-mono text-xs text-slate-500"),
            Td(A(s.name or "—", href=f"/students/{quote(s.regno, safe='')}",
                 cls="text-blue-600 hover:text-blue-700 font-medium")),
            Td(gpa_cell(s.gpa)),
            Td(f"{s.percentile:.1f}%" if s.percentile else "—", cls="text-slate-500"),
            Td(str(s.arrear_count), cls="text-red-600 font-semibold") if s.arrear_count else Td("0"),
            Td(A("Profile →", href=f"/students/{quote(s.regno, safe='')}", cls="text-xs text-blue-600 hover:underline font-medium")),
        ))

    return layout("Rankings", "rankings", Div(
        Div(
            H1("Class Rankings", cls="text-2xl font-bold text-slate-800"),
            P(f"Based on credit-weighted GPA under Regulations 2024 · {len(ranked)} students ranked",
              cls="text-slate-500 text-sm mt-1"),
            cls="mb-6"
        ),

        Div(
            stat_card("Total Ranked", str(len(ranked)), "#3b82f6"),
            stat_card("Top GPA", fmt_gpa(ranked[0].gpa) if ranked else "—", "#16a34a"),
            stat_card("Median GPA", fmt_gpa(ca.median_gpa), "#64748b"),
            stat_card("Lowest GPA", fmt_gpa(ranked[-1].gpa) if ranked else "—", "#dc2626"),
            cls="grid grid-cols-2 lg:grid-cols-4 gap-4 mb-6"
        ),

        data_table(
            ["Rank", "Reg No", "Name", "GPA", "Percentile", "Arrears", "Action"],
            rows,
            table_id="rankings-table"
        ),

        cls="max-w-5xl mx-auto"
    ), ca=ca)


def page_ai_insights(ca: ClassAnalysis) -> Tuple:
    ai = generate_class_ai_insight(ca)

    problems = [s for s in ca.subjects if s.arrear_count > 0]
    problems.sort(key=lambda s: s.arrear_count, reverse=True)

    return layout("AI Insights", "ai-insights", Div(
        Div(
            H1("AI Class Analysis", cls="text-2xl font-bold text-slate-800"),
            P("Advisory output from local analysis and LLM review. Verify against official records.",
              cls="text-slate-500 text-sm mt-1"),
            cls="mb-6"
        ),

        md_block(ai["text"], ai["live"]),

        Div(
            Div(
                Div(
                    Span("✓", cls="text-green-500 text-lg mr-2"),
                    Span("Strengths Summary", cls="text-sm font-semibold text-slate-700"),
                    cls="flex items-center mb-3"
                ),
                P(f"Pass rate: {fmt_np_pct(ca.pass_rate)}", cls="text-sm text-slate-600 mb-1"),
                P(f"Class GPA Average: {fmt_gpa(ca.class_gpa)}", cls="text-sm text-slate-600 mb-1"),
                P(f"Cleared students: {ca.cleared_count} of {ca.student_count}", cls="text-sm text-slate-600"),
                cls="card p-5"
            ),
            Div(
                Div(
                    Span("⚠", cls="text-red-500 text-lg mr-2"),
                    Span("Problem Subjects (Arrears)", cls="text-sm font-semibold text-slate-700"),
                    cls="flex items-center mb-3"
                ),
                *[Div(
                    A(s.subject, href=f"/subjects/{quote(s.subject, safe='')}", cls="font-medium text-slate-700 hover:text-blue-600"),
                    Span(f"{s.arrear_count} Arrear", cls="text-red-600 font-semibold text-sm ml-2"),
                    Span(f"({fmt_np_pct(s.u_pct)})", cls="text-slate-400 text-sm ml-1"),
                    cls="flex items-center py-1.5"
                ) for s in problems[:5]] or [
                    P("No subjects with arrears recorded.", cls="text-sm text-green-600")
                ],
                cls="card p-5"
            ),
            cls="grid grid-cols-1 lg:grid-cols-2 gap-6 mb-6"
        ),

        Div(
            Div(
                Span("→", cls="text-blue-500 text-lg mr-2"),
                Span("Recommended Actions", cls="text-sm font-semibold text-slate-700"),
                cls="flex items-center mb-3"
            ),
            P(f"1. Give targeted support to {ca.multiple_u_count} student(s) with multiple arrears.",
              cls="text-sm text-slate-600 mb-1"),
            P("2. Schedule remedial tutorial sessions for high failure concentration subjects.",
              cls="text-sm text-slate-600 mb-1"),
            P(f"3. Engage {ca.sa_student_count} SA, {ca.wd_student_count} WD, and {ca.malpractice_student_count} Malpractice students "
              "through the appropriate academic office.", cls="text-sm text-slate-600"),
            cls="card p-5"
        ),

        cls="max-w-4xl mx-auto"
    ), ca=ca)


def page_reports(ca: ClassAnalysis) -> Tuple:
    return layout("Reports", "reports", Div(
        Div(
            H1("Faculty Academic Report Center", cls="text-2xl font-bold text-slate-800"),
            P("Purpose-built PDF reports tailored for Class Coordinators, Subject Teachers, and Parent Meetings.",
              cls="text-slate-500 text-sm mt-1"),
            cls="mb-8"
        ),

        Div(
            # Card 1: Class Coordinator Report
            Div(
                Div(
                    Span("📊", cls="text-3xl mb-3 block"),
                    Span("FOR CLASS COORDINATOR / HOD", cls="text-[10px] font-bold tracking-wider uppercase bg-blue-100 text-blue-800 px-2 py-0.5 rounded mb-2 inline-block"),
                    H3("Class Coordinator Report", cls="text-base font-semibold text-slate-800 mb-2"),
                    P("Comprehensive class overview including executive cover page, grade distribution, "
                      "subject rankings, risk list, and AI advisory.",
                      cls="text-sm text-slate-500 mb-4 leading-relaxed"),
                    Div(
                        Form(Button("📄 Download Class Coordinator PDF", type="submit",
                                    cls="w-full px-4 py-2.5 text-sm font-medium bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors mb-2"),
                             action="/report/class", method="POST"),
                        A("🌐 Open Interactive Web Report", href="/reports/interactive",
                          cls="block w-full text-center px-4 py-2.5 text-sm font-semibold bg-indigo-50 hover:bg-indigo-100 text-indigo-700 rounded-lg transition-colors"),
                    ),
                ),
                cls="card p-6 border-t-4 border-t-blue-600"
            ),

            # Card 2: Subject Faculty Report
            Div(
                Div(
                    Span("📚", cls="text-3xl mb-3 block"),
                    Span("FOR SUBJECT TEACHER", cls="text-[10px] font-bold tracking-wider uppercase bg-emerald-100 text-emerald-800 px-2 py-0.5 rounded mb-2 inline-block"),
                    H3("Subject Faculty Report", cls="text-base font-semibold text-slate-800 mb-2"),
                    P("Detailed course performance metrics including grade distribution, failure concentration, "
                      "top students, arrear list, and remedial action items.",
                      cls="text-sm text-slate-500 mb-4 leading-relaxed"),
                    Form(
                        Select(*[Option(s.subject, value=s.subject) for s in ca.subjects],
                               name="subject", aria_label="Select Subject for Report",
                               cls="w-full px-3 py-2 text-sm border border-slate-200 rounded-lg mb-3 "
                                   "focus:ring-2 focus:ring-blue-500 focus:border-blue-500"),
                        Button("Generate Subject Faculty PDF", type="submit",
                               cls="w-full px-4 py-2.5 text-sm font-medium bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors"),
                        action="/report/subject-select", method="POST",
                    ),
                ),
                cls="card p-6 border-t-4 border-t-emerald-600"
            ),

            # Card 3: Parent Meeting (PTM) Report
            Div(
                Div(
                    Span("👨‍👩‍👦", cls="text-3xl mb-3 block"),
                    Span("FOR PARENT-TEACHER MEETING", cls="text-[10px] font-bold tracking-wider uppercase bg-indigo-100 text-indigo-800 px-2 py-0.5 rounded mb-2 inline-block"),
                    H3("Parent Meeting (PTM) Report", cls="text-base font-semibold text-slate-800 mb-2"),
                    P("Individual student report highlighting strengths, challenges, failed courses, "
                      "and structured discussion points for parent meetings.",
                      cls="text-sm text-slate-500 mb-4 leading-relaxed"),
                    Form(
                        Input(type="text", name="regno", placeholder="Enter student register number", aria_label="Student Register Number for Report",
                               cls="w-full px-3 py-2 text-sm border border-slate-200 rounded-lg mb-3 "
                                   "focus:ring-2 focus:ring-blue-500 focus:border-blue-500"),
                        Button("Generate PTM Report PDF", type="submit",
                               cls="w-full px-4 py-2.5 text-sm font-medium bg-indigo-600 hover:bg-indigo-700 text-white rounded-lg transition-colors"),
                        action="/report/student-select", method="POST",
                    ),
                ),
                cls="card p-6 border-t-4 border-t-indigo-600"
            ),

            cls="grid grid-cols-1 md:grid-cols-3 gap-6"
        ),

        cls="max-w-6xl mx-auto"
    ), ca=ca)


# =============================================================================
# 13) ROUTES & FASTHTML APPLICATION SETUP
# =============================================================================

app = FastHTML(
    hdrs=(
        Meta(charset="utf-8"),
        Meta(name="viewport", content="width=device-width, initial-scale=1"),
        Link(rel="stylesheet", href="https://cdn.jsdelivr.net/npm/tailwindcss@2.2.19/dist/tailwind.min.css"),
        Link(rel="preconnect", href="https://fonts.googleapis.com"),
        Link(rel="preconnect", href="https://fonts.gstatic.com", crossorigin=""),
        Link(rel="stylesheet", href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600;700&display=swap"),
        Style("""
            :root {
                --blue-primary: #0052cc;
                --blue-hover: #0043a8;
                --blue-dark: #0033a0;
                --blue-light: #edf4ff;
                --blue-mid: #2563eb;
                --blue-border: #bfdbfe;
                --bg-canvas: #f4f8fc;
                --bg-card: #ffffff;
                --border-card: #e8eef8;
                --shadow-card: 0 10px 30px -5px rgba(0, 82, 204, 0.04), 0 4px 12px -2px rgba(0,0,0,0.02);
                --shadow-card-hover: 0 16px 36px -4px rgba(0, 82, 204, 0.12);
                --text-heading: #0b192c;
                --text-body: #334155;
                --text-muted: #64748b;
                --text-subtle: #94a3b8;
            }

            * { box-sizing: border-box; }
            body {
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                background-color: #f4f8fc;
                background-image: 
                    radial-gradient(ellipse 90% 45% at 50% -10%, rgba(219, 234, 254, 0.85), rgba(244, 248, 252, 0)),
                    radial-gradient(circle at 15% 25%, rgba(219, 234, 254, 0.5), transparent 450px),
                    radial-gradient(circle at 85% 25%, rgba(219, 234, 254, 0.5), transparent 450px),
                    linear-gradient(180deg, #edf4ff 0%, #f4f8fc 380px, #f8fafc 100%);
                background-attachment: fixed;
                color: var(--text-body);
                line-height: 1.6;
                -webkit-font-smoothing: antialiased;
                -moz-osx-font-smoothing: grayscale;
            }

            /* Technical & Data Typography */
            .font-mono, th, td, .badge, .grade-badge, .stat-value {
                font-family: 'JetBrains Mono', monospace;
            }

            /* Headings */
            h1 { font-size: 2rem; font-weight: 800; color: var(--text-heading); line-height: 1.2; letter-spacing: -0.03em; }
            h2 { font-size: 1.35rem; font-weight: 700; color: #1e293b; line-height: 1.3; }
            h3 { font-size: 1.05rem; font-weight: 600; color: #334155; line-height: 1.4; }
            h4 { font-size: 0.875rem; font-weight: 600; color: #475569; line-height: 1.5; }

            /* ============ CARDS ============ */
            .card {
                background: var(--bg-card);
                border: 1px solid var(--border-card);
                border-radius: 1rem;
                box-shadow: var(--shadow-card);
                transition: all 0.22s cubic-bezier(0.16, 1, 0.3, 1);
            }
            .card:hover {
                transform: translateY(-2px);
                box-shadow: var(--shadow-card-hover);
                border-color: #c3d4f4;
            }

            /* Stat Cards */
            .stat-card {
                background: var(--bg-card);
                border: 1px solid var(--border-card);
                border-radius: 1rem;
                padding: 1.25rem 1.5rem;
                position: relative;
                overflow: hidden;
                box-shadow: var(--shadow-card);
                transition: all 0.22s cubic-bezier(0.16, 1, 0.3, 1);
            }
            .stat-card::before {
                content: '';
                position: absolute;
                top: 0; left: 0; right: 0;
                height: 3px;
                background: var(--accent-color, var(--blue-primary));
            }
            .stat-card:hover {
                transform: translateY(-2px);
                box-shadow: var(--shadow-card-hover);
            }
            .stat-card:hover::before { height: 4px; }

            /* ============ SIDEBAR ============ */
            .sidebar {
                background: #ffffff;
                border-right: 1px solid #e5eaf3;
                box-shadow: 2px 0 20px rgba(0,0,0,0.04);
                display: flex;
                flex-direction: column;
            }
            .sidebar-brand {
                padding: 1.125rem 1rem 0.875rem;
                border-bottom: 1px solid #f0f4fa;
                flex-shrink: 0;
            }
            .sidebar-brand-logo {
                width: 54px;
                height: 54px;
                object-fit: contain;
                border-radius: 50%;
                border: 2px solid #e5eaf3;
                background: #fff;
                display: block;
                margin: 0 auto 0.5rem;
                box-shadow: 0 2px 8px rgba(26,86,219,0.1);
            }
            .sidebar-brand-name {
                font-size: 0.75rem;
                font-weight: 800;
                color: #1a56db;
                letter-spacing: 0.07em;
                text-align: center;
                display: block;
                line-height: 1.2;
            }
            .sidebar-brand-sub {
                font-size: 0.6rem;
                font-weight: 600;
                color: #475569;
                letter-spacing: 0.04em;
                text-align: center;
                display: block;
                line-height: 1.4;
                margin-top: 0.1rem;
            }
            .sidebar-brand-city {
                font-size: 0.58rem;
                color: #94a3b8;
                text-align: center;
                font-weight: 500;
                display: block;
                margin-top: 0.1rem;
                margin-bottom: 0.5rem;
            }
            .sidebar-portal-label {
                background: #ebf5ff;
                border: 1px solid #a4cafe;
                border-radius: 0.5rem;
                padding: 0.35rem 0.6rem;
                text-align: center;
            }
            .portal-name {
                font-size: 0.58rem;
                font-weight: 800;
                color: #1a56db;
                text-transform: uppercase;
                letter-spacing: 0.06em;
                display: block;
                line-height: 1.3;
            }
            .portal-ver {
                font-size: 0.54rem;
                color: #64748b;
                display: block;
                margin-top: 0.1rem;
            }
            .sidebar-nav-section {
                padding: 0.75rem 0.75rem 0.5rem;
                flex: 1;
                overflow-y: auto;
            }
            .sidebar-nav-label {
                font-size: 0.54rem;
                font-weight: 700;
                letter-spacing: 0.1em;
                color: #94a3b8;
                text-transform: uppercase;
                padding: 0 0.5rem;
                margin-bottom: 0.375rem;
                display: block;
            }
            .sidebar-link {
                display: flex;
                align-items: center;
                gap: 0.6rem;
                padding: 0.575rem 0.75rem;
                font-size: 0.8125rem;
                font-weight: 500;
                color: #64748b;
                border-radius: 0.625rem;
                transition: all 0.16s ease;
                margin-bottom: 0.125rem;
                border: 1px solid transparent;
                position: relative;
            }
            .sidebar-link svg {
                width: 1rem;
                height: 1rem;
                flex-shrink: 0;
                transition: all 0.16s ease;
                opacity: 0.7;
            }
            .sidebar-link:hover:not(.active) {
                background: #f1f5f9;
                color: #0f172a;
            }
            .sidebar-link:hover:not(.active) svg { opacity: 1; }
            .sidebar-link.active {
                background: #ebf5ff;
                color: #1a56db;
                font-weight: 700;
                border-color: #a4cafe;
                box-shadow: 0 2px 8px rgba(26,86,219,0.08);
            }
            .sidebar-link.active svg { opacity: 1; color: #1a56db; }
            .sidebar-link.active::before {
                content: '';
                position: absolute;
                left: -0.75rem;
                top: 20%;
                height: 60%;
                width: 3px;
                background: #1a56db;
                border-radius: 0 3px 3px 0;
            }
            .sidebar-text { flex: 1; }
            .sidebar-footer {
                border-top: 1px solid #f0f4fa;
                padding: 0.75rem;
                flex-shrink: 0;
            }
            .sidebar-user {
                display: flex;
                align-items: center;
                gap: 0.6rem;
                padding: 0.5rem 0.625rem;
                border-radius: 0.625rem;
                background: #f8fafc;
                border: 1px solid #e5eaf3;
            }
            .sidebar-avatar {
                width: 2rem;
                height: 2rem;
                border-radius: 50%;
                background: linear-gradient(135deg, #1a56db, #3b82f6);
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 0.75rem;
                font-weight: 700;
                color: #fff;
                flex-shrink: 0;
            }
            .sidebar-user-name { font-size: 0.75rem; font-weight: 600; color: #0f172a; display: block; line-height: 1.3; }
            .sidebar-user-role { font-size: 0.65rem; color: #64748b; display: block; }

            /* ============ MOBILE HEADER ============ */
            .mobile-header-bar {
                background: #ffffff;
                border-bottom: 1px solid #e5eaf3;
                box-shadow: 0 2px 12px rgba(0,0,0,0.04);
            }

            /* ============ DATA TABLES ============ */
            table {
                border-collapse: separate;
                border-spacing: 0;
                width: 100%;
                border-radius: 0.875rem;
                overflow: hidden;
                border: 1px solid var(--border-card);
                box-shadow: var(--shadow-card);
            }
            thead th {
                background: #f8fafc;
                border-bottom: 1px solid var(--border-card);
                font-weight: 700;
                font-size: 0.72rem;
                text-transform: uppercase;
                letter-spacing: 0.08em;
                color: #475569;
                padding: 0.875rem 1rem;
                text-align: left;
                white-space: nowrap;
            }
            tbody td {
                padding: 0.875rem 1rem;
                border-bottom: 1px solid #f1f5f9;
                font-size: 0.875rem;
                color: var(--text-body);
                background: #ffffff;
                transition: all 0.12s ease;
            }
            tbody tr { transition: all 0.12s ease; }
            tbody tr:hover td { background: #f0f6ff; color: #0f172a; }
            tbody tr:last-child td { border-bottom: none; }

            /* ============ FILTER BUTTONS ============ */
            .filter-btn {
                padding: 0.4rem 0.875rem;
                font-size: 0.75rem;
                font-weight: 600;
                border-radius: 9999px;
                transition: all 0.18s ease;
                cursor: pointer;
                font-family: 'JetBrains Mono', monospace;
                display: inline-flex;
                align-items: center;
                gap: 0.4rem;
            }
            .filter-btn:not(.active) {
                background: #fff;
                border: 1px solid #e5eaf3;
                color: #64748b;
                box-shadow: 0 1px 2px rgba(0,0,0,0.04);
            }
            .filter-btn:not(.active):hover {
                background: #f8fafc;
                color: #0f172a;
                border-color: #94a3b8;
                transform: translateY(-1px);
            }
            .filter-btn.active {
                background: var(--blue-primary);
                color: #fff;
                font-weight: 700;
                border: 1px solid var(--blue-primary);
                box-shadow: 0 4px 12px rgba(26,86,219,0.3);
            }

            /* ============ LINKS ============ */
            a { color: inherit; text-decoration: none; transition: all 0.15s ease; }
            a.text-blue, a.text-blue-600 { color: var(--blue-primary); }
            a.text-blue:hover, a.text-blue-600:hover { color: var(--blue-hover); }

            /* ============ BUTTONS ============ */
            button, .btn {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                background: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 0.5rem;
                font-weight: 600;
                color: #1e293b;
                box-shadow: 0 1px 2px rgba(0,0,0,0.05);
                transition: all 0.16s ease;
                cursor: pointer;
            }
            button:hover, .btn:hover {
                background: #f8fafc;
                border-color: #94a3b8;
                color: #0f172a;
                transform: translateY(-1px);
                box-shadow: 0 3px 8px rgba(0,0,0,0.08);
            }
            button:active, .btn:active { transform: translateY(0); }
            button[type="submit"], .btn-primary, .bg-blue-600 {
                background: linear-gradient(160deg, #2563eb 0%, #1a56db 100%) !important;
                border: 1px solid #1a56db !important;
                color: #ffffff !important;
                box-shadow: 0 4px 14px rgba(26,86,219,0.35) !important;
            }
            button[type="submit"]:hover, .btn-primary:hover, .bg-blue-600:hover {
                background: linear-gradient(160deg, #1e429f 0%, #1e3a8a 100%) !important;
                box-shadow: 0 6px 18px rgba(26,86,219,0.45) !important;
                transform: translateY(-1px) !important;
            }
            button[type="submit"]:active {
                transform: translateY(0) !important;
            }

            /* ============ FORM INPUTS & FANCY FILE PICKERS ============ */
            input, select, textarea {
                background: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 0.5rem;
                color: #0f172a;
                box-shadow: 0 1px 2px rgba(0,0,0,0.04);
                transition: all 0.18s ease;
            }
            input:focus, select:focus, textarea:focus {
                background: #fff;
                border-color: var(--blue-primary);
                box-shadow: 0 0 0 3px rgba(26,86,219,0.12);
                outline: none;
                color: #0f172a;
            }
            *:focus-visible {
                outline: 2px solid var(--blue-primary);
                outline-offset: 2px;
                border-radius: 6px;
            }

            /* Fancy File Input Buttons */
            input[type="file"] {
                cursor: pointer;
            }
            input[type="file"]::file-selector-button,
            input[type="file"]::-webkit-file-upload-button {
                background: linear-gradient(135deg, #1a56db 0%, #2563eb 100%);
                color: #ffffff !important;
                border: none;
                padding: 0.45rem 1rem;
                border-radius: 0.45rem;
                font-weight: 700;
                font-size: 0.75rem;
                letter-spacing: 0.02em;
                cursor: pointer;
                box-shadow: 0 2px 8px rgba(26, 86, 219, 0.25);
                transition: all 0.2s cubic-bezier(0.16, 1, 0.3, 1);
                margin-right: 0.75rem;
            }
            input[type="file"]::file-selector-button:hover,
            input[type="file"]::-webkit-file-upload-button:hover {
                background: linear-gradient(135deg, #1e429f 0%, #1d4ed8 100%);
                box-shadow: 0 4px 12px rgba(26, 86, 219, 0.38);
                transform: translateY(-1px);
            }
            .fancy-file-dropzone {
                border: 2px dashed #cbd5e1;
                border-radius: 0.875rem;
                padding: 0.875rem 1rem;
                background: #f8fafc;
                transition: all 0.22s cubic-bezier(0.16, 1, 0.3, 1);
                cursor: pointer;
            }
            .fancy-file-dropzone:hover, .fancy-file-dropzone:focus-within {
                border-color: #1a56db;
                background: #ebf5ff;
                box-shadow: 0 4px 16px rgba(26, 86, 219, 0.08);
            }

            /* ============ BADGES ============ */
            .badge {
                display: inline-flex;
                align-items: center;
                padding: 0.275rem 0.7rem;
                font-size: 0.72rem;
                font-weight: 700;
                border-radius: 9999px;
                line-height: 1;
                border: 1px solid transparent;
            }
            .badge-green { background: #dcfce7; color: #15803d; border-color: #bbf7d0; }
            .badge-red { background: #fee2e2; color: #b91c1c; border-color: #fecaca; }
            .badge-amber { background: #fef3c7; color: #b45309; border-color: #fde68a; }
            .badge-purple { background: #f3e8ff; color: #7e22ce; border-color: #e9d5ff; }
            .badge-indigo { background: #e0e7ff; color: #4338ca; border-color: #c7d2fe; }
            .badge-slate { background: #f1f5f9; color: #475569; border-color: #e2e8f0; }
            .badge-blue { background: #ebf5ff; color: #1a56db; border-color: #a4cafe; }

            /* ============ GRADE BADGES ============ */
            .grade-badge {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                min-width: 2.2rem;
                padding: 0.25rem 0.55rem;
                font-size: 0.75rem;
                font-weight: 800;
                border-radius: 0.45rem;
                color: #fff;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                transition: transform 0.15s ease;
            }
            .grade-badge:hover { transform: scale(1.08); }
            .grade-O { background: #16a34a; }
            .grade-APLUS { background: #22c55e; }
            .grade-A { background: #10b981; }
            .grade-BPLUS { background: #1a56db; }
            .grade-B { background: #3b82f6; }
            .grade-C { background: #64748b; }
            .grade-U { background: #ef4444; }
            .grade-RA { background: #dc2626; }
            .grade-SA { background: #f59e0b; }
            .grade-WD { background: #a855f7; }
            .grade-MM { background: #7c3aed; }
            .grade-WH2 { background: #6b21a8; }

            /* ============ ALERTS ============ */
            .alert {
                padding: 0.875rem 1.125rem;
                border-radius: 0.75rem;
                font-size: 0.875rem;
                font-weight: 600;
                display: flex;
                align-items: center;
                gap: 0.625rem;
                margin-bottom: 1.25rem;
            }
            .alert-blue { background: #ebf5ff; color: #1e40af; border: 1px solid #a4cafe; }
            .alert-green { background: #f0fdf4; color: #166534; border: 1px solid #bbf7d0; }
            .alert-red { background: #fef2f2; color: #991b1b; border: 1px solid #fecaca; }
            .alert-amber { background: #fffbeb; color: #92400e; border: 1px solid #fde68a; }

            /* ============ AI / PROSE ============ */
            .prose { font-size: 0.875rem; line-height: 1.75; color: var(--text-body); }
            .prose li { margin: 0.4rem 0; list-style: disc; padding-left: 1.25rem; }
            .prose p { margin: 0.6rem 0; }
            .prose h4 { margin: 1.1rem 0 0.4rem 0; font-weight: 700; color: #0f172a; }
            .prose strong { font-weight: 700; color: var(--blue-primary); }

            /* ============ CHART CONTAINER ============ */
            .chart-container {
                background: #fff;
                border: 1px solid var(--border-card);
                border-radius: 1rem;
                padding: 1.25rem;
                min-height: 20rem;
                box-shadow: var(--shadow-card);
                transition: all 0.22s ease;
            }
            .chart-container:hover {
                box-shadow: var(--shadow-card-hover);
                border-color: #c3d4f4;
            }
            .chart-container .js-plotly-plot { width: 100% !important; }

            /* ============ EMPTY STATE / LOADING ============ */
            .empty-state {
                text-align: center;
                padding: 3.5rem 1.5rem;
                color: #64748b;
            }
            .loading { display: flex; align-items: center; justify-content: center; padding: 3rem; }
            .spinner {
                width: 2.25rem;
                height: 2.25rem;
                border: 3px solid #e0f2fe;
                border-top-color: var(--blue-primary);
                border-radius: 50%;
                animation: spin 0.8s cubic-bezier(0.6, 0.2, 0.4, 0.8) infinite;
            }
            @keyframes spin { to { transform: rotate(360deg); } }

            /* ============ SCROLLBAR ============ */
            ::-webkit-scrollbar { width: 7px; height: 7px; }
            ::-webkit-scrollbar-track { background: #f1f5f9; }
            ::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 4px; }
            ::-webkit-scrollbar-thumb:hover { background: var(--blue-primary); }

            /* ============ RESPONSIVE / MOBILE ============ */
            @media (max-width: 1023px) {
                .mobile-nav {
                    position: fixed;
                    top: 0; left: 0; right: 0;
                    z-index: 50;
                    background: #ffffff;
                    border-bottom: 1px solid #e5eaf3;
                    box-shadow: 0 4px 16px rgba(0,0,0,0.06);
                }
            }
            @media (min-width: 1024px) {
                .mobile-nav { display: none; }
                .main-content { margin-left: 16rem; }
            }

            /* ============ UTILITY OVERRIDES ============ */
            .text-slate-800, .text-slate-900 { color: #0f172a !important; }
            .text-white { color: #0f172a !important; }
            .sidebar .text-white { color: #0f172a !important; }
            .text-slate-700 { color: #1e293b !important; }
            .text-slate-600 { color: #334155 !important; }
            .text-slate-500, .text-slate-400 { color: #64748b !important; }
            .bg-white { background: #ffffff !important; }
            .bg-slate-50 { background: #f8fafc !important; }
            .bg-slate-100 { background: #f1f5f9 !important; }
            .border-slate-200, .border-slate-300 { border-color: #e5eaf3 !important; }

            /* ============================================================
               DARK MODE
               Applied via class="dark" on <body>
            ============================================================ */
            body.dark {
                --blue-primary: #4d94ff;
                --blue-hover: #6ba8ff;
                --blue-light: #1a2845;
                --blue-border: #2d4a7a;
                --bg-canvas: #0d1117;
                --bg-card: #161b22;
                --border-card: #21262d;
                --shadow-card: 0 4px 16px rgba(0,0,0,0.4), 0 1px 4px rgba(0,0,0,0.3);
                --shadow-card-hover: 0 12px 28px rgba(0,0,0,0.6);
                --text-heading: #f0f6ff;
                --text-body: #c9d1d9;
                --text-muted: #8b949e;
                --text-subtle: #6e7681;
                background-color: #0d1117;
                background-image:
                    radial-gradient(ellipse 90% 40% at 50% -5%, rgba(30, 60, 120, 0.55), rgba(13, 17, 23, 0)),
                    radial-gradient(circle at 15% 25%, rgba(26, 40, 80, 0.4), transparent 400px),
                    radial-gradient(circle at 85% 25%, rgba(26, 40, 80, 0.4), transparent 400px),
                    linear-gradient(180deg, #111827 0%, #0d1117 400px, #0a0e15 100%);
                color: var(--text-body);
            }

            /* Dark: Headings */
            body.dark h1, body.dark h2, body.dark h3, body.dark h4 { color: var(--text-heading); }

            /* Dark: Cards */
            body.dark .card {
                background: var(--bg-card);
                border-color: var(--border-card);
            }
            body.dark .card:hover { border-color: #2d4a7a; }
            body.dark .stat-card { background: var(--bg-card); border-color: var(--border-card); }

            /* Dark: Sidebar */
            body.dark .sidebar {
                background: #111827;
                border-right-color: #21262d;
                box-shadow: 2px 0 20px rgba(0,0,0,0.5);
            }
            body.dark .sidebar-brand { border-bottom-color: #21262d; }
            body.dark .sidebar-brand-logo { border-color: #21262d; background: #1a2236; }
            body.dark .sidebar-brand-name { color: #4d94ff; }
            body.dark .sidebar-brand-sub { color: #8b949e; }
            body.dark .sidebar-brand-city { color: #6e7681; }
            body.dark .sidebar-portal-label { background: #1a2845; border-color: #2d4a7a; }
            body.dark .portal-name { color: #4d94ff; }
            body.dark .portal-ver { color: #8b949e; }
            body.dark .sidebar-nav-label { color: #6e7681; }
            body.dark .sidebar-link { color: #8b949e; }
            body.dark .sidebar-link:hover:not(.active) { background: #1a2236; color: #e6edf3; }
            body.dark .sidebar-link.active {
                background: #1a2845;
                color: #4d94ff;
                border-color: #2d4a7a;
            }
            body.dark .sidebar-link.active::before { background: #4d94ff; }
            body.dark .sidebar-footer { border-top-color: #21262d; }
            body.dark .sidebar-user { background: #161b22; border-color: #21262d; }
            body.dark .sidebar-user-name { color: #e6edf3; }
            body.dark .sidebar-user-role { color: #8b949e; }

            /* Dark: Mobile Header */
            body.dark .mobile-header-bar {
                background: #111827;
                border-bottom-color: #21262d;
                box-shadow: 0 2px 12px rgba(0,0,0,0.5);
            }
            body.dark .mobile-nav {
                background: #111827;
                border-bottom-color: #21262d;
            }

            /* Dark: Main content area */
            body.dark .main-content { background: transparent; }
            body.dark .bg-slate-50 { background: #111827 !important; }
            body.dark .bg-white { background: #161b22 !important; }
            body.dark .bg-slate-100 { background: #1a2236 !important; }
            body.dark .border-slate-200, body.dark .border-slate-300 { border-color: #21262d !important; }
            body.dark .text-slate-900, body.dark .text-slate-800 { color: #e6edf3 !important; }
            body.dark .text-slate-700 { color: #c9d1d9 !important; }
            body.dark .text-slate-600 { color: #b1bac4 !important; }
            body.dark .text-slate-500, body.dark .text-slate-400 { color: #8b949e !important; }
            body.dark .text-white { color: #ffffff !important; }

            /* Dark: Tables */
            body.dark table { border-color: #21262d; }
            body.dark thead th { background: #1a2236; color: #8b949e; border-bottom-color: #21262d; }
            body.dark tbody td { background: #161b22; color: #c9d1d9; border-bottom-color: #21262d; }
            body.dark tbody tr:hover td { background: #1a2845; color: #e6edf3; }

            /* Dark: Inputs */
            body.dark input, body.dark select, body.dark textarea {
                background: #1a2236;
                border-color: #30363d;
                color: #e6edf3;
            }
            body.dark input:focus, body.dark select:focus, body.dark textarea:focus {
                background: #1d2a3e;
                border-color: #4d94ff;
                box-shadow: 0 0 0 3px rgba(77, 148, 255, 0.15);
                color: #e6edf3;
            }
            body.dark .fancy-file-dropzone {
                background: #1a2236;
                border-color: #30363d;
            }
            body.dark .fancy-file-dropzone:hover, body.dark .fancy-file-dropzone:focus-within {
                border-color: #4d94ff;
                background: #1a2845;
                box-shadow: 0 4px 16px rgba(77, 148, 255, 0.12);
            }

            /* Dark: Buttons */
            body.dark button:not([type="submit"]):not(.btn-primary):not(.bg-blue-600):not(.filter-btn) {
                background: #1a2236;
                border-color: #30363d;
                color: #c9d1d9;
            }
            body.dark button:not([type="submit"]):not(.btn-primary):not(.bg-blue-600):not(.filter-btn):hover {
                background: #21262d;
                border-color: #4d94ff;
                color: #e6edf3;
            }
            body.dark .filter-btn:not(.active) {
                background: #1a2236;
                border-color: #30363d;
                color: #8b949e;
            }
            body.dark .filter-btn:not(.active):hover {
                background: #21262d;
                border-color: #4d94ff;
                color: #e6edf3;
            }

            /* Dark: Badges */
            body.dark .badge-green { background: #0d2a1a; color: #56d364; border-color: #196132; }
            body.dark .badge-red { background: #2d0f0f; color: #ff7b7b; border-color: #6d1a1a; }
            body.dark .badge-amber { background: #2a1f08; color: #e3b341; border-color: #5e3f08; }
            body.dark .badge-blue { background: #1a2845; color: #4d94ff; border-color: #2d4a7a; }
            body.dark .badge-slate { background: #1a2236; color: #8b949e; border-color: #21262d; }
            body.dark .badge-purple { background: #21103a; color: #c084fc; border-color: #4a1880; }
            body.dark .badge-indigo { background: #1a1f45; color: #818cf8; border-color: #2d3578; }

            /* Dark: Alerts */
            body.dark .alert-blue { background: #1a2845; color: #93c5fd; border-color: #2d4a7a; }
            body.dark .alert-green { background: #0d2a1a; color: #56d364; border-color: #196132; }
            body.dark .alert-red { background: #2d0f0f; color: #ff7b7b; border-color: #6d1a1a; }
            body.dark .alert-amber { background: #2a1f08; color: #e3b341; border-color: #5e3f08; }

            /* Dark: Chart containers */
            body.dark .chart-container { background: #161b22; border-color: #21262d; }
            body.dark .chart-container:hover { border-color: #2d4a7a; }

            /* Dark: Scrollbar */
            body.dark ::-webkit-scrollbar-track { background: #0d1117; }
            body.dark ::-webkit-scrollbar-thumb { background: #30363d; }
            body.dark ::-webkit-scrollbar-thumb:hover { background: #4d94ff; }

            /* Dark: Header bar divider */
            body.dark [class*="border-slate-200"] { border-color: rgba(33, 38, 45, 0.8) !important; }

            /* Theme toggle icon transition */
            #theme-toggle-icon { transition: transform 0.4s cubic-bezier(0.34, 1.56, 0.64, 1), opacity 0.2s ease; }
            body.dark #theme-toggle-btn { background: #1a2236; border-color: #30363d; color: #e6edf3; }
        """),
        Script("""
            // Apply dark mode immediately before paint to avoid flash
            (function() {
                var saved = localStorage.getItem('theme');
                var prefersDark = window.matchMedia('(prefers-color-scheme: dark)').matches;
                if (saved === 'dark' || (!saved && prefersDark)) {
                    document.body ? document.body.classList.add('dark') : document.addEventListener('DOMContentLoaded', function(){ document.body.classList.add('dark'); });
                }
            })();

            function toggleTheme() {
                var body = document.body;
                var isDark = body.classList.toggle('dark');
                localStorage.setItem('theme', isDark ? 'dark' : 'light');

                // Update all toggle button icons
                document.querySelectorAll('#theme-toggle-icon').forEach(function(el) {
                    if (isDark) {
                        el.innerHTML = '<path stroke-linecap="round" stroke-linejoin="round" d="M20.354 15.354A9 9 0 018.646 3.646 9.003 9.003 0 0012 21a9.003 9.003 0 008.354-5.646z"/>';
                        el.setAttribute('fill', 'none');
                    } else {
                        el.innerHTML = '<path stroke-linecap="round" stroke-linejoin="round" d="M12 3v1m0 16v1m9-9h-1M4 12H3m15.364 6.364l-.707-.707M6.343 6.343l-.707-.707m12.728 0l-.707.707M6.343 17.657l-.707.707M16 12a4 4 0 11-8 0 4 4 0 018 0z"/>';
                    }
                });
            }

            // Sync icon state on load
            document.addEventListener('DOMContentLoaded', function() {
                var isDark = document.body.classList.contains('dark');
                document.querySelectorAll('#theme-toggle-icon').forEach(function(el) {
                    if (isDark) {
                        el.innerHTML = '<path stroke-linecap="round" stroke-linejoin="round" d="M20.354 15.354A9 9 0 018.646 3.646 9.003 9.003 0 0012 21a9.003 9.003 0 008.354-5.646z"/>';
                        el.setAttribute('fill', 'none');
                    }
                });
            });
        """),
    ),
    exception_handlers={404: lambda req, exc: page_upload()},
)


@app.get("/favicon.ico")
def route_favicon():
    return Response(status_code=204)


@app.get("/logo")
def route_logo():
    logo_path = os.path.join(os.path.dirname(__file__), "sara.webp")
    try:
        with open(logo_path, "rb") as f:
            data = f.read()
        return Response(content=data, media_type="image/webp", headers={"Cache-Control": "public, max-age=86400"})
    except FileNotFoundError:
        return Response(status_code=404)


@app.get("/")
def route_root():
    return page_upload()



@app.get("/upload")
def route_upload_get():
    return page_upload()


def page_pdf_preview() -> Tuple:
    """PDF Preflight Extraction Review & Inspection Page."""
    pdf_report: PDFExtractionReport = SESSION.get("preview_pdf_report") or PDFExtractionReport()
    reconcil: Optional[ReconciliationReport] = SESSION.get("reconciliation_report")
    filename = SESSION.get("preview_pdf_filename", "coe_result.pdf")
    meta = pdf_report.doc_metadata

    # 1. Document Metadata Card
    doc_meta_card = Div(
        H3("COE PDF Document Metadata", cls="text-sm font-bold text-slate-800 mb-3"),
        Div(
            Div(Span("Institution:", cls="text-xs text-slate-400 block"), Span(meta.institution, cls="text-sm font-medium text-slate-800")),
            Div(Span("Programme / Dept:", cls="text-xs text-slate-400 block"), Span(f"{meta.programme} ({meta.department})", cls="text-sm font-medium text-slate-800")),
            Div(Span("Regulation & Semester:", cls="text-xs text-slate-400 block"), Span(f"{meta.regulation} · {meta.semester}", cls="text-sm font-medium text-slate-800")),
            Div(Span("Exam Session & Date:", cls="text-xs text-slate-400 block"), Span(f"{meta.exam_session} · {meta.publication_date}", cls="text-sm font-medium text-slate-800")),
            Div(Span("Doc Type & Pages:", cls="text-xs text-slate-400 block"), Span(f"{meta.document_type} ({meta.page_count} Pages)", cls="text-sm font-medium text-slate-800")),
            cls="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 gap-4"
        ),
        cls="card p-5 mb-6"
    )

    # 2. Multi-Semester Cohort & Reappearance Summary Card
    multi_sem_card = None
    if getattr(pdf_report, "detected_semesters", None) and len(pdf_report.detected_semesters) > 1:
        sem_badges = []
        for s_info in pdf_report.detected_semesters:
            if s_info["is_primary"]:
                sem_badges.append(Div(
                    Div(
                        Span("🎯 TARGET REGULAR COHORT", cls="text-[10px] font-extrabold text-blue-700 bg-blue-100 px-2 py-0.5 rounded-full inline-block mb-1.5"),
                        H4(f"{s_info['semester_label']} (Active Analysis)", cls="text-sm font-bold text-slate-900"),
                        P(f"{s_info['student_count']} Students · {s_info['subject_count']} Subjects · {s_info['record_count']} Grade Records", cls="text-xs text-slate-600 mt-0.5"),
                        P(f"Authoritative Courses: {', '.join(s_info['subject_codes'][:8])}", cls="text-[11px] font-mono text-blue-800 mt-1"),
                    ),
                    cls="p-3.5 bg-blue-50/90 border border-blue-200 rounded-xl flex-1 shadow-xs"
                ))
            else:
                sem_badges.append(Div(
                    Div(
                        Span("📚 OPTIONAL PREVIOUS SEMESTER REFERENCE", cls="text-[10px] font-extrabold text-purple-700 bg-purple-100 px-2 py-0.5 rounded-full inline-block mb-1.5"),
                        H4(f"{s_info['semester_label']} (Arrear / Reappearance)", cls="text-sm font-bold text-slate-900"),
                        P(f"{s_info['student_count']} Students · {s_info['subject_count']} Subjects · {s_info['record_count']} Grade Records", cls="text-xs text-slate-600 mt-0.5"),
                        P(f"Reappearance Courses: {', '.join(s_info['subject_codes'][:8])}", cls="text-[11px] font-mono text-purple-800 mt-1"),
                    ),
                    cls="p-3.5 bg-purple-50/90 border border-purple-200 rounded-xl flex-1 shadow-xs"
                ))
        
        multi_sem_card = Div(
            Div(
                Span("📑", cls="text-xl mr-2.5"),
                Div(
                    H3("Multi-Semester COE Result Document Detected & Segregated", cls="text-sm font-bold text-slate-900"),
                    P("This PDF contains both regular semester results and earlier semester arrear reappearance exams across multiple pages. The engine has cleanly segregated the primary regular semester for 100% conflict-free analytics, while attaching all previous semester records as an optional backlog reference.", cls="text-xs text-slate-600 mt-0.5"),
                ),
                cls="flex items-start mb-3.5"
            ),
            Div(*sem_badges, cls="flex flex-col sm:flex-row gap-3"),
            cls="card p-5 mb-6 border-l-4 border-l-purple-600 bg-gradient-to-r from-purple-50/40 via-slate-50 to-blue-50/40"
        )

    # 3. Data Quality & Confidence Stats
    conf_pct = int(pdf_report.overall_confidence * 100)
    conf_color = "#16a34a" if conf_pct >= 85 else "#d97706"

    stats_cards = Div(
        stat_card("Extraction Confidence", f"{conf_pct}%", conf_color),
        stat_card("Detected Students", str(pdf_report.student_count), "#3b82f6"),
        stat_card("Detected Subjects", str(pdf_report.subject_count), "#8b5cf6"),
        stat_card("Validated Result Cells", str(pdf_report.result_cell_count), "#16a34a"),
        stat_card("Quarantined Tokens", str(pdf_report.unknown_token_count), "#dc2626" if pdf_report.unknown_token_count > 0 else "#64748b"),
        cls="grid grid-cols-2 lg:grid-cols-5 gap-4 mb-6"
    )

    # 3. Raw Extraction Inspector
    inspector_rows = []
    for item in pdf_report.raw_inspector_items[:60]:
        badge_cls = "bg-green-100 text-green-800" if item["confidence"] == "HIGH" else "bg-amber-100 text-amber-800"
        page_num = item['source_page']
        raw_escaped = item['raw_text'].replace("'", "\\'").replace('"', '\\"')
        reg_v = item["parsed_regno"]
        name_v = item["parsed_name"]
        subj_v = item["parsed_subject"]
        grade_v = item["parsed_grade"]
        conf_v = item["confidence"]
        inspector_rows.append(Tr(
            Td(
                Div(
                    Span(f"Page {page_num}", cls="text-xs font-mono font-bold text-slate-700 block mb-0.5"),
                    A(
                        f"📄 View Page {page_num}",
                        href="#",
                        onclick=f"alert('Source Document Provenance Traceability\\n\\nRecord: {reg_v} ({name_v})\\nSubject: {subj_v} -> Grade: {grade_v}\\nConfidence: {conf_v}\\n\\nSource Location: COE PDF Page {page_num}\\nRaw Text Line:\\n{raw_escaped}'); return false;",
                        cls="inline-block px-2 py-0.5 text-[10px] font-semibold bg-blue-50 text-blue-700 hover:bg-blue-100 rounded border border-blue-200 transition-colors"
                    )
                ),
                cls="px-3 py-2 border-b"
            ),
            Td(item["raw_text"], cls="px-3 py-2 text-xs font-mono text-slate-600 border-b truncate max-w-xs"),
            Td(item["parsed_regno"], cls="px-3 py-2 text-xs font-mono font-bold text-slate-900 border-b"),
            Td(item["parsed_name"], cls="px-3 py-2 text-xs font-medium text-slate-800 border-b"),
            Td(item["parsed_subject"], cls="px-3 py-2 text-xs text-slate-700 border-b"),
            Td(Span(item["parsed_grade"], cls=f"px-2 py-0.5 text-xs font-bold rounded {grade_badge_cls(item['parsed_grade'])}"), cls="px-3 py-2 border-b"),
            Td(Span(item["confidence"], cls=f"px-2 py-0.5 text-[10px] font-semibold rounded {badge_cls}"), cls="px-3 py-2 border-b"),
        ))

    inspector_card = Div(
        H3("🔍 Raw PDF Extraction Inspector (Page Provenance Traceability)", cls="text-sm font-bold text-slate-800 mb-2"),
        P("Inspect exact page source and coordinates for parsed student grade cells. Click [View Page X] to verify source line:", cls="text-xs text-slate-500 mb-3"),
        Div(
            Table(
                Thead(Tr(
                    Th("Source Location", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    Th("Raw Text Line", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    Th("Parsed RegNo", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    Th("Parsed Name", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    Th("Subject", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    Th("Grade", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    Th("Confidence", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                )),
                Tbody(*inspector_rows),
                cls="w-full border border-slate-200 rounded-lg overflow-hidden"
            ),
            cls="overflow-x-auto max-h-96 mb-6"
        ),
        cls="card p-5 mb-6"
    )

    # 4. PDF vs Excel Reconciliation Section (If available)
    reconcil_card = None
    if reconcil:
        reconcil_rows = []
        for m in reconcil.mismatched_records:
            reconcil_rows.append(Tr(
                Td(m["register_number"], cls="px-3 py-2 text-xs font-mono font-bold text-slate-900 border-b"),
                Td(m["student_name"], cls="px-3 py-2 text-xs text-slate-800 border-b"),
                Td(m["subject"], cls="px-3 py-2 text-xs text-slate-700 border-b"),
                Td(Span(m["pdf_grade"], cls="px-2 py-0.5 text-xs font-bold bg-blue-100 text-blue-800 rounded"), cls="px-3 py-2 border-b"),
                Td(Span(m["excel_grade"], cls="px-2 py-0.5 text-xs font-bold bg-amber-100 text-amber-800 rounded"), cls="px-3 py-2 border-b"),
                Td(Span("🚨 MISMATCH", cls="px-2 py-0.5 text-xs font-bold bg-red-100 text-red-800 rounded"), cls="px-3 py-2 border-b"),
            ))

        reconcil_card = Div(
            Div(
                Span("🔄", cls="text-xl mr-2"),
                H3("PDF vs Excel Reconciliation Report", cls="text-sm font-bold text-slate-800"),
                Span(f"{reconcil.matched_count} Matched · {reconcil.mismatched_count} Mismatches", cls="ml-auto text-xs badge badge-slate"),
                cls="flex items-center mb-3"
            ),
            P("COE PDF is treated as the primary authoritative source. Mismatches require faculty verification:", cls="text-xs text-slate-500 mb-3"),
            Div(
                Table(
                    Thead(Tr(
                        Th("Register No", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Student Name", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Subject", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("COE PDF Grade (Authoritative)", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Excel Grade", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Discrepancy Status", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    )),
                    Tbody(*reconcil_rows),
                    cls="w-full border border-slate-200 rounded-lg overflow-hidden"
                ) if reconcil_rows else P("✓ 100% Match! PDF and Excel result grades match perfectly.", cls="text-xs font-semibold text-green-700 bg-green-50 p-3 rounded"),
                cls="overflow-x-auto mb-4"
            ),
            cls="card p-5 border-l-4 border-l-blue-600 mb-6"
        )

    return layout("PDF Result Preflight Review", "upload", Div(
        Div(
            H1("COE PDF Analysis & Data Quality Review", cls="text-2xl font-bold text-slate-800 mb-1"),
            P(f"File: {html.escape(filename)} · Source: COE Digital Result PDF", cls="text-sm text-slate-500 mb-6"),
        ),
        doc_meta_card,
        stats_cards,
        reconcil_card if reconcil_card else None,
        inspector_card,
        Form(
            Div(
                A("← Upload Different File", href="/upload", cls="px-4 py-2.5 text-sm font-medium text-slate-600 hover:text-slate-800"),
                Div(
                    A("📊 Convert to Department Excel", href="/pdf-to-excel",
                      cls="px-4 py-2.5 text-sm font-semibold text-blue-700 bg-blue-50 hover:bg-blue-100 border border-blue-200 rounded-lg transition-colors mr-3"),
                    Button("🔎 Continue to Academic Analysis", type="submit",
                           cls="px-6 py-2.5 text-sm font-semibold bg-blue-600 hover:bg-blue-700 text-white rounded-lg transition-colors shadow-sm"),
                    cls="flex items-center"
                ),
                cls="flex items-center justify-between border-t pt-4"
            ),
            action="/confirm-pdf", method="POST",
            cls="card p-6 mb-8"
        ),
        cls="max-w-5xl mx-auto"
    ))


def _get_pdf_to_excel_context() -> Tuple[Optional["PDFExtractionReport"], Optional["ClassAnalysis"]]:
    """
    Build (or reuse a cached) ClassAnalysis straight from the preflight PDF extraction,
    without requiring the faculty to go through /confirm-pdf and the normal analytics
    workflow first. Reuses pdf_records_to_dataframe + compute_class_analysis -- the
    same trusted GPA/credit engine used everywhere else in the app.
    """
    pdf_report: Optional[PDFExtractionReport] = SESSION.get("preview_pdf_report")
    if not pdf_report or not pdf_report.records:
        return None, None
    filename = SESSION.get("preview_pdf_filename", "coe_result.pdf")
    cache_key = (filename, len(pdf_report.records), id(pdf_report))
    cached = SESSION.get("pdf_to_excel_cache")
    if cached and cached[0] == cache_key:
        return pdf_report, cached[1]
    df = pdf_records_to_dataframe(pdf_report.records)
    prev_recs = getattr(pdf_report, "prev_records", None)
    ca = compute_class_analysis(df, filename, prev_records=prev_recs)
    ca.subject_mappings = build_subject_mapping_log(pdf_report.records)
    ca.quarantined_tokens = pdf_report.quarantined_tokens
    ca.format_detected = "pdf"
    ca.metadata["source_page_count"] = pdf_report.doc_metadata.page_count
    ca.metadata["extraction_confidence"] = pdf_report.overall_confidence
    SESSION["pdf_to_excel_cache"] = (cache_key, ca)
    return pdf_report, ca


def page_pdf_to_excel(pdf_report: "PDFExtractionReport", ca: "ClassAnalysis", filename: str) -> Tuple:
    """COE PDF -> Department Excel conversion page: metadata, staff mapping, IA marks upload, validation, download."""
    meta = pdf_report.doc_metadata
    mappings = list(ca.subject_mappings or [])
    staff_directory = SESSION.get("staff_directory", {})
    ia_marks_store = SESSION.get("ia_marks_directory", {})

    meta_card = card(
        H3("COE PDF → Department Excel", cls="text-lg font-bold text-slate-800 mb-1"),
        P("Convert the official COE result PDF into the department's analysis workbook.", cls="text-sm text-slate-500 mb-4"),
        Div(
            Div(Span("Institution:", cls="text-xs text-slate-400 block"), Span(meta.institution, cls="text-sm font-medium text-slate-800")),
            Div(Span("Programme:", cls="text-xs text-slate-400 block"), Span(meta.programme, cls="text-sm font-medium text-slate-800")),
            Div(Span("Semester:", cls="text-xs text-slate-400 block"), Span(meta.semester, cls="text-sm font-medium text-slate-800")),
            Div(Span("Academic Session:", cls="text-xs text-slate-400 block"), Span(f"{meta.academic_year} / {meta.exam_session}", cls="text-sm font-medium text-slate-800")),
            Div(Span("Students:", cls="text-xs text-slate-400 block"), Span(str(ca.student_count), cls="text-sm font-medium text-slate-800")),
            Div(Span("Subjects:", cls="text-xs text-slate-400 block"), Span(str(ca.subject_count), cls="text-sm font-medium text-slate-800")),
            Div(Span("Result Cells:", cls="text-xs text-slate-400 block"),
                Span(f"{ca.student_count * ca.subject_count} expected · {ca.record_count} extracted", cls="text-sm font-medium text-slate-800")),
            cls="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 gap-4"
        ),
        cls="p-5 mb-6"
    )

    rows = []
    for i, m in enumerate(mappings, start=1):
        code = m["course_code"]
        prefill = resolve_staff_for_code(code, staff_directory)
        status = "✓" if not m.get("unresolved") else "⚠"
        rows.append(Tr(
            Td(str(i), cls="px-3 py-2 text-xs text-slate-600 border-b"),
            Td(code, cls="px-3 py-2 text-xs font-mono font-bold text-slate-900 border-b"),
            Td(m["official_subject_name"], cls="px-3 py-2 text-xs text-slate-800 border-b"),
            Td(Input(type="text", name=f"staff__{code}", value=prefill, placeholder="Enter staff name",
                     cls="w-full px-2 py-1 text-xs border border-slate-300 rounded"), cls="px-3 py-2 border-b"),
            Td(str(m["credits"]), cls="px-3 py-2 text-xs text-slate-600 border-b text-center"),
            Td(status, cls="px-3 py-2 text-xs border-b text-center"),
        ))

    staff_form = Form(
        Div(
            H3("Subject Staff Mapping", cls="text-sm font-bold text-slate-800 mb-1"),
            (Div(
                Span("⚠️ ACTION REQUIRED: Please review the extracted/entered staff names below and click 'Save Staff Names' to verify.", 
                     cls="text-xs font-bold text-amber-800"),
                cls="mb-3 p-3 bg-amber-50 border border-amber-200 rounded-lg"
            ) if not SESSION.get("staff_verified", False) else Div(
                Span("✓ Staff names manually verified.", cls="text-xs font-bold text-green-800"),
                cls="mb-3 p-3 bg-green-50 border border-green-200 rounded-lg"
            )),
            P("Course codes come from the PDF and subject names come from the R2024 syllabus catalog — the only "
              "manual field required is Staff Name. Leave a field blank to flag it as not entered rather than "
              "guessing.", cls="text-xs text-slate-500 mb-3"),
            Div(
                Table(
                    Thead(Tr(
                        Th("S.No", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Course Code", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Subject Name", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Staff Name", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Credits", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                        Th("Status", cls="px-3 py-2 text-left text-xs font-semibold text-slate-600 bg-slate-50 border-b"),
                    )),
                    Tbody(*rows),
                    cls="w-full border border-slate-200 rounded-lg overflow-hidden"
                ),
                cls="overflow-x-auto mb-4"
            ),
            Div(
                Button("💾 Save Staff Names", type="submit",
                       cls="px-4 py-2 text-sm font-semibold bg-blue-600 hover:bg-blue-700 text-white rounded-lg"),
                cls="flex items-center justify-end gap-3"
            ),
            cls="p-5"
        ),
        action="/pdf-to-excel/save-staff", method="POST", cls="card mb-3"
    )

    clear_form = Form(
        Button("Clear saved staff mappings", type="submit", cls="text-xs text-red-600 hover:text-red-700 underline"),
        action="/pdf-to-excel/clear-mappings", method="POST", cls="mb-6"
    )

    # Internal Assessment Marks Upload Panel for Analysis 4
    ia1_count = len(ia_marks_store.get("ia1", {}))
    ia2_count = len(ia_marks_store.get("ia2", {}))
    ia3_count = len(ia_marks_store.get("ia3", {}))
    total_ia_students = len(set().union(
        ia_marks_store.get("ia1", {}).keys(),
        ia_marks_store.get("ia2", {}).keys(),
        ia_marks_store.get("ia3", {}).keys()
    ))
    student_meta_store = SESSION.get("student_meta_directory", {})
    gq_count = 0
    mq_count = 0
    for s in ca.students:
        q = (
            student_meta_store.get(s.regno, {}).get("quota") or
            s.meta.get("quota") or
            ia_marks_store.get("ia1", {}).get(s.regno, {}).get("quota") or
            ia_marks_store.get("ia2", {}).get(s.regno, {}).get("quota") or
            ia_marks_store.get("ia3", {}).get(s.regno, {}).get("quota") or
            "GQ"
        ).strip().upper()
        if q == "MQ":
            mq_count += 1
        else:
            gq_count += 1

    ia_upload_form = Form(
        Div(
            Div(
                Div(
                    H3("Analysis 4 — Internal Assessment (Cycle Test) Mark Sheets", cls="text-sm font-bold text-slate-800 mb-1"),
                    P("Upload Internal Assessment mark sheets to populate IAT 1, IAT 2, and IAT 3 marks, Quota, and faculty names of failed students in Analysis 4:", cls="text-xs text-slate-500"),
                    cls="flex-1"
                ),
                Div(
                    Span(f"✓ {total_ia_students} Total Students Loaded" if total_ia_students else "No IA Data Loaded",
                         cls=f"text-xs font-bold px-3 py-1 rounded-full {('bg-green-100 text-green-800 border border-green-300' if total_ia_students else 'bg-slate-100 text-slate-500 border border-slate-200')}"),
                    cls="flex-shrink-0"
                ),
                cls="flex flex-col sm:flex-row sm:items-center justify-between gap-2 mb-4"
            ),

            # Status Summary Badges
            Div(
                Span(f"IAT 1: {ia1_count} loaded", cls=f"text-[11px] font-semibold px-2.5 py-1 rounded {('bg-green-50 text-green-700 border border-green-200' if ia1_count else 'bg-slate-50 text-slate-400 border border-slate-200')}"),
                Span(f"IAT 2: {ia2_count} loaded", cls=f"text-[11px] font-semibold px-2.5 py-1 rounded {('bg-green-50 text-green-700 border border-green-200' if ia2_count else 'bg-slate-50 text-slate-400 border border-slate-200')}"),
                Span(f"IAT 3: {ia3_count} loaded", cls=f"text-[11px] font-semibold px-2.5 py-1 rounded {('bg-green-50 text-green-700 border border-green-200' if ia3_count else 'bg-slate-50 text-slate-400 border border-slate-200')}"),
                Span(f"Quotas: {gq_count} GQ · {mq_count} MQ", cls="text-[11px] font-semibold px-2.5 py-1 rounded bg-indigo-50 text-indigo-700 border border-indigo-200"),
                cls="flex flex-wrap gap-2 mb-5"
            ),

            # Upload Type Tab Controller
            Div(
                # Option 1: Combined Section-wise Files (Recommended)
                Div(
                    Div(
                        Span("⚡ Option 1: Combined Section-Wise Reports (Recommended)", cls="text-xs font-bold text-indigo-900 block mb-1"),
                        P("Upload all-cycle-test section mark sheets (e.g. II A IA.html, II B.html). Select all section files together — extracts all tests (IAT 1, IAT 2, IAT 3), Quota, and faculty names at once:", cls="text-xs text-slate-600 mb-2"),
                        Input(type="file", name="combined_files", accept=".html,.htm,.mhtml,.mht,.xlsx,.xls,.csv", multiple=True,
                              cls="w-full text-xs text-slate-500 file:mr-2 file:py-1.5 file:px-3 file:rounded-md file:border-0 file:text-xs file:font-semibold file:bg-indigo-100 file:text-indigo-800 hover:file:bg-indigo-200"),
                        cls="card p-4 bg-indigo-50/40 border border-indigo-200 rounded-lg mb-4"
                    ),
                ),

                # Option 2: Individual Cycle Test Files (Collapsible / Alternative)
                Div(
                    Span("📁 Option 2: Individual Assessment Files (Alternative)", cls="text-xs font-bold text-slate-700 block mb-2"),
                    Div(
                        Div(
                            Label("IAT 1 / Cycle Test 1", cls="block text-xs font-semibold text-slate-700 mb-1"),
                            Input(type="file", name="ia1_file", accept=".mhtml,.mht,.html,.htm,.xlsx,.xls,.csv", multiple=True,
                                  cls="w-full text-xs text-slate-500 file:mr-1 file:py-1 file:px-2 file:rounded file:border-0 file:text-xs file:bg-slate-100"),
                            cls="p-2.5 bg-slate-50 border border-slate-200 rounded"
                        ),
                        Div(
                            Label("IAT 2 / Cycle Test 2", cls="block text-xs font-semibold text-slate-700 mb-1"),
                            Input(type="file", name="ia2_file", accept=".mhtml,.mht,.html,.htm,.xlsx,.xls,.csv", multiple=True,
                                  cls="w-full text-xs text-slate-500 file:mr-1 file:py-1 file:px-2 file:rounded file:border-0 file:text-xs file:bg-slate-100"),
                            cls="p-2.5 bg-slate-50 border border-slate-200 rounded"
                        ),
                        Div(
                            Label("IAT 3 / Cycle Test 3", cls="block text-xs font-semibold text-slate-700 mb-1"),
                            Input(type="file", name="ia3_file", accept=".mhtml,.mht,.html,.htm,.xlsx,.xls,.csv", multiple=True,
                                  cls="w-full text-xs text-slate-500 file:mr-1 file:py-1 file:px-2 file:rounded file:border-0 file:text-xs file:bg-slate-100"),
                            cls="p-2.5 bg-slate-50 border border-slate-200 rounded"
                        ),
                        cls="grid grid-cols-1 sm:grid-cols-3 gap-2.5"
                    ),
                    cls="card p-4 bg-slate-50/70 border border-slate-200 rounded-lg mb-4"
                ),
            ),
            
            Div(
                (Form(Button("Clear saved IA marks", type="submit", cls="text-xs text-red-600 hover:text-red-700 underline mr-auto"),
                      action="/pdf-to-excel/clear-ia", method="POST") if total_ia_students else None),
                Button("📤 Upload & Process Mark Sheets", type="submit",
                       cls="px-5 py-2.5 text-sm font-semibold bg-indigo-600 hover:bg-indigo-700 text-white rounded-lg transition-colors shadow-xs"),
                cls="flex items-center justify-between mt-2"
            ),
            cls="p-5"
        ),
        action="/pdf-to-excel/upload-ia", method="POST", enctype="multipart/form-data", cls="card mb-6 border-l-4 border-l-indigo-600"
    )

    ok, issues = validate_export_dataset(ca)
    if ok:
        validation_body = P("✓ Dataset validated — ready for export.", cls="text-sm font-semibold text-green-700")
    else:
        validation_body = Div(
            P(f"⚠ {len(issues)} issue(s) found — download is blocked until resolved:",
              cls="text-sm font-semibold text-red-700 mb-2"),
            *[Div(P(i, cls="text-xs text-red-700 whitespace-pre-wrap"), cls="mb-2 pl-3 border-l-2 border-red-300")
              for i in issues]
        )
    validation_card = card(H3("Data Validation", cls="text-sm font-bold text-slate-800 mb-2"), validation_body, cls="p-5 mb-6")

    actions = Div(
        A("← Upload Different File", href="/upload", cls="px-4 py-2.5 text-sm font-medium text-slate-600 hover:text-slate-800"),
        (A("📥 Download Department Excel", href="/pdf-to-excel/download",
           cls="px-6 py-2.5 text-sm font-semibold bg-green-600 hover:bg-green-700 text-white rounded-lg transition-colors shadow-sm")
         if ok else
         Span("Resolve validation issues above to enable download", cls="px-4 py-2.5 text-sm font-medium text-red-600")),
        cls="flex items-center justify-between border-t pt-4"
    )

    return layout("COE PDF → Department Excel", "upload", Div(
        Div(
            H1("COE PDF → Department Excel", cls="text-2xl font-bold text-slate-800 mb-1"),
            P(f"File: {html.escape(filename)} · Convert the official COE result PDF into the department's analysis workbook.",
              cls="text-sm text-slate-500 mb-6"),
        ),
        meta_card,
        staff_form,
        clear_form,
        ia_upload_form,
        validation_card,
        card(actions, cls="p-5"),
        cls="max-w-6xl mx-auto"
    ))


# =============================================================================
# 13.5) SYLLABUS CATALOG MANAGEMENT API ROUTES
# =============================================================================

from starlette.responses import JSONResponse

@app.post("/api/syllabus/extract-draft")
async def api_extract_syllabus_draft(request):
    try:
        form = await request.form()
        file = form.get("file_syllabus")
        dept_code = str(form.get("department_code") or "AI_DS").strip().upper()
        reg_code = str(form.get("regulation_code") or "R2028").strip().upper()

        if not file:
            return JSONResponse({"ok": False, "error": "No syllabus file uploaded."}, status_code=400)

        file_bytes = await file.read()
        filename = getattr(file, "filename", "") or "syllabus.pdf"

        if filename.lower().endswith(".json"):
            try:
                data = json.loads(file_bytes.decode("utf-8"))
                courses = data.get("courses", [])
                return JSONResponse({
                    "ok": True,
                    "department_code": dept_code,
                    "regulation_code": reg_code,
                    "courses": courses,
                    "count": len(courses)
                })
            except Exception as je:
                return JSONResponse({"ok": False, "error": f"Invalid JSON format: {je}"}, status_code=400)

        try:
            draft = extract_syllabus_pdf(file_bytes, department_code_hint=dept_code, regulation_hint=reg_code)
            courses = draft.get("courses", [])
            return JSONResponse({
                "ok": True,
                "department_code": dept_code,
                "regulation_code": reg_code,
                "courses": courses,
                "count": len(courses)
            })
        except Exception as pe:
            return JSONResponse({"ok": False, "error": f"Syllabus extraction failed: {pe}"}, status_code=400)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@app.post("/api/syllabus/confirm-and-save")
async def api_confirm_and_save_syllabus(request):
    try:
        body = await request.body()
        payload = json.loads(body.decode("utf-8"))
        dept_code = str(payload.get("department_code", "")).strip().upper()
        reg_code = str(payload.get("regulation_code", "")).strip().upper()
        courses = payload.get("courses", [])
        actor = str(payload.get("actor", "Faculty / Staff Confirmation")).strip()

        if not dept_code or not reg_code:
            return JSONResponse({"ok": False, "error": "Department code and regulation name are required."}, status_code=400)

        if not courses:
            return JSONResponse({"ok": False, "error": "Course list cannot be empty."}, status_code=400)

        dept_name = None
        programme = None
        for d in get_registered_departments():
            if d.get("code") == dept_code:
                dept_name = d.get("name")
                programme = d.get("programme")
                break

        res = save_and_register_syllabus(
            department_code=dept_code,
            regulation_code=reg_code,
            courses=courses,
            department_name=dept_name,
            programme=programme,
            actor=actor
        )

        _build_syllabus_indexes()

        return JSONResponse(res)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@app.post("/upload-pdf")
async def route_upload_pdf(request):
    try:
        form = await request.form()
        file = form.get("file_pdf")
        if file is None:
            push_alert("No PDF file selected.", "red")
            return RedirectResponse("/upload", status_code=303)

        raw_filename = getattr(file, "filename", "") or "result.pdf"
        filename = os.path.basename(raw_filename)
        filename = re.sub(r"[^\w\.\-]", "_", filename)

        if not filename.lower().endswith(".pdf"):
            push_alert("Only .pdf files are accepted in Mode A.", "red")
            return RedirectResponse("/upload", status_code=303)

        pdf_bytes = await file.read()
        if not pdf_bytes:
            push_alert("Uploaded PDF file is empty.", "red")
            return RedirectResponse("/upload", status_code=303)

        dept = str(form.get("department", "") or "").strip()
        sem_raw = str(form.get("semester", "") or "").strip()
        reg = str(form.get("regulation", "") or "").strip() or "R2024"
        sem = int(sem_raw) if sem_raw and sem_raw.isdigit() else None

        analysis_ctx = {
            "department": dept,
            "semester": sem,
            "regulation": reg
        }
        SESSION["analysis_context"] = analysis_ctx

        pdf_report = extract_coe_pdf(pdf_bytes, filename, analysis_context=analysis_ctx)
        if not pdf_report.ok:
            push_alert(pdf_report.fatal_error, "red")
            return RedirectResponse("/upload", status_code=303)

        SESSION["preview_pdf_bytes"] = pdf_bytes
        SESSION["preview_pdf_filename"] = filename
        SESSION["preview_pdf_report"] = pdf_report
        SESSION["reconciliation_report"] = None

        # Extract previous batch PDF if provided or use segregated prev_records from multi-semester PDF
        file_prev = form.get("file_prev_pdf")
        prev_pdf_report = None
        prev_ca = None
        if file_prev and getattr(file_prev, "filename", ""):
            prev_bytes = await file_prev.read()
            if prev_bytes:
                p_filename = os.path.basename(getattr(file_prev, "filename", "prev_result.pdf"))
                p_filename = re.sub(r"[^\w\.\-]", "_", p_filename)
                prev_pdf_report = extract_coe_pdf(prev_bytes, p_filename, analysis_context=analysis_ctx)
                if prev_pdf_report.ok and prev_pdf_report.records:
                    p_df = pdf_records_to_dataframe(prev_pdf_report.records)
                    prev_ca = compute_class_analysis(p_df, p_filename)
                    prev_ca.subject_mappings = build_subject_mapping_log(prev_pdf_report.records)

        SESSION["prev_pdf_report"] = prev_pdf_report
        SESSION["prev_ca"] = prev_ca

        return page_pdf_preview()
    except Exception as e:
        push_alert(f"PDF extraction error: {e}", "red")
        return RedirectResponse("/upload", status_code=303)


@app.post("/upload-dual")
async def route_upload_dual(request):
    try:
        form = await request.form()
        file_pdf = form.get("file_pdf")
        file_excel = form.get("file_excel")
        if not file_pdf or not file_excel:
            push_alert("Please select both PDF and Excel files for dual reconciliation.", "red")
            return RedirectResponse("/upload", status_code=303)

        pdf_bytes = await file_pdf.read()
        excel_bytes = await file_excel.read()

        pdf_filename = os.path.basename(getattr(file_pdf, "filename", "") or "result.pdf")
        excel_filename = os.path.basename(getattr(file_excel, "filename", "") or "result.xlsx")

        dept = str(form.get("department", "") or "").strip()
        sem_raw = str(form.get("semester", "") or "").strip()
        reg = str(form.get("regulation", "") or "").strip() or "R2024"
        sem = int(sem_raw) if sem_raw and sem_raw.isdigit() else None

        analysis_ctx = {
            "department": dept,
            "semester": sem,
            "regulation": reg
        }
        SESSION["analysis_context"] = analysis_ctx

        pdf_report = extract_coe_pdf(pdf_bytes, pdf_filename, analysis_context=analysis_ctx)
        if not pdf_report.ok:
            push_alert(f"PDF error: {pdf_report.fatal_error}", "red")
            return RedirectResponse("/upload", status_code=303)

        excel_res = validate_and_clean(excel_bytes, excel_filename)
        if not excel_res.ok:
            push_alert(f"Excel error: {excel_res.report.fatal_error}", "red")
            return RedirectResponse("/upload", status_code=303)

        reconcil_report = reconcile_pdf_and_excel(pdf_report.records, excel_res.records)

        SESSION["preview_pdf_bytes"] = pdf_bytes
        SESSION["preview_pdf_filename"] = pdf_filename
        SESSION["preview_pdf_report"] = pdf_report
        SESSION["preview_raw_bytes"] = excel_bytes
        SESSION["preview_filename"] = excel_filename
        SESSION["preview_report"] = excel_res.report
        SESSION["reconciliation_report"] = reconcil_report

        push_alert(f"Dual Reconciliation complete: {reconcil_report.matched_count} matched, {reconcil_report.mismatched_count} mismatches.", "blue")
        return page_pdf_preview()
    except Exception as e:
        push_alert(f"Dual upload error: {e}", "red")
        return RedirectResponse("/upload", status_code=303)


@app.post("/confirm-pdf")
async def route_confirm_pdf(request):
    try:
        start_t = time.time()
        pdf_report: PDFExtractionReport = SESSION.get("preview_pdf_report")
        filename = SESSION.get("preview_pdf_filename", "result.pdf")
        if not pdf_report or not pdf_report.records:
            push_alert("No extracted PDF records found to analyze.", "red")
            return RedirectResponse("/upload", status_code=303)

        df = pdf_records_to_dataframe(pdf_report.records)
        prev_recs = getattr(pdf_report, "prev_records", None)
        ca = compute_class_analysis(df, filename, prev_records=prev_recs)

        ca.subject_mappings = build_subject_mapping_log(pdf_report.records)
        ca.quarantined_tokens = pdf_report.quarantined_tokens
        ca.format_detected = "pdf"
        ca.metadata["source_page_count"] = pdf_report.doc_metadata.page_count
        ca.metadata["extraction_confidence"] = pdf_report.overall_confidence

        SESSION["records"] = df
        SESSION["analytics"] = ca
        SESSION["file_name"] = filename
        SESSION["ptm_briefs"] = {}
        SESSION["analysis_duration"] = round(time.time() - start_t, 2)

        push_alert(
            f"Successfully analyzed COE PDF ({pdf_report.doc_metadata.programme}, {pdf_report.doc_metadata.semester}): "
            f"Processed {ca.student_count} students across {ca.subject_count} subjects in {SESSION['analysis_duration']}s.",
            "green"
        )
        return RedirectResponse("/dashboard", status_code=303)
    except Exception as e:
        push_alert(f"PDF confirm error: {e}", "red")
        return RedirectResponse("/upload", status_code=303)


@app.get("/pdf-to-excel")
def route_pdf_to_excel():
    pdf_report, ca = _get_pdf_to_excel_context()
    if not pdf_report or not ca:
        push_alert("Upload a COE PDF first to convert it to a department Excel workbook.", "amber")
        return RedirectResponse("/upload", status_code=303)
    return page_pdf_to_excel(pdf_report, ca, SESSION.get("preview_pdf_filename", "coe_result.pdf"))


@app.post("/pdf-to-excel/save-staff")
async def route_pdf_to_excel_save_staff(request):
    pdf_report, ca = _get_pdf_to_excel_context()
    if not pdf_report or not ca:
        push_alert("Upload a COE PDF first to convert it to a department Excel workbook.", "amber")
        return RedirectResponse("/upload", status_code=303)
    form = await request.form()
    directory = SESSION.setdefault("staff_directory", {})
    for m in (ca.subject_mappings or []):
        code = m["course_code"]
        key = f"staff__{code}"
        if key in form:
            val = str(form[key]).strip()
            directory[code] = val
            base_code = re.sub(r"([A-Z])$", "", code) if re.search(r"^(?:24)?[A-Z]{2,4}\d{3}[A-Z]$", code) else code
            if base_code != code:
                directory[base_code] = val
    SESSION["staff_verified"] = True
    push_alert("Staff names saved and verified.", "green")
    return page_pdf_to_excel(pdf_report, ca, SESSION.get("preview_pdf_filename", "coe_result.pdf"))


@app.post("/pdf-to-excel/clear-mappings")
def route_pdf_to_excel_clear_mappings():
    SESSION["staff_directory"] = {}
    push_alert("Saved staff mappings cleared.", "blue")
    return RedirectResponse("/pdf-to-excel", status_code=303)


@app.post("/pdf-to-excel/upload-ia")
async def route_pdf_to_excel_upload_ia(request):
    pdf_report, ca = _get_pdf_to_excel_context()
    if not pdf_report or not ca:
        push_alert("Upload a COE PDF first to convert it to a department Excel workbook.", "amber")
        return RedirectResponse("/upload", status_code=303)

    form = await request.form()
    ia_store = SESSION.setdefault("ia_marks_directory", {"ia1": {}, "ia2": {}, "ia3": {}})
    staff_directory = SESSION.setdefault("staff_directory", {})
    student_meta_dir = SESSION.setdefault("student_meta_directory", {})
    uploaded_counts = []
    has_new_staff = False

    def _merge_marks_records(dest: Dict[str, Dict[str, Any]], src: Dict[str, Dict[str, Any]]):
        for regno, sdata in src.items():
            if regno not in dest:
                dest[regno] = {
                    "name": sdata.get("name", ""),
                    "batch_no": sdata.get("batch_no", ""),
                    "quota": sdata.get("quota", "GQ"),
                    "marks": dict(sdata.get("marks", {}))
                }
            else:
                cur = dest[regno]
                for field in ("name", "batch_no", "quota"):
                    val = sdata.get(field)
                    if val and (not cur.get(field) or cur.get(field) in ("NA", "N/A", "GQ")):
                        cur[field] = val
                cur.setdefault("marks", {}).update(sdata.get("marks", {}))

    # 1. Process Combined Section Mark Sheets (e.g. II A IA.html, II B.html)
    combined_files = form.getlist("combined_files")
    comb_files_processed = 0
    comb_students_loaded = set()

    for file_obj in combined_files:
        if file_obj and getattr(file_obj, "filename", ""):
            raw_bytes = await file_obj.read()
            if raw_bytes:
                fname = getattr(file_obj, "filename", "combined_marks.html")
                multi_marks, _titles, parsed_staff, stu_meta = parse_combined_ia_marks_content(raw_bytes, fname)
                if any(multi_marks.values()):
                    comb_files_processed += 1
                    for t_key, t_map in multi_marks.items():
                        if t_map:
                            dest_map = ia_store.setdefault(t_key, {})
                            _merge_marks_records(dest_map, t_map)
                            comb_students_loaded.update(t_map.keys())

                    for regno, smeta in stu_meta.items():
                        student_meta_dir.setdefault(regno, {}).update(smeta)
                        # Also update ca.students meta
                        for s in ca.students:
                            if s.regno == regno:
                                s.meta.update(smeta)

                    if parsed_staff:
                        for code, staff_name in parsed_staff.items():
                            staff_name = staff_name.strip()
                            if not staff_name:
                                continue
                            base_code = re.sub(r"([A-Z])$", "", code) if re.search(r"^(?:\d{2})?[A-Z]{2,4}\d{3,4}[A-Z]$", code) else code
                            for target_code in set([code, base_code]):
                                existing_staff = staff_directory.get(target_code, "").strip()
                                if not existing_staff:
                                    staff_directory[target_code] = staff_name
                                    has_new_staff = True
                                else:
                                    existing_names = [s.strip() for s in existing_staff.split("&")]
                                    if staff_name not in existing_names:
                                        staff_directory[target_code] = f"{existing_staff} & {staff_name}"
                                        has_new_staff = True

    if comb_files_processed > 0:
        uploaded_counts.append(f"Combined Section Reports ({len(comb_students_loaded)} students across {comb_files_processed} section file(s))")

    # 2. Process Individual IA Files (ia1_file, ia2_file, ia3_file)
    for test_key, field_name in [
        ("ia1", "ia1_file"),
        ("ia2", "ia2_file"),
        ("ia3", "ia3_file")
    ]:
        file_objs = form.getlist(field_name)
        processed_files_count = 0

        for file_obj in file_objs:
            if file_obj and getattr(file_obj, "filename", ""):
                raw_bytes = await file_obj.read()
                if raw_bytes:
                    fname = getattr(file_obj, "filename", "marks.mhtml")
                    # Check if uploaded file in individual box is actually a combined report
                    multi_marks, _titles, parsed_staff, stu_meta = parse_combined_ia_marks_content(raw_bytes, fname)
                    if any(multi_marks.values()):
                        for t_k, t_map in multi_marks.items():
                            if t_map:
                                dest_map = ia_store.setdefault(t_k, {})
                                _merge_marks_records(dest_map, t_map)
                        if stu_meta:
                            for regno, smeta in stu_meta.items():
                                student_meta_dir.setdefault(regno, {}).update(smeta)
                                for s in ca.students:
                                    if s.regno == regno:
                                        s.meta.update(smeta)
                        processed_files_count += 1
                    else:
                        parsed_marks, _titles, parsed_staff = parse_ia_marks_content(raw_bytes, fname, target_test_key=test_key)
                        if parsed_marks:
                            dest_map = ia_store.setdefault(test_key, {})
                            _merge_marks_records(dest_map, parsed_marks)
                            processed_files_count += 1

                    if parsed_staff:
                        for code, staff_name in parsed_staff.items():
                            staff_name = staff_name.strip()
                            if not staff_name:
                                continue
                            base_code = re.sub(r"([A-Z])$", "", code) if re.search(r"^(?:\d{2})?[A-Z]{2,4}\d{3,4}[A-Z]$", code) else code
                            for target_code in set([code, base_code]):
                                existing_staff = staff_directory.get(target_code, "").strip()
                                if not existing_staff:
                                    staff_directory[target_code] = staff_name
                                    has_new_staff = True
                                else:
                                    existing_names = [s.strip() for s in existing_staff.split("&")]
                                    if staff_name not in existing_names:
                                        staff_directory[target_code] = f"{existing_staff} & {staff_name}"
                                        has_new_staff = True

        if processed_files_count > 0:
            uploaded_counts.append(f"{test_key.upper()} ({len(ia_store.get(test_key, {}))} students from {processed_files_count} file(s))")

    if uploaded_counts:
        push_alert(f"Successfully processed IA mark sheets: {', '.join(uploaded_counts)}.", "green")
        if has_new_staff:
            SESSION["staff_verified"] = False
            push_alert("Staff names extracted from mark sheet(s). Please review them below and click 'Save Staff Names' to confirm.", "amber")
    else:
        push_alert("No valid IA mark sheet files were uploaded.", "amber")

    return page_pdf_to_excel(pdf_report, ca, SESSION.get("preview_pdf_filename", "coe_result.pdf"))


@app.post("/pdf-to-excel/clear-ia")
def route_pdf_to_excel_clear_ia():
    SESSION["ia_marks_directory"] = {"ia1": {}, "ia2": {}, "ia3": {}}
    push_alert("Saved Internal Assessment marks cleared.", "blue")
    return RedirectResponse("/pdf-to-excel", status_code=303)


@app.post("/pdf-to-excel/save-student-meta")
async def route_pdf_to_excel_save_student_meta(request):
    pdf_report, ca = _get_pdf_to_excel_context()
    if not pdf_report or not ca:
        push_alert("Upload a COE PDF first to convert it to a department Excel workbook.", "amber")
        return RedirectResponse("/upload", status_code=303)
    form = await request.form()
    student_meta_dir = SESSION.setdefault("student_meta_directory", {})
    action = str(form.get("bulk_action") or "").strip()

    if action == "default_all_gq":
        for s in ca.students:
            cur_q = student_meta_dir.get(s.regno, {}).get("quota", "")
            if not cur_q or cur_q in ("NA", "N/A"):
                student_meta_dir.setdefault(s.regno, {})["quota"] = "GQ"
                s.meta["quota"] = "GQ"
        push_alert("All missing/unassigned student quotas set to GQ (Government Quota).", "green")
    else:
        updated = 0
        for s in ca.students:
            q_key = f"quota__{s.regno}"
            if q_key in form:
                val = str(form[q_key]).strip().upper()
                if val in ("GQ", "MQ"):
                    student_meta_dir.setdefault(s.regno, {})["quota"] = val
                    s.meta["quota"] = val
                    updated += 1
        push_alert(f"Saved quota metadata for {updated} students.", "green")

    return page_pdf_to_excel(pdf_report, ca, SESSION.get("preview_pdf_filename", "coe_result.pdf"))


@app.get("/pdf-to-excel/download")
def route_pdf_to_excel_download():
    pdf_report, ca = _get_pdf_to_excel_context()
    if not pdf_report or not ca:
        push_alert("Upload a COE PDF first to convert it to a department Excel workbook.", "amber")
        return RedirectResponse("/upload", status_code=303)
    ok, issues = validate_export_dataset(ca)
    if not ok:
        push_alert("Export validation failed: " + "; ".join(issues), "red")
        return RedirectResponse("/pdf-to-excel", status_code=303)
    staff_map = SESSION.get("staff_directory", {})
    ia_store = SESSION.get("ia_marks_directory", {})
    student_meta_dir = SESSION.get("student_meta_directory", {})
    filename = SESSION.get("preview_pdf_filename", "coe_result.pdf")
    prev_ca = SESSION.get("prev_ca")
    prev_pdf_report = SESSION.get("prev_pdf_report")
    mentor_map = SESSION.get("mentor_directory", {})
    xlsx_bytes = build_department_excel(
        ca, pdf_report, staff_map, filename, ia_store,
        prev_ca=prev_ca, prev_pdf_report=prev_pdf_report,
        mentor_map=mentor_map, student_meta_map=student_meta_dir
    )
    fname = department_excel_filename(pdf_report)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



@app.post("/upload-preview")
async def route_upload_preview(request):
    try:
        form = await request.form()
        file = form.get("file")
        if file is None:
            push_alert("No file selected.", "red")
            return RedirectResponse("/upload", status_code=303)

        raw_filename = getattr(file, "filename", "") or ""
        filename = os.path.basename(raw_filename)
        filename = re.sub(r"[^\w\.\-]", "_", filename)

        if not filename.lower().endswith(".xlsx"):
            push_alert("Only .xlsx files are accepted.", "red")
            return RedirectResponse("/upload", status_code=303)

        max_bytes = CFG["max_upload_mb"] * 1024 * 1024
        chunks = []
        total_size = 0
        while True:
            chunk = await file.read(64 * 1024)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > max_bytes:
                push_alert(f"File exceeds maximum allowed size of {CFG['max_upload_mb']} MB.", "red")
                return RedirectResponse("/upload", status_code=303)
            chunks.append(chunk)

        data = b"".join(chunks)
        if not data:
            push_alert("Uploaded file is empty.", "red")
            return RedirectResponse("/upload", status_code=303)

        df, report = _read_workbook(data)
        if report.has_fatal():
            push_alert(report.fatal_error, "red")
            return RedirectResponse("/upload", status_code=303)

        cols = [str(c) for c in df.columns.tolist()]
        mapping = _columns_to_targets(cols)
        report.mapped_columns = mapping

        full_res = validate_and_clean(data, filename, mapping)

        SESSION["preview_raw_bytes"] = data
        SESSION["preview_filename"] = filename
        SESSION["preview_cols"] = cols
        SESSION["preview_report"] = full_res.report

        return page_upload_mapping()
    except Exception as e:
        push_alert(f"File analysis error: {e}", "red")
        return RedirectResponse("/upload", status_code=303)


@app.post("/upload-confirm")
async def route_upload_confirm(request):
    try:
        start_t = time.time()
        data = SESSION.get("preview_raw_bytes")
        filename = SESSION.get("preview_filename", "result.xlsx")
        if not data:
            push_alert("Upload session expired. Please re-upload your file.", "amber")
            return RedirectResponse("/upload", status_code=303)

        form = await request.form()
        custom_mapping = {}
        for f in REQUIRED_FIELDS + OPTIONAL_FIELDS:
            val = form.get(f"map_{f}", "").strip()
            if val:
                custom_mapping[f] = val

        result = validate_and_clean(data, filename, custom_mapping)

        if not result.ok:
            SESSION["records"] = None
            SESSION["analytics"] = None
            SESSION["validation"] = result.report
            push_alert(f"Validation error: {result.report.fatal_error}", "red")
            return RedirectResponse("/upload", status_code=303)

        ca = compute_class_analysis(result.records, filename)
        ca.subject_mappings = result.report.subject_mappings
        ca.quarantined_tokens = result.report.quarantined_tokens
        ca.copy_paste_cleaned_count = result.report.copy_paste_cleaned_count
        ca.format_detected = result.report.format_detected

        SESSION["records"] = result.records
        SESSION["analytics"] = ca
        SESSION["file_name"] = filename
        SESSION["validation"] = result.report
        SESSION["ptm_briefs"] = {}
        SESSION["analysis_duration"] = round(time.time() - start_t, 2)

        msg = f"Successfully processed {result.report.valid_records} records for {ca.student_count} students across {ca.subject_count} subjects in {SESSION['analysis_duration']}s."
        if result.report.copy_paste_cleaned_count > 0:
            msg += f" (Sanitized {result.report.copy_paste_cleaned_count} PDF copy-paste formatting anomalies)."
        push_alert(msg, "green")
        return RedirectResponse("/dashboard", status_code=303)
    except Exception as e:
        push_alert(f"Processing error: {e}", "red")
        return RedirectResponse("/upload", status_code=303)


@app.post("/alias-override")
async def route_alias_override(request):
    try:
        form = await request.form()
        raw_token = form.get("raw_token", "").strip()
        canonical_target = form.get("canonical_target", "").strip()
        if raw_token and canonical_target:
            overrides = SESSION.get("custom_alias_overrides", {})
            overrides[raw_token.upper()] = canonical_target
            SESSION["custom_alias_overrides"] = overrides
            push_alert(f"Added custom alias override: '{raw_token}' → '{canonical_target}'. Re-analyzing spreadsheet...", "green")
            
            data = SESSION.get("preview_raw_bytes")
            filename = SESSION.get("preview_filename", "result.xlsx")
            if data:
                full_res = validate_and_clean(data, filename, overrides)
                SESSION["preview_report"] = full_res.report
        return RedirectResponse("/upload-preview", status_code=303)
    except Exception as e:
        push_alert(f"Alias override error: {e}", "red")
        return RedirectResponse("/upload-preview", status_code=303)


@app.post("/reset")
def route_reset():
    SESSION["records"] = None
    SESSION["analytics"] = None
    SESSION["validation"] = None
    SESSION["preview_raw_bytes"] = None
    SESSION["ptm_briefs"] = {}
    SESSION["staff_directory"] = {}
    SESSION["staff_verified"] = False
    push_alert("Session cleared. Upload a new file to begin.", "blue")
    return RedirectResponse("/", status_code=303)


@app.get("/dashboard")
def route_dashboard():
    if not session_ready():
        push_alert("No data loaded. Please upload a semester result first.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_dashboard(SESSION["analytics"])


@app.get("/students")
def route_students():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_students(SESSION["analytics"])


@app.get("/students/search")
def route_students_search(q: str = ""):
    if not session_ready():
        return NotStr("")
    ca = SESSION["analytics"]
    q_lower = unquote(q).lower().strip()
    results = []
    for s in ca.students:
        if q_lower and q_lower not in s.regno.lower() and q_lower not in (s.name or "").lower():
            continue
        results.append(Tr(
            Td(str(s.rank or "—"), cls="font-medium text-slate-600"),
            Td(s.regno, cls="font-mono text-xs text-slate-500"),
            Td(A(s.name or "—", href=f"/students/{quote(s.regno, safe='')}",
                 cls="text-blue-600 hover:text-blue-700 font-medium")),
            Td(gpa_cell(s.gpa)),
            Td(str(s.arrear_count), cls="text-red-600 font-semibold") if s.arrear_count else Td("0"),
            Td(status_badge(s.attention)),
            Td(f"{s.percentile:.1f}%" if s.percentile else "—", cls="text-slate-500"),
            data_row_type="student-row",
            data_status=s.attention,
            data_high=str(s.is_high_performer).lower(),
        ))
    if not results:
        return NotStr(f"<tr><td colspan='7' class='text-center py-6 text-slate-400'>No student matching '{html.escape(q)}' found.</td></tr>")
    return NotStr("".join(str(r) for r in results))


@app.get("/students/{regno}")
def route_student_detail(regno: str):
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_student_detail(SESSION["analytics"], regno)


@app.post("/student/{regno}/ptm")
def route_ptm(regno: str):
    if not session_ready():
        return NotStr("<p class='text-red-600 text-sm'>No data loaded.</p>")
    ca = SESSION["analytics"]
    s = get_student(ca, regno)
    if s is None:
        return NotStr(f"<p class='text-red-600 text-sm'>Student {html.escape(regno)} not found.</p>")
    ptm = generate_ptm_brief(s, ca)
    SESSION.setdefault("ptm_briefs", {})[s.regno] = ptm
    return Div(
        P(ai_source_banner(ptm["live"]), cls="text-xs text-slate-400 italic mb-2"),
        NotStr(md_to_html(ptm["text"])),
        cls="prose prose-sm max-w-none"
    )


@app.get("/subjects")
def route_subjects():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_subjects(SESSION["analytics"])


@app.get("/subjects/{subject}")
def route_subject_detail(subject: str):
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_subject_detail(SESSION["analytics"], subject)


@app.get("/attention")
def route_attention():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_attention(SESSION["analytics"])


@app.get("/rankings")
def route_rankings():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_rankings(SESSION["analytics"])


@app.get("/ai-insights")
def route_ai_insights():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_ai_insights(SESSION["analytics"])


@app.get("/ai-insights/class-summary")
def route_class_ai_summary():
    if not session_ready():
        return Div()
    ca = SESSION["analytics"]
    ai = generate_class_ai_insight(ca)
    return Div(
        H3("AI Academic Advisory Insight", cls="text-sm font-semibold text-slate-700 mb-3"),
        md_block(ai["text"], ai["live"]),
        id="class-ai-advisory"
    )


@app.get("/reports")
def route_reports():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_reports(SESSION["analytics"])


@app.get("/reports/interactive")
def route_reports_interactive():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    return page_interactive_report(SESSION["analytics"])


@app.post("/report/class")
def route_report_class():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    rid = create_report_id("class", {"class_analysis": SESSION["analytics"]})
    return RedirectResponse(f"/download-pdf?report_id={rid}", status_code=303)


@app.post("/report/subject/{subject}")
def route_report_subject(subject: str):
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    rid = create_report_id("subject", {"class_analysis": SESSION["analytics"], "subject": subject})
    return RedirectResponse(f"/download-pdf?report_id={rid}", status_code=303)


@app.post("/report/subject-select")
async def route_report_subject_select(request):
    form = await request.form()
    subject = form.get("subject", "")
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    rid = create_report_id("subject", {"class_analysis": SESSION["analytics"], "subject": subject})
    return RedirectResponse(f"/download-pdf?report_id={rid}", status_code=303)


@app.post("/report/student/{regno}")
def route_report_student(regno: str):
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    rid = create_report_id("student", {"class_analysis": SESSION["analytics"], "regno": regno})
    return RedirectResponse(f"/download-pdf?report_id={rid}", status_code=303)


@app.post("/report/student-select")
async def route_report_student_select(request):
    form = await request.form()
    regno = form.get("regno", "").strip()
    if not regno:
        push_alert("Please enter a register number.", "amber")
        return RedirectResponse("/reports", status_code=303)
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    ca = SESSION["analytics"]
    s = get_student(ca, regno)
    if not s:
        push_alert(f"Student '{html.escape(regno)}' not found in the dataset.", "red")
        return RedirectResponse("/reports", status_code=303)
    rid = create_report_id("student", {"class_analysis": ca, "regno": s.regno})
    return RedirectResponse(f"/download-pdf?report_id={rid}", status_code=303)


@app.get("/download-class-excel")
def route_download_class_excel():
    if not session_ready():
        push_alert("No data loaded.", "amber")
        return RedirectResponse("/", status_code=303)
    ca = SESSION["analytics"]
    ok, issues = validate_export_dataset(ca)
    if not ok:
        push_alert("Export validation failed: " + "; ".join(issues), "red")
        return RedirectResponse("/dashboard", status_code=303)
    xlsx_bytes = build_class_analysis_excel(ca)
    fname = re.sub(r"[^\w\.\-]", "_", f"class_analysis_{ca.file_name or 'report'}.xlsx")
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/download-pdf")
def route_download_pdf(report_id: str = ""):
    if not report_id or not is_valid_uuid(report_id):
        push_alert("Invalid report link.", "red")
        return RedirectResponse("/reports", status_code=303)

    entry = get_report(report_id)
    if not entry:
        push_alert("Report expired or not found. Please regenerate.", "red")
        return RedirectResponse("/reports", status_code=303)

    pdf_bytes = generate_report_pdf(report_id)
    if not pdf_bytes:
        push_alert("PDF generation failed. Ensure report data is valid.", "red")
        return RedirectResponse("/reports", status_code=303)

    kind = entry["kind"]
    ca = entry["payload"].get("class_analysis")
    if kind == "class":
        fname = f"class_report_{ca.file_name if ca else 'report'}.pdf"
    elif kind == "subject":
        subj = entry["payload"].get("subject", "subject")
        fname = f"subject_report_{subj}.pdf"
    else:
        regno = entry["payload"].get("regno", "student")
        fname = f"student_report_{regno}.pdf"

    safe_fname = re.sub(r"[^\w\.\-]", "_", fname)

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_fname}"'},
    )


# =============================================================================
# 14) APPLICATION STARTUP
# =============================================================================

if __name__ == "__main__":
    import os
    import uvicorn

    port = int(os.environ.get("PORT", 8000))

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=port
    )