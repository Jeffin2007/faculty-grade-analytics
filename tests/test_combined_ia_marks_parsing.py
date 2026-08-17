import unittest
import os
import sys
import pandas as pd
sys.path.insert(0, os.path.abspath("."))
import app
from app import (
    parse_combined_ia_marks_content,
    parse_ia_marks_content,
    clean_staff_name,
    resolve_staff_for_code,
    build_department_excel,
    compute_class_analysis,
)

class TestCombinedIAMarksParsing(unittest.TestCase):
    def setUp(self):
        self.ia_a_path = "f:/mark_analysis/II A IA.html"
        self.ia_b_path = "f:/mark_analysis/II B.html"

    def test_clean_staff_name(self):
        self.assertEqual(clean_staff_name("gohila priyadharshini7130"), "Gohila Priyadharshini")
        self.assertEqual(clean_staff_name("suganyadevi-eee"), "Suganyadevi")
        self.assertEqual(clean_staff_name("subhashini-mat"), "Subhashini")
        self.assertEqual(clean_staff_name("murali-phy"), "Murali")
        self.assertEqual(clean_staff_name("thangam-tamil"), "Thangam")

    def test_resolve_staff_for_code(self):
        staff_map = {
            "24CS201A": "Gohila Priyadharshini",
            "24CS201B": "Mangalambigai",
            "24MA201": "Subhashini & Arunkumar"
        }
        self.assertEqual(resolve_staff_for_code("24CS201A", staff_map), "Gohila Priyadharshini")
        self.assertEqual(resolve_staff_for_code("24CS201", staff_map), "Gohila Priyadharshini & Mangalambigai")
        self.assertEqual(resolve_staff_for_code("CS201", staff_map), "Gohila Priyadharshini & Mangalambigai")
        self.assertEqual(resolve_staff_for_code("24MA201A", staff_map), "Subhashini & Arunkumar")

    def test_parse_section_a_html(self):
        if not os.path.exists(self.ia_a_path):
            self.skipTest("II A IA.html not found")
        with open(self.ia_a_path, "rb") as f:
            multi_marks, titles, staff, meta = parse_combined_ia_marks_content(f.read(), "II A IA.html")

        self.assertEqual(len(multi_marks["ia1"]), 63)
        self.assertEqual(len(multi_marks["ia2"]), 63)
        self.assertEqual(len(multi_marks["ia3"]), 63)

        # Verify staff names
        self.assertEqual(staff.get("24CS201A"), "Gohila Priyadharshini")
        self.assertEqual(staff.get("24EN201A"), "Vijayarenganayaki")
        self.assertEqual(staff.get("24ES201A"), "Suganyadevi")
        self.assertEqual(staff.get("24MA201A"), "Subhashini")
        self.assertEqual(staff.get("24PH201A"), "Murali")
        self.assertEqual(staff.get("24TA201A"), "Thangam")

        # Verify sample student marks (813825243001)
        s1 = multi_marks["ia1"].get("813825243001")
        self.assertIsNotNone(s1)
        self.assertEqual(s1["marks"].get("24CS201A"), "96")
        self.assertEqual(s1["marks"].get("24EN201A"), "68")
        self.assertEqual(s1["quota"], "GQ")

    def test_parse_section_b_html(self):
        if not os.path.exists(self.ia_b_path):
            self.skipTest("II B.html not found")
        with open(self.ia_b_path, "rb") as f:
            multi_marks, titles, staff, meta = parse_combined_ia_marks_content(f.read(), "II B.html")

        self.assertEqual(len(multi_marks["ia1"]), 41)
        self.assertEqual(len(multi_marks["ia2"]), 41)
        self.assertEqual(len(multi_marks["ia3"]), 41)

        # Verify Section B staff names
        self.assertEqual(staff.get("24CS201A"), "Mangalambigai")
        self.assertEqual(staff.get("24EN201A"), "Lakshmi")
        self.assertEqual(staff.get("24ES201A"), "Vigneshwaran")
        self.assertEqual(staff.get("24MA201A"), "Arunkumar")
        self.assertEqual(staff.get("24PH201A"), "Senthilkumar")

        # Verify sample student marks (813825243065)
        s65 = multi_marks["ia1"].get("813825243065")
        self.assertIsNotNone(s65)
        self.assertEqual(s65["marks"].get("24CS201A"), "92")
        self.assertEqual(s65["quota"], "GQ")

    def test_multi_section_merge_and_excel_generation(self):
        if not os.path.exists(self.ia_a_path) or not os.path.exists(self.ia_b_path):
            self.skipTest("Section HTML files not found")

        ia_store = {"ia1": {}, "ia2": {}, "ia3": {}}
        staff_directory = {}

        for p in [self.ia_a_path, self.ia_b_path]:
            with open(p, "rb") as f:
                multi_marks, titles, staff, meta = parse_combined_ia_marks_content(f.read(), os.path.basename(p))
                for tk, tdata in multi_marks.items():
                    ia_store.setdefault(tk, {}).update(tdata)
                for code, sname in staff.items():
                    if code not in staff_directory:
                        staff_directory[code] = sname
                    else:
                        existing = [x.strip() for x in staff_directory[code].split("&")]
                        if sname not in existing:
                            staff_directory[code] = f"{staff_directory[code]} & {sname}"

        # Total students merged: 63 + 41 = 104
        self.assertEqual(len(ia_store["ia1"]), 104)
        self.assertEqual(staff_directory.get("24CS201A"), "Gohila Priyadharshini & Mangalambigai")

        # Build mock class analysis with failed student to verify Analysis 4
        records = [
            {"regno": "813825243001", "name": "ABINASH. S", "subject": "Data Structures", "course_code": "24CS201A", "credits": 4.0, "grade": "RA", "points": 0.0},
            {"regno": "813825243001", "name": "ABINASH. S", "subject": "English", "course_code": "24EN201A", "credits": 3.0, "grade": "A", "points": 8.0},
        ]
        df = pd.DataFrame(records)
        ca = compute_class_analysis(df, "mock.pdf")
        ca.subject_mappings = [
            {"course_code": "24CS201A", "official_subject_name": "Data Structures", "credits": 4.0},
            {"course_code": "24EN201A", "official_subject_name": "English", "credits": 3.0},
        ]

        from app import PDFExtractionReport, DocumentMetadata
        rep = PDFExtractionReport(doc_metadata=DocumentMetadata(
            institution="Saranathan College of Engineering",
            department="Artificial Intelligence and Data Science",
            programme="B.Tech. AI & DS",
            semester="II",
            academic_year="2025-2026",
            exam_session="April/May 2026"
        ))

        xlsx_bytes = build_department_excel(ca, rep, staff_directory, "mock.pdf", ia_marks_dir=ia_store)
        self.assertGreater(len(xlsx_bytes), 1000)

        # Inspect generated openpyxl sheet Analysis 4_New
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        self.assertIn("Analysis 4_New", wb.sheetnames)
        ws4 = wb["Analysis 4_New"]
        
        # Check student row contains IAT 1 mark 96, blank MUT 1, IAT 2 mark 86, blank MUT 2, IAT 3 mark 94 for 24CS201A
        found = False
        for r in range(1, ws4.max_row + 1):
            if ws4.cell(row=r, column=6).value == "813825243001":
                found = True
                self.assertEqual(ws4.cell(row=r, column=8).value, "GQ") # Quota
                self.assertEqual(ws4.cell(row=r, column=9).value, "96") # IAT 1
                self.assertIn(ws4.cell(row=r, column=10).value, ("", None)) # MUT 1 is blank
                self.assertEqual(ws4.cell(row=r, column=11).value, "86") # IAT 2
                self.assertIn(ws4.cell(row=r, column=12).value, ("", None)) # MUT 2 is blank
                self.assertEqual(ws4.cell(row=r, column=13).value, "94") # IAT 3
                break
        self.assertTrue(found, "Student 813825243001 should be present in Analysis 4_New")

    def test_base_code_matching_in_excel(self):
        if not os.path.exists(self.ia_a_path) or not os.path.exists(self.ia_b_path):
            self.skipTest("Section HTML files not found")

        ia_store = {"ia1": {}, "ia2": {}, "ia3": {}}
        staff_directory = {}

        for p in [self.ia_a_path, self.ia_b_path]:
            with open(p, "rb") as f:
                multi_marks, titles, staff, meta = parse_combined_ia_marks_content(f.read(), os.path.basename(p))
                for tk, tdata in multi_marks.items():
                    ia_store.setdefault(tk, {}).update(tdata)
                for code, sname in staff.items():
                    if code not in staff_directory:
                        staff_directory[code] = sname
                    else:
                        existing = [x.strip() for x in staff_directory[code].split("&")]
                        if sname not in existing:
                            staff_directory[code] = f"{staff_directory[code]} & {sname}"

        # Test case: Official COE PDF has base course code '24CS201' (not '24CS201A')
        records = [
            {"regno": "813825243001", "name": "ABINASH. S", "subject": "Data Structures", "course_code": "24CS201", "credits": 4.0, "grade": "RA", "points": 0.0},
        ]
        df = pd.DataFrame(records)
        ca = compute_class_analysis(df, "mock.pdf")
        ca.subject_mappings = [
            {"course_code": "24CS201", "official_subject_name": "Data Structures", "credits": 4.0},
        ]

        from app import PDFExtractionReport, DocumentMetadata
        rep = PDFExtractionReport(doc_metadata=DocumentMetadata(
            institution="Saranathan College of Engineering",
            department="Artificial Intelligence and Data Science",
            programme="B.Tech. AI & DS",
            semester="II",
            academic_year="2025-2026",
            exam_session="April/May 2026"
        ))

        xlsx_bytes = build_department_excel(ca, rep, staff_directory, "mock.pdf", ia_marks_dir=ia_store)
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        
        # Verify Analysis 1 has staff name for 24CS201
        ws1 = wb["Analysis 1_New"]
        a1_staff_found = any("Gohila Priyadharshini" in str(ws1.cell(row=r, column=3).value or "") for r in range(1, ws1.max_row + 1))
        self.assertTrue(a1_staff_found, "Analysis 1 should contain resolved faculty name for 24CS201")

        # Verify Analysis 4 has staff name and marks for 24CS201
        ws4 = wb["Analysis 4_New"]
        found = False
        for r in range(1, ws4.max_row + 1):
            if ws4.cell(row=r, column=6).value == "813825243001":
                found = True
                self.assertEqual(ws4.cell(row=r, column=2).value, "Gohila Priyadharshini & Mangalambigai")
                self.assertEqual(ws4.cell(row=r, column=9).value, "96")
                self.assertIn(ws4.cell(row=r, column=10).value, ("", None))
                self.assertIn(ws4.cell(row=r, column=12).value, ("", None))
                break
        self.assertTrue(found, "Student 813825243001 should be present in Analysis 4_New for 24CS201")

    def test_semester_code_and_department_excel_filename(self):
        from app import _semester_short_code, department_excel_filename, PDFExtractionReport, DocumentMetadata
        self.assertEqual(_semester_short_code("II"), "SEM_II")
        self.assertEqual(_semester_short_code("Semester II"), "SEM_II")
        self.assertEqual(_semester_short_code("Semester 2"), "SEM_II")
        self.assertEqual(_semester_short_code("Semester I"), "SEM_I")
        self.assertEqual(_semester_short_code("1"), "SEM_I")
        self.assertEqual(_semester_short_code("Semester IV"), "SEM_IV")
        self.assertEqual(_semester_short_code("Semester 4"), "SEM_IV")

        rep = PDFExtractionReport(doc_metadata=DocumentMetadata(
            department="Artificial Intelligence and Data Science",
            programme="B.Tech. AI & DS",
            semester="Semester II",
            academic_year="2025-2026",
            exam_session="Nov / Dec 2025"
        ))
        fname = department_excel_filename(rep)
        self.assertEqual(fname, "AI_DS_NOV_DEC_2025_SEM_II_Result_Analysis.xlsx")

        rep2 = PDFExtractionReport(doc_metadata=DocumentMetadata(
            department="Computer Science and Engineering",
            programme="B.E. Computer Science and Engineering",
            semester="Semester I",
            academic_year="2025-2026",
            exam_session="Nov / Dec 2025"
        ))
        self.assertEqual(department_excel_filename(rep2), "CSE_NOV_DEC_2025_SEM_I_Result_Analysis.xlsx")

    def test_quota_fallback_resolution_without_na(self):
        # Build mock analysis with 2 failed students: 1 in IA store, 1 not in IA store
        records = [
            {"regno": "813825243001", "name": "ABINASH. S", "subject": "Data Structures", "course_code": "24CS201A", "credits": 4.0, "grade": "RA", "points": 0.0},
            {"regno": "813825243064", "name": "NIKIL BHARATHI R", "subject": "Data Structures", "course_code": "24CS201A", "credits": 4.0, "grade": "RA", "points": 0.0},
        ]
        df = pd.DataFrame(records)
        ca = compute_class_analysis(df, "mock.pdf")
        ca.subject_mappings = [
            {"course_code": "24CS201A", "official_subject_name": "Data Structures", "credits": 4.0},
        ]

        from app import PDFExtractionReport, DocumentMetadata
        rep = PDFExtractionReport(doc_metadata=DocumentMetadata(
            institution="Saranathan College of Engineering",
            department="Artificial Intelligence and Data Science",
            programme="B.Tech. AI & DS",
            semester="II",
            academic_year="2025-2026",
            exam_session="Nov / Dec 2025"
        ))

        # IA store only has 813825243001 (MQ), 813825243064 is missing
        ia_store = {
            "ia1": {"813825243001": {"quota": "MQ", "marks": {"24CS201A": "88"}}},
            "ia2": {},
            "ia3": {}
        }
        staff_directory = {"24CS201A": "Staff Member"}

        xlsx_bytes = build_department_excel(ca, rep, staff_directory, "mock.pdf", ia_marks_dir=ia_store)
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

        # Check Analysis 3_New
        ws3 = wb["Analysis 3_New"]
        q_3001 = None
        q_3064 = None
        for r in range(1, ws3.max_row + 1):
            if ws3.cell(row=r, column=2).value == "813825243001":
                q_3001 = ws3.cell(row=r, column=4).value
            elif ws3.cell(row=r, column=2).value == "813825243064":
                q_3064 = ws3.cell(row=r, column=4).value

        self.assertEqual(q_3001, "MQ")
        self.assertEqual(q_3064, "GQ")  # Defaults to GQ instead of NA

        # Check Analysis 4_New
        ws4 = wb["Analysis 4_New"]
        q4_3001 = None
        q4_3064 = None
        for r in range(1, ws4.max_row + 1):
            if ws4.cell(row=r, column=6).value == "813825243001":
                q4_3001 = ws4.cell(row=r, column=8).value
            elif ws4.cell(row=r, column=6).value == "813825243064":
                q4_3064 = ws4.cell(row=r, column=8).value

        self.assertEqual(q4_3001, "MQ")
        self.assertEqual(q4_3064, "GQ")  # Defaults to GQ instead of NA

    def test_universal_course_codes_and_separate_a5_model(self):
        # Sample HTML representing another department (e.g. ECE / 4-digit codes) with A1, A2, A3, A5
        html_content = """
        <html><body>
        <table><tr><td>Classwise Mark Summary</td></tr></table>
        <table>
          <tr>
            <th rowspan="2">S.no</th><th rowspan="2">Regno</th><th rowspan="2">Batchno</th><th rowspan="2">Name</th><th rowspan="2">Quota</th><th rowspan="2">ArrearCount</th><th rowspan="2">CGPA</th>
            <th colspan="4">EC3351-karthik-ece</th>
            <th colspan="4">MA3151-saravanan-mat</th>
          </tr>
          <tr>
            <th>A1</th><th>A2</th><th>A3</th><th>A5</th>
            <th>A1</th><th>A2</th><th>A3</th><th>A5</th>
          </tr>
          <tr>
            <td>1</td><td>813821106001</td><td>21001</td><td>ANITHA R</td><td>MQ</td><td>0</td><td>8.5</td>
            <td>85</td><td>78</td><td>90</td><td>95</td>
            <td>70</td><td>65</td><td>80</td><td></td>
          </tr>
        </table>
        </body></html>
        """
        multi_marks, titles, staff, meta = parse_combined_ia_marks_content(html_content.encode("utf-8"), "ECE_Marks.html")
        self.assertIn("ia1", multi_marks)
        self.assertIn("ia2", multi_marks)
        self.assertIn("ia3", multi_marks)
        self.assertIn("ia4", multi_marks)

        # Verify A3 (90) is NOT overwritten by A5 (95)
        self.assertEqual(multi_marks["ia3"]["813821106001"]["marks"]["EC3351"], "90")
        self.assertEqual(multi_marks["ia4"]["813821106001"]["marks"]["EC3351"], "95")
        self.assertEqual(multi_marks["ia1"]["813821106001"]["marks"]["MA3151"], "70")

        # Verify staff resolution
        self.assertEqual(staff.get("EC3351"), "Karthik")
        self.assertEqual(staff.get("MA3151"), "Saravanan")

        # Verify quota
        self.assertEqual(meta["813821106001"]["quota"], "MQ")


if __name__ == "__main__":
    unittest.main()
