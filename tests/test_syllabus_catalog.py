"""
tests/test_syllabus_catalog.py
================================================================================
Comprehensive Test Suite for Dynamic Syllabus Catalog Architecture
================================================================================
Tests:
1. Catalog Loading & Manifest (AI_DS R2024: 56 courses, SHA-256, versioning, status)
2. Normalization & O(1) Indexing (hyphens, spaces, uppercase, no reordering)
3. Course Code & Alias Resolution (priority: overrides -> code -> alias -> name -> fuzzy)
4. Context-Aware Disambiguation (Semester & Department context)
5. Catalog Invalidation & Hot Reloading
6. Validation Gate (Duplicate code detection, malformed values)
7. Syllabus PDF Extraction & Draft Generation
8. Draft Approval, Publishing, Version Archiving & Rollback
9. Regression Testing on PDF Extraction, Excel Generation, and Metadata Provenance
"""

import io
import json
import os
import shutil
import tempfile
import unittest

from syllabus.loader import (
    get_catalog,
    get_registered_departments,
    load_registry,
    normalize_course_code,
    resolve_course,
    invalidate_catalog_cache,
    CatalogIndex,
)
from syllabus_tools.validator import validate_catalog_schema, CatalogValidationError
from syllabus_tools.extractor import extract_syllabus_pdf
from syllabus_tools.publisher import publish_syllabus_draft, rollback_catalog
import app


class TestSyllabusCatalogLoader(unittest.TestCase):

    def setUp(self):
        invalidate_catalog_cache()

    def test_registry_loading(self):
        registry = load_registry()
        self.assertIn("departments", registry)
        depts = get_registered_departments()
        self.assertTrue(len(depts) >= 1)
        ai_ds = next((d for d in depts if d["code"] == "AI_DS"), None)
        self.assertIsNotNone(ai_ds)
        self.assertTrue(ai_ds["catalog_available"])
        self.assertIn("R2024", ai_ds["available_regulations"])

    def test_ai_ds_r2024_catalog_loading_and_manifest(self):
        cat = get_catalog("AI_DS", "R2024")
        self.assertIsNotNone(cat)
        self.assertEqual(cat.metadata.department_code, "AI_DS")
        self.assertEqual(cat.metadata.regulation, "R2024")
        self.assertEqual(cat.metadata.course_count, 135)
        self.assertEqual(len(cat.courses), 135)
        self.assertEqual(len(cat.metadata.sha256_hash), 64)
        self.assertEqual(cat.metadata.status, "ACTIVE")

    def test_course_code_normalization(self):
        # Must strip non-alphanumeric and uppercase without reordering
        self.assertEqual(normalize_course_code("24-AD-401"), "24AD401")
        self.assertEqual(normalize_course_code("24 AD 401"), "24AD401")
        self.assertEqual(normalize_course_code("24ad401"), "24AD401")
        self.assertEqual(normalize_course_code(" 24_ad_401 "), "24AD401")
        self.assertNotEqual(normalize_course_code("24AD401"), "AD24401")

    def test_exact_course_code_resolution(self):
        cat = get_catalog("AI_DS", "R2024")
        name, code, credits_val, sem, cat_name, conf, amb = resolve_course("24AD401", catalog=cat)
        self.assertEqual(code, "24AD401")
        self.assertEqual(name, "Machine Learning")
        self.assertEqual(credits_val, 3.0)
        self.assertEqual(sem, 4)
        self.assertEqual(conf, 1.0)
        self.assertFalse(amb)

        # Test hyphenated raw input
        name2, code2, cred2, sem2, _, conf2, amb2 = resolve_course("24-AD-404", catalog=cat)
        self.assertEqual(code2, "24AD404")
        self.assertEqual(name2, "Principles to Operating System")
        self.assertEqual(cred2, 4.0)
        self.assertEqual(conf2, 1.0)
        self.assertFalse(amb2)

    def test_alias_resolution(self):
        cat = get_catalog("AI_DS", "R2024")
        # Test short alias ML -> Machine Learning
        name, code, cred, sem, _, conf, amb = resolve_course("ML", catalog=cat)
        self.assertEqual(code, "24AD401")
        self.assertEqual(name, "Machine Learning")
        self.assertEqual(conf, 1.0)

        # Test DSA LAB -> 24AD311
        name_lab, code_lab, cred_lab, sem_lab, _, conf_lab, _ = resolve_course("DSA LAB", catalog=cat)
        self.assertEqual(code_lab, "24AD311")
        self.assertEqual(cred_lab, 1.5)
        self.assertEqual(sem_lab, 3)

    def test_context_aware_semester_disambiguation(self):
        cat = get_catalog("AI_DS", "R2024")
        # Professional English exists in Sem 1 (24EN101) and Sem 2 (24EN201)
        res_sem1 = resolve_course("Professional English", catalog=cat, context={"department": "AI_DS", "semester": 1, "regulation": "R2024"})
        self.assertEqual(res_sem1[1], "24EN101")
        self.assertEqual(res_sem1[3], 1)

        res_sem2 = resolve_course("Professional English", catalog=cat, context={"department": "AI_DS", "semester": 2, "regulation": "R2024"})
        self.assertEqual(res_sem2[1], "24EN201")
        self.assertEqual(res_sem2[3], 2)

    def test_fuzzy_matching_and_unresolved(self):
        cat = get_catalog("AI_DS", "R2024")
        # Close typo should fuzzy match
        name, code, cred, sem, _, conf, _ = resolve_course("Machne Learnin", catalog=cat)
        self.assertEqual(code, "24AD401")
        self.assertTrue(conf >= 0.80)

        # Completely unknown subject
        name_u, code_u, cred_u, sem_u, _, conf_u, amb_u = resolve_course("XYZ_COMPLETELY_RANDOM_999", catalog=cat)
        self.assertTrue(amb_u)
        self.assertTrue(conf_u < 0.60)

    def test_cache_invalidation(self):
        cat1 = get_catalog("AI_DS", "R2024")
        self.assertIsNotNone(cat1)
        invalidate_catalog_cache("AI_DS", "R2024")
        cat2 = get_catalog("AI_DS", "R2024")
        self.assertIsNotNone(cat2)

    def test_get_expected_subjects_and_arrear_knowledge(self):
        from syllabus.loader import get_expected_subjects
        # For AI&DS Semester 3
        manifest = get_expected_subjects("AI_DS", "R2024", 3)
        self.assertTrue(manifest["ok"])
        self.assertEqual(manifest["department"], "AI_DS")
        self.assertEqual(manifest["semester"], 3)
        self.assertEqual(len(manifest["current_semester_courses"]), 10)
        self.assertEqual(len(manifest["arrear_courses"]), 19)
        self.assertEqual(manifest["arrear_semesters_loaded"], [1, 2])

        # Verify arrear course codes
        sem3_codes = [c["code"] for c in manifest["current_semester_courses"]]
        self.assertIn("24AD301", sem3_codes)
        self.assertIn("24MA301", sem3_codes)

        arrear_codes = [c["code"] for c in manifest["arrear_courses"]]
        self.assertIn("24MA101", arrear_codes)
        self.assertIn("24EN201", arrear_codes)


class TestSyllabusTools(unittest.TestCase):

    def setUp(self):
        test_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "syllabus", "test_eng")
        if os.path.isdir(test_dir):
            shutil.rmtree(test_dir)
        invalidate_catalog_cache()

    def tearDown(self):
        test_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "syllabus", "test_eng")
        if os.path.isdir(test_dir):
            shutil.rmtree(test_dir)
        reg_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "syllabus", "registry.json")
        if os.path.isfile(reg_file):
            with open(reg_file, "r", encoding="utf-8") as f:
                rdata = json.load(f)
            rdata["departments"] = [d for d in rdata.get("departments", []) if d.get("code") != "TEST_ENG"]
            with open(reg_file, "w", encoding="utf-8") as f:
                json.dump(rdata, f, indent=2, ensure_ascii=False)
        invalidate_catalog_cache()

    def test_validator_duplicate_detection(self):
        bad_data = {
            "department": "Test Dept",
            "department_code": "TEST",
            "regulation": "R2024",
            "courses": [
                {"code": "24CS101", "name": "Programming in C", "credits": 3.0, "semester": 1, "aliases": []},
                {"code": "24-CS-101", "name": "Another Course", "credits": 3.0, "semester": 1, "aliases": []},
            ]
        }
        is_valid, issues = validate_catalog_schema(bad_data)
        self.assertFalse(is_valid)
        self.assertTrue(any("Duplicate course code" in issue for issue in issues))

    def test_validator_invalid_credits_and_semester(self):
        bad_data = {
            "department": "Test Dept",
            "department_code": "TEST",
            "regulation": "R2024",
            "courses": [
                {"code": "24CS101", "name": "Programming in C", "credits": -5.0, "semester": 15, "aliases": []},
            ]
        }
        is_valid, issues = validate_catalog_schema(bad_data)
        self.assertFalse(is_valid)
        self.assertTrue(any("invalid credits" in issue for issue in issues))
        self.assertTrue(any("invalid semester" in issue for issue in issues))

    def test_pdf_extractor_and_draft_creation(self):
        sample_pdf = os.path.join(os.path.dirname(os.path.dirname(__file__)), "B.E. AIML - R2024 (09-06-2026).pdf")
        if os.path.isfile(sample_pdf):
            with tempfile.TemporaryDirectory() as tmpdir:
                res = extract_syllabus_pdf(sample_pdf, department_code_hint="AIML", regulation_hint="R2024", output_dir=tmpdir)
                self.assertIn("draft_id", res)
                self.assertIn("draft_data", res)
                self.assertTrue(res["draft_data"]["total_courses_extracted"] > 0)
                self.assertTrue(os.path.isfile(res["draft_file"]))

    def test_publishing_version_preservation_and_rollback(self):
        # Create a mock catalog draft
        mock_draft = {
            "draft_id": "test_mock_1",
            "catalog_version": "1.0",
            "department": "Test Engineering",
            "department_code": "TEST_ENG",
            "programme": "B.E. Test Engineering",
            "regulation": "R2024",
            "academic_structure": {"semesters": 8},
            "courses": [
                {
                    "code": "24TE101",
                    "name": "Fundamentals of Testing",
                    "credits": 3.0,
                    "semester": 1,
                    "type": "THEORY",
                    "category": "Sem 1-4 Foundation",
                    "aliases": ["TEST101", "FT"]
                }
            ]
        }

        # Publish v1
        pub1 = publish_syllabus_draft(mock_draft, actor="Test Faculty")
        self.assertEqual(pub1["status"], "PUBLISHED")
        self.assertEqual(pub1["catalog_version"], "1.0")
        self.assertEqual(len(pub1["catalog_hash"]), 64)

        # Update draft and publish v2
        mock_draft["courses"].append({
            "code": "24TE201",
            "name": "Advanced Testing",
            "credits": 4.0,
            "semester": 2,
            "type": "THEORY",
            "category": "Sem 1-4 Foundation",
            "aliases": ["TEST201"]
        })
        pub2 = publish_syllabus_draft(mock_draft, actor="Test Faculty")
        self.assertEqual(pub2["status"], "PUBLISHED")
        self.assertEqual(pub2["catalog_version"], "2.0")
        self.assertIsNotNone(pub2["archived_prior"])

        # Verify loading published catalog
        cat = get_catalog("TEST_ENG", "R2024")
        self.assertIsNotNone(cat)
        self.assertEqual(len(cat.courses), 2)
        self.assertEqual(cat.metadata.catalog_version, "2.0")

        # Test Rollback
        rollback_res = rollback_catalog("TEST_ENG", "R2024", pub2["archived_prior"])
        self.assertEqual(rollback_res["status"], "ROLLED_BACK")
        cat_rolled = get_catalog("TEST_ENG", "R2024")
        self.assertEqual(len(cat_rolled.courses), 1)


class TestPortalRegressions(unittest.TestCase):

    def test_app_resolve_subject_info(self):
        # Direct call to app.resolve_subject_info
        name, code, cred, sem, cat, conf, amb = app.resolve_subject_info("24AD401")
        self.assertEqual(code, "24AD401")
        self.assertEqual(name, "Machine Learning")
        self.assertEqual(cred, 3.0)
        self.assertEqual(sem, 4)
        self.assertEqual(conf, 1.0)
        self.assertFalse(amb)

    def test_build_subject_mapping_log(self):
        rec1 = app.StudentResultRecord(
            register_number="813824104001",
            student_name="ALICE",
            subject_code="24AD401",
            subject_name="Machine Learning",
            credits=3.0,
            result_status="A+",
            source_page=1
        )
        rec2 = app.StudentResultRecord(
            register_number="813824104002",
            student_name="BOB",
            subject_code="24AD401",
            subject_name="Machine Learning",
            credits=3.0,
            result_status="O",
            source_page=1
        )
        mappings = app.build_subject_mapping_log([rec1, rec2])
        self.assertEqual(len(mappings), 1)
        self.assertEqual(mappings[0]["course_code"], "24AD401")
        self.assertEqual(mappings[0]["official_subject_name"], "Machine Learning")
        self.assertEqual(mappings[0]["resolution_method"], "EXACT_COURSE_CODE")

    def test_class_analysis_and_gpa(self):
        records = [
            app.StudentResultRecord(
                register_number="813824104001",
                student_name="ALICE",
                subject_code="24AD401",
                subject_name="Machine Learning",
                credits=3.0,
                result_status="A+",
                source_page=1
            ),
            app.StudentResultRecord(
                register_number="813824104001",
                student_name="ALICE",
                subject_code="24AD402",
                subject_name="Data Science and Exploratory Data Analysis",
                credits=3.0,
                result_status="O",
                source_page=1
            ),
        ]
        df = app.pdf_records_to_dataframe(records)
        ca = app.compute_class_analysis(df, "test.pdf")
        self.assertEqual(ca.student_count, 1)
        self.assertEqual(ca.subject_count, 2)
        alice = ca.students[0]
        # (3*9 + 3*10) / 6 = 57 / 6 = 9.50
        self.assertEqual(alice.gpa, 9.5)
        self.assertEqual(alice.arrear_count, 0)

    def test_elective_choice_split_and_arrear_prevention(self):
        import pandas as pd
        records_data = [
            # Alice takes Choice A (24AD901), does not take Choice B
            {"regno": "813824104001", "name": "Alice", "course_code": "24AD501", "subject": "Deep Learning", "credits": 3.0, "grade": "A", "points": 8.0},
            {"regno": "813824104001", "name": "Alice", "course_code": "24AD901", "subject": "Natural Language Processing", "credits": 3.0, "grade": "A+", "points": 9.0},
            # Bob takes Choice B (24AD902), does not take Choice A
            {"regno": "813824104002", "name": "Bob", "course_code": "24AD501", "subject": "Deep Learning", "credits": 3.0, "grade": "A", "points": 8.0},
            {"regno": "813824104002", "name": "Bob", "course_code": "24AD902", "subject": "Computer Vision", "credits": 3.0, "grade": "O", "points": 10.0},
        ]
        df = pd.DataFrame(records_data)
        ca = app.compute_class_analysis(df, "elective_test.pdf")

        # Alice checks
        alice = ca.students[0]
        alice_courses = [c.course_code for c in alice.courses]
        self.assertIn("24AD901", alice_courses)
        self.assertNotIn("24AD902", alice_courses)
        self.assertEqual(alice.arrear_count, 0)
        self.assertEqual(alice.attention, app.STATUS_CLEARED)

        # Bob checks
        bob = ca.students[1]
        bob_courses = [c.course_code for c in bob.courses]
        self.assertIn("24AD902", bob_courses)
        self.assertNotIn("24AD901", bob_courses)
        self.assertEqual(bob.arrear_count, 0)
        self.assertEqual(bob.attention, app.STATUS_CLEARED)

        # Subject-level checks: Choice A student count must be 1, not 2
        subj_a = next(s for s in ca.subjects if s.course_code == "24AD901")
        self.assertEqual(subj_a.student_count, 1)
        self.assertEqual(subj_a.pass_count, 1)
        self.assertEqual(subj_a.pass_pct, 100.0)

    def test_r24_r26_code_equivalence_and_analysis_2(self):
        import pandas as pd
        import openpyxl
        import io

        # 1. Test resolution equivalence
        m1, code1, cred1, sem1, cat1, conf1, amb1 = resolve_course("24CH401", context={"department": "CIVIL", "regulation": "R2024", "semester": 4})
        m2, code2, cred2, sem2, cat2, conf2, amb2 = resolve_course("24CH401A", context={"department": "CIVIL", "regulation": "R2026", "semester": 4})
        self.assertEqual(m1, m2)
        self.assertEqual(code1, "24CH401")
        self.assertEqual(code2, "24CH401")
        self.assertEqual(conf1, 1.0)
        self.assertEqual(conf2, 1.0)

        # 2. Test Analysis 2_New match
        prev_df = pd.DataFrame([
            {"regno": "813824103001", "name": "Old_1", "course_code": "24CH401", "subject": "Environmental Science and Engineering", "credits": 3.0, "grade": "A", "points": 8.0},
            {"regno": "813824103002", "name": "Old_2", "course_code": "24CH401", "subject": "Environmental Science and Engineering", "credits": 3.0, "grade": "U", "points": 0.0},
        ])
        prev_ca = app.compute_class_analysis(prev_df, "prev.pdf")

        curr_df = pd.DataFrame([
            {"regno": "813826103001", "name": "Curr_1", "course_code": "24CH401A", "subject": "Environmental Science and Engineering", "credits": 3.0, "grade": "A+", "points": 9.0},
        ])
        curr_ca = app.compute_class_analysis(curr_df, "curr.pdf")
        curr_ca.subject_mappings = [{"course_code": "24CH401A", "official_subject_name": "Environmental Science and Engineering", "credits": 3.0}]

        pdf_rep = app.PDFExtractionReport()
        pdf_rep.doc_metadata.department = "Department of Civil Engineering"
        pdf_rep.doc_metadata.programme = "B.E. Civil Engineering"
        pdf_rep.doc_metadata.semester = "Semester IV"

        excel_bytes = app.build_department_excel(
            ca=curr_ca,
            pdf_report=pdf_rep,
            staff_map={"24CH401A": "Dr. Smith"},
            prev_ca=prev_ca
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws2 = wb["Analysis 2_New"]

        found = False
        for r in range(1, ws2.max_row + 1):
            if ws2.cell(row=r, column=3).value == "24CH401A":
                self.assertEqual(ws2.cell(row=r, column=1).value, 50.0)
                self.assertIn("24CH401", str(ws2.cell(row=r, column=2).value))
                self.assertEqual(ws2.cell(row=r, column=6).value, 100.0)
                found = True
                break
        self.assertTrue(found)


if __name__ == "__main__":
    unittest.main()


